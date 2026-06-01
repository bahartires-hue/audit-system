#!/usr/bin/env python3
"""
إصلاح روابط صور ملف سلة Excel محلياً — بدون نشر على Render.

الاستخدام:
  python scripts/fix_salla_image_links.py ^
    --salla exports/salla_products_ready.xlsx ^
    --csv exports/tire_products.csv ^
    --out exports/salla_products_fixed.xlsx

يستبدل روابط Cloudinary .webp (أو الفارغة) بـ:
  1) source_image_url من CSV إن وُجد
  2) أو تحويل Cloudinary إلى f_jpg (قد لا يفيد إن الملف أصلاً placeholder)
"""
from __future__ import annotations

import argparse
import re
from pathlib import Path
from typing import Any, Dict, Optional

import pandas as pd
import requests
from openpyxl import load_workbook

_MIN_BYTES = 6000


def _norm_title(s: str) -> str:
    return re.sub(r"\s+", " ", str(s or "").strip().lower())


def _cloudinary_force_jpg(url: str) -> str:
    u = (url or "").strip()
    if "res.cloudinary.com" not in u or "/upload/" not in u:
        return u
    if "/f_jpg/" in u or "/f_auto/" in u:
        return u
    return u.replace("/upload/", "/upload/f_jpg/", 1)


def _url_has_image(url: str) -> bool:
    if not url.startswith("http"):
        return False
    try:
        r = requests.head(url, timeout=15, allow_redirects=True)
        if r.status_code >= 400:
            r = requests.get(url, timeout=20, stream=True, headers={"User-Agent": "Mozilla/5.0"})
        size = int(r.headers.get("Content-Length") or 0)
        if size and size < _MIN_BYTES:
            return False
        if not size:
            chunk = next(r.iter_content(8192), b"")
            return len(chunk) >= _MIN_BYTES
        return True
    except Exception:
        return False


def _load_source_map(csv_path: Path) -> Dict[str, str]:
    df = pd.read_csv(csv_path, encoding="utf-8-sig")
    out: Dict[str, str] = {}
    for _, row in df.iterrows():
        src = str(row.get("source_image_url") or row.get("image_url") or "").strip()
        if not src.startswith("http"):
            continue
        for key in (
            row.get("name"),
            row.get("product_title"),
            row.get("original_name"),
        ):
            k = _norm_title(str(key or ""))
            if k:
                out[k] = src
    return out


def _pick_image(
    current: str,
    title: str,
    source_map: Dict[str, str],
    *,
    check_size: bool,
) -> tuple[str, str]:
    """يعيد (رابط جديد, سبب التغيير)."""
    cur = (current or "").strip()
    key = _norm_title(title)

    def ok(u: str) -> bool:
        return bool(u) and (not check_size or _url_has_image(u))

    if cur and cur.lower().endswith(".webp"):
        alt = source_map.get(key, "")
        if ok(alt):
            return alt, "replaced_webp_with_csv_source"
        jpg = _cloudinary_force_jpg(cur)
        if ok(jpg):
            return jpg, "cloudinary_f_jpg"
        return "", "dropped_bad_webp"

    if cur and "res.cloudinary.com" in cur and not ok(cur):
        alt = source_map.get(key, "")
        if ok(alt):
            return alt, "replaced_small_cloudinary"
        return "", "dropped_small_cloudinary"

    if not cur:
        alt = source_map.get(key, "")
        if ok(alt):
            return alt, "filled_from_csv_source"
        return "", "still_empty"

    return cur, "unchanged"


def fix_salla_xlsx(
    salla_path: Path,
    out_path: Path,
    csv_path: Optional[Path],
    *,
    check_size: bool,
) -> Dict[str, int]:
    source_map = _load_source_map(csv_path) if csv_path and csv_path.exists() else {}

    wb = load_workbook(salla_path)
    ws = wb.active
    header_row = 1
    for r in range(1, min(20, ws.max_row) + 1):
        c1 = str(ws.cell(row=r, column=1).value or "").strip()
        c2 = str(ws.cell(row=r, column=2).value or "").strip()
        if c1 in {"النوع", "النوع "} and c2 == "أسم المنتج":
            header_row = r
            break

    col_title = col_image = None
    col_promo = None
    for c in range(1, ws.max_column + 1):
        h = str(ws.cell(row=header_row, column=c).value or "").strip()
        if h in {"أسم المنتج", "اسم المنتج"}:
            col_title = c
        if h == "صورة المنتج":
            col_image = c
        if h == "العنوان الترويجي":
            col_promo = c

    if not col_title or not col_image:
        raise SystemExit("لم أجد أعمدة أسم المنتج أو صورة المنتج في ملف سلة.")

    stats: Dict[str, int] = {}
    for r in range(header_row + 1, ws.max_row + 1):
        title = str(ws.cell(row=r, column=col_title).value or "").strip()
        if not title:
            continue
        cur_img = str(ws.cell(row=r, column=col_image).value or "").strip()
        new_img, reason = _pick_image(cur_img, title, source_map, check_size=check_size)
        if new_img != cur_img:
            ws.cell(row=r, column=col_image).value = new_img or None
            if col_promo and not new_img:
                promo = str(ws.cell(row=r, column=col_promo).value or "").strip()
                if "needs_image" not in promo:
                    ws.cell(row=r, column=col_promo).value = (
                        f"{promo} - needs_image".strip(" -") if promo else "needs_image"
                    )
        stats[reason] = stats.get(reason, 0) + 1

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
    return stats


def main() -> None:
    p = argparse.ArgumentParser(description="إصلاح روابط صور ملف سلة محلياً")
    p.add_argument("--salla", required=True, help="مسار salla_products_ready.xlsx")
    p.add_argument("--csv", default="", help="مسار tire_products.csv (اختياري لكن مُفضّل)")
    p.add_argument("--out", default="", help="ملف الإخراج (افتراضي: *_fixed.xlsx)")
    p.add_argument(
        "--check-size",
        action="store_true",
        help="فحص حجم الصورة عبر HTTP (أبطأ لكن أدق)",
    )
    args = p.parse_args()

    salla_path = Path(args.salla).resolve()
    csv_path = Path(args.csv).resolve() if args.csv else None
    out_path = (
        Path(args.out).resolve()
        if args.out
        else salla_path.with_name(salla_path.stem + "_fixed.xlsx")
    )

    stats = fix_salla_xlsx(salla_path, out_path, csv_path, check_size=args.check_size)
    print(f"تم الحفظ: {out_path}")
    print("ملخص التغييرات:")
    for k, v in sorted(stats.items(), key=lambda x: -x[1]):
        print(f"  {k}: {v}")


if __name__ == "__main__":
    main()
