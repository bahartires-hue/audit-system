from __future__ import annotations
import uuid
import datetime as dt
from fastapi.testclient import TestClient
from openpyxl import load_workbook
import io
from app.auth_core import hash_password
from app.db import SessionLocal
from app.main import app
from app.models import User

CSRF_TOKEN = "test-csrf-token-payroll"


def _csrf_headers():
    return {"X-CSRF-Token": CSRF_TOKEN}


def _login():
    uname = f"hrpayroll_{uuid.uuid4().hex[:8]}"
    db = SessionLocal()
    try:
        db.add(User(id=uuid.uuid4().hex, username=uname, email=f"{uname}@example.com",
                     password_hash=hash_password("RightPass123"), is_active=1, plan_name="free", role_name="admin", is_admin=1))
        db.commit()
    finally:
        db.close()
    client = TestClient(app, base_url="https://testserver")
    client.cookies.set("auditflow_csrf", CSRF_TOKEN)
    res = client.post("/auth/login", json={"username": uname, "password": "RightPass123"}, headers=_csrf_headers())
    assert res.status_code == 200, res.text
    return client


def test_payroll_recalculate_and_payslip_exports():
    c = _login()
    h = {"X-CSRF-Token": c.cookies.get("auditflow_csrf", domain="testserver.local")}

    r = c.post("/api/hr/employees", json={
        "name": "منى القحطاني", "nationality": "سعودية", "national_id": "3033445566",
        "position": "مطورة", "department": "التقنية", "branch_name": "الرئيسي",
        "basic_salary": 9000, "hire_date": "2022-01-01",
    }, headers=h)
    assert r.status_code == 200, r.text
    emp_id = r.json()["id"]

    pr = c.post("/api/hr/payrolls", json={
        "employee_id": emp_id, "month": 5, "year": 2026,
        "extra_allowances": [{"label": "مكافأة أداء", "amount": 500}],
        "extra_deductions": [{"label": "غرامة تأخير", "amount": 100}],
        "notes": "راتب تجريبي",
    }, headers=h)
    assert pr.status_code == 200, pr.text
    payroll = pr.json()
    assert payroll["net_salary"] == 9000 + 500 - 100
    assert any(a["label"] == "مكافأة أداء" for a in payroll["allowances"])
    assert any(d["label"] == "غرامة تأخير" for d in payroll["deductions"])

    dd = c.post("/api/hr/deductions", json={
        "employee_id": emp_id, "deduction_type": "جزاء", "amount": 250, "reason": "مخالفة", "date": "2026-05-10",
    }, headers=h)
    assert dd.status_code == 200, dd.text

    recalc = c.post(f"/api/hr/payrolls/{payroll['id']}/recalculate", headers=h)
    assert recalc.status_code == 200, recalc.text
    recalced = recalc.json()
    assert recalced["id"] != payroll["id"]
    assert any(a["label"] == "مكافأة أداء" and a["amount"] == 500 for a in recalced["allowances"])
    assert any(d["label"] == "غرامة تأخير" and d["amount"] == 100 for d in recalced["deductions"])
    assert any(d["label"] == "جزاء" and d["amount"] == 250 for d in recalced["deductions"])
    assert recalced["net_salary"] == 9000 + 500 - 100 - 250
    assert recalced["notes"] == "راتب تجريبي"

    old_get = c.get(f"/api/hr/payrolls/{payroll['id']}", headers=h)
    assert old_get.status_code == 404

    pdf = c.get(f"/api/hr/payrolls/{recalced['id']}/payslip/pdf", headers=h)
    assert pdf.status_code == 200
    assert pdf.headers["content-type"].startswith("application/pdf")
    assert len(pdf.content) > 500

    xl = c.get(f"/api/hr/payrolls/{recalced['id']}/payslip/excel", headers=h)
    assert xl.status_code == 200
    wb = load_workbook(io.BytesIO(xl.content))
    ws = wb.active
    values = [row[1] for row in ws.iter_rows(min_row=2, values_only=True)]
    labels = [row[0] for row in ws.iter_rows(min_row=2, values_only=True)]
    assert "منى القحطاني" in values
    assert any("مكافأة أداء" in (lbl or "") for lbl in labels)
    assert any("جزاء" in (lbl or "") for lbl in labels)

    pay = c.post(f"/api/hr/payrolls/{recalced['id']}/pay", headers=h)
    assert pay.status_code == 200, pay.text
    recalc_after_pay = c.post(f"/api/hr/payrolls/{recalced['id']}/recalculate", headers=h)
    assert recalc_after_pay.status_code == 400
