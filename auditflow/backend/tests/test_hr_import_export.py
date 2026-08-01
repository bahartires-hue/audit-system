from __future__ import annotations
import io
import uuid
import datetime as dt
from fastapi.testclient import TestClient
from openpyxl import Workbook, load_workbook
from app.auth_core import hash_password
from app.db import SessionLocal
from app.main import app
from app.models import User

CSRF_TOKEN = "test-csrf-token-import"


def _csrf_headers():
    return {"X-CSRF-Token": CSRF_TOKEN}


def _login():
    uname = f"hrimport_{uuid.uuid4().hex[:8]}"
    db = SessionLocal()
    try:
        db.add(User(id=uuid.uuid4().hex, username=uname, email=f"{uname}@example.com",
                     password_hash=hash_password("RightPass123"), is_active=1, plan_name="free", role_name="admin", is_admin=1))
        db.commit()
    finally:
        db.close()
    client = TestClient(app, base_url="https://testserver")
    client.cookies.set("auditflow_csrf", CSRF_TOKEN)
    res = client.post("/auth/login", json={"username": uname, "password": "RightPass123"}, headers=_csrf_headers())
    assert res.status_code == 200, res.text
    return client


EMPLOYEE_IMPORT_HEADERS = [
    "الاسم", "الرقم الوظيفي", "الفرع", "القسم", "الوظيفة", "الجنسية", "الجنس",
    "تاريخ الميلاد", "الجوال", "البريد الإلكتروني", "العنوان", "رقم الهوية",
    "تاريخ انتهاء الهوية/الإقامة", "رقم الجواز", "تاريخ انتهاء الجواز",
    "رقم رخصة العمل", "تاريخ انتهاء رخصة العمل", "رقم التأمين الطبي",
    "تاريخ انتهاء التأمين الطبي", "تاريخ التعيين", "نوع العقد", "بداية العقد",
    "نهاية العقد", "الراتب الأساسي", "بدل السكن", "بدل النقل", "البدلات الأخرى",
    "نسبة العمولة", "الراتب البنكي", "اسم البنك", "رقم الآيبان", "المدير المباشر",
    "الحالة الوظيفية", "ملاحظات",
]


def _build_import_workbook():
    wb = Workbook()
    ws = wb.active
    ws.append(EMPLOYEE_IMPORT_HEADERS)
    ws.append([
        "سارة العتيبي", "EMP-9001", "فرع جدة", "المبيعات", "مندوب مبيعات", "سعودية", "أنثى",
        "1995-03-20", "0559876543", "sara@example.com", "جدة - حي الشاطئ", "2022334455",
        "2028-01-01", "B7654321", "2030-01-01",
        "RP-1122", "2027-06-01", "MI-3344",
        "2027-07-01", "2024-02-01", "دائم", "2024-02-01",
        "", 5500, 400, 200, 100,
        1.5, 6200, "بنك الأهلي", "SA1111111111111111111111", "MGR-001",
        "نشط", "ملاحظة اختبار",
    ])
    ws.append([
        "خالد المطيري", "MGR-001", "فرع جدة", "المبيعات", "مدير المبيعات", "سعودي", "ذكر",
        "1985-01-01", "0501112222", "khaled@example.com", "جدة", "1099887766",
        "2027-05-01", "C1122334", "2029-08-01",
        "", "", "",
        "", "2020-01-01", "دائم", "2020-01-01",
        "", 12000, 0, 0, 0,
        "", "", "", "", "",
        "نشط", "",
    ])
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def test_employee_import_writes_through_full_360_profile():
    c = _login()
    h = {"X-CSRF-Token": c.cookies.get("auditflow_csrf", domain="testserver.local")}

    buf = _build_import_workbook()
    r = c.post(
        "/api/hr/employees/import",
        files={"file": ("import.xlsx", buf, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        headers=h,
    )
    assert r.status_code == 200, r.text
    result = r.json()
    assert result["added"] == 2, result
    assert result["updated"] == 0, result
    assert result["errors"] == [], result
    assert result["warnings"] == [], result

    lst = c.get("/api/hr/employees", headers=h)
    assert lst.status_code == 200
    employees = lst.json() if isinstance(lst.json(), list) else lst.json().get("items", [])
    sara = next(e for e in employees if e["employee_number"] == "EMP-9001")
    khaled = next(e for e in employees if e["employee_number"] == "MGR-001")

    assert sara["gender"] == "female"
    assert sara["address"] == "جدة - حي الشاطئ"
    assert sara["bank_name"] == "بنك الأهلي"
    assert sara["iban"] == "SA1111111111111111111111"
    assert sara["bank_salary"] == 6200
    assert sara["national_id"] == "2022334455"
    assert sara["nationality"] == "سعودية"
    assert sara["branch_name"] == "فرع جدة"
    assert sara["department"] == "المبيعات"
    assert sara["position"] == "مندوب مبيعات"
    assert sara["basic_salary"] == 5500

    assert sara["manager_id"] == khaled["id"]
    sara_full = c.get(f"/api/hr/employees/{sara['id']}", headers=h)
    assert sara_full.status_code == 200
    assert sara_full.json()["manager_name"] == "خالد المطيري"

    assert sara["residency_expiry"] is not None and sara["residency_expiry"].startswith("2028-01-01")
    assert sara["passport_number"] == "B7654321"
    assert sara["passport_expiry"] is not None and sara["passport_expiry"].startswith("2030-01-01")

    docs = c.get(f"/api/hr/employees/{sara['id']}/documents", headers=h)
    assert docs.status_code == 200
    doc_list = docs.json() if isinstance(docs.json(), list) else docs.json().get("items", [])
    doc_types = {d["doc_type"]: d for d in doc_list}
    assert "رخصة عمل" in doc_types, doc_types
    assert doc_types["رخصة عمل"]["doc_number"] == "RP-1122"
    assert "تأمين طبي" in doc_types
    assert doc_types["تأمين طبي"]["doc_number"] == "MI-3344"

    contracts = c.get("/api/hr/contracts", params={"employee_id": sara["id"]}, headers=h)
    assert contracts.status_code == 200
    c_list = contracts.json() if isinstance(contracts.json(), list) else contracts.json().get("items", [])
    assert len(c_list) == 1, c_list
    assert c_list[0]["contract_type"] == "دائم"
    assert c_list[0]["housing_allowance"] == 400
    assert c_list[0]["transport_allowance"] == 200
    assert c_list[0]["other_allowances"] == 100

    summary = c.get(f"/api/hr/employees/{sara['id']}/profile-summary", headers=h)
    assert summary.status_code == 200
    qi = summary.json()["quick_info"]
    assert qi["name"] == "سارة العتيبي"
    assert qi["manager_name"] == "خالد المطيري"
    assert qi["contract_status"] == "ساري"

    buf2 = _build_import_workbook()
    r2 = c.post(
        "/api/hr/employees/import",
        files={"file": ("import2.xlsx", buf2, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        headers=h,
    )
    assert r2.status_code == 200, r2.text
    result2 = r2.json()
    assert result2["added"] == 0, result2
    assert result2["updated"] == 2, result2
    contracts_after = c.get("/api/hr/contracts", params={"employee_id": sara["id"]}, headers=h)
    c_list_after = contracts_after.json() if isinstance(contracts_after.json(), list) else contracts_after.json().get("items", [])
    assert len(c_list_after) == 1, c_list_after

    exp = c.get("/api/hr/employees/export", headers=h)
    assert exp.status_code == 200
    wb = load_workbook(io.BytesIO(exp.content))
    ws = wb.active
    header_row = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    assert header_row == EMPLOYEE_IMPORT_HEADERS, header_row
    rows_by_empnum = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        rows_by_empnum[row[1]] = row
    sara_row = rows_by_empnum["EMP-9001"]
    assert sara_row[0] == "سارة العتيبي"
    assert sara_row[6] == "أنثى"
    assert sara_row[15] == "RP-1122"
    assert sara_row[17] == "MI-3344"
    assert sara_row[20] == "دائم"
    assert sara_row[31] == "خالد المطيري"
    assert sara_row[32] == "نشط"


def test_import_rejects_bad_gender_value_with_clear_error():
    c = _login()
    h = {"X-CSRF-Token": c.cookies.get("auditflow_csrf", domain="testserver.local")}
    wb = Workbook()
    ws = wb.active
    ws.append(EMPLOYEE_IMPORT_HEADERS)
    row = ["موظف تجريبي"] + [""] * (len(EMPLOYEE_IMPORT_HEADERS) - 1)
    row[6] = "غير معروف"
    ws.append(row)
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    r = c.post(
        "/api/hr/employees/import",
        files={"file": ("bad.xlsx", buf, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        headers=h,
    )
    assert r.status_code == 200
    result = r.json()
    assert result["added"] == 0
    assert len(result["errors"]) == 1
    assert "الجنس" in result["errors"][0]["reason"]
