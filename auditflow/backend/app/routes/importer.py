from __future__ import annotations

import json
import re
import time
from pathlib import Path
from urllib.parse import urljoin, urlparse

import requests
from bs4 import BeautifulSoup
from fastapi import APIRouter, Form, HTTPException
from fastapi.responses import FileResponse, JSONResponse
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

router = APIRouter(prefix="/importer", tags=["Importer"])

EXPORT_DIR = Path("auditflow/exports")
EXPORT_DIR.mkdir(parents=True, exist_ok=True)

LATEST_PRODUCTS_JSON = EXPORT_DIR / "latest_scraped_products.json"
SALLA_XLSX_PATH = EXPORT_DIR / "salla_products_ready.xlsx"

HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) "
        "Chrome/124.0 Safari/537.36"
    ),
    "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,image/webp,*/*;q=0.8",
    "Accept-Language": "ar,en-US;q=0.9,en;q=0.8",
    "Connection": "keep-alive",
}

BAD_TEXT = [
    "وقود أقل",
    "راحة عالية",
    "إضافة إلى السلة",
    "اضف للسلة",
    "للإطار الواحد",
    "الضمان",
    "النقشة",
    "سنة الصنع",
]

BAD_LINK_PARTS = [
    "/cdn-cgi/",
    "email-protection",
    "add-to-cart",
    "product-category",
    "/cart",
    "/checkout",
    "/my-account",
    "mailto:",
    "tel:",
]

SALLA_HEADERS = [
    "النوع",
    "أسم المنتج",
    "تصنيف المنتج",
    "صورة المنتج",
    "وصف صورة المنتج",
    "نوع المنتج",
    "سعر المنتج",
    "الوصف",
    "هل يتطلب شحن؟",
    "رمز المنتج sku",
    "سعر التكلفة",
    "السعر المخفض",
    "تاريخ بداية التخفيض",
    "تاريخ نهاية التخفيض",
    "اقصي كمية لكل عميل",
    "إخفاء خيار تحديد الكمية",
    "اضافة صورة عند الطلب",
    "الوزن",
    "الباركود",
    "الكمية",
]


def fetch_html(url: str) -> tuple[str, str, int]:
    try:
        r = requests.get(url, headers=HEADERS, timeout=35, allow_redirects=True)
        print("FETCH URL =", url)
        print("FINAL URL =", r.url)
        print("STATUS =", r.status_code)
        print("HTML LENGTH =", len(r.text))
        print("HTML START =", r.text[:300])
        return r.text, str(r.url), r.status_code
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"فشل فتح الرابط: {e}")


def normalize_space(text: str) -> str:
    return re.sub(r"\s+", " ", str(text or "")).strip()


def clean_price(text: str) -> str:
    text = str(text or "")
    text = text.replace(",", "")
    nums = re.findall(r"\d+(?:\.\d+)?", text)
    valid = []
    for n in nums:
        try:
            v = float(n)
            if v >= 100:
                valid.append(v)
        except Exception:
            pass
    if not valid:
        return ""
    return str(int(valid[0])) if valid[0].is_integer() else str(valid[0])


def extract_all_prices(text: str) -> list[float]:
    text = str(text or "").replace(",", "")
    nums = re.findall(r"\d+(?:\.\d+)?", text)
    prices = []
    for n in nums:
        try:
            v = float(n)
            if 100 <= v <= 5000:
                prices.append(v)
        except Exception:
            pass
    return prices


def is_real_tire_name(name: str, brand: str = "") -> bool:
    name = normalize_space(name)
    if len(name) < 8:
        return False
    if any(x in name for x in BAD_TEXT):
        return False
    has_size = bool(re.search(r"\b\d{3}/\d{2}\s*R\d{2}\b", name, re.I)) or bool(
        re.search(r"\b\d{3}\s*R\d{2}\b", name, re.I)
    )
    if not has_size:
        return False
    if brand and brand.lower() not in name.lower():
        return False
    return True


def extract_size(name: str) -> str:
    m = re.search(r"(\d{3}/\d{2}\s*R\d{2})", name, re.I)
    if m:
        return m.group(1).replace(" ", "").upper()
    m = re.search(r"(\d{3}\s*R\d{2})", name, re.I)
    if m:
        return m.group(1).replace(" ", "").upper()
    return ""


def extract_load_speed(name: str) -> str:
    m = re.search(r"\b(\d{2,3}(?:/\d{2,3})?[A-Z])\b", name, re.I)
    return m.group(1).upper() if m else ""


def extract_year(text: str) -> str:
    m = re.search(r"سنة\s*الصنع\s*[:：]?\s*(20\d{2})", text)
    if m:
        return m.group(1)
    m = re.search(r"\b(20\d{2})\b", text)
    if m:
        return m.group(1)
    return ""


def extract_warranty(text: str) -> str:
    m = re.search(r"الضمان\s*[:：]?\s*([^\n\r|]+)", text)
    return normalize_space(m.group(1)) if m else ""


def extract_pattern(text: str) -> str:
    m = re.search(r"النقشة\s*[:：]?\s*([A-Za-z0-9\u0600-\u06FF\s\-]+)", text)
    return normalize_space(m.group(1)) if m else ""


def make_sku(name: str) -> str:
    sku = re.sub(r"[^A-Za-z0-9]+", "-", name).strip("-").upper()
    return sku[:60]


def find_image(card, page_url: str) -> str:
    img = card.select_one("img")
    if not img:
        return ""
    src = img.get("data-src") or img.get("data-lazy-src") or img.get("src") or ""
    if not src:
        srcset = img.get("srcset") or ""
        if srcset:
            src = srcset.split(",")[0].strip().split(" ")[0]
    return urljoin(page_url, src) if src else ""


def find_product_name_from_card(card, brand: str) -> str:
    candidates = []
    for selector in ["h1", "h2", "h3", ".woocommerce-loop-product__title", ".product-title", ".wd-entities-title", ".title", "a"]:
        for el in card.select(selector):
            txt = normalize_space(el.get_text(" ", strip=True))
            if txt:
                candidates.append(txt)
    for txt in candidates:
        if is_real_tire_name(txt, brand):
            return txt
    full_text = card.get_text("\n", strip=True)
    lines = [normalize_space(x) for x in full_text.splitlines() if normalize_space(x)]
    for line in lines:
        if is_real_tire_name(line, brand):
            return line
    return ""


def find_price_from_card(card) -> str:
    for selector in [".price .amount", ".amount", ".price", ".woocommerce-Price-amount", "[class*='price']"]:
        el = card.select_one(selector)
        if el:
            price = clean_price(el.get_text(" ", strip=True))
            if price:
                return price
    text = card.get_text(" ", strip=True)
    m = re.search(r"(\d{3,5})\s+(\d{3,5})\s+للإطار الواحد", text)
    if m:
        return m.group(1)
    prices = extract_all_prices(text)
    if prices:
        return str(int(prices[0]))
    return ""


def card_to_product(card, page_url: str, brand: str, category: str) -> dict | None:
    text = card.get_text("\n", strip=True)
    name = find_product_name_from_card(card, brand)
    if not name:
        return None
    price = find_price_from_card(card)
    if not price:
        return None
    size = extract_size(name)
    load_speed = extract_load_speed(name)
    year = extract_year(text)
    warranty = extract_warranty(text)
    pattern = extract_pattern(text)
    image = find_image(card, page_url)
    description_parts = [
        name,
        f"المقاس: {size}" if size else "",
        f"الحمولة/السرعة: {load_speed}" if load_speed else "",
        f"سنة الصنع: {year}" if year else "",
        f"الضمان: {warranty}" if warranty else "",
        f"النقشة: {pattern}" if pattern else "",
    ]
    description = " | ".join([x for x in description_parts if x])
    return {
        "name": name,
        "brand": brand or "",
        "category": category or "إطارات",
        "size": size,
        "load_speed": load_speed,
        "price": price,
        "old_price": "",
        "year": year,
        "warranty": warranty,
        "pattern": pattern,
        "image": image,
        "description": description,
        "sku": make_sku(name),
        "source_url": page_url,
    }


def extract_cards(soup: BeautifulSoup):
    selectors = ["li.product", ".product", ".type-product", ".product-grid-item", ".wd-product", ".product-small", "[class*='product']"]
    cards = []
    seen = set()
    for selector in selectors:
        for el in soup.select(selector):
            key = id(el)
            txt = el.get_text(" ", strip=True)
            if key in seen:
                continue
            if len(txt) < 20:
                continue
            seen.add(key)
            cards.append(el)
    return cards


def parse_products_from_html(html: str, page_url: str, brand: str, category: str) -> list[dict]:
    soup = BeautifulSoup(html, "html.parser")
    products = []
    cards = extract_cards(soup)
    print("CARDS FOUND =", len(cards))
    for card in cards:
        item = card_to_product(card, page_url, brand, category)
        if item:
            products.append(item)
    products = clean_products(products, brand)
    print("PRODUCTS AFTER CARD PARSE =", len(products))
    return products


def clean_products(products: list[dict], brand: str = "") -> list[dict]:
    clean = {}
    for p in products:
        name = normalize_space(p.get("name", ""))
        if not is_real_tire_name(name, brand):
            continue
        price = clean_price(p.get("price", ""))
        if not price:
            continue
        p["name"] = name
        p["price"] = price
        p["size"] = p.get("size") or extract_size(name)
        p["load_speed"] = p.get("load_speed") or extract_load_speed(name)
        p["sku"] = p.get("sku") or make_sku(name)
        clean[name] = p
    return list(clean.values())


def build_page_url(category_url: str, page: int) -> str:
    base = category_url.rstrip("/")
    if page <= 1:
        return base + "/"
    return f"{base}/page/{page}/"


def scrape_tireex_category(category_url: str, brand: str = "", category: str = "إطارات", max_pages: int = 5) -> list[dict]:
    all_products = []
    for page in range(1, max_pages + 1):
        page_url = build_page_url(category_url, page)
        html, final_url, status = fetch_html(page_url)
        if status >= 400:
            print("SKIP PAGE STATUS =", status)
            continue
        products = parse_products_from_html(html, final_url, brand, category)
        print("PAGE PRODUCTS COUNT =", len(products), "PAGE =", page_url)
        if not products and page > 1:
            break
        all_products.extend(products)
        time.sleep(1)
    final_products = clean_products(all_products, brand)
    print("FINAL PRODUCTS COUNT =", len(final_products))
    if not final_products:
        raise HTTPException(status_code=422, detail="فشل الجلب: لم يتم استخراج أي منتج من الصفحة المحددة.")
    return final_products


def save_latest_products(products: list[dict]):
    if not products:
        raise HTTPException(status_code=422, detail="لا يمكن حفظ منتجات فارغة.")
    with open(LATEST_PRODUCTS_JSON, "w", encoding="utf-8") as f:
        json.dump(products, f, ensure_ascii=False, indent=2)
    print("SAVED PRODUCTS COUNT =", len(products))
    print("SAVED JSON PATH =", str(LATEST_PRODUCTS_JSON))


def load_latest_products() -> list[dict]:
    if not LATEST_PRODUCTS_JSON.exists():
        raise HTTPException(status_code=422, detail="لا توجد منتجات محفوظة للتصدير. نفذ الجلب أولاً.")
    with open(LATEST_PRODUCTS_JSON, "r", encoding="utf-8") as f:
        products = json.load(f)
    if not products:
        raise HTTPException(status_code=422, detail="ملف المنتجات المحفوظ فارغ. نفذ الجلب مرة أخرى.")
    return products


def build_salla_xlsx(products: list[dict], output_path: Path):
    if not products:
        raise HTTPException(status_code=422, detail="لا يمكن إنشاء ملف Excel بدون منتجات.")
    wb = Workbook()
    ws = wb.active
    ws.title = "products"
    ws.append(SALLA_HEADERS)
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1F4E78")
        cell.alignment = Alignment(horizontal="center", vertical="center")
    for p in products:
        row = [
            "منتج",
            p.get("name", ""),
            p.get("category", "إطارات"),
            p.get("image", ""),
            p.get("name", ""),
            "منتج جاهز",
            p.get("price", ""),
            p.get("description", ""),
            "نعم",
            p.get("sku", ""),
            "",
            "",
            "",
            "",
            "",
            "لا",
            "لا",
            "",
            "",
            "100",
        ]
        ws.append(row)
    for col in ws.columns:
        max_len = 0
        col_letter = col[0].column_letter
        for cell in col:
            value = str(cell.value or "")
            if len(value) > max_len:
                max_len = len(value)
        ws.column_dimensions[col_letter].width = min(max_len + 3, 45)
    wb.save(output_path)
    print("EXPORT PRODUCTS COUNT =", len(products))
    print("EXPORT FILE PATH =", str(output_path))


@router.post("/scrape")
def scrape_importer(url: str = Form(...), brand: str = Form(""), category: str = Form("إطارات"), max_pages: int = Form(1)):
    if not url.startswith("http"):
        raise HTTPException(status_code=400, detail="الرابط غير صحيح.")
    max_pages = max(1, min(int(max_pages), 10))
    products = scrape_tireex_category(
        category_url=url,
        brand=brand.strip(),
        category=category.strip() or "إطارات",
        max_pages=max_pages,
    )
    save_latest_products(products)
    return JSONResponse({"ok": True, "count": len(products), "products": products})


@router.get("/products")
def get_latest_products():
    products = load_latest_products()
    return {"ok": True, "count": len(products), "products": products}


@router.get("/salla-xlsx")
def download_salla_xlsx():
    products = load_latest_products()
    if not products:
        raise HTTPException(status_code=422, detail="لا توجد منتجات للتصدير.")
    build_salla_xlsx(products, SALLA_XLSX_PATH)
    return FileResponse(
        path=str(SALLA_XLSX_PATH),
        filename="salla_products_ready.xlsx",
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )

