from __future__ import annotations

import datetime as dt
import uuid
from typing import Any, Optional

import io

from fastapi import APIRouter, Body, File, HTTPException, Query, Request, UploadFile
from fastapi.responses import StreamingResponse
from openpyxl import Workbook, load_workbook
from pydantic import BaseModel, Field

from ..auth_core import log_event, require_csrf, require_user
from ..db import SessionLocal
from ..models import BranchTransfer, BranchTransferLine, Item, Purchase, PurchaseLine, Sale, SaleLine, StockAdjustment, StockMovement
from ..models import Branch, Customer, Supplier, Category, Unit
from ..models import SuspendedSale, SuspendedSaleLine
from ..models import ReturnLine, ReturnTxn

router = APIRouter(prefix="/api/trade", tags=["trade"])


class ItemCreate(BaseModel):
    code: str = ""
    barcode: str = ""
    name: str = Field(min_length=1, max_length=200)
    category: str = "rim"
    category_id: Optional[str] = None
    brand: str = ""
    size: str = ""
    pcd: str = ""
    color: str = ""
    item_condition: str = ""
    location: str = ""
    unit: str = "قطعة"
    unit_id: Optional[str] = None
    is_set: bool = False
    min_qty: float = 0.0
    default_sale_price: float = 0.0
    is_price_tax_inclusive: bool = False
    is_taxable: bool = True
    tax_rate: float = 15.0
    is_active: bool = True
    notes: str = ""
    image_url: str = ""
    default_purchase_price: float = 0.0


class ItemUpdate(BaseModel):
    code: Optional[str] = None
    barcode: Optional[str] = None
    name: Optional[str] = None
    category: Optional[str] = None
    category_id: Optional[str] = None
    brand: Optional[str] = None
    size: Optional[str] = None
    pcd: Optional[str] = None
    color: Optional[str] = None
    item_condition: Optional[str] = None
    location: Optional[str] = None
    unit: Optional[str] = None
    unit_id: Optional[str] = None
    is_set: Optional[bool] = None
    min_qty: Optional[float] = None
    default_sale_price: Optional[float] = None
    is_price_tax_inclusive: Optional[bool] = None
    is_taxable: Optional[bool] = None
    tax_rate: Optional[float] = None
    is_active: Optional[bool] = None
    notes: Optional[str] = None
    image_url: Optional[str] = None
    default_purchase_price: Optional[float] = None


class PurchaseLineIn(BaseModel):
    item_id: str
    qty: float = Field(gt=0)
    unit_cost: float = Field(ge=0)
    extra_cost: float = Field(ge=0, default=0)


class PurchaseCreate(BaseModel):
    invoice_no: str = Field(min_length=1, max_length=80)
    supplier_name: str = Field(min_length=1, max_length=200)
    purchase_date: str
    branch_id: str = ""
    supplier_id: str = ""
    payment_type: str = "cash"
    tax_amount: float = Field(ge=0, default=0)
    discount: float = Field(ge=0, default=0)
    paid_amount: float = Field(ge=0, default=0)
    notes: str = ""
    lines: list[PurchaseLineIn]


class SaleLineIn(BaseModel):
    item_id: str
    qty: float = Field(gt=0)
    sale_price: float = Field(ge=0)
    tax_amount: float = Field(ge=0, default=0)


class SaleCreate(BaseModel):
    invoice_no: str = Field(min_length=1, max_length=80)
    customer_name: str = Field(min_length=1, max_length=200)
    sale_date: str
    payment_type: str = "cash"
    customer_id: str = ""
    customer_tax_no: str = ""
    customer_phone: str = ""
    customer_address: str = ""
    branch_id: str = ""
    discount: float = Field(ge=0, default=0)
    paid_amount: float = Field(ge=0, default=0)
    notes: str = ""
    seller_name: str = ""
    branch_name: str = ""
    allow_negative_stock: bool = False
    lines: list[SaleLineIn]


class SaleUpdate(SaleCreate):
    pass


class SaleSuspendCreate(SaleCreate):
    pass


class ReturnLineIn(BaseModel):
    item_id: str
    qty: float = Field(gt=0)
    unit_price: float = Field(ge=0, default=0)
    unit_cost: float = Field(ge=0, default=0)


class SaleReturnCreate(BaseModel):
    sale_id: str
    return_date: str
    reason: str = ""
    lines: list[ReturnLineIn]


class PurchaseReturnCreate(BaseModel):
    purchase_id: str
    return_date: str
    reason: str = ""
    lines: list[ReturnLineIn]


class StockAdjustCreate(BaseModel):
    item_id: str
    adjust_date: str
    qty_after: float
    reason: str = Field(min_length=1, max_length=500)


class BranchTransferLineIn(BaseModel):
    from_item_id: str
    to_item_id: str
    qty: float = Field(gt=0)


class BranchTransferCreate(BaseModel):
    transfer_no: str = Field(min_length=1, max_length=80)
    transfer_date: str
    from_branch_id: str
    to_branch_id: str
    notes: str = ""
    lines: list[BranchTransferLineIn]


def _parse_date(v: str, field_name: str) -> dt.datetime:
    try:
        return dt.datetime.strptime((v or "").strip(), "%Y-%m-%d")
    except Exception:
        raise HTTPException(400, f"{field_name} يجب أن يكون بصيغة YYYY-MM-DD")


def _can_override_stock(user: Any) -> bool:
    if int(getattr(user, "is_admin", 0) or 0) == 1:
        return True
    role = (getattr(user, "role_name", "") or "").strip().lower()
    return role in {"admin", "manager"}


def _fmt_date(v: Any) -> str:
    if isinstance(v, dt.datetime):
        return v.strftime("%Y-%m-%d")
    if isinstance(v, dt.date):
        return v.strftime("%Y-%m-%d")
    s = str(v or "").strip()
    return s[:10] if s else ""


def _available_qty(db: Any, user_id: str, item_id: str) -> float:
    rows = (
        db.query(StockMovement)
        .filter(StockMovement.user_id == user_id, StockMovement.item_id == item_id)
        .all()
    )
    return round(sum(float(x.qty_in or 0.0) - float(x.qty_out or 0.0) for x in rows), 4)


def _avg_cost(db: Any, user_id: str, item_id: str) -> float:
    rows = (
        db.query(StockMovement)
        .filter(
            StockMovement.user_id == user_id,
            StockMovement.item_id == item_id,
            StockMovement.qty_in > 0,
        )
        .all()
    )
    qty = sum(float(x.qty_in or 0.0) for x in rows)
    amt = sum(float(x.qty_in or 0.0) * float(x.unit_cost or 0.0) for x in rows)
    if qty <= 0:
        return 0.0
    return round(amt / qty, 4)


def _returned_qty(db: Any, user_id: str, reference_type: str, reference_id: str, item_id: str) -> float:
    rows = (
        db.query(ReturnLine, ReturnTxn)
        .join(ReturnTxn, ReturnTxn.id == ReturnLine.return_id)
        .filter(
            ReturnTxn.user_id == user_id,
            ReturnTxn.reference_type == reference_type,
            ReturnTxn.reference_id == reference_id,
            ReturnLine.item_id == item_id,
        )
        .all()
    )
    return round(sum(float(ln.qty or 0.0) for ln, _ in rows), 4)


def _safe_div(a: float, b: float) -> float:
    aa = float(a or 0.0)
    bb = float(b or 0.0)
    if abs(bb) < 1e-9:
        return 0.0
    return aa / bb


def _reverse_sale_effects(db: Any, user_id: str, sale_id: str) -> None:
    sale_lines = db.query(SaleLine).filter(SaleLine.sale_id == sale_id).all()
    item_ids = {x.item_id for x in sale_lines}
    items = db.query(Item).filter(Item.user_id == user_id, Item.id.in_(list(item_ids))).all() if item_ids else []
    item_by_id = {x.id: x for x in items}
    for ln in sale_lines:
        item = item_by_id.get(ln.item_id)
        if item:
            item.quantity = round(float(item.quantity or 0.0) + float(ln.qty or 0.0), 4)
    for mv in db.query(StockMovement).filter(
        StockMovement.user_id == user_id,
        StockMovement.reference_type == "sale",
        StockMovement.reference_id == sale_id,
    ).all():
        db.delete(mv)
    for ln in sale_lines:
        db.delete(ln)


class CategoryCreate(BaseModel):
    code: str = ""
    name: str = Field(min_length=1, max_length=200)
    notes: str = ""


class UnitCreate(BaseModel):
    code: str = ""
    name: str = Field(min_length=1, max_length=200)
    notes: str = ""


@router.get("/categories")
def list_categories(request: Request):
    db = SessionLocal()
    try:
        user = require_user(db, request)
        rows = db.query(Category).filter(Category.user_id == user.id).order_by(Category.name.asc()).all()
        return {"categories": [{"id": r.id, "code": r.code, "name": r.name, "notes": r.notes or ""} for r in rows]}
    finally:
        db.close()


@router.post("/categories")
def create_category(request: Request, body: CategoryCreate = Body(...)):
    db = SessionLocal()
    try:
        require_csrf(request)
        user = require_user(db, request)
        name = body.name.strip()
        if not name:
            raise HTTPException(400, "اسم التصنيف مطلوب")
        dup = db.query(Category).filter(Category.user_id == user.id, Category.name == name).first()
        if dup:
            raise HTTPException(400, "التصنيف موجود مسبقاً")
        code = (body.code or "").strip() or ("CAT-" + uuid.uuid4().hex[:8].upper())
        rec = Category(id=uuid.uuid4().hex, user_id=user.id, code=code, name=name, notes=(body.notes or "").strip() or None)
        db.add(rec)
        db.commit()
        return {"id": rec.id, "code": rec.code, "name": rec.name}
    finally:
        db.close()


@router.delete("/categories/{category_id}")
def delete_category(category_id: str, request: Request):
    db = SessionLocal()
    try:
        require_csrf(request)
        user = require_user(db, request)
        rec = db.query(Category).filter(Category.user_id == user.id, Category.id == category_id).first()
        if not rec:
            raise HTTPException(404, "التصنيف غير موجود")
        in_use = db.query(Item).filter(Item.user_id == user.id, Item.category_id == category_id).first()
        if in_use:
            raise HTTPException(400, "لا يمكن حذف التصنيف لارتباطه بأصناف")
        db.delete(rec)
        db.commit()
        return {"deleted": True}
    finally:
        db.close()


@router.get("/units")
def list_units(request: Request):
    db = SessionLocal()
    try:
        user = require_user(db, request)
        rows = db.query(Unit).filter(Unit.user_id == user.id).order_by(Unit.name.asc()).all()
        return {"units": [{"id": r.id, "code": r.code, "name": r.name, "notes": r.notes or ""} for r in rows]}
    finally:
        db.close()


@router.post("/units")
def create_unit(request: Request, body: UnitCreate = Body(...)):
    db = SessionLocal()
    try:
        require_csrf(request)
        user = require_user(db, request)
        name = body.name.strip()
        if not name:
            raise HTTPException(400, "اسم الوحدة مطلوب")
        dup = db.query(Unit).filter(Unit.user_id == user.id, Unit.name == name).first()
        if dup:
            raise HTTPException(400, "الوحدة موجودة مسبقاً")
        code = (body.code or "").strip() or ("UNT-" + uuid.uuid4().hex[:8].upper())
        rec = Unit(id=uuid.uuid4().hex, user_id=user.id, code=code, name=name, notes=(body.notes or "").strip() or None)
        db.add(rec)
        db.commit()
        return {"id": rec.id, "code": rec.code, "name": rec.name}
    finally:
        db.close()


@router.delete("/units/{unit_id}")
def delete_unit(unit_id: str, request: Request):
    db = SessionLocal()
    try:
        require_csrf(request)
        user = require_user(db, request)
        rec = db.query(Unit).filter(Unit.user_id == user.id, Unit.id == unit_id).first()
        if not rec:
            raise HTTPException(404, "الوحدة غير موجودة")
        in_use = db.query(Item).filter(Item.user_id == user.id, Item.unit_id == unit_id).first()
        if in_use:
            raise HTTPException(400, "لا يمكن حذف الوحدة لارتباطها بأصناف")
        db.delete(rec)
        db.commit()
        return {"deleted": True}
    finally:
        db.close()


def _norm_rate(tax_rate, is_taxable) -> float:
    if not is_taxable:
        return 0.0
    r = float(tax_rate or 0.0)
    if r > 1:
        r = r / 100.0
    return r


@router.get("/items")
def list_items(
    request: Request,
    q: str = Query(""),
    category: str = Query(""),
    is_active: str = Query(""),
):
    db = SessionLocal()
    try:
        user = require_user(db, request)
        query = db.query(Item).filter(Item.user_id == user.id)
        cat = (category or "").strip().lower()
        if cat:
            query = query.filter(Item.category == cat)
        active_flag = (is_active or "").strip().lower()
        if active_flag in {"1", "true", "yes", "active"}:
            query = query.filter(Item.is_active == 1)
        elif active_flag in {"0", "false", "no", "inactive"}:
            query = query.filter(Item.is_active == 0)
        qn = (q or "").strip().lower()
        if qn:
            rows = query.order_by(Item.created_at.desc()).all()
            rows = [r for r in rows if qn in (f"{r.code} {r.barcode or ''} {r.name} {r.brand or ''} {r.size or ''}").lower()]
        else:
            rows = query.order_by(Item.created_at.desc()).all()
        cat_ids = {r.category_id for r in rows if r.category_id}
        unit_ids = {getattr(r, "unit_id", None) for r in rows if getattr(r, "unit_id", None)}
        cat_names = {}
        if cat_ids:
            for c in db.query(Category).filter(Category.id.in_(cat_ids)).all():
                cat_names[c.id] = c.name
        unit_names = {}
        if unit_ids:
            for u in db.query(Unit).filter(Unit.id.in_(unit_ids)).all():
                unit_names[u.id] = u.name

        def _row_json(r):
            qty = round(float(r.quantity or 0.0), 4)
            min_qty = round(float(r.min_qty or 0.0), 4)
            status = "out" if qty <= 0 else ("low" if qty <= min_qty else "available")
            rate = _norm_rate(r.tax_rate, r.is_taxable)
            excl = round(float(r.default_sale_price or 0.0), 2)
            incl = round(excl * (1 + rate), 2)
            return {
                "id": r.id,
                "code": r.code,
                "barcode": r.barcode or "",
                "name": r.name,
                "category": r.category,
                "category_id": r.category_id or "",
                "category_name": cat_names.get(r.category_id, ""),
                "brand": r.brand or "",
                "size": r.size or "",
                "pcd": r.pcd or "",
                "color": r.color or "",
                "item_condition": r.item_condition or "",
                "location": r.location or "",
                "unit": r.unit or "قطعة",
                "unit_id": getattr(r, "unit_id", None) or "",
                "unit_name": unit_names.get(getattr(r, "unit_id", None), ""),
                "branch_id": r.branch_id or "",
                "is_set": bool(int(r.is_set or 0)),
                "is_unique": bool(int(r.is_unique or 0)),
                "needs_service": bool(int(r.needs_service or 0)),
                "quantity": qty,
                "min_qty": min_qty,
                "status": status,
                "default_sale_price": excl,
                "sale_price_incl_tax": incl,
                "is_price_tax_inclusive": bool(int(getattr(r, "is_price_tax_inclusive", 0) or 0)),
                "is_taxable": bool(int(r.is_taxable or 0)),
                "tax_rate": round(float(r.tax_rate or 0.0), 4),
                "is_active": bool(int(r.is_active or 0)),
                "last_cost": round(float(r.last_cost or 0.0), 2),
                "stock_value": round(qty * round(float(r.last_cost or 0.0), 4), 2),
                "notes": r.notes or "",
                "image_url": r.image_url or "",
            }

        return {"items": [_row_json(r) for r in rows]}
    finally:
        db.close()


@router.post("/items")
def create_item(request: Request, body: ItemCreate = Body(...)):
    db = SessionLocal()
    try:
        require_csrf(request)
        user = require_user(db, request)
        code = (body.code or "").strip()
        if not code:
            for _ in range(20):
                candidate = _new_item_code()
                if not db.query(Item).filter(Item.user_id == user.id, Item.code == candidate).first():
                    code = candidate
                    break
            if not code:
                raise HTTPException(400, "تعذّر توليد كود صنف فريد، حاول مرة أخرى")
        elif db.query(Item).filter(Item.user_id == user.id, Item.code == code).first():
            raise HTTPException(400, "كود الصنف مستخدم مسبقاً")

        category_name = (body.category or "rim").strip().lower()
        category_id = (body.category_id or "").strip() or None
        if category_id:
            cat = db.query(Category).filter(Category.user_id == user.id, Category.id == category_id).first()
            if not cat:
                raise HTTPException(400, "التصنيف غير موجود")
            category_name = cat.name

        unit_name = (body.unit or "قطعة").strip() or "قطعة"
        unit_id = (body.unit_id or "").strip() or None
        if unit_id:
            un = db.query(Unit).filter(Unit.user_id == user.id, Unit.id == unit_id).first()
            if not un:
                raise HTTPException(400, "الوحدة غير موجودة")
            unit_name = un.name

        is_taxable = 1 if body.is_taxable else 0
        tax_rate = round(abs(float(body.tax_rate or 0.0)), 4)
        raw_price = round(abs(float(body.default_sale_price or 0.0)), 2)
        rate = _norm_rate(tax_rate, is_taxable)
        if body.is_price_tax_inclusive and rate > 0:
            sale_price_excl = round(raw_price / (1 + rate), 2)
        else:
            sale_price_excl = raw_price

        rec = Item(
            id=uuid.uuid4().hex,
            user_id=user.id,
            code=code,
            barcode=(body.barcode or "").strip() or None,
            name=body.name.strip(),
            category=category_name,
            category_id=category_id,
            brand=(body.brand or "").strip() or None,
            size=(body.size or "").strip() or None,
            pcd=(body.pcd or "").strip() or None,
            color=(body.color or "").strip() or None,
            item_condition=(body.item_condition or "").strip() or None,
            location=(body.location or "").strip() or None,
            unit=unit_name,
            unit_id=unit_id,
            is_set=1 if body.is_set else 0,
            quantity=0.0,
            min_qty=round(abs(float(body.min_qty or 0.0)), 4),
            default_sale_price=sale_price_excl,
            is_price_tax_inclusive=1 if body.is_price_tax_inclusive else 0,
            is_taxable=is_taxable,
            tax_rate=tax_rate,
            is_active=1 if body.is_active else 0,
            last_cost=round(abs(float(body.default_purchase_price or 0.0)), 2),
            notes=(body.notes or "").strip() or None,
            image_url=(body.image_url or "").strip() or None,
        )
        db.add(rec)
        db.commit()
        return {"id": rec.id, "code": rec.code, "name": rec.name}
    finally:
        db.close()


@router.patch("/items/{item_id}")
def update_item(item_id: str, request: Request, body: ItemUpdate = Body(...)):
    db = SessionLocal()
    try:
        require_csrf(request)
        user = require_user(db, request)
        rec = db.query(Item).filter(Item.user_id == user.id, Item.id == item_id).first()
        if not rec:
            raise HTTPException(404, "الصنف غير موجود")
        if body.code is not None:
            code = body.code.strip()
            if not code:
                raise HTTPException(400, "code لا يمكن أن يكون فارغاً")
            dup = db.query(Item).filter(Item.user_id == user.id, Item.code == code, Item.id != rec.id).first()
            if dup:
                raise HTTPException(400, "كود الصنف مستخدم مسبقاً")
            rec.code = code
        for k in ("barcode", "name", "category", "brand", "size", "pcd", "color", "item_condition", "location", "unit", "notes", "image_url"):
            v = getattr(body, k)
            if v is not None:
                vv = str(v).strip()
                if k == "unit":
                    setattr(rec, k, vv if vv else "قطعة")
                else:
                    setattr(rec, k, vv if vv else None)
        if body.category_id is not None:
            cid = body.category_id.strip() or None
            if cid:
                cat = db.query(Category).filter(Category.user_id == user.id, Category.id == cid).first()
                if not cat:
                    raise HTTPException(400, "التصنيف غير موجود")
                rec.category_id = cid
                rec.category = cat.name
            else:
                rec.category_id = None
        if body.unit_id is not None:
            uid = body.unit_id.strip() or None
            if uid:
                un = db.query(Unit).filter(Unit.user_id == user.id, Unit.id == uid).first()
                if not un:
                    raise HTTPException(400, "الوحدة غير موجودة")
                rec.unit_id = uid
                rec.unit = un.name
            else:
                rec.unit_id = None
        if body.name is not None and not (body.name or "").strip():
            raise HTTPException(400, "name لا يمكن أن يكون فارغاً")
        if body.name is not None:
            rec.name = body.name.strip()
        if body.is_set is not None:
            rec.is_set = 1 if body.is_set else 0
        if body.min_qty is not None:
            rec.min_qty = round(abs(float(body.min_qty or 0.0)), 4)
        if body.is_taxable is not None:
            rec.is_taxable = 1 if body.is_taxable else 0
        if body.tax_rate is not None:
            rec.tax_rate = round(abs(float(body.tax_rate or 0.0)), 4)
        if body.is_price_tax_inclusive is not None:
            rec.is_price_tax_inclusive = 1 if body.is_price_tax_inclusive else 0
        if body.default_sale_price is not None:
            raw_price = round(abs(float(body.default_sale_price or 0.0)), 2)
            incl_flag = bool(body.is_price_tax_inclusive) if body.is_price_tax_inclusive is not None else bool(int(rec.is_price_tax_inclusive or 0))
            rate = _norm_rate(rec.tax_rate, rec.is_taxable)
            if incl_flag and rate > 0:
                rec.default_sale_price = round(raw_price / (1 + rate), 2)
            else:
                rec.default_sale_price = raw_price
        if body.is_active is not None:
            rec.is_active = 1 if body.is_active else 0
        if body.default_purchase_price is not None:
            rec.last_cost = round(abs(float(body.default_purchase_price or 0.0)), 2)
        db.commit()
        return {"ok": True}
    finally:
        db.close()


@router.delete("/items/{item_id}")
def delete_item(item_id: str, request: Request):
    db = SessionLocal()
    try:
        require_csrf(request)
        user = require_user(db, request)
        rec = db.query(Item).filter(Item.user_id == user.id, Item.id == item_id).first()
        if not rec:
            raise HTTPException(404, "الصنف غير موجود")
        has_purchase = db.query(PurchaseLine).filter(PurchaseLine.item_id == rec.id).first()
        has_sale = db.query(SaleLine).filter(SaleLine.item_id == rec.id).first()
        if has_purchase or has_sale:
            raise HTTPException(400, "لا يمكن حذف الصنف لوجود حركات مرتبطة")
        db.delete(rec)
        db.commit()
        return {"deleted": True}
    finally:
        db.close()


@router.post("/purchases")
def create_purchase(request: Request, body: PurchaseCreate = Body(...)):
    db = SessionLocal()
    try:
        require_csrf(request)
        user = require_user(db, request)
        p_date = _parse_date(body.purchase_date, "purchase_date")
        if not body.lines:
            raise HTTPException(400, "الفاتورة تحتاج بنوداً")
        item_ids = {x.item_id for x in body.lines}
        items = db.query(Item).filter(Item.user_id == user.id, Item.id.in_(list(item_ids))).all()
        if len(items) != len(item_ids):
            raise HTTPException(400, "يوجد صنف غير صالح")
        item_by_id = {x.id: x for x in items}
        supplier_name = body.supplier_name.strip()
        if (body.supplier_id or "").strip():
            s = db.query(Supplier).filter(Supplier.user_id == user.id, Supplier.id == body.supplier_id.strip()).first()
            if s:
                supplier_name = s.name
        branch_id = (body.branch_id or "").strip() or None
        if branch_id and not db.query(Branch).filter(Branch.user_id == user.id, Branch.id == branch_id).first():
            raise HTTPException(400, "الفرع غير صالح")

        rec = Purchase(
            id=uuid.uuid4().hex,
            user_id=user.id,
            branch_id=branch_id,
            supplier_id=(body.supplier_id or "").strip() or None,
            invoice_no=body.invoice_no.strip(),
            supplier_name=supplier_name,
            purchase_date=p_date,
            payment_type=(body.payment_type or "cash").strip().lower(),
            tax_amount=round(abs(float(body.tax_amount or 0.0)), 2),
            discount=round(abs(float(body.discount or 0.0)), 2),
            paid_amount=round(abs(float(body.paid_amount or 0.0)), 2),
            due_amount=0.0,
            notes=(body.notes or "").strip() or None,
            total_amount=0.0,
        )
        db.add(rec)
        db.flush()
        total = 0.0
        for ln in body.lines:
            qty = round(abs(float(ln.qty or 0.0)), 4)
            unit_cost = round(abs(float(ln.unit_cost or 0.0)), 4)
            extra = round(abs(float(ln.extra_cost or 0.0)), 4)
            total_cost = round(qty * unit_cost + extra, 2)
            if qty <= 0:
                continue
            db.add(PurchaseLine(purchase_id=rec.id, item_id=ln.item_id, qty=qty, unit_cost=unit_cost, extra_cost=extra, total_cost=total_cost))
            eff_unit = round(total_cost / qty, 4)
            db.add(
                StockMovement(
                    id=uuid.uuid4().hex,
                    user_id=user.id,
                    item_id=ln.item_id,
                    movement_type="purchase",
                    qty_in=qty,
                    qty_out=0.0,
                    unit_cost=eff_unit,
                    reference_type="purchase",
                    reference_id=rec.id,
                    movement_date=p_date,
                )
            )
            item = item_by_id[ln.item_id]
            item.quantity = round(float(item.quantity or 0.0) + qty, 4)
            item.last_cost = eff_unit
            total += total_cost
        gross = round(total + rec.tax_amount, 2)
        rec.total_amount = round(max(0.0, gross - rec.discount), 2)
        rec.due_amount = round(max(0.0, rec.total_amount - rec.paid_amount), 2)
        db.commit()
        log_event(db, "trade.purchase.create", user.id, {"purchase_id": rec.id, "invoice_no": rec.invoice_no, "total_amount": rec.total_amount})
        return {"id": rec.id, "total_amount": rec.total_amount, "due_amount": rec.due_amount}
    finally:
        db.close()


@router.get("/purchases")
def list_purchases(request: Request, limit: int = Query(200, ge=1, le=1000)):
    db = SessionLocal()
    try:
        user = require_user(db, request)
        rows = db.query(Purchase).filter(Purchase.user_id == user.id).order_by(Purchase.purchase_date.desc(), Purchase.created_at.desc()).limit(limit).all()
        return {
            "items": [
                {
                    "id": x.id,
                    "invoice_no": x.invoice_no,
                    "supplier_name": x.supplier_name,
                    "purchase_date": _fmt_date(x.purchase_date),
                    "payment_type": x.payment_type or "cash",
                    "tax_amount": round(float(x.tax_amount or 0.0), 2),
                    "discount": round(float(x.discount or 0.0), 2),
                    "paid_amount": round(float(x.paid_amount or 0.0), 2),
                    "due_amount": round(float(x.due_amount or 0.0), 2),
                    "total_amount": round(float(x.total_amount or 0.0), 2),
                }
                for x in rows
            ]
        }
    finally:
        db.close()


@router.post("/sales")
def create_sale(request: Request, body: SaleCreate = Body(...)):
    db = SessionLocal()
    try:
        require_csrf(request)
        user = require_user(db, request)
        s_date = _parse_date(body.sale_date, "sale_date")
        if not body.lines:
            raise HTTPException(400, "فاتورة البيع تحتاج بنوداً")
        item_ids = {x.item_id for x in body.lines}
        items = db.query(Item).filter(Item.user_id == user.id, Item.id.in_(list(item_ids))).all()
        if len(items) != len(item_ids):
            raise HTTPException(400, "يوجد صنف غير صالح")
        item_by_id = {x.id: x for x in items}
        pay_type = (body.payment_type or "cash").strip().lower()
        customer_name = body.customer_name.strip()
        if pay_type == "credit":
            if not customer_name and not (body.customer_id or "").strip():
                raise HTTPException(400, "في البيع الآجل يجب تحديد عميل")
            if customer_name in {"عميل نقدي", "cash customer"} and not (body.customer_id or "").strip():
                raise HTTPException(400, "في البيع الآجل يجب إدخال اسم عميل صحيح")
        elif not customer_name:
            customer_name = "عميل نقدي"
        customer_tax_no = (body.customer_tax_no or "").strip()
        customer_phone = (body.customer_phone or "").strip()
        customer_address = (body.customer_address or "").strip()
        if (body.customer_id or "").strip():
            c = db.query(Customer).filter(Customer.user_id == user.id, Customer.id == body.customer_id.strip()).first()
            if c:
                customer_name = c.name
                customer_phone = (c.phone or "").strip()
                customer_address = (c.address or "").strip()
        branch_id = (body.branch_id or "").strip() or None
        if branch_id and not db.query(Branch).filter(Branch.user_id == user.id, Branch.id == branch_id).first():
            raise HTTPException(400, "الفرع غير صالح")

        rec = Sale(
            id=uuid.uuid4().hex,
            user_id=user.id,
            branch_id=branch_id,
            customer_id=(body.customer_id or "").strip() or None,
            invoice_no=body.invoice_no.strip(),
            customer_name=customer_name,
            customer_tax_no=customer_tax_no or None,
            customer_phone=customer_phone or None,
            customer_address=customer_address or None,
            sale_date=s_date,
            payment_type=pay_type,
            discount=round(abs(float(body.discount or 0.0)), 2),
            paid_amount=round(abs(float(body.paid_amount or 0.0)), 2),
            tax_amount=0.0,
            due_amount=0.0,
            notes=(body.notes or "").strip() or None,
            seller_name=(body.seller_name or "").strip() or None,
            branch_name=(body.branch_name or "").strip() or None,
            total_amount=0.0,
        )
        db.add(rec)
        db.flush()
        total = 0.0
        for ln in body.lines:
            qty = round(abs(float(ln.qty or 0.0)), 4)
            sale_price = round(abs(float(ln.sale_price or 0.0)), 4)
            tax = round(abs(float(ln.tax_amount or 0.0)), 2)
            if qty <= 0:
                continue
            item = item_by_id[ln.item_id]
            available = round(float(item.quantity or 0.0), 4)
            allow_negative_stock = bool(body.allow_negative_stock and _can_override_stock(user))
            if available + 0.0001 < qty and not allow_negative_stock:
                raise HTTPException(400, f"المخزون غير كافٍ للصنف: {item.name}")
            unit_cost = _avg_cost(db, user.id, ln.item_id)
            cost_total = round(unit_cost * qty, 2)
            line_total = round(sale_price * qty + tax, 2)
            profit = round((sale_price * qty) - cost_total, 2)
            db.add(SaleLine(sale_id=rec.id, item_id=ln.item_id, qty=qty, sale_price=sale_price, tax_amount=tax, cost_price=cost_total, profit=profit))
            db.add(
                StockMovement(
                    id=uuid.uuid4().hex,
                    user_id=user.id,
                    item_id=ln.item_id,
                    movement_type="sale",
                    qty_in=0.0,
                    qty_out=qty,
                    unit_cost=unit_cost,
                    reference_type="sale",
                    reference_id=rec.id,
                    movement_date=s_date,
                )
            )
            item.quantity = round(available - qty, 4)
            total += line_total
        rec.total_amount = round(max(0.0, total - rec.discount), 2)
        rec.tax_amount = round(sum(float(x.tax_amount or 0.0) for x in db.query(SaleLine).filter(SaleLine.sale_id == rec.id).all()), 2)
        rec.due_amount = round(max(0.0, rec.total_amount - rec.paid_amount), 2)
        db.commit()
        log_event(db, "trade.sale.create", user.id, {"sale_id": rec.id, "invoice_no": rec.invoice_no, "total_amount": rec.total_amount})
        return {"id": rec.id, "total_amount": rec.total_amount}
    finally:
        db.close()


@router.get("/sales")
def list_sales(request: Request, limit: int = Query(200, ge=1, le=1000)):
    db = SessionLocal()
    try:
        user = require_user(db, request)
        rows = db.query(Sale).filter(Sale.user_id == user.id).order_by(Sale.sale_date.desc(), Sale.created_at.desc()).limit(limit).all()
        return {
            "items": [
                {
                    "id": x.id,
                    "invoice_no": x.invoice_no,
                    "customer_name": x.customer_name,
                    "customer_tax_no": x.customer_tax_no or "",
                    "customer_phone": x.customer_phone or "",
                    "customer_address": x.customer_address or "",
                    "sale_date": _fmt_date(x.sale_date),
                    "payment_type": x.payment_type,
                    "paid_amount": round(float(x.paid_amount or 0.0), 2),
                    "due_amount": round(float(x.due_amount or 0.0), 2),
                    "total_amount": round(float(x.total_amount or 0.0), 2),
                }
                for x in rows
            ]
        }
    finally:
        db.close()


@router.get("/sales/suspended")
def list_suspended_sales(request: Request, limit: int = Query(200, ge=1, le=1000)):
    db = SessionLocal()
    try:
        user = require_user(db, request)
        rows = (
            db.query(SuspendedSale)
            .filter(SuspendedSale.user_id == user.id)
            .order_by(SuspendedSale.sale_date.desc(), SuspendedSale.created_at.desc())
            .limit(limit)
            .all()
        )
        return {
            "items": [
                {
                    "id": x.id,
                    "invoice_no": x.invoice_no,
                    "customer_name": x.customer_name,
                    "sale_date": _fmt_date(x.sale_date),
                    "payment_type": x.payment_type,
                    "discount": round(float(x.discount or 0.0), 2),
                    "paid_amount": round(float(x.paid_amount or 0.0), 2),
                }
                for x in rows
            ]
        }
    finally:
        db.close()


@router.get("/sales/suspended/{suspended_id}")
def suspended_sale_details(suspended_id: str, request: Request):
    db = SessionLocal()
    try:
        user = require_user(db, request)
        s = db.query(SuspendedSale).filter(SuspendedSale.user_id == user.id, SuspendedSale.id == suspended_id).first()
        if not s:
            raise HTTPException(404, "الفاتورة المعلقة غير موجودة")
        lines = db.query(SuspendedSaleLine, Item).join(Item, Item.id == SuspendedSaleLine.item_id).filter(SuspendedSaleLine.suspended_sale_id == s.id).all()
        return {
            "id": s.id,
            "invoice_no": s.invoice_no,
            "customer_name": s.customer_name,
            "sale_date": _fmt_date(s.sale_date),
            "payment_type": s.payment_type,
            "discount": round(float(s.discount or 0.0), 2),
            "paid_amount": round(float(s.paid_amount or 0.0), 2),
            "notes": s.notes or "",
            "seller_name": s.seller_name or "",
            "branch_name": s.branch_name or "",
            "lines": [
                {
                    "item_id": item.id,
                    "item_code": item.code,
                    "item_name": item.name,
                    "qty": round(float(ln.qty or 0.0), 4),
                    "sale_price": round(float(ln.sale_price or 0.0), 2),
                    "tax_amount": round(float(ln.tax_amount or 0.0), 2),
                }
                for ln, item in lines
            ],
        }
    finally:
        db.close()


@router.post("/sales/suspend")
def suspend_sale(request: Request, body: SaleSuspendCreate = Body(...)):
    db = SessionLocal()
    try:
        require_csrf(request)
        user = require_user(db, request)
        s_date = _parse_date(body.sale_date, "sale_date")
        if not body.lines:
            raise HTTPException(400, "الفاتورة تحتاج بنوداً")
        rec = SuspendedSale(
            id=uuid.uuid4().hex,
            user_id=user.id,
            invoice_no=body.invoice_no.strip(),
            customer_name=body.customer_name.strip(),
            sale_date=s_date,
            payment_type=(body.payment_type or "cash").strip().lower(),
            discount=round(abs(float(body.discount or 0.0)), 2),
            paid_amount=round(abs(float(body.paid_amount or 0.0)), 2),
            notes=(body.notes or "").strip() or None,
            seller_name=(body.seller_name or "").strip() or None,
            branch_name=(body.branch_name or "").strip() or None,
        )
        db.add(rec)
        db.flush()
        for ln in body.lines:
            db.add(
                SuspendedSaleLine(
                    suspended_sale_id=rec.id,
                    item_id=ln.item_id,
                    qty=round(abs(float(ln.qty or 0.0)), 4),
                    sale_price=round(abs(float(ln.sale_price or 0.0)), 4),
                    tax_amount=round(abs(float(ln.tax_amount or 0.0)), 2),
                )
            )
        db.commit()
        return {"id": rec.id, "suspended": True}
    finally:
        db.close()


@router.post("/sales/suspended/{suspended_id}/checkout")
def checkout_suspended_sale(suspended_id: str, request: Request):
    db = SessionLocal()
    try:
        require_csrf(request)
        user = require_user(db, request)
        s = db.query(SuspendedSale).filter(SuspendedSale.user_id == user.id, SuspendedSale.id == suspended_id).first()
        if not s:
            raise HTTPException(404, "الفاتورة المعلقة غير موجودة")
        lines = db.query(SuspendedSaleLine).filter(SuspendedSaleLine.suspended_sale_id == s.id).all()
        body = SaleCreate(
            invoice_no=s.invoice_no,
            customer_name=s.customer_name,
            sale_date=s.sale_date.strftime("%Y-%m-%d"),
            payment_type=s.payment_type,
            customer_id="",
            branch_id="",
            discount=round(float(s.discount or 0.0), 2),
            paid_amount=round(float(s.paid_amount or 0.0), 2),
            notes=s.notes or "",
            seller_name=s.seller_name or "",
            branch_name=s.branch_name or "",
            lines=[
                SaleLineIn(
                    item_id=x.item_id,
                    qty=round(float(x.qty or 0.0), 4),
                    sale_price=round(float(x.sale_price or 0.0), 4),
                    tax_amount=round(float(x.tax_amount or 0.0), 2),
                )
                for x in lines
            ],
        )
        res = create_sale(request, body)
        for x in lines:
            db.delete(x)
        db.delete(s)
        db.commit()
        return {"checked_out": True, "sale_id": res.get("id"), "total_amount": res.get("total_amount")}
    finally:
        db.close()


@router.delete("/sales/suspended/{suspended_id}")
def delete_suspended_sale(suspended_id: str, request: Request):
    db = SessionLocal()
    try:
        require_csrf(request)
        user = require_user(db, request)
        s = db.query(SuspendedSale).filter(SuspendedSale.user_id == user.id, SuspendedSale.id == suspended_id).first()
        if not s:
            raise HTTPException(404, "الفاتورة المعلقة غير موجودة")
        for x in db.query(SuspendedSaleLine).filter(SuspendedSaleLine.suspended_sale_id == s.id).all():
            db.delete(x)
        db.delete(s)
        db.commit()
        return {"deleted": True}
    finally:
        db.close()


@router.get("/sales/export")
def export_sales(request: Request):
    db = SessionLocal()
    try:
        user = require_user(db, request)
        rows = db.query(Sale).filter(Sale.user_id == user.id).order_by(Sale.sale_date.desc()).all()
        wb = Workbook()
        ws = wb.active
        ws.title = "المبيعات"
        ws.append(["رقم الفاتورة", "العميل", "التاريخ", "طريقة الدفع", "الخصم", "الضريبة", "الإجمالي", "المدفوع", "المتبقي"])
        for r in rows:
            ws.append(
                [
                    r.invoice_no,
                    r.customer_name,
                    _fmt_date(r.sale_date),
                    r.payment_type,
                    round(float(r.discount or 0.0), 2),
                    round(float(r.tax_amount or 0.0), 2),
                    round(float(r.total_amount or 0.0), 2),
                    round(float(r.paid_amount or 0.0), 2),
                    round(float(r.due_amount or 0.0), 2),
                ]
            )
        return _xlsx_response(wb, "sales.xlsx")
    finally:
        db.close()


@router.get("/sales/{sale_id}")
def sale_details(sale_id: str, request: Request):
    db = SessionLocal()
    try:
        user = require_user(db, request)
        sale = db.query(Sale).filter(Sale.user_id == user.id, Sale.id == sale_id).first()
        if not sale:
            raise HTTPException(404, "فاتورة البيع غير موجودة")
        lines = (
            db.query(SaleLine, Item)
            .join(Item, Item.id == SaleLine.item_id)
            .filter(SaleLine.sale_id == sale.id)
            .order_by(SaleLine.id.asc())
            .all()
        )
        return {
            "id": sale.id,
            "invoice_no": sale.invoice_no,
            "customer_name": sale.customer_name,
            "customer_tax_no": sale.customer_tax_no or "",
            "customer_phone": sale.customer_phone or "",
            "customer_address": sale.customer_address or "",
            "sale_date": _fmt_date(sale.sale_date),
            "payment_type": sale.payment_type,
            "discount": round(float(sale.discount or 0.0), 2),
            "paid_amount": round(float(sale.paid_amount or 0.0), 2),
            "due_amount": round(float(sale.due_amount or 0.0), 2),
            "total_amount": round(float(sale.total_amount or 0.0), 2),
            "seller_name": sale.seller_name or "",
            "branch_name": sale.branch_name or "",
            "lines": [
                {
                    "item_id": item.id,
                    "item_code": item.code,
                    "item_name": item.name,
                    "qty": round(float(ln.qty or 0.0), 4),
                    "sale_price": round(float(ln.sale_price or 0.0), 2),
                    "tax_amount": round(float(ln.tax_amount or 0.0), 2),
                    "line_total": round(float(ln.sale_price or 0.0) * float(ln.qty or 0.0) + float(ln.tax_amount or 0.0), 2),
                }
                for ln, item in lines
            ],
        }
    finally:
        db.close()


@router.get("/returns")
def list_returns(request: Request, limit: int = Query(200, ge=1, le=1000)):
    db = SessionLocal()
    try:
        user = require_user(db, request)
        rows = (
            db.query(ReturnTxn)
            .filter(ReturnTxn.user_id == user.id)
            .order_by(ReturnTxn.return_date.desc(), ReturnTxn.created_at.desc())
            .limit(limit)
            .all()
        )
        return {
            "items": [
                {
                    "id": x.id,
                    "return_type": x.return_type,
                    "invoice_no": x.invoice_no,
                    "return_date": x.return_date.strftime("%Y-%m-%d"),
                    "customer_name": x.customer_name or "",
                    "supplier_name": x.supplier_name or "",
                    "reason": x.reason or "",
                    "total_amount": round(float(x.total_amount or 0.0), 2),
                }
                for x in rows
            ]
        }
    finally:
        db.close()


@router.post("/returns/sale")
def create_sale_return(request: Request, body: SaleReturnCreate = Body(...)):
    db = SessionLocal()
    try:
        require_csrf(request)
        user = require_user(db, request)
        rec_sale = db.query(Sale).filter(Sale.user_id == user.id, Sale.id == body.sale_id).first()
        if not rec_sale:
            raise HTTPException(404, "فاتورة البيع غير موجودة")
        r_date = _parse_date(body.return_date, "return_date")
        sale_lines = db.query(SaleLine).filter(SaleLine.sale_id == rec_sale.id).all()
        sold_by_item = {}
        for ln in sale_lines:
            sold_by_item[ln.item_id] = round(float(sold_by_item.get(ln.item_id, 0.0)) + float(ln.qty or 0.0), 4)
        ret = ReturnTxn(
            id=uuid.uuid4().hex,
            user_id=user.id,
            return_type="sale_return",
            reference_type="sale",
            reference_id=rec_sale.id,
            invoice_no=f"SR-{int(dt.datetime.utcnow().timestamp())}",
            return_date=r_date,
            customer_name=rec_sale.customer_name,
            reason=(body.reason or "").strip() or None,
            total_amount=0.0,
        )
        db.add(ret)
        db.flush()
        total = 0.0
        for ln in body.lines:
            sold_qty = round(float(sold_by_item.get(ln.item_id, 0.0)), 4)
            already = _returned_qty(db, user.id, "sale", rec_sale.id, ln.item_id)
            can_return = round(max(0.0, sold_qty - already), 4)
            qty = round(abs(float(ln.qty or 0.0)), 4)
            if qty <= 0:
                continue
            if can_return + 0.0001 < qty:
                raise HTTPException(400, "كمية المرتجع أكبر من المتاح للصنف")
            item = db.query(Item).filter(Item.user_id == user.id, Item.id == ln.item_id).first()
            if not item:
                raise HTTPException(400, "صنف المرتجع غير صالح")
            unit_cost = round(abs(float(ln.unit_cost or _avg_cost(db, user.id, ln.item_id))), 4)
            unit_price = round(abs(float(ln.unit_price or 0.0)), 4)
            line_total = round(qty * unit_price, 2)
            db.add(ReturnLine(return_id=ret.id, item_id=ln.item_id, qty=qty, unit_price=unit_price, unit_cost=unit_cost, line_total=line_total))
            db.add(
                StockMovement(
                    id=uuid.uuid4().hex,
                    user_id=user.id,
                    item_id=ln.item_id,
                    movement_type="sale_return",
                    qty_in=qty,
                    qty_out=0.0,
                    unit_cost=unit_cost,
                    reference_type="sale_return",
                    reference_id=ret.id,
                    movement_date=r_date,
                    notes=ret.reason or None,
                )
            )
            item.quantity = round(float(item.quantity or 0.0) + qty, 4)
            total += line_total
        ret.total_amount = round(total, 2)
        db.commit()
        log_event(db, "trade.sale_return.create", user.id, {"return_id": ret.id, "reference_sale_id": rec_sale.id, "total_amount": ret.total_amount})
        return {"id": ret.id, "total_amount": ret.total_amount}
    finally:
        db.close()


@router.post("/returns/purchase")
def create_purchase_return(request: Request, body: PurchaseReturnCreate = Body(...)):
    db = SessionLocal()
    try:
        require_csrf(request)
        user = require_user(db, request)
        rec_purchase = db.query(Purchase).filter(Purchase.user_id == user.id, Purchase.id == body.purchase_id).first()
        if not rec_purchase:
            raise HTTPException(404, "فاتورة الشراء غير موجودة")
        r_date = _parse_date(body.return_date, "return_date")
        p_lines = db.query(PurchaseLine).filter(PurchaseLine.purchase_id == rec_purchase.id).all()
        bought_by_item = {}
        for ln in p_lines:
            bought_by_item[ln.item_id] = round(float(bought_by_item.get(ln.item_id, 0.0)) + float(ln.qty or 0.0), 4)
        ret = ReturnTxn(
            id=uuid.uuid4().hex,
            user_id=user.id,
            return_type="purchase_return",
            reference_type="purchase",
            reference_id=rec_purchase.id,
            invoice_no=f"PR-{int(dt.datetime.utcnow().timestamp())}",
            return_date=r_date,
            supplier_name=rec_purchase.supplier_name,
            reason=(body.reason or "").strip() or None,
            total_amount=0.0,
        )
        db.add(ret)
        db.flush()
        total = 0.0
        for ln in body.lines:
            bought_qty = round(float(bought_by_item.get(ln.item_id, 0.0)), 4)
            already = _returned_qty(db, user.id, "purchase", rec_purchase.id, ln.item_id)
            can_return = round(max(0.0, bought_qty - already), 4)
            qty = round(abs(float(ln.qty or 0.0)), 4)
            if qty <= 0:
                continue
            if can_return + 0.0001 < qty:
                raise HTTPException(400, "كمية مرتجع الشراء أكبر من المتاح")
            item = db.query(Item).filter(Item.user_id == user.id, Item.id == ln.item_id).first()
            if not item:
                raise HTTPException(400, "صنف المرتجع غير صالح")
            available = round(float(item.quantity or 0.0), 4)
            if available + 0.0001 < qty:
                raise HTTPException(400, "المخزون الحالي لا يسمح بمرتجع الشراء")
            unit_cost = round(abs(float(ln.unit_cost or item.last_cost or 0.0)), 4)
            unit_price = round(abs(float(ln.unit_price or unit_cost)), 4)
            line_total = round(qty * unit_price, 2)
            db.add(ReturnLine(return_id=ret.id, item_id=ln.item_id, qty=qty, unit_price=unit_price, unit_cost=unit_cost, line_total=line_total))
            db.add(
                StockMovement(
                    id=uuid.uuid4().hex,
                    user_id=user.id,
                    item_id=ln.item_id,
                    movement_type="purchase_return",
                    qty_in=0.0,
                    qty_out=qty,
                    unit_cost=unit_cost,
                    reference_type="purchase_return",
                    reference_id=ret.id,
                    movement_date=r_date,
                    notes=ret.reason or None,
                )
            )
            item.quantity = round(available - qty, 4)
            total += line_total
        ret.total_amount = round(total, 2)
        db.commit()
        log_event(db, "trade.purchase_return.create", user.id, {"return_id": ret.id, "reference_purchase_id": rec_purchase.id, "total_amount": ret.total_amount})
        return {"id": ret.id, "total_amount": ret.total_amount}
    finally:
        db.close()


@router.put("/sales/{sale_id}")
def update_sale(sale_id: str, request: Request, body: SaleUpdate = Body(...)):
    db = SessionLocal()
    try:
        require_csrf(request)
        user = require_user(db, request)
        sale = db.query(Sale).filter(Sale.user_id == user.id, Sale.id == sale_id).first()
        if not sale:
            raise HTTPException(404, "فاتورة البيع غير موجودة")
        s_date = _parse_date(body.sale_date, "sale_date")
        if not body.lines:
            raise HTTPException(400, "فاتورة البيع تحتاج بنوداً")

        _reverse_sale_effects(db, user.id, sale.id)

        item_ids = {x.item_id for x in body.lines}
        items = db.query(Item).filter(Item.user_id == user.id, Item.id.in_(list(item_ids))).all()
        if len(items) != len(item_ids):
            raise HTTPException(400, "يوجد صنف غير صالح")
        item_by_id = {x.id: x for x in items}

        pay_type = (body.payment_type or "cash").strip().lower()
        customer_name = body.customer_name.strip()
        if pay_type == "credit":
            if not customer_name and not (body.customer_id or "").strip():
                raise HTTPException(400, "في البيع الآجل يجب تحديد عميل")
            if customer_name in {"عميل نقدي", "cash customer"} and not (body.customer_id or "").strip():
                raise HTTPException(400, "في البيع الآجل يجب إدخال اسم عميل صحيح")
        elif not customer_name:
            customer_name = "عميل نقدي"

        sale.invoice_no = body.invoice_no.strip()
        sale.customer_name = customer_name
        sale.customer_tax_no = (body.customer_tax_no or "").strip() or None
        sale.customer_phone = (body.customer_phone or "").strip() or None
        sale.customer_address = (body.customer_address or "").strip() or None
        if (body.customer_id or "").strip():
            c = db.query(Customer).filter(Customer.user_id == user.id, Customer.id == body.customer_id.strip()).first()
            if c:
                sale.customer_name = c.name
                sale.customer_id = c.id
                sale.customer_phone = (c.phone or "").strip() or None
                sale.customer_address = (c.address or "").strip() or None
        sale.sale_date = s_date
        sale.payment_type = pay_type
        sale.discount = round(abs(float(body.discount or 0.0)), 2)
        sale.paid_amount = round(abs(float(body.paid_amount or 0.0)), 2)
        sale.notes = (body.notes or "").strip() or None
        sale.seller_name = (body.seller_name or "").strip() or None
        sale.branch_name = (body.branch_name or "").strip() or None
        sale.branch_id = (body.branch_id or "").strip() or None

        total = 0.0
        for ln in body.lines:
            qty = round(abs(float(ln.qty or 0.0)), 4)
            sale_price = round(abs(float(ln.sale_price or 0.0)), 4)
            tax = round(abs(float(ln.tax_amount or 0.0)), 2)
            if qty <= 0:
                continue
            item = item_by_id[ln.item_id]
            available = round(float(item.quantity or 0.0), 4)
            allow_negative_stock = bool(body.allow_negative_stock and _can_override_stock(user))
            if available + 0.0001 < qty and not allow_negative_stock:
                raise HTTPException(400, f"المخزون غير كافٍ للصنف: {item.name}")
            unit_cost = _avg_cost(db, user.id, ln.item_id)
            cost_total = round(unit_cost * qty, 2)
            line_total = round(sale_price * qty + tax, 2)
            profit = round((sale_price * qty) - cost_total, 2)
            db.add(
                SaleLine(
                    sale_id=sale.id,
                    item_id=ln.item_id,
                    qty=qty,
                    sale_price=sale_price,
                    tax_amount=tax,
                    cost_price=cost_total,
                    profit=profit,
                )
            )
            db.add(
                StockMovement(
                    id=uuid.uuid4().hex,
                    user_id=user.id,
                    item_id=ln.item_id,
                    movement_type="sale",
                    qty_in=0.0,
                    qty_out=qty,
                    unit_cost=unit_cost,
                    reference_type="sale",
                    reference_id=sale.id,
                    movement_date=s_date,
                )
            )
            item.quantity = round(available - qty, 4)
            total += line_total
        sale.total_amount = round(max(0.0, total - sale.discount), 2)
        sale.tax_amount = round(sum(float(x.tax_amount or 0.0) for x in db.query(SaleLine).filter(SaleLine.sale_id == sale.id).all()), 2)
        sale.due_amount = round(max(0.0, sale.total_amount - sale.paid_amount), 2)
        db.commit()
        log_event(db, "trade.sale.update", user.id, {"sale_id": sale.id, "invoice_no": sale.invoice_no, "total_amount": sale.total_amount})
        return {"id": sale.id, "total_amount": sale.total_amount, "updated": True}
    finally:
        db.close()


@router.delete("/sales/{sale_id}")
def delete_sale(sale_id: str, request: Request):
    db = SessionLocal()
    try:
        require_csrf(request)
        user = require_user(db, request)
        sale = db.query(Sale).filter(Sale.user_id == user.id, Sale.id == sale_id).first()
        if not sale:
            raise HTTPException(404, "فاتورة البيع غير موجودة")
        _reverse_sale_effects(db, user.id, sale.id)
        db.delete(sale)
        db.commit()
        log_event(db, "trade.sale.delete", user.id, {"sale_id": sale_id, "invoice_no": sale.invoice_no})
        return {"deleted": True}
    finally:
        db.close()


@router.get("/inventory")
def inventory_report(request: Request):
    db = SessionLocal()
    try:
        user = require_user(db, request)
        rows = db.query(Item).filter(Item.user_id == user.id).order_by(Item.code.asc(), Item.created_at.asc()).all()
        out = []
        for x in rows:
            qty = round(float(x.quantity or 0.0), 4)
            cost = round(float(x.last_cost or 0.0), 2)
            out.append(
                {
                    "id": x.id,
                    "code": x.code,
                    "name": x.name,
                    "category": x.category,
                    "item_condition": x.item_condition or "",
                    "location": x.location or "",
                    "quantity": qty,
                    "cost_price": cost,
                    "sale_price": round(float(x.default_sale_price or 0.0), 2),
                    "stock_value": round(qty * cost, 2),
                }
            )
        return {"items": out}
    finally:
        db.close()


@router.post("/inventory/adjust")
def inventory_adjust(request: Request, body: StockAdjustCreate = Body(...)):
    db = SessionLocal()
    try:
        require_csrf(request)
        user = require_user(db, request)
        item = db.query(Item).filter(Item.user_id == user.id, Item.id == body.item_id).first()
        if not item:
            raise HTTPException(404, "الصنف غير موجود")
        adj_date = _parse_date(body.adjust_date, "adjust_date")
        before = round(float(item.quantity or 0.0), 4)
        after = round(float(body.qty_after or 0.0), 4)
        diff = round(after - before, 4)
        if abs(diff) < 0.0001:
            raise HTTPException(400, "لا يوجد فرق لتسجيله")
        item.quantity = after
        db.add(
            StockAdjustment(
                id=uuid.uuid4().hex,
                user_id=user.id,
                item_id=item.id,
                branch_id=item.branch_id,
                adjust_date=adj_date,
                qty_before=before,
                qty_after=after,
                difference=diff,
                reason=(body.reason or "").strip(),
            )
        )
        db.add(
            StockMovement(
                id=uuid.uuid4().hex,
                user_id=user.id,
                item_id=item.id,
                movement_type="adjust",
                qty_in=diff if diff > 0 else 0.0,
                qty_out=abs(diff) if diff < 0 else 0.0,
                unit_cost=round(float(item.last_cost or 0.0), 4),
                reference_type="adjust",
                reference_id=item.id,
                movement_date=adj_date,
                notes=(body.reason or "").strip() or None,
            )
        )
        db.commit()
        log_event(db, "trade.inventory.adjust", user.id, {"item_id": item.id, "qty_before": before, "qty_after": after, "difference": diff})
        return {"ok": True, "qty_before": before, "qty_after": after, "difference": diff}
    finally:
        db.close()


@router.get("/reports/profit")
def profit_report(request: Request, date_from: str = Query(""), date_to: str = Query("")):
    db = SessionLocal()
    try:
        user = require_user(db, request)
        df = _parse_date(date_from, "date_from") if (date_from or "").strip() else None
        dt_to = _parse_date(date_to, "date_to") if (date_to or "").strip() else None
        q = db.query(Sale, SaleLine, Item).join(SaleLine, SaleLine.sale_id == Sale.id).join(Item, Item.id == SaleLine.item_id).filter(Sale.user_id == user.id, Item.user_id == user.id)
        if df:
            q = q.filter(Sale.sale_date >= df)
        if dt_to:
            q = q.filter(Sale.sale_date <= dt_to)
        rows = q.all()
        items = []
        sales_total = 0.0
        cost_total = 0.0
        profit_total = 0.0
        for sale, ln, item in rows:
            net = round(float(ln.sale_price or 0.0) * float(ln.qty or 0.0), 2)
            cost = round(float(ln.cost_price or 0.0), 2)
            pr = round(float(ln.profit or 0.0), 2)
            sales_total += net
            cost_total += cost
            profit_total += pr
            items.append(
                {
                    "date": sale.sale_date.strftime("%Y-%m-%d"),
                    "invoice_no": sale.invoice_no,
                    "item": item.name,
                    "qty": round(float(ln.qty or 0.0), 4),
                    "net_sales": net,
                    "cost_total": cost,
                    "profit": pr,
                }
            )
        returns_q = db.query(ReturnTxn, ReturnLine).join(ReturnLine, ReturnLine.return_id == ReturnTxn.id).filter(ReturnTxn.user_id == user.id, ReturnTxn.return_type == "sale_return")
        if df:
            returns_q = returns_q.filter(ReturnTxn.return_date >= df)
        if dt_to:
            returns_q = returns_q.filter(ReturnTxn.return_date <= dt_to)
        return_rows = returns_q.all()
        returned_sales = round(sum(float(ln.line_total or 0.0) for _, ln in return_rows), 2)
        returned_cost = round(sum(float(ln.unit_cost or 0.0) * float(ln.qty or 0.0) for _, ln in return_rows), 2)
        sales_total = round(max(0.0, sales_total - returned_sales), 2)
        cost_total = round(max(0.0, cost_total - returned_cost), 2)
        profit_total = round(sales_total - cost_total, 2)
        margin = round((profit_total / sales_total) * 100.0, 2) if sales_total > 0 else 0.0
        return {
            "items": items,
            "totals": {
                "net_sales": round(sales_total, 2),
                "cost_total": round(cost_total, 2),
                "profit": round(profit_total, 2),
                "returned_sales": returned_sales,
                "returned_cost": returned_cost,
                "profit_margin_pct": margin,
            },
        }
    finally:
        db.close()


@router.get("/reports/top-items")
def top_items_report(request: Request, date_from: str = Query(""), date_to: str = Query(""), limit: int = Query(10, ge=1, le=100)):
    db = SessionLocal()
    try:
        user = require_user(db, request)
        df = _parse_date(date_from, "date_from") if (date_from or "").strip() else None
        dt_to = _parse_date(date_to, "date_to") if (date_to or "").strip() else None
        q = (
            db.query(Sale, SaleLine, Item)
            .join(SaleLine, SaleLine.sale_id == Sale.id)
            .join(Item, Item.id == SaleLine.item_id)
            .filter(Sale.user_id == user.id, Item.user_id == user.id)
        )
        if df:
            q = q.filter(Sale.sale_date >= df)
        if dt_to:
            q = q.filter(Sale.sale_date <= dt_to)
        rows = q.all()
        agg: dict[str, dict[str, Any]] = {}
        for _, ln, item in rows:
            cur = agg.setdefault(item.id, {"item_id": item.id, "item": item.name, "qty": 0.0, "sales": 0.0, "profit": 0.0})
            qty = round(float(ln.qty or 0.0), 4)
            sales = round(float(ln.sale_price or 0.0) * qty, 2)
            cur["qty"] = round(float(cur["qty"]) + qty, 4)
            cur["sales"] = round(float(cur["sales"]) + sales, 2)
            cur["profit"] = round(float(cur["profit"]) + float(ln.profit or 0.0), 2)
        data = list(agg.values())
        best = sorted(data, key=lambda x: (float(x["qty"]), float(x["sales"])), reverse=True)[:limit]
        worst = sorted(data, key=lambda x: (float(x["qty"]), float(x["sales"])))[:limit]
        return {"best_selling": best, "worst_selling": worst}
    finally:
        db.close()


@router.get("/reports/profit-by-item")
def profit_by_item_report(request: Request, date_from: str = Query(""), date_to: str = Query("")):
    db = SessionLocal()
    try:
        user = require_user(db, request)
        df = _parse_date(date_from, "date_from") if (date_from or "").strip() else None
        dt_to = _parse_date(date_to, "date_to") if (date_to or "").strip() else None
        q = (
            db.query(Sale, SaleLine, Item)
            .join(SaleLine, SaleLine.sale_id == Sale.id)
            .join(Item, Item.id == SaleLine.item_id)
            .filter(Sale.user_id == user.id, Item.user_id == user.id)
        )
        if df:
            q = q.filter(Sale.sale_date >= df)
        if dt_to:
            q = q.filter(Sale.sale_date <= dt_to)
        rows = q.all()
        agg: dict[str, dict[str, Any]] = {}
        for _, ln, item in rows:
            cur = agg.setdefault(item.id, {"item_id": item.id, "item": item.name, "qty": 0.0, "net_sales": 0.0, "cost_total": 0.0, "profit": 0.0})
            qty = round(float(ln.qty or 0.0), 4)
            net_sales = round(float(ln.sale_price or 0.0) * qty, 2)
            cost_total = round(float(ln.cost_price or 0.0), 2)
            cur["qty"] = round(float(cur["qty"]) + qty, 4)
            cur["net_sales"] = round(float(cur["net_sales"]) + net_sales, 2)
            cur["cost_total"] = round(float(cur["cost_total"]) + cost_total, 2)
            cur["profit"] = round(float(cur["profit"]) + float(ln.profit or 0.0), 2)
        items = sorted(list(agg.values()), key=lambda x: float(x["profit"]), reverse=True)
        return {"items": items}
    finally:
        db.close()


@router.get("/reports/stock-status")
def stock_status_report(request: Request, slow_days: int = Query(45, ge=1, le=365)):
    db = SessionLocal()
    try:
        user = require_user(db, request)
        items = db.query(Item).filter(Item.user_id == user.id).order_by(Item.code.asc(), Item.created_at.asc()).all()
        now = dt.datetime.utcnow()
        low = []
        out = []
        slow = []
        for x in items:
            qty = round(float(x.quantity or 0.0), 4)
            min_qty = round(float(x.min_qty or 0.0), 4)
            row = {
                "id": x.id,
                "code": x.code,
                "name": x.name,
                "quantity": qty,
                "min_qty": min_qty,
                "last_cost": round(float(x.last_cost or 0.0), 2),
                "sale_price": round(float(x.default_sale_price or 0.0), 2),
            }
            if qty <= 0:
                out.append(row)
            if qty <= min_qty:
                low.append(row)
            last_sale = (
                db.query(StockMovement)
                .filter(StockMovement.user_id == user.id, StockMovement.item_id == x.id, StockMovement.movement_type == "sale")
                .order_by(StockMovement.movement_date.desc(), StockMovement.created_at.desc())
                .first()
            )
            if not last_sale or (now - last_sale.movement_date).days >= slow_days:
                row["last_sale_date"] = last_sale.movement_date.strftime("%Y-%m-%d") if last_sale and last_sale.movement_date else ""
                slow.append(row)
        return {"low_stock": low, "out_of_stock": out, "slow_moving": slow}
    finally:
        db.close()


@router.get("/branch-transfers")
def list_branch_transfers(request: Request, limit: int = Query(200, ge=1, le=1000)):
    db = SessionLocal()
    try:
        user = require_user(db, request)
        rows = (
            db.query(BranchTransfer)
            .filter(BranchTransfer.user_id == user.id)
            .order_by(BranchTransfer.transfer_date.desc(), BranchTransfer.created_at.desc())
            .limit(limit)
            .all()
        )
        items = []
        for t in rows:
            fb = db.query(Branch).filter(Branch.id == t.from_branch_id).first()
            tb = db.query(Branch).filter(Branch.id == t.to_branch_id).first()
            items.append(
                {
                    "id": t.id,
                    "transfer_no": t.transfer_no,
                    "transfer_date": t.transfer_date.strftime("%Y-%m-%d"),
                    "from_branch": (fb.name if fb else ""),
                    "to_branch": (tb.name if tb else ""),
                    "notes": t.notes or "",
                }
            )
        return {"items": items}
    finally:
        db.close()


@router.post("/branch-transfers")
def create_branch_transfer(request: Request, body: BranchTransferCreate = Body(...)):
    db = SessionLocal()
    try:
        require_csrf(request)
        user = require_user(db, request)
        if not body.lines:
            raise HTTPException(400, "سند التحويل يحتاج بنوداً")
        t_date = _parse_date(body.transfer_date, "transfer_date")
        if body.from_branch_id == body.to_branch_id:
            raise HTTPException(400, "لا يمكن التحويل لنفس الفرع")
        from_b = db.query(Branch).filter(Branch.user_id == user.id, Branch.id == body.from_branch_id).first()
        to_b = db.query(Branch).filter(Branch.user_id == user.id, Branch.id == body.to_branch_id).first()
        if not from_b or not to_b:
            raise HTTPException(400, "بيانات الفرع غير صحيحة")
        rec = BranchTransfer(
            id=uuid.uuid4().hex,
            user_id=user.id,
            transfer_no=body.transfer_no.strip(),
            transfer_date=t_date,
            from_branch_id=from_b.id,
            to_branch_id=to_b.id,
            notes=(body.notes or "").strip() or None,
        )
        db.add(rec)
        db.flush()
        for ln in body.lines:
            qty = round(abs(float(ln.qty or 0.0)), 4)
            if qty <= 0:
                continue
            from_item = db.query(Item).filter(Item.user_id == user.id, Item.id == ln.from_item_id).first()
            to_item = db.query(Item).filter(Item.user_id == user.id, Item.id == ln.to_item_id).first()
            if not from_item or not to_item:
                raise HTTPException(400, "أحد الأصناف في التحويل غير صالح")
            if from_item.branch_id != from_b.id or to_item.branch_id != to_b.id:
                raise HTTPException(400, "الصنف غير مرتبط بالفرع المختار")
            available = round(float(from_item.quantity or 0.0), 4)
            if available + 0.0001 < qty:
                raise HTTPException(400, f"المخزون غير كافٍ للصنف: {from_item.name}")
            from_item.quantity = round(available - qty, 4)
            to_item.quantity = round(float(to_item.quantity or 0.0) + qty, 4)
            unit_cost = round(float(from_item.last_cost or 0.0), 4)
            db.add(BranchTransferLine(transfer_id=rec.id, from_item_id=from_item.id, to_item_id=to_item.id, qty=qty))
            db.add(
                StockMovement(
                    id=uuid.uuid4().hex,
                    user_id=user.id,
                    item_id=from_item.id,
                    movement_type="transfer_out",
                    qty_in=0.0,
                    qty_out=qty,
                    unit_cost=unit_cost,
                    reference_type="branch_transfer",
                    reference_id=rec.id,
                    movement_date=t_date,
                    notes=rec.notes or f"تحويل إلى {to_b.name}",
                )
            )
            db.add(
                StockMovement(
                    id=uuid.uuid4().hex,
                    user_id=user.id,
                    item_id=to_item.id,
                    movement_type="transfer_in",
                    qty_in=qty,
                    qty_out=0.0,
                    unit_cost=unit_cost,
                    reference_type="branch_transfer",
                    reference_id=rec.id,
                    movement_date=t_date,
                    notes=rec.notes or f"تحويل من {from_b.name}",
                )
            )
        db.commit()
        return {"id": rec.id, "transfer_no": rec.transfer_no}
    finally:
        db.close()


@router.get("/reports/item-movement")
def item_movement_report(
    request: Request,
    item_id: str = Query(...),
    date_from: str = Query(""),
    date_to: str = Query(""),
):
    db = SessionLocal()
    try:
        user = require_user(db, request)
        item = db.query(Item).filter(Item.user_id == user.id, Item.id == item_id).first()
        if not item:
            raise HTTPException(404, "الصنف غير موجود")
        df = _parse_date(date_from, "date_from") if (date_from or "").strip() else None
        dt_to = _parse_date(date_to, "date_to") if (date_to or "").strip() else None
        q = db.query(StockMovement).filter(StockMovement.user_id == user.id, StockMovement.item_id == item_id)
        if df:
            q = q.filter(StockMovement.movement_date >= df)
        if dt_to:
            q = q.filter(StockMovement.movement_date <= dt_to)
        rows = q.order_by(StockMovement.movement_date.asc(), StockMovement.created_at.asc()).all()
        items = []
        qty_in = 0.0
        qty_out = 0.0
        running_balance = round(float(item.quantity or 0.0) - (sum(float(r.qty_in or 0.0) - float(r.qty_out or 0.0) for r in rows)), 4)
        for x in rows:
            qi = round(float(x.qty_in or 0.0), 4)
            qo = round(float(x.qty_out or 0.0), 4)
            qty_in += qi
            qty_out += qo
            running_balance = round(running_balance + qi - qo, 4)
            items.append(
                {
                    "date": x.movement_date.strftime("%Y-%m-%d"),
                    "movement_type": x.movement_type,
                    "qty_in": qi,
                    "qty_out": qo,
                    "unit_cost": round(float(x.unit_cost or 0.0), 2),
                    "reference_type": x.reference_type or "",
                    "reference_id": x.reference_id or "",
                    "balance_after": running_balance,
                }
            )
        return {
            "item": {"id": item.id, "code": item.code, "name": item.name},
            "items": items,
            "summary": {
                "opening_qty": round(float(item.quantity or 0.0) - (qty_in - qty_out), 4),
                "qty_in": round(qty_in, 4),
                "qty_out": round(qty_out, 4),
                "balance_qty": round(float(item.quantity or 0.0), 4),
            },
        }
    finally:
        db.close()


@router.get("/reports/tax-return")
def tax_return_report(request: Request, date_from: str = Query(""), date_to: str = Query("")):
    db = SessionLocal()
    try:
        user = require_user(db, request)
        df = _parse_date(date_from, "date_from") if (date_from or "").strip() else None
        dt_to = _parse_date(date_to, "date_to") if (date_to or "").strip() else None

        sales_q = db.query(Sale).filter(Sale.user_id == user.id)
        purchases_q = db.query(Purchase).filter(Purchase.user_id == user.id)
        sale_ret_q = db.query(ReturnTxn).filter(ReturnTxn.user_id == user.id, ReturnTxn.return_type == "sale_return")
        purchase_ret_q = db.query(ReturnTxn).filter(ReturnTxn.user_id == user.id, ReturnTxn.return_type == "purchase_return")
        if df:
            sales_q = sales_q.filter(Sale.sale_date >= df)
            purchases_q = purchases_q.filter(Purchase.purchase_date >= df)
            sale_ret_q = sale_ret_q.filter(ReturnTxn.return_date >= df)
            purchase_ret_q = purchase_ret_q.filter(ReturnTxn.return_date >= df)
        if dt_to:
            sales_q = sales_q.filter(Sale.sale_date <= dt_to)
            purchases_q = purchases_q.filter(Purchase.purchase_date <= dt_to)
            sale_ret_q = sale_ret_q.filter(ReturnTxn.return_date <= dt_to)
            purchase_ret_q = purchase_ret_q.filter(ReturnTxn.return_date <= dt_to)

        sales_rows = sales_q.all()
        purchase_rows = purchases_q.all()
        sale_returns = sale_ret_q.all()
        purchase_returns = purchase_ret_q.all()

        sales_taxable_before = round(sum(max(0.0, float(x.total_amount or 0.0) - float(x.tax_amount or 0.0)) for x in sales_rows), 2)
        sales_tax = round(sum(float(x.tax_amount or 0.0) for x in sales_rows), 2)
        sales_total_with_tax = round(sum(float(x.total_amount or 0.0) for x in sales_rows), 2)

        purchases_taxable_before = round(sum(max(0.0, float(x.total_amount or 0.0) - float(x.tax_amount or 0.0)) for x in purchase_rows), 2)
        purchases_tax = round(sum(float(x.tax_amount or 0.0) for x in purchase_rows), 2)
        purchases_total_with_tax = round(sum(float(x.total_amount or 0.0) for x in purchase_rows), 2)

        sales_by_id = {x.id: x for x in sales_rows}
        purchases_by_id = {x.id: x for x in purchase_rows}

        sale_ret_total = 0.0
        sale_ret_tax = 0.0
        for r in sale_returns:
            amt = float(r.total_amount or 0.0)
            sale_ret_total += amt
            ref = sales_by_id.get(r.reference_id)
            if ref:
                ratio = _safe_div(float(ref.tax_amount or 0.0), max(0.0, float(ref.total_amount or 0.0)))
                sale_ret_tax += round(amt * ratio, 2)

        purchase_ret_total = 0.0
        purchase_ret_tax = 0.0
        for r in purchase_returns:
            amt = float(r.total_amount or 0.0)
            purchase_ret_total += amt
            ref = purchases_by_id.get(r.reference_id)
            if ref:
                ratio = _safe_div(float(ref.tax_amount or 0.0), max(0.0, float(ref.total_amount or 0.0)))
                purchase_ret_tax += round(amt * ratio, 2)

        sale_ret_total = round(sale_ret_total, 2)
        sale_ret_tax = round(sale_ret_tax, 2)
        sale_ret_before = round(max(0.0, sale_ret_total - sale_ret_tax), 2)

        purchase_ret_total = round(purchase_ret_total, 2)
        purchase_ret_tax = round(purchase_ret_tax, 2)
        purchase_ret_before = round(max(0.0, purchase_ret_total - purchase_ret_tax), 2)

        net_sales_before = round(max(0.0, sales_taxable_before - sale_ret_before), 2)
        net_sales_tax = round(max(0.0, sales_tax - sale_ret_tax), 2)

        net_purchases_before = round(max(0.0, purchases_taxable_before - purchase_ret_before), 2)
        net_purchases_tax = round(max(0.0, purchases_tax - purchase_ret_tax), 2)

        net_tax_due = round(net_sales_tax - net_purchases_tax, 2)

        settings_row = db.query(AppSetting).filter(AppSetting.key == f"cashierko_settings:{user.id}").first()
        settings = settings_row.value_json if settings_row and isinstance(settings_row.value_json, dict) else {}

        return {
            "company": {
                "shop_name": settings.get("shop_name", "SmartPOS"),
                "tax_number": settings.get("tax_number", ""),
                "address": settings.get("address", ""),
            },
            "period": {"date_from": _fmt_date(df) if df else "", "date_to": _fmt_date(dt_to) if dt_to else ""},
            "output_vat": {
                "sales_before_tax": sales_taxable_before,
                "sales_tax": sales_tax,
                "sales_with_tax": sales_total_with_tax,
                "returns_before_tax": sale_ret_before,
                "returns_tax": sale_ret_tax,
                "net_sales_before_tax": net_sales_before,
                "net_sales_tax": net_sales_tax,
            },
            "input_vat": {
                "purchases_before_tax": purchases_taxable_before,
                "purchases_tax": purchases_tax,
                "purchases_with_tax": purchases_total_with_tax,
                "returns_before_tax": purchase_ret_before,
                "returns_tax": purchase_ret_tax,
                "net_purchases_before_tax": net_purchases_before,
                "net_purchases_tax": net_purchases_tax,
            },
            "net_vat": {
                "net_tax_due": net_tax_due,
                "status": "payable" if net_tax_due >= 0 else "credit",
                "payable_amount": net_tax_due if net_tax_due >= 0 else 0.0,
                "credit_amount": abs(net_tax_due) if net_tax_due < 0 else 0.0,
            },
            "generated_at": dt.datetime.utcnow().strftime("%Y-%m-%d %H:%M:%S"),
        }
    finally:
        db.close()


@router.get("/dashboard-summary")
def dashboard_summary(request: Request, date_from: str = Query(""), date_to: str = Query("")):
    db = SessionLocal()
    try:
        user = require_user(db, request)
        df = _parse_date(date_from, "date_from") if (date_from or "").strip() else None
        dt_to = _parse_date(date_to, "date_to") if (date_to or "").strip() else None

        sales_q = db.query(Sale).filter(Sale.user_id == user.id)
        purchases_q = db.query(Purchase).filter(Purchase.user_id == user.id)
        if df:
            sales_q = sales_q.filter(Sale.sale_date >= df)
            purchases_q = purchases_q.filter(Purchase.purchase_date >= df)
        if dt_to:
            sales_q = sales_q.filter(Sale.sale_date <= dt_to)
            purchases_q = purchases_q.filter(Purchase.purchase_date <= dt_to)

        sales_rows = sales_q.all()
        purchase_rows = purchases_q.all()
        items = db.query(Item).filter(Item.user_id == user.id).all()

        total_sales = round(sum(float(x.total_amount or 0.0) for x in sales_rows), 2)
        total_purchases = round(sum(float(x.total_amount or 0.0) for x in purchase_rows), 2)
        sales_tax = round(sum(float(x.tax_amount or 0.0) for x in sales_rows), 2)
        purchases_tax = round(sum(float(x.tax_amount or 0.0) for x in purchase_rows), 2)
        tax_due = round(sales_tax - purchases_tax, 2)
        total_inventory_value = round(
            sum(round(float(x.quantity or 0.0), 4) * round(float(x.last_cost or 0.0), 4) for x in items),
            2,
        )
        total_items_qty = round(sum(float(x.quantity or 0.0) for x in items), 4)

        low_stock_items = []
        out_of_stock_count = 0
        for x in items:
            qty = round(float(x.quantity or 0.0), 4)
            min_qty = round(float(x.min_qty or 0.0), 4)
            if qty <= 0:
                out_of_stock_count += 1
            if qty <= min_qty:
                low_stock_items.append({"id": x.id, "code": x.code, "name": x.name, "quantity": qty, "min_qty": min_qty})

        sale_ids = [s.id for s in sales_rows]
        profit_total = 0.0
        if sale_ids:
            lines = db.query(SaleLine).filter(SaleLine.sale_id.in_(sale_ids)).all()
            profit_total = round(sum(float(ln.profit or 0.0) for ln in lines), 2)

        top_q = (
            db.query(Sale, SaleLine, Item)
            .join(SaleLine, SaleLine.sale_id == Sale.id)
            .join(Item, Item.id == SaleLine.item_id)
            .filter(Sale.user_id == user.id, Item.user_id == user.id)
        )
        if df:
            top_q = top_q.filter(Sale.sale_date >= df)
        if dt_to:
            top_q = top_q.filter(Sale.sale_date <= dt_to)
        agg: dict[str, dict[str, Any]] = {}
        for _, ln, item in top_q.all():
            cur = agg.setdefault(item.id, {"item": item.name, "code": item.code, "qty": 0.0, "sales": 0.0})
            qty = round(float(ln.qty or 0.0), 4)
            cur["qty"] = round(float(cur["qty"]) + qty, 4)
            cur["sales"] = round(float(cur["sales"]) + round(float(ln.sale_price or 0.0) * qty, 2), 2)
        top_items = sorted(agg.values(), key=lambda x: (float(x["qty"]), float(x["sales"])), reverse=True)[:5]

        latest_invoices = [
            {
                "id": s.id,
                "invoice_no": s.invoice_no,
                "customer_name": s.customer_name,
                "date": s.sale_date.strftime("%Y-%m-%d") if s.sale_date else "",
                "total_amount": round(float(s.total_amount or 0.0), 2),
            }
            for s in sorted(sales_rows, key=lambda x: x.created_at or x.sale_date, reverse=True)[:5]
        ]

        return {
            "totals": {
                "sales": total_sales,
                "purchases": total_purchases,
                "inventory_value": total_inventory_value,
                "inventory_qty": total_items_qty,
                "item_count": len(items),
                "invoice_count": len(sales_rows),
                "profit": profit_total,
                "tax_due": tax_due,
                "low_stock_count": len(low_stock_items),
                "out_of_stock_count": out_of_stock_count,
            },
            "low_stock_items": low_stock_items[:10],
            "top_items": top_items,
            "latest_invoices": latest_invoices,
        }
    finally:
        db.close()


# ============================================================
# استيراد وتصدير Excel — الأصناف، المشتريات، المبيعات، التقارير
# ============================================================

def _xlsx_response(wb: Workbook, filename: str) -> StreamingResponse:
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    headers = {"Content-Disposition": f'attachment; filename="{filename}"'}
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers=headers,
    )


def _new_item_code() -> str:
    return "ITM-" + uuid.uuid4().hex[:10].upper()


def _find_header_index(header: list[str], *names: str) -> int:
    for n in names:
        for i, h in enumerate(header):
            if h == n:
                return i
    return -1


class StocktakeApplyLine(BaseModel):
    item_id: str
    counted_qty: float = Field(ge=0)


class StocktakeApply(BaseModel):
    reason: str = ""
    lines: list[StocktakeApplyLine]


@router.get("/stocktake/template")
def stocktake_template(request: Request):
    db = SessionLocal()
    try:
        user = require_user(db, request)
        rows = db.query(Item).filter(Item.user_id == user.id, Item.is_active == 1).order_by(Item.code.asc()).all()
        wb = Workbook()
        ws = wb.active
        ws.title = "الجرد"
        ws.append(["الكود", "اسم الصنف (للعرض فقط)", "الكمية بالنظام (للعرض فقط)", "الكمية الفعلية (عدّها هنا)"])
        for r in rows:
            ws.append([r.code, r.name, round(float(r.quantity or 0.0), 4), ""])
        return _xlsx_response(wb, "stocktake_template.xlsx")
    finally:
        db.close()


@router.post("/stocktake/preview")
async def stocktake_preview(request: Request, file: UploadFile = File(...)):
    db = SessionLocal()
    try:
        user = require_user(db, request)
        content = await file.read()
        wb = load_workbook(io.BytesIO(content), data_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            raise HTTPException(400, "الملف فارغ")
        header = [str(c or "").strip() for c in rows[0]]
        idx_code = _find_header_index(header, "الكود", "code")
        idx_counted = _find_header_index(header, "الكمية الفعلية (عدّها هنا)", "الكمية الفعلية", "counted_qty")
        if idx_code == -1 or idx_counted == -1:
            raise HTTPException(400, "لم يتم العثور على أعمدة الكود/الكمية الفعلية في الملف")
        items_by_code = {x.code: x for x in db.query(Item).filter(Item.user_id == user.id).all()}
        out = []
        errors = []
        for rn, row in enumerate(rows[1:], start=2):
            code = str(row[idx_code] or "").strip() if idx_code < len(row) and row[idx_code] is not None else ""
            if not code:
                continue
            if idx_counted >= len(row) or row[idx_counted] is None or str(row[idx_counted]).strip() == "":
                continue
            item = items_by_code.get(code)
            if not item:
                errors.append(f"صف {rn}: الكود {code} غير موجود")
                continue
            try:
                counted = round(abs(float(row[idx_counted])), 4)
            except (TypeError, ValueError):
                errors.append(f"صف {rn}: قيمة كمية غير صالحة")
                continue
            system_qty = round(float(item.quantity or 0.0), 4)
            out.append({
                "item_id": item.id,
                "code": item.code,
                "name": item.name,
                "system_qty": system_qty,
                "counted_qty": counted,
                "difference": round(counted - system_qty, 4),
            })
        return {"rows": out, "errors": errors}
    finally:
        db.close()


@router.post("/stocktake/apply")
def stocktake_apply(request: Request, body: StocktakeApply = Body(...)):
    db = SessionLocal()
    try:
        require_csrf(request)
        user = require_user(db, request)
        if not body.lines:
            raise HTTPException(400, "لا توجد بنود لتطبيقها")
        now = dt.datetime.utcnow()
        reason = (body.reason or "").strip() or "جرد مخزون"
        applied = 0
        skipped = 0
        results = []
        for ln in body.lines:
            item = db.query(Item).filter(Item.user_id == user.id, Item.id == ln.item_id).first()
            if not item:
                skipped += 1
                continue
            before = round(float(item.quantity or 0.0), 4)
            after = round(abs(float(ln.counted_qty or 0.0)), 4)
            diff = round(after - before, 4)
            if abs(diff) < 0.0001:
                skipped += 1
                continue
            item.quantity = after
            db.add(StockAdjustment(id=uuid.uuid4().hex, user_id=user.id, item_id=item.id, branch_id=item.branch_id, adjust_date=now, qty_before=before, qty_after=after, difference=diff, reason=reason))
            db.add(StockMovement(id=uuid.uuid4().hex, user_id=user.id, item_id=item.id, movement_type="adjust", qty_in=diff if diff > 0 else 0.0, qty_out=abs(diff) if diff < 0 else 0.0, unit_cost=round(float(item.last_cost or 0.0), 4), reference_type="stocktake", reference_id=item.id, movement_date=now, notes=reason))
            applied += 1
            results.append({"item_id": item.id, "code": item.code, "qty_before": before, "qty_after": after, "difference": diff})
        db.commit()
        log_event(db, "trade.stocktake.apply", user.id, {"applied": applied, "skipped": skipped})
        return {"applied": applied, "skipped": skipped, "results": results}
    finally:
        db.close()


@router.get("/items/export")
def export_items(request: Request, template: str = Query("full")):
    db = SessionLocal()
    try:
        user = require_user(db, request)
        rows = db.query(Item).filter(Item.user_id == user.id).order_by(Item.code.asc(), Item.created_at.desc()).all()
        wb = Workbook()
        ws = wb.active
        tpl = (template or "full").strip().lower()
        if tpl == "price":
            ws.title = "تحديث الأسعار"
            ws.append(["الكود", "اسم الصنف (للعرض فقط)", "سعر الشراء", "سعر البيع (غير شامل الضريبة)"])
            for r in rows:
                ws.append([r.code, r.name, round(float(r.last_cost or 0.0), 2), round(float(r.default_sale_price or 0.0), 2)])
            fname = "items_price_template.xlsx"
        elif tpl == "qty":
            ws.title = "تحديث الكميات"
            ws.append(["الكود", "اسم الصنف (للعرض فقط)", "الكمية الفعلية"])
            for r in rows:
                ws.append([r.code, r.name, round(float(r.quantity or 0.0), 4)])
            fname = "items_qty_template.xlsx"
        else:
            ws.title = "الأصناف"
            ws.append(["الكود", "اسم الصنف", "التصنيف", "الوحدة", "سعر الشراء", "سعر البيع (غير شامل الضريبة)", "نسبة الضريبة %", "خاضع للضريبة (1/0)", "الحد الأدنى للكمية", "ملاحظات"])
            for r in rows:
                ws.append(
                    [
                        r.code,
                        r.name,
                        r.category or "",
                        r.unit or "قطعة",
                        round(float(r.last_cost or 0.0), 2),
                        round(float(r.default_sale_price or 0.0), 2),
                        round(float(r.tax_rate or 0.0), 4),
                        1 if r.is_taxable else 0,
                        round(float(r.min_qty or 0.0), 4),
                        r.notes or "",
                    ]
                )
            fname = "items_full_template.xlsx"
        return _xlsx_response(wb, fname)
    finally:
        db.close()


@router.post("/items/import")
async def import_items(request: Request, file: UploadFile = File(...), template: str = Query("full")):
    db = SessionLocal()
    try:
        require_csrf(request)
        user = require_user(db, request)
        content = await file.read()
        wb = load_workbook(io.BytesIO(content), data_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            raise HTTPException(400, "الملف فارغ")
        header = [str(c or "").strip() for c in rows[0]]
        idx_code = _find_header_index(header, "الكود", "code")
        tpl = (template or "full").strip().lower()
        created = 0
        updated = 0
        errors: list[str] = []

        if tpl == "price":
            idx_purchase = _find_header_index(header, "سعر الشراء", "purchase_price")
            idx_sale = _find_header_index(header, "سعر البيع (غير شامل الضريبة)", "سعر البيع", "sale_price")
            if idx_code == -1:
                raise HTTPException(400, "لم يتم العثور على عمود الكود في الملف")
            for rn, row in enumerate(rows[1:], start=2):
                code = str(row[idx_code] or "").strip() if idx_code < len(row) and row[idx_code] is not None else ""
                if not code:
                    continue
                rec = db.query(Item).filter(Item.user_id == user.id, Item.code == code).first()
                if not rec:
                    errors.append(f"صف {rn}: الكود {code} غير موجود")
                    continue
                try:
                    if idx_purchase != -1 and idx_purchase < len(row) and row[idx_purchase] is not None:
                        rec.last_cost = round(abs(float(row[idx_purchase])), 2)
                    if idx_sale != -1 and idx_sale < len(row) and row[idx_sale] is not None:
                        rec.default_sale_price = round(abs(float(row[idx_sale])), 2)
                except (TypeError, ValueError):
                    errors.append(f"صف {rn}: قيمة سعر غير صالحة")
                    continue
                updated += 1
            db.commit()
            log_event(db, "trade.items.import_price", user.id, {"updated": updated, "errors": len(errors)})
            return {"created": 0, "updated": updated, "errors": errors}

        if tpl == "qty":
            idx_qty = _find_header_index(header, "الكمية الفعلية", "الكمية", "quantity")
            if idx_code == -1 or idx_qty == -1:
                raise HTTPException(400, "لم يتم العثور على أعمدة الكود/الكمية في الملف")
            now = dt.datetime.utcnow()
            for rn, row in enumerate(rows[1:], start=2):
                code = str(row[idx_code] or "").strip() if idx_code < len(row) and row[idx_code] is not None else ""
                if not code:
                    continue
                rec = db.query(Item).filter(Item.user_id == user.id, Item.code == code).first()
                if not rec:
                    errors.append(f"صف {rn}: الكود {code} غير موجود")
                    continue
                try:
                    new_qty = round(abs(float(row[idx_qty])), 4) if row[idx_qty] is not None else None
                except (TypeError, ValueError):
                    errors.append(f"صف {rn}: قيمة كمية غير صالحة")
                    continue
                if new_qty is None:
                    continue
                before = round(float(rec.quantity or 0.0), 4)
                diff = round(new_qty - before, 4)
                if abs(diff) < 0.0001:
                    continue
                rec.quantity = new_qty
                db.add(StockAdjustment(id=uuid.uuid4().hex, user_id=user.id, item_id=rec.id, branch_id=rec.branch_id, adjust_date=now, qty_before=before, qty_after=new_qty, difference=diff, reason="استيراد الكميات من Excel"))
                db.add(StockMovement(id=uuid.uuid4().hex, user_id=user.id, item_id=rec.id, movement_type="adjust", qty_in=diff if diff > 0 else 0.0, qty_out=abs(diff) if diff < 0 else 0.0, unit_cost=round(float(rec.last_cost or 0.0), 4), reference_type="adjust", reference_id=rec.id, movement_date=now, notes="استيراد Excel"))
                updated += 1
            db.commit()
            log_event(db, "trade.items.import_qty", user.id, {"updated": updated, "errors": len(errors)})
            return {"created": 0, "updated": updated, "errors": errors}

        idx_name = _find_header_index(header, "اسم الصنف", "الاسم", "name")
        idx_category = _find_header_index(header, "التصنيف", "category")
        idx_unit = _find_header_index(header, "الوحدة", "unit")
        idx_purchase = _find_header_index(header, "سعر الشراء", "سعر الشراء الافتراضي", "purchase_price")
        idx_sale = _find_header_index(header, "سعر البيع (غير شامل الضريبة)", "سعر البيع الافتراضي", "سعر البيع", "sale_price")
        idx_rate = _find_header_index(header, "نسبة الضريبة %", "نسبة الضريبة", "tax_rate")
        idx_taxable = _find_header_index(header, "خاضع للضريبة (1/0)", "خاضع للضريبة", "is_taxable")
        idx_minqty = _find_header_index(header, "الحد الأدنى للكمية", "min_qty")
        idx_notes = _find_header_index(header, "ملاحظات", "notes")
        if idx_name == -1:
            raise HTTPException(400, "لم يتم العثور على عمود اسم الصنف في الملف")

        def _cell(row, idx, cast=None):
            if idx == -1 or idx >= len(row) or row[idx] is None:
                return None
            v = row[idx]
            if cast is None:
                return str(v).strip()
            try:
                return cast(v)
            except (TypeError, ValueError):
                return None

        for rn, row in enumerate(rows[1:], start=2):
            name = _cell(row, idx_name)
            if not name:
                continue
            code = _cell(row, idx_code) or ""
            existing = db.query(Item).filter(Item.user_id == user.id, Item.code == code).first() if code else None
            purchase_price = _cell(row, idx_purchase, float) or 0.0
            sale_price = _cell(row, idx_sale, float) or 0.0
            rate = _cell(row, idx_rate, float)
            taxable_raw = _cell(row, idx_taxable, float)
            min_qty = _cell(row, idx_minqty, float) or 0.0
            category = _cell(row, idx_category) or "rim"
            unit = _cell(row, idx_unit) or "قطعة"
            notes = _cell(row, idx_notes) or ""
            if existing:
                existing.name = name
                existing.category = category
                existing.unit = unit
                existing.last_cost = round(abs(purchase_price), 2)
                existing.default_sale_price = round(abs(sale_price), 2)
                if rate is not None:
                    existing.tax_rate = round(abs(rate), 4)
                if taxable_raw is not None:
                    existing.is_taxable = 1 if taxable_raw else 0
                existing.min_qty = round(abs(min_qty), 4)
                if notes:
                    existing.notes = notes
                updated += 1
            else:
                new_code = code
                if not new_code:
                    for _ in range(20):
                        cand = _new_item_code()
                        if not db.query(Item).filter(Item.user_id == user.id, Item.code == cand).first():
                            new_code = cand
                            break
                elif db.query(Item).filter(Item.user_id == user.id, Item.code == new_code).first():
                    errors.append(f"صف {rn}: الكود {new_code} مستخدم مسبقاً")
                    continue
                db.add(
                    Item(
                        id=uuid.uuid4().hex,
                        user_id=user.id,
                        code=new_code,
                        name=name,
                        category=category,
                        unit=unit,
                        quantity=0.0,
                        min_qty=round(abs(min_qty), 4),
                        default_sale_price=round(abs(sale_price), 2),
                        is_taxable=1 if (taxable_raw is None or taxable_raw) else 0,
                        tax_rate=round(abs(rate), 4) if rate is not None else 15.0,
                        is_active=1,
                        last_cost=round(abs(purchase_price), 2),
                        notes=notes or None,
                    )
                )
                created += 1
        db.commit()
        log_event(db, "trade.items.import", user.id, {"created": created, "updated": updated, "errors": len(errors)})
        return {"created": created, "updated": updated, "errors": errors}
    finally:
        db.close()


@router.get("/purchases/export")
def export_purchases(request: Request):
    db = SessionLocal()
    try:
        user = require_user(db, request)
        rows = db.query(Purchase).filter(Purchase.user_id == user.id).order_by(Purchase.purchase_date.desc()).all()
        wb = Workbook()
        ws = wb.active
        ws.title = "المشتريات"
        ws.append(["رقم الفاتورة", "المورد", "التاريخ", "طريقة الدفع", "الضريبة", "الخصم", "المدفوع", "المتبقي", "الإجمالي"])
        for r in rows:
            ws.append(
                [
                    r.invoice_no,
                    r.supplier_name,
                    _fmt_date(r.purchase_date),
                    r.payment_type,
                    round(float(r.tax_amount or 0.0), 2),
                    round(float(r.discount or 0.0), 2),
                    round(float(r.paid_amount or 0.0), 2),
                    round(float(r.due_amount or 0.0), 2),
                    round(float(r.total_amount or 0.0), 2),
                ]
            )
        ws2 = wb.create_sheet("تفاصيل الأصناف")
        ws2.append(["رقم الفاتورة", "الصنف", "الكمية", "سعر الشراء", "الإجمالي"])
        lines = (
            db.query(PurchaseLine, Item, Purchase)
            .join(Item, Item.id == PurchaseLine.item_id)
            .join(Purchase, Purchase.id == PurchaseLine.purchase_id)
            .filter(Purchase.user_id == user.id)
            .all()
        )
        for ln, item, pur in lines:
            ws2.append(
                [
                    pur.invoice_no,
                    item.name,
                    round(float(ln.qty or 0.0), 4),
                    round(float(ln.unit_cost or 0.0), 2),
                    round(float(ln.total_cost or 0.0), 2),
                ]
            )
        return _xlsx_response(wb, "purchases.xlsx")
    finally:
        db.close()


@router.post("/purchases/import")
async def import_purchases(request: Request, file: UploadFile = File(...)):
    db = SessionLocal()
    try:
        require_csrf(request)
        user = require_user(db, request)
        content = await file.read()
        wb = load_workbook(io.BytesIO(content), data_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            raise HTTPException(400, "الملف فارغ")
        header = [str(c or "").strip() for c in rows[0]]
        idx_item = _find_header_index(header, "الصنف", "اسم الصنف", "الكود", "item")
        idx_qty = _find_header_index(header, "الكمية", "qty")
        idx_price = _find_header_index(header, "سعر الشراء", "price")
        idx_date = _find_header_index(header, "التاريخ", "date")
        idx_notes = _find_header_index(header, "الملاحظات", "ملاحظات", "notes")
        if idx_item == -1 or idx_qty == -1 or idx_price == -1:
            raise HTTPException(400, "الأعمدة المطلوبة: الصنف، الكمية، سعر الشراء")
        created = 0
        errors: list[str] = []
        for rn, row in enumerate(rows[1:], start=2):
            raw_item = str(row[idx_item] or "").strip() if idx_item < len(row) and row[idx_item] is not None else ""
            if not raw_item:
                continue
            item = (
                db.query(Item)
                .filter(Item.user_id == user.id)
                .filter((Item.name == raw_item) | (Item.code == raw_item))
                .first()
            )
            if not item:
                errors.append(f"صف {rn}: الصنف '{raw_item}' غير موجود")
                continue
            try:
                qty = round(abs(float(row[idx_qty])), 4) if idx_qty < len(row) else 0.0
                price = round(abs(float(row[idx_price])), 4) if idx_price < len(row) else 0.0
            except (TypeError, ValueError):
                errors.append(f"صف {rn}: كمية أو سعر غير صالح")
                continue
            if qty <= 0:
                errors.append(f"صف {rn}: الكمية يجب أن تكون أكبر من صفر")
                continue
            raw_date = row[idx_date] if idx_date != -1 and idx_date < len(row) else None
            p_date = None
            if raw_date:
                try:
                    if hasattr(raw_date, "strftime"):
                        p_date = raw_date if isinstance(raw_date, dt.datetime) else dt.datetime.combine(raw_date, dt.time())
                    else:
                        p_date = _parse_date(str(raw_date), "date")
                except Exception:
                    p_date = None
            if p_date is None:
                p_date = dt.datetime.utcnow()
            notes_val = ""
            if idx_notes != -1 and idx_notes < len(row) and row[idx_notes] is not None:
                notes_val = str(row[idx_notes]).strip()
            rec = Purchase(
                id=uuid.uuid4().hex,
                user_id=user.id,
                invoice_no=f"IMP-{uuid.uuid4().hex[:8].upper()}",
                supplier_name="مستورد من Excel",
                purchase_date=p_date,
                payment_type="cash",
                tax_amount=0.0,
                discount=0.0,
                paid_amount=0.0,
                due_amount=0.0,
                notes=notes_val or None,
                total_amount=round(qty * price, 2),
            )
            db.add(rec)
            db.flush()
            db.add(
                PurchaseLine(
                    purchase_id=rec.id,
                    item_id=item.id,
                    qty=qty,
                    unit_cost=price,
                    extra_cost=0.0,
                    total_cost=round(qty * price, 2),
                )
            )
            db.add(
                StockMovement(
                    id=uuid.uuid4().hex,
                    user_id=user.id,
                    item_id=item.id,
                    movement_type="purchase",
                    qty_in=qty,
                    qty_out=0.0,
                    unit_cost=price,
                    reference_type="purchase",
                    reference_id=rec.id,
                    movement_date=p_date,
                )
            )
            item.quantity = round(float(item.quantity or 0.0) + qty, 4)
            item.last_cost = price
            created += 1
        db.commit()
        log_event(db, "trade.purchases.import", user.id, {"created": created, "errors": len(errors)})
        return {"created": created, "errors": errors}
    finally:
        db.close()


@router.get("/reports/item-movement/export")
def export_item_movement(request: Request, item_id: str = Query(...), date_from: str = Query(""), date_to: str = Query("")):
    data = item_movement_report(request, item_id=item_id, date_from=date_from, date_to=date_to)
    wb = Workbook()
    ws = wb.active
    ws.title = "حركة الصنف"
    ws.append(["الصنف", data["item"]["name"]])
    ws.append([])
    ws.append(["التاريخ", "نوع الحركة", "وارد", "صادر", "تكلفة الوحدة", "الرصيد بعد الحركة"])
    for x in data["items"]:
        ws.append([x["date"], x["movement_type"], x["qty_in"], x["qty_out"], x["unit_cost"], x["balance_after"]])
    return _xlsx_response(wb, f"item_movement_{item_id}.xlsx")


@router.get("/reports/export-all")
def export_all_reports(request: Request, date_from: str = Query(""), date_to: str = Query("")):
    profit_data = profit_report(request, date_from=date_from, date_to=date_to)
    top_data = top_items_report(request, date_from=date_from, date_to=date_to, limit=20)
    stock_data = stock_status_report(request)
    inv_data = inventory_report(request)

    db = SessionLocal()
    try:
        user = require_user(db, request)
        purchases = db.query(Purchase).filter(Purchase.user_id == user.id).order_by(Purchase.purchase_date.desc()).all()
    finally:
        db.close()

    wb = Workbook()
    ws1 = wb.active
    ws1.title = "المخزون الحالي"
    ws1.append(["الكود", "اسم الصنف", "الكمية", "تكلفة الوحدة", "سعر البيع", "قيمة المخزون"])
    for r in inv_data["items"]:
        ws1.append([r["code"], r["name"], r["quantity"], r["cost_price"], r["sale_price"], r["stock_value"]])

    ws2 = wb.create_sheet("المبيعات والأرباح")
    ws2.append(["التاريخ", "رقم الفاتورة", "الصنف", "الكمية", "صافي المبيعات", "التكلفة", "الربح"])
    for r in profit_data["items"]:
        ws2.append([r["date"], r["invoice_no"], r["item"], r["qty"], r["net_sales"], r["cost_total"], r["profit"]])
    ws2.append([])
    t = profit_data["totals"]
    ws2.append(["الإجمالي", "", "", "", t["net_sales"], t["cost_total"], t["profit"]])

    ws3 = wb.create_sheet("المشتريات")
    ws3.append(["رقم الفاتورة", "المورد", "التاريخ", "الإجمالي"])
    for r in purchases:
        ws3.append([r.invoice_no, r.supplier_name, _fmt_date(r.purchase_date), round(float(r.total_amount or 0.0), 2)])

    ws4 = wb.create_sheet("الأكثر والأقل مبيعاً")
    ws4.append(["الأكثر مبيعاً"])
    ws4.append(["الصنف", "الكمية", "المبيعات", "الربح"])
    for r in top_data["best_selling"]:
        ws4.append([r["item"], r["qty"], r["sales"], r["profit"]])
    ws4.append([])
    ws4.append(["الأقل مبيعاً"])
    ws4.append(["الصنف", "الكمية", "المبيعات", "الربح"])
    for r in top_data["worst_selling"]:
        ws4.append([r["item"], r["qty"], r["sales"], r["profit"]])

    ws5 = wb.create_sheet("المخزون المنخفض")
    ws5.append(["الكود", "اسم الصنف", "الكمية", "الحد الأدنى"])
    for r in stock_data["low_stock"]:
        ws5.append([r["code"], r["name"], r["quantity"], r["min_qty"]])

    return _xlsx_response(wb, "all_reports.xlsx")
