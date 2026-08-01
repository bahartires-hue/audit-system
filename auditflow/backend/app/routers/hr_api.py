from __future__ import annotations

import datetime as dt
import io
import json
import os as _os
import uuid
from urllib.parse import quote as _quote
from pathlib import Path as _Path
from typing import Optional

from fastapi import APIRouter, Body, File, Form, HTTPException, Query, Request, UploadFile
from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from pydantic import BaseModel, Field

from ..auth_core import log_event, require_csrf, require_user
from ..db import SessionLocal
from ..models import (
    Employee,
    EmployeeContract,
    EmployeeDocument,
    EmployeeDocumentRenewal,
    Advance,
    DeductionRecord,
    AllowanceRecord,
    CommissionRecord,
    LeaveRequest,
    Attendance,
    EndOfService,
    HRConfigItem,
    Payroll,
    PayrollAllowance,
    PayrollDeduction,
    AppSetting,
    CompanyProfile,
    WithdrawalRecord,
    TravelRecord,
    CustodyRecord,
    AuditLog,
)
from ..pdf_utils import generate_pdf_report

router = APIRouter(prefix="/api/hr", tags=["hr"])

_BASE_DIR = _Path(__file__).resolve().parents[3]
_data_root = (_os.getenv("AUDITFLOW_DATA_ROOT") or "").strip()
_UPLOAD_DIR = (_Path(_data_root) / "uploads") if _data_root else (_BASE_DIR / "uploads")
_HR_DOCS_DIR = _UPLOAD_DIR / "hr_documents"
_HR_CONTRACTS_DIR = _UPLOAD_DIR / "hr_contracts"
_HR_PHOTOS_DIR = _UPLOAD_DIR / "hr_photos"
for _d in (_HR_DOCS_DIR, _HR_CONTRACTS_DIR, _HR_PHOTOS_DIR):
    _d.mkdir(parents=True, exist_ok=True)

_ALLOWED_DOC_SUFFIXES = (".pdf", ".png", ".jpg", ".jpeg")
_ALLOWED_IMAGE_SUFFIXES = (".png", ".jpg", ".jpeg", ".webp")
_MAX_DOC_MB = 10


def _parse_simple_date(s, field_name: str):
    s = (s or "").strip()
    if not s:
        return None
    try:
        return dt.datetime.strptime(s, "%Y-%m-%d")
    except ValueError:
        raise HTTPException(400, f"{field_name} يجب أن يكون بصيغة YYYY-MM-DD")


def _fmt_date(d):
    return d.strftime("%Y-%m-%d") if d else ""


def _fmt_dt(d):
    return d.strftime("%Y-%m-%d %H:%M") if d else ""


def _days_left(expiry):
    if not expiry:
        return None
    today = dt.datetime.utcnow().date()
    exp = expiry.date() if hasattr(expiry, "date") else expiry
    return (exp - today).days


def _next_employee_number(db, user_id: str) -> str:
    count = db.query(Employee).filter(Employee.user_id == user_id).count()
    return f"EMP-{count + 1:04d}"


def _xlsx_response(headers, rows, filename):
    wb = Workbook()
    ws = wb.active
    ws.append(headers)
    for row in rows:
        ws.append(row)
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    ascii_fallback = "export.xlsx"
    disposition = f"attachment; filename=\"{ascii_fallback}\"; filename*=UTF-8''{_quote(filename, safe='')}"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": disposition},
    )


def _pdf_meta(db, user):
    company = db.query(CompanyProfile).filter(CompanyProfile.id == "default").first()
    company_name = (company.trade_name or company.company_name) if company else None
    logo_url = company.logo_url if company else None
    generated_by = getattr(user, "username", None) or getattr(user, "email", None)
    return company_name, logo_url, generated_by


STATUS_LABELS = {
    "active": "نشط",
    "on_leave": "إجازة",
    "suspended": "موقوف",
    "resigned": "مستقيل",
    "terminated": "منتهي الخدمة",
}


# ============================================================
# الإعدادات (الأقسام / الوظائف / أنواع الإجازات / البدلات / الاستقطاعات / الفروع / الضريبة / العملة)
# ============================================================

class ConfigItemCreate(BaseModel):
    category: str
    name: str = Field(min_length=1, max_length=120)


@router.get("/config")
def list_config(request: Request, category: str = Query("")):
    db = SessionLocal()
    try:
        user = require_user(db, request)
        q = db.query(HRConfigItem).filter(HRConfigItem.user_id == user.id)
        if category:
            q = q.filter(HRConfigItem.category == category)
        rows = q.order_by(HRConfigItem.name.asc()).all()
        return {"items": [{"id": r.id, "category": r.category, "name": r.name} for r in rows]}
    finally:
        db.close()


@router.post("/config")
def create_config(request: Request, body: ConfigItemCreate = Body(...)):
    db = SessionLocal()
    try:
        require_csrf(request)
        user = require_user(db, request)
        rec = HRConfigItem(id=uuid.uuid4().hex, user_id=user.id, category=body.category.strip(), name=body.name.strip())
        db.add(rec)
        db.commit()
        return {"id": rec.id, "category": rec.category, "name": rec.name}
    finally:
        db.close()


@router.delete("/config/{item_id}")
def delete_config(item_id: str, request: Request):
    db = SessionLocal()
    try:
        require_csrf(request)
        user = require_user(db, request)
        rec = db.query(HRConfigItem).filter(HRConfigItem.user_id == user.id, HRConfigItem.id == item_id).first()
        if not rec:
            raise HTTPException(404, "العنصر غير موجود")
        db.delete(rec)
        db.commit()
        return {"deleted": True}
    finally:
        db.close()


class HRSettingsIn(BaseModel):
    tax_percentage: float = 0.0
    tax_enabled: bool = False
    currency: str = "SAR"


def _get_hr_settings(db) -> dict:
    row = db.query(AppSetting).filter(AppSetting.key == "hr.settings").first()
    if row and isinstance(row.value_json, dict):
        return {"tax_percentage": 0.0, "tax_enabled": False, "currency": "SAR", **row.value_json}
    return {"tax_percentage": 0.0, "tax_enabled": False, "currency": "SAR"}


@router.get("/settings")
def get_hr_settings(request: Request):
    db = SessionLocal()
    try:
        require_user(db, request)
        return _get_hr_settings(db)
    finally:
        db.close()


@router.patch("/settings")
def update_hr_settings(request: Request, body: HRSettingsIn = Body(...)):
    db = SessionLocal()
    try:
        require_csrf(request)
        require_user(db, request)
        row = db.query(AppSetting).filter(AppSetting.key == "hr.settings").first()
        value = {"tax_percentage": body.tax_percentage, "tax_enabled": body.tax_enabled, "currency": body.currency}
        if row:
            row.value_json = value
            row.updated_at = dt.datetime.utcnow()
        else:
            row = AppSetting(key="hr.settings", value_json=value)
            db.add(row)
        db.commit()
        return value
    finally:
        db.close()



# ============================================================
# الموظفون
# ============================================================

class EmployeeCreate(BaseModel):
    name: str = Field(min_length=1, max_length=200)
    nationality: str = ""
    national_id: str = ""
    residency_expiry: str = ""
    passport_number: str = ""
    passport_expiry: str = ""
    position: str = ""
    department: str = ""
    branch_name: str = ""
    basic_salary: float = 0.0
    commission_percentage: Optional[float] = None
    hire_date: str = ""
    birth_date: str = ""
    phone: str = ""
    email: str = ""
    status: str = "active"
    notes: str = ""
    gender: str = ""
    address: str = ""
    manager_id: str = ""
    bank_name: str = ""
    iban: str = ""
    bank_salary: Optional[float] = None


class EmployeeUpdate(BaseModel):
    name: Optional[str] = None
    nationality: Optional[str] = None
    national_id: Optional[str] = None
    residency_expiry: Optional[str] = None
    passport_number: Optional[str] = None
    passport_expiry: Optional[str] = None
    position: Optional[str] = None
    department: Optional[str] = None
    branch_name: Optional[str] = None
    basic_salary: Optional[float] = None
    commission_percentage: Optional[float] = None
    hire_date: Optional[str] = None
    birth_date: Optional[str] = None
    phone: Optional[str] = None
    email: Optional[str] = None
    status: Optional[str] = None
    is_active: Optional[bool] = None
    notes: Optional[str] = None
    gender: Optional[str] = None
    address: Optional[str] = None
    manager_id: Optional[str] = None
    bank_name: Optional[str] = None
    iban: Optional[str] = None
    bank_salary: Optional[float] = None


def _years_of_service(hire_date) -> Optional[float]:
    if not hire_date:
        return None
    days = (dt.datetime.utcnow() - hire_date).days
    if days < 0:
        return None
    return round(days / 365.25, 1)


def _employee_out(e: Employee, db=None) -> dict:
    residency_days = _days_left(e.residency_expiry)
    passport_days = _days_left(e.passport_expiry)
    manager_name = ""
    if db is not None and e.manager_id:
        mgr = db.query(Employee).filter(Employee.id == e.manager_id).first()
        manager_name = mgr.name if mgr else ""
    return {
        "id": e.id,
        "employee_number": e.employee_number or "",
        "name": e.name,
        "nationality": e.nationality or "",
        "national_id": e.national_id or "",
        "residency_expiry": _fmt_date(e.residency_expiry),
        "residency_days_left": residency_days,
        "passport_number": e.passport_number or "",
        "passport_expiry": _fmt_date(e.passport_expiry),
        "passport_days_left": passport_days,
        "position": e.position or "",
        "department": e.department or "",
        "branch_name": e.branch_name or "",
        "basic_salary": round(float(e.basic_salary or 0.0), 2),
        "commission_percentage": e.commission_percentage,
        "hire_date": _fmt_date(e.hire_date),
        "birth_date": _fmt_date(e.birth_date),
        "phone": e.phone or "",
        "email": e.email or "",
        "status": e.status or "active",
        "status_label": STATUS_LABELS.get(e.status or "active", e.status or ""),
        "photo_url": e.photo_url or "",
        "is_active": bool(int(e.is_active or 0)),
        "notes": e.notes or "",
        "gender": e.gender or "",
        "address": e.address or "",
        "manager_id": e.manager_id or "",
        "manager_name": manager_name,
        "bank_name": e.bank_name or "",
        "iban": e.iban or "",
        "bank_salary": e.bank_salary,
        "years_of_service": _years_of_service(e.hire_date),
    }


@router.get("/employees")
def list_employees(request: Request, q: str = Query(""), department: str = Query(""), branch_name: str = Query(""), status: str = Query("")):
    db = SessionLocal()
    try:
        user = require_user(db, request)
        rows = db.query(Employee).filter(Employee.user_id == user.id).order_by(Employee.created_at.desc()).all()
        qn = (q or "").strip().lower()
        if qn:
            rows = [r for r in rows if qn in (r.name or "").lower() or qn in (r.employee_number or "").lower()]
        if department:
            rows = [r for r in rows if (r.department or "") == department]
        if branch_name:
            rows = [r for r in rows if (r.branch_name or "") == branch_name]
        if status:
            rows = [r for r in rows if (r.status or "active") == status]
        out = []
        for r in rows:
            item = _employee_out(r)
            contract = (
                db.query(EmployeeContract)
                .filter(EmployeeContract.employee_id == r.id, EmployeeContract.status == "active")
                .order_by(EmployeeContract.created_at.desc())
                .first()
            )
            item["contract_status"] = "ساري" if contract else "لا يوجد"
            out.append(item)
        return {"items": out}
    finally:
        db.close()


@router.post("/employees")
def create_employee(request: Request, body: EmployeeCreate = Body(...)):
    db = SessionLocal()
    try:
        require_csrf(request)
        user = require_user(db, request)
        rec = Employee(
            id=uuid.uuid4().hex,
            user_id=user.id,
            employee_number=_next_employee_number(db, user.id),
            name=body.name.strip(),
            nationality=(body.nationality or "").strip() or None,
            national_id=(body.national_id or "").strip() or None,
            # ملاحظة معمارية: الإقامة/الجواز لم تعد تُدخل هنا مباشرة — تبويب الوثائق
            # في ملف الموظف هو المصدر الوحيد لها، وتتم مزامنتها تلقائيًا من هناك
            # عبر _sync_identity_docs_from_documents (Single Source of Truth).
            position=(body.position or "").strip() or None,
            department=(body.department or "").strip() or None,
            branch_name=(body.branch_name or "").strip() or None,
            basic_salary=body.basic_salary or 0.0,
            commission_percentage=body.commission_percentage,
            hire_date=_parse_simple_date(body.hire_date, "تاريخ التعيين"),
            birth_date=_parse_simple_date(body.birth_date, "تاريخ الميلاد"),
            phone=(body.phone or "").strip() or None,
            email=(body.email or "").strip() or None,
            status=body.status or "active",
            is_active=1 if (body.status or "active") == "active" else 0,
            notes=(body.notes or "").strip() or None,
            gender=(body.gender or "").strip() or None,
            address=(body.address or "").strip() or None,
            manager_id=(body.manager_id or "").strip() or None,
            bank_name=(body.bank_name or "").strip() or None,
            iban=(body.iban or "").strip() or None,
            bank_salary=body.bank_salary,
        )
        db.add(rec)
        db.commit()
        log_event(db, "hr.employee.create", user.id, {"employee_id": rec.id}, employee_id=rec.id)
        return _employee_out(rec, db)
    finally:
        db.close()


EMPLOYEE_IMPORT_HEADERS = [
    "الاسم", "الرقم الوظيفي", "الفرع", "القسم", "الوظيفة", "الجنسية", "الجنس",
    "تاريخ الميلاد", "الجوال", "البريد الإلكتروني", "العنوان", "رقم الهوية",
    "تاريخ انتهاء الهوية/الإقامة", "رقم الجواز", "تاريخ انتهاء الجواز",
    "رقم رخصة العمل", "تاريخ انتهاء رخصة العمل", "رقم التأمين الطبي",
    "تاريخ انتهاء التأمين الطبي", "تاريخ التعيين", "نوع العقد", "بداية العقد",
    "نهاية العقد", "الراتب الأساسي", "بدل السكن", "بدل النقل", "البدلات الأخرى",
    "نسبة العمولة", "الراتب البنكي", "اسم البنك", "رقم الآيبان", "المدير المباشر",
    "الحالة الوظيفية", "ملاحظات",
]

_STATUS_LABEL_TO_VALUE = {v: k for k, v in STATUS_LABELS.items()}
GENDER_LABELS = {"male": "ذكر", "female": "أنثى"}
_GENDER_LABEL_TO_VALUE = {v: k for k, v in GENDER_LABELS.items()}


def _cell_str(v) -> str:
    if v is None:
        return ""
    if isinstance(v, dt.datetime):
        return v.strftime("%Y-%m-%d")
    return str(v).strip()


def _parse_import_date(v, field_name: str):
    if v is None or (isinstance(v, str) and not v.strip()):
        return None
    if isinstance(v, dt.datetime):
        return v
    s = str(v).strip()
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y", "%Y/%m/%d"):
        try:
            return dt.datetime.strptime(s, fmt)
        except ValueError:
            continue
    raise ValueError(f"{field_name} بصيغة غير صحيحة (استخدم YYYY-MM-DD): {s}")


def _parse_import_status(v):
    s = _cell_str(v)
    if not s:
        return "active"
    if s in STATUS_LABELS:
        return s
    if s in _STATUS_LABEL_TO_VALUE:
        return _STATUS_LABEL_TO_VALUE[s]
    raise ValueError(f"قيمة الحالة غير معروفة: {s}")


def _parse_import_gender(v):
    s = _cell_str(v)
    if not s:
        return None
    if s in GENDER_LABELS:
        return s
    if s in _GENDER_LABEL_TO_VALUE:
        return _GENDER_LABEL_TO_VALUE[s]
    raise ValueError(f"قيمة الجنس غير معروفة (استخدم ذكر/أنثى): {s}")


def _upsert_generic_document(db, user_id: str, employee_id: str, doc_type_label: str, doc_number: str, expiry_date):
    """للوثائق التي لا يوجد لها عمود مخصص على الموظف (رخصة عمل / تأمين طبي) —
    نفس مبدأ عدم تكرار البيانات: تُكتب فقط في جدول الوثائق (مصدر واحد للحقيقة)،
    وتظهر تلقائيًا في تبويب الوثائق وتنبيهات الانتهاء دون أي إدخال إضافي."""
    if not expiry_date and not (doc_number or "").strip():
        return
    existing = (
        db.query(EmployeeDocument)
        .filter(EmployeeDocument.employee_id == employee_id, EmployeeDocument.doc_type == doc_type_label)
        .order_by(EmployeeDocument.created_at.desc())
        .first()
    )
    if existing:
        if expiry_date:
            existing.expiry_date = expiry_date
        if (doc_number or "").strip():
            existing.doc_number = doc_number.strip()
    else:
        db.add(EmployeeDocument(
            id=uuid.uuid4().hex, user_id=user_id, employee_id=employee_id,
            doc_type=doc_type_label, doc_number=(doc_number or "").strip() or None,
            expiry_date=expiry_date,
        ))


def _latest_document_by_type(db, employee_id: str, doc_type_label: str):
    return (
        db.query(EmployeeDocument)
        .filter(EmployeeDocument.employee_id == employee_id, EmployeeDocument.doc_type == doc_type_label)
        .order_by(EmployeeDocument.created_at.desc())
        .first()
    )


def _latest_contract_for_export(db, employee_id: str):
    return (
        db.query(EmployeeContract)
        .filter(EmployeeContract.employee_id == employee_id)
        .order_by(EmployeeContract.created_at.desc())
        .first()
    )


@router.get("/employees/import-template")
def download_employees_import_template(request: Request):
    db = SessionLocal()
    try:
        require_user(db, request)
        example = [
            "محمد أحمد", "EMP-0001", "الفرع الرئيسي", "المالية", "محاسب", "سعودي", "ذكر",
            "1990-05-10", "0501234567", "employee@example.com", "الرياض - حي النخيل", "1012345678",
            "2027-01-01", "A1234567", "2029-05-01",
            "RP-778899", "2027-03-01", "MI-556677",
            "2027-02-01", "2024-01-01", "دائم", "2024-01-01",
            "", 6000, 500, 300, 0,
            2.5, 6800, "بنك الراجحي", "SA0000000000000000000000", "",
            "نشط", "",
        ]
        return _xlsx_response(EMPLOYEE_IMPORT_HEADERS, [example], "قالب_استيراد_الموظفين.xlsx")
    finally:
        db.close()


@router.get("/employees/export")
def export_employees(request: Request, q: str = Query(""), department: str = Query(""), branch_name: str = Query(""), status: str = Query("")):
    db = SessionLocal()
    try:
        user = require_user(db, request)
        rows = db.query(Employee).filter(Employee.user_id == user.id).order_by(Employee.created_at.desc()).all()
        qn = (q or "").strip().lower()
        if qn:
            rows = [r for r in rows if qn in (r.name or "").lower() or qn in (r.employee_number or "").lower()]
        if department:
            rows = [r for r in rows if (r.department or "") == department]
        if branch_name:
            rows = [r for r in rows if (r.branch_name or "") == branch_name]
        if status:
            rows = [r for r in rows if (r.status or "active") == status]
        name_by_id = {
            e_id: e_name for e_id, e_name in
            db.query(Employee.id, Employee.name).filter(Employee.user_id == user.id).all()
        }
        out_rows = []
        for e in rows:
            contract = _latest_contract_for_export(db, e.id)
            work_permit = _latest_document_by_type(db, e.id, "رخصة عمل")
            medical_ins = _latest_document_by_type(db, e.id, "تأمين طبي")
            out_rows.append([
                e.name or "",
                e.employee_number or "",
                e.branch_name or "",
                e.department or "",
                e.position or "",
                e.nationality or "",
                GENDER_LABELS.get(e.gender or "", e.gender or ""),
                _fmt_date(e.birth_date),
                e.phone or "",
                e.email or "",
                e.address or "",
                e.national_id or "",
                _fmt_date(e.residency_expiry),
                e.passport_number or "",
                _fmt_date(e.passport_expiry),
                work_permit.doc_number if work_permit else "",
                _fmt_date(work_permit.expiry_date) if work_permit else "",
                medical_ins.doc_number if medical_ins else "",
                _fmt_date(medical_ins.expiry_date) if medical_ins else "",
                _fmt_date(e.hire_date),
                contract.contract_type if contract else "",
                _fmt_date(contract.start_date) if contract else "",
                _fmt_date(contract.end_date) if contract else "",
                round(float(e.basic_salary or 0.0), 2),
                round(float(contract.housing_allowance), 2) if contract else 0,
                round(float(contract.transport_allowance), 2) if contract else 0,
                round(float(contract.other_allowances), 2) if contract else 0,
                e.commission_percentage,
                round(float(e.bank_salary), 2) if e.bank_salary is not None else "",
                e.bank_name or "",
                e.iban or "",
                name_by_id.get(e.manager_id, "") if e.manager_id else "",
                STATUS_LABELS.get(e.status or "active", e.status or ""),
                e.notes or "",
            ])
        log_event(db, "hr.employee.export", user.id, {"count": len(out_rows)})
        return _xlsx_response(EMPLOYEE_IMPORT_HEADERS, out_rows, "الموظفون.xlsx")
    finally:
        db.close()


@router.post("/employees/import")
async def import_employees(request: Request, file: UploadFile = File(...)):
    db = SessionLocal()
    try:
        require_csrf(request)
        user = require_user(db, request)
        original = file.filename or "import.xlsx"
        if not original.lower().endswith((".xlsx", ".xlsm")):
            raise HTTPException(400, "يجب أن يكون الملف بصيغة Excel (.xlsx)")
        content = await file.read()
        if len(content) > 15 * 1024 * 1024:
            raise HTTPException(400, "حجم الملف أكبر من 15 ميجابايت")
        try:
            from openpyxl import load_workbook
            wb = load_workbook(io.BytesIO(content), data_only=True)
            ws = wb.active
        except Exception:
            raise HTTPException(400, "تعذر قراءة ملف Excel — تأكد أنه بصيغة .xlsx صحيحة")

        added = 0
        updated = 0
        errors = []
        warnings = []
        pending_manager_links = []  # (target_employee_id, manager_cell_value, row_num)
        row_num = 1
        for raw_row in ws.iter_rows(min_row=2, values_only=True):
            row_num += 1
            if raw_row is None or all(c is None or (isinstance(c, str) and not c.strip()) for c in raw_row):
                continue
            cells = list(raw_row) + [None] * max(0, 34 - len(raw_row))
            try:
                name = _cell_str(cells[0])
                if not name:
                    raise ValueError("اسم الموظف مطلوب")
                employee_number = _cell_str(cells[1])
                branch_name = _cell_str(cells[2])
                department = _cell_str(cells[3])
                position = _cell_str(cells[4])
                nationality = _cell_str(cells[5])
                gender = _parse_import_gender(cells[6])
                birth_date = _parse_import_date(cells[7], "تاريخ الميلاد")
                phone = _cell_str(cells[8])
                email = _cell_str(cells[9])
                address = _cell_str(cells[10])
                national_id = _cell_str(cells[11])
                residency_expiry = _parse_import_date(cells[12], "تاريخ انتهاء الهوية/الإقامة")
                passport_number = _cell_str(cells[13])
                passport_expiry = _parse_import_date(cells[14], "تاريخ انتهاء الجواز")
                work_permit_number = _cell_str(cells[15])
                work_permit_expiry = _parse_import_date(cells[16], "تاريخ انتهاء رخصة العمل")
                medical_ins_number = _cell_str(cells[17])
                medical_ins_expiry = _parse_import_date(cells[18], "تاريخ انتهاء التأمين الطبي")
                hire_date = _parse_import_date(cells[19], "تاريخ التعيين")
                contract_type = _cell_str(cells[20])
                contract_start = _parse_import_date(cells[21], "بداية العقد")
                contract_end = _parse_import_date(cells[22], "نهاية العقد")
                salary_raw = cells[23]
                try:
                    basic_salary = float(salary_raw) if salary_raw not in (None, "") else 0.0
                except (TypeError, ValueError):
                    raise ValueError(f"الراتب الأساسي غير رقمي: {salary_raw}")

                def _num(idx, label):
                    raw = cells[idx]
                    if raw in (None, ""):
                        return 0.0
                    try:
                        return float(raw)
                    except (TypeError, ValueError):
                        raise ValueError(f"{label} غير رقمي: {raw}")

                housing_allowance = _num(24, "بدل السكن")
                transport_allowance = _num(25, "بدل النقل")
                other_allowances = _num(26, "البدلات الأخرى")
                comm_raw = cells[27]
                try:
                    commission_percentage = float(comm_raw) if comm_raw not in (None, "") else None
                except (TypeError, ValueError):
                    raise ValueError(f"نسبة العمولة غير رقمية: {comm_raw}")
                bank_salary_raw = cells[28]
                try:
                    bank_salary = float(bank_salary_raw) if bank_salary_raw not in (None, "") else None
                except (TypeError, ValueError):
                    raise ValueError(f"الراتب البنكي غير رقمي: {bank_salary_raw}")
                bank_name = _cell_str(cells[29])
                iban = _cell_str(cells[30])
                manager_cell = _cell_str(cells[31])
                status = _parse_import_status(cells[32])
                notes = _cell_str(cells[33])
            except ValueError as ve:
                errors.append({"row": row_num, "reason": str(ve)})
                continue

            existing = None
            if national_id:
                existing = db.query(Employee).filter(Employee.user_id == user.id, Employee.national_id == national_id).first()
            if not existing and employee_number:
                existing = db.query(Employee).filter(Employee.user_id == user.id, Employee.employee_number == employee_number).first()

            try:
                target_employee_id = existing.id if existing else uuid.uuid4().hex
                if existing:
                    existing.name = name
                    existing.nationality = nationality or None
                    existing.national_id = national_id or existing.national_id
                    # ملاحظة معمارية: لا نكتب على existing.residency_expiry/passport_* مباشرة —
                    # نكتب وثيقة حقيقية في جدول الوثائق ثم نزامن (Single Source of Truth).
                    existing.position = position or None
                    existing.department = department or None
                    existing.branch_name = branch_name or None
                    existing.basic_salary = basic_salary
                    existing.commission_percentage = commission_percentage
                    existing.hire_date = hire_date
                    existing.birth_date = birth_date or existing.birth_date
                    existing.phone = phone or None
                    existing.email = email or None
                    existing.address = address or existing.address
                    existing.gender = gender or existing.gender
                    existing.bank_name = bank_name or existing.bank_name
                    existing.iban = iban or existing.iban
                    if bank_salary is not None:
                        existing.bank_salary = bank_salary
                    existing.status = status
                    existing.is_active = 1 if status == "active" else 0
                    existing.notes = notes or None
                    updated += 1
                else:
                    rec = Employee(
                        id=target_employee_id,
                        user_id=user.id,
                        employee_number=employee_number or _next_employee_number(db, user.id),
                        name=name,
                        nationality=nationality or None,
                        national_id=national_id or None,
                        position=position or None,
                        department=department or None,
                        branch_name=branch_name or None,
                        basic_salary=basic_salary,
                        commission_percentage=commission_percentage,
                        hire_date=hire_date,
                        birth_date=birth_date,
                        phone=phone or None,
                        email=email or None,
                        address=address or None,
                        gender=gender,
                        bank_name=bank_name or None,
                        iban=iban or None,
                        bank_salary=bank_salary,
                        status=status,
                        is_active=1 if status == "active" else 0,
                        notes=notes or None,
                    )
                    db.add(rec)
                    added += 1
                db.commit()

                if residency_expiry:
                    _upsert_identity_document(db, user.id, target_employee_id, "إقامة", "", residency_expiry)
                if passport_expiry or passport_number:
                    _upsert_identity_document(db, user.id, target_employee_id, "جواز", passport_number, passport_expiry)
                if work_permit_expiry or work_permit_number:
                    _upsert_generic_document(db, user.id, target_employee_id, "رخصة عمل", work_permit_number, work_permit_expiry)
                if medical_ins_expiry or medical_ins_number:
                    _upsert_generic_document(db, user.id, target_employee_id, "تأمين طبي", medical_ins_number, medical_ins_expiry)
                db.commit()
                _sync_identity_docs_from_documents(db, target_employee_id)

                # عقد العمل: يُنشأ تلقائيًا من الاستيراد فقط إذا لا يوجد عقد ساري بالفعل لهذا
                # الموظف (لتفادي تكرار العقود عند إعادة استيراد نفس الملف)؛ وإلا فالتعديل على
                # عقد قائم يتم من داخل ملف الموظف نفسه وليس عبر إعادة الاستيراد.
                if contract_type or contract_start or contract_end:
                    has_active_contract = db.query(EmployeeContract).filter(
                        EmployeeContract.employee_id == target_employee_id,
                        EmployeeContract.status == "active",
                    ).first()
                    if not has_active_contract:
                        db.add(EmployeeContract(
                            id=uuid.uuid4().hex, user_id=user.id, employee_id=target_employee_id,
                            contract_type=contract_type or None,
                            start_date=contract_start, end_date=contract_end,
                            salary=basic_salary,
                            housing_allowance=housing_allowance,
                            transport_allowance=transport_allowance,
                            other_allowances=other_allowances,
                            status="active",
                        ))
                        db.commit()

                if manager_cell:
                    pending_manager_links.append((target_employee_id, manager_cell, row_num))
            except Exception as ex:
                db.rollback()
                errors.append({"row": row_num, "reason": f"تعذر الحفظ: {ex}"})

        # ربط المدير المباشر: يتم بعد معالجة كل الصفوف حتى لو ظهر المدير في صف لاحق بنفس الملف
        for target_employee_id, manager_cell, row_num in pending_manager_links:
            manager_rec = db.query(Employee).filter(
                Employee.user_id == user.id, Employee.employee_number == manager_cell, Employee.id != target_employee_id,
            ).first()
            if not manager_rec:
                manager_rec = db.query(Employee).filter(
                    Employee.user_id == user.id, Employee.name == manager_cell, Employee.id != target_employee_id,
                ).first()
            if manager_rec:
                emp_rec = db.query(Employee).filter(Employee.id == target_employee_id).first()
                if emp_rec:
                    emp_rec.manager_id = manager_rec.id
                    db.commit()
            else:
                warnings.append({"row": row_num, "reason": f"تعذر إيجاد المدير المباشر المطابق: {manager_cell}"})

        log_event(db, "hr.employee.import", user.id, {"added": added, "updated": updated, "errors": len(errors), "warnings": len(warnings)})
        return {"added": added, "updated": updated, "errors": errors, "warnings": warnings, "total_rows": row_num - 1}
    finally:
        db.close()


@router.get("/employees/{employee_id}")
def get_employee(employee_id: str, request: Request):
    db = SessionLocal()
    try:
        user = require_user(db, request)
        rec = db.query(Employee).filter(Employee.user_id == user.id, Employee.id == employee_id).first()
        if not rec:
            raise HTTPException(404, "الموظف غير موجود")
        return _employee_out(rec, db)
    finally:
        db.close()


TIMELINE_ICONS = {
    "hr.employee.create": "👤", "hr.employee.update": "✏️", "hr.employee.salary_change": "💵",
    "hr.contract.create": "📄", "hr.advance.create": "💰", "hr.withdrawal.create": "🧾",
    "hr.deduction.create": "➖", "hr.allowance.create": "➕", "hr.commission.create": "📈",
    "hr.leave.create": "🏖️", "hr.leave.decision": "✅", "hr.payroll.create": "🧮",
    "hr.payroll.pay": "💳", "hr.payroll.delete": "🗑️", "hr.end_of_service.create": "🏁",
    "hr.document.create": "📎", "hr.document.file_upload": "📎", "hr.document.delete": "🗑️",
    "hr.document.renew": "🔄", "hr.travel.create": "✈️", "hr.travel.return": "🛬",
    "hr.custody.create": "📦", "hr.custody.return": "📦",
}


def _timeline_label(action: str, meta: dict) -> str:
    if action == "hr.employee.create":
        return "تم تعيين الموظف"
    if action == "hr.employee.update":
        return "تم تحديث بيانات الموظف"
    if action == "hr.employee.salary_change":
        return f"تم تعديل الراتب من {meta.get('old_salary')} إلى {meta.get('new_salary')}"
    if action == "hr.contract.create":
        return "تم إضافة عقد جديد"
    if action == "hr.advance.create":
        return "تم تسجيل سلفة"
    if action == "hr.withdrawal.create":
        return "تم تسجيل سحبية"
    if action == "hr.deduction.create":
        return "تم تسجيل استقطاع"
    if action == "hr.allowance.create":
        return "تم تسجيل بدل"
    if action == "hr.commission.create":
        return "تم تسجيل عمولة"
    if action == "hr.leave.create":
        return "تم تقديم طلب إجازة"
    if action == "hr.leave.decision":
        return "تم اعتماد الإجازة" if meta.get("decision") == "accepted" else "تم رفض طلب الإجازة"
    if action == "hr.payroll.create":
        return "تم إنشاء مسير راتب"
    if action == "hr.payroll.pay":
        return "تم صرف الراتب"
    if action == "hr.payroll.delete":
        return "تم حذف مسير راتب"
    if action == "hr.end_of_service.create":
        return "تم إنهاء خدمة الموظف"
    if action == "hr.document.create":
        return f"تم رفع وثيقة ({meta.get('doc_type', '')})"
    if action == "hr.document.file_upload":
        return "تم رفع ملف للوثيقة"
    if action == "hr.document.delete":
        return "تم حذف وثيقة"
    if action == "hr.document.renew":
        return f"تم تجديد وثيقة ({meta.get('doc_type', '')})"
    if action == "hr.travel.create":
        return "تم تسجيل سفر"
    if action == "hr.travel.return":
        return "تم تسجيل العودة من السفر"
    if action == "hr.custody.create":
        return f"تم تسليم عهدة: {meta.get('item_name', '')}"
    if action == "hr.custody.return":
        cond = meta.get("condition")
        return {"returned": "تم إرجاع العهدة", "lost": "تم تسجيل العهدة كمفقودة", "damaged": "تم تسجيل العهدة كتالفة"}.get(cond, "تم إرجاع العهدة")
    return action


@router.get("/employees/{employee_id}/timeline")
def employee_timeline(employee_id: str, request: Request, limit: int = Query(200, ge=1, le=1000)):
    db = SessionLocal()
    try:
        user = require_user(db, request)
        employee = db.query(Employee).filter(Employee.user_id == user.id, Employee.id == employee_id).first()
        if not employee:
            raise HTTPException(404, "الموظف غير موجود")
        rows = db.query(AuditLog).filter(
            AuditLog.employee_id == employee_id, AuditLog.user_id == user.id,
        ).order_by(AuditLog.created_at.desc()).limit(limit).all()
        items = []
        for r in rows:
            meta = r.meta_json or {}
            items.append({
                "id": r.id, "action": r.action, "icon": TIMELINE_ICONS.get(r.action, "•"),
                "label": _timeline_label(r.action, meta), "meta": meta,
                "created_at": _fmt_dt(r.created_at),
            })
        return {"items": items}
    finally:
        db.close()


@router.get("/employees/{employee_id}/profile-summary")
def employee_profile_summary(employee_id: str, request: Request):
    """بطاقة المعلومات السريعة + البطاقات الإحصائية أعلى ملف الموظف — كل الأرقام محسوبة لحظياً من القاعدة (لا تخزين مستقل)."""
    db = SessionLocal()
    try:
        user = require_user(db, request)
        e = db.query(Employee).filter(Employee.user_id == user.id, Employee.id == employee_id).first()
        if not e:
            raise HTTPException(404, "الموظف غير موجود")

        contract = db.query(EmployeeContract).filter(
            EmployeeContract.employee_id == employee_id, EmployeeContract.status == "active",
        ).order_by(EmployeeContract.created_at.desc()).first()

        last_attendance = db.query(Attendance).filter(Attendance.employee_id == employee_id).order_by(Attendance.date.desc()).first()
        last_leave = db.query(LeaveRequest).filter(LeaveRequest.employee_id == employee_id).order_by(LeaveRequest.created_at.desc()).first()
        last_advance = db.query(Advance).filter(Advance.employee_id == employee_id).order_by(Advance.created_at.desc()).first()
        last_payroll = db.query(Payroll).filter(Payroll.employee_id == employee_id).order_by(Payroll.year.desc(), Payroll.month.desc()).first()

        quick_info = _employee_out(e, db)
        quick_info.update({
            "contract_status": "ساري" if contract else "لا يوجد",
            "contract_type": contract.contract_type if contract else None,
            "contract_end_date": _fmt_date(contract.end_date) if contract else None,
            "residency_status": "منتهية" if (e.residency_expiry and (_days_left(e.residency_expiry) or 0) < 0) else ("سارية" if e.residency_expiry else "غير مسجّلة"),
            "passport_status": "منتهٍ" if (e.passport_expiry and (_days_left(e.passport_expiry) or 0) < 0) else ("ساري" if e.passport_expiry else "غير مسجّل"),
            "last_attendance_date": _fmt_date(last_attendance.date) if last_attendance else None,
            "last_leave": {"type": last_leave.leave_type, "start_date": _fmt_date(last_leave.start_date), "status": last_leave.status} if last_leave else None,
            "last_advance": {"amount": round(float(last_advance.amount or 0), 2), "date": _fmt_date(last_advance.date), "status": last_advance.status} if last_advance else None,
            "last_net_salary": round(float(last_payroll.net_salary or 0), 2) if last_payroll else None,
            "last_salary_period": f"{MONTHS_AR[last_payroll.month-1]} {last_payroll.year}" if last_payroll else None,
        })

        active_advances = db.query(Advance).filter(Advance.employee_id == employee_id, Advance.status == "active").all()
        all_advances = db.query(Advance).filter(Advance.employee_id == employee_id).all()
        pending_withdrawals = db.query(WithdrawalRecord).filter(WithdrawalRecord.employee_id == employee_id, WithdrawalRecord.status == "pending").all()
        pending_deductions = db.query(DeductionRecord).filter(DeductionRecord.employee_id == employee_id, DeductionRecord.applied == 0).all()
        pending_allowances = db.query(AllowanceRecord).filter(AllowanceRecord.employee_id == employee_id, AllowanceRecord.applied == 0).all()
        pending_commissions = db.query(CommissionRecord).filter(CommissionRecord.employee_id == employee_id, CommissionRecord.applied == 0).all()
        all_leaves = db.query(LeaveRequest).filter(LeaveRequest.employee_id == employee_id, LeaveRequest.status == "accepted").all()
        this_year = dt.datetime.utcnow().year
        leave_days_this_year = sum(l.days_count for l in all_leaves if l.start_date and l.start_date.year == this_year)
        annual_leave_days = contract.annual_leave_days if contract else 21
        absent_count = db.query(Attendance).filter(Attendance.employee_id == employee_id, Attendance.is_absent == 1).count()
        late_count = db.query(Attendance).filter(Attendance.employee_id == employee_id, Attendance.late_minutes > 0).count()
        all_payrolls = db.query(Payroll).filter(Payroll.employee_id == employee_id).all()
        total_salaries_paid = round(sum(float(p.net_salary or 0) for p in all_payrolls if p.status == "paid"), 2)
        gratuity_estimate = _compute_gratuity(e.basic_salary or 0.0, e.hire_date, dt.datetime.utcnow()) if e.hire_date else 0.0
        active_custody_count = db.query(CustodyRecord).filter(CustodyRecord.employee_id == employee_id, CustodyRecord.status == "assigned").count()
        travel_count = db.query(TravelRecord).filter(TravelRecord.employee_id == employee_id).count()

        stats = {
            "advances_total": round(sum(float(a.amount or 0) for a in all_advances), 2),
            "advances_count": len(all_advances),
            "advances_remaining": round(sum(float(a.remaining_amount or 0) for a in active_advances), 2),
            "withdrawals_pending_total": round(sum(float(w.amount or 0) for w in pending_withdrawals), 2),
            "withdrawals_pending_count": len(pending_withdrawals),
            "deductions_pending_total": round(sum(float(d.amount or 0) for d in pending_deductions), 2),
            "allowances_pending_total": round(sum(float(a.amount or 0) for a in pending_allowances), 2),
            "commissions_pending_total": round(sum(float(c.commission_value or 0) for c in pending_commissions), 2),
            "leave_days_used_this_year": leave_days_this_year,
            "leave_balance_days": max(annual_leave_days - leave_days_this_year, 0),
            "absent_days_count": absent_count,
            "late_count": late_count,
            "total_salaries_paid": total_salaries_paid,
            "eos_gratuity_estimate": round(gratuity_estimate, 2),
            "active_custody_count": active_custody_count,
            "travel_count": travel_count,
        }

        return {"quick_info": quick_info, "stats": stats}
    finally:
        db.close()


@router.patch("/employees/{employee_id}")
def update_employee(employee_id: str, request: Request, body: EmployeeUpdate = Body(...)):
    db = SessionLocal()
    try:
        require_csrf(request)
        user = require_user(db, request)
        rec = db.query(Employee).filter(Employee.user_id == user.id, Employee.id == employee_id).first()
        if not rec:
            raise HTTPException(404, "الموظف غير موجود")
        data = body.dict(exclude_unset=True)
        # ملاحظة معمارية: "passport_number" و"residency_expiry" و"passport_expiry" أُزيلت
        # عمدًا من التحديث المباشر — تبويب الوثائق هو المصدر الوحيد لها الآن (Single Source
        # of Truth)، وتتم مزامنتها تلقائيًا عبر _sync_identity_docs_from_documents.
        simple_fields = ["name", "nationality", "national_id", "position", "department", "branch_name", "phone", "email", "notes", "gender", "address", "manager_id", "bank_name", "iban"]
        for f in simple_fields:
            if f in data and data[f] is not None:
                setattr(rec, f, (data[f] or "").strip() or None)
        salary_changed = False
        old_salary = rec.basic_salary
        if "basic_salary" in data and data["basic_salary"] is not None:
            if round(float(data["basic_salary"]), 2) != round(float(old_salary or 0.0), 2):
                salary_changed = True
            rec.basic_salary = data["basic_salary"]
        if "bank_salary" in data:
            rec.bank_salary = data["bank_salary"]
        if "commission_percentage" in data:
            rec.commission_percentage = data["commission_percentage"]
        if "hire_date" in data:
            rec.hire_date = _parse_simple_date(data["hire_date"], "تاريخ التعيين")
        if "birth_date" in data:
            rec.birth_date = _parse_simple_date(data["birth_date"], "تاريخ الميلاد")
        if "status" in data and data["status"]:
            rec.status = data["status"]
            rec.is_active = 1 if data["status"] == "active" else 0
        if "is_active" in data and data["is_active"] is not None:
            rec.is_active = 1 if data["is_active"] else 0
        db.commit()
        if salary_changed:
            log_event(db, "hr.employee.salary_change", user.id, {"employee_id": rec.id, "old_salary": round(float(old_salary or 0.0), 2), "new_salary": round(float(rec.basic_salary or 0.0), 2)}, employee_id=rec.id)
        else:
            log_event(db, "hr.employee.update", user.id, {"employee_id": rec.id}, employee_id=rec.id)
        return _employee_out(rec, db)
    finally:
        db.close()


@router.delete("/employees/{employee_id}")
def delete_employee(employee_id: str, request: Request):
    db = SessionLocal()
    try:
        require_csrf(request)
        user = require_user(db, request)
        rec = db.query(Employee).filter(Employee.user_id == user.id, Employee.id == employee_id).first()
        if not rec:
            raise HTTPException(404, "الموظف غير موجود")
        has_payroll = db.query(Payroll).filter(Payroll.employee_id == employee_id).first()
        if has_payroll:
            raise HTTPException(400, "لا يمكن حذف موظف لديه سجل رواتب — يمكنك تغيير حالته إلى (منتهي الخدمة) بدلاً من الحذف")
        db.delete(rec)
        db.commit()
        log_event(db, "hr.employee.delete", user.id, {"employee_id": employee_id})
        return {"deleted": True}
    finally:
        db.close()


@router.post("/employees/{employee_id}/photo")
async def upload_employee_photo(employee_id: str, request: Request, file: UploadFile = File(...)):
    db = SessionLocal()
    try:
        require_csrf(request)
        user = require_user(db, request)
        rec = db.query(Employee).filter(Employee.user_id == user.id, Employee.id == employee_id).first()
        if not rec:
            raise HTTPException(404, "الموظف غير موجود")
        original = file.filename or "photo"
        lower = original.lower()
        if not any(lower.endswith(s) for s in _ALLOWED_IMAGE_SUFFIXES):
            raise HTTPException(400, "صيغة الصورة يجب أن تكون PNG أو JPG أو WEBP")
        content = await file.read()
        if len(content) > _MAX_DOC_MB * 1024 * 1024:
            raise HTTPException(400, f"حجم الصورة أكبر من {_MAX_DOC_MB} ميجابايت")
        suffix = _Path(original).suffix.lower() or ".jpg"
        saved_name = f"{uuid.uuid4().hex}{suffix}"
        with open(_HR_PHOTOS_DIR / saved_name, "wb") as f:
            f.write(content)
        rec.photo_url = f"/uploads/hr_photos/{saved_name}"
        db.commit()
        return _employee_out(rec)
    finally:
        db.close()



# ============================================================
# العقود
# ============================================================

def _contract_out(c: EmployeeContract) -> dict:
    return {
        "id": c.id,
        "employee_id": c.employee_id,
        "contract_type": c.contract_type or "",
        "start_date": _fmt_date(c.start_date),
        "end_date": _fmt_date(c.end_date),
        "end_days_left": _days_left(c.end_date),
        "salary": round(float(c.salary or 0.0), 2),
        "housing_allowance": round(float(c.housing_allowance or 0.0), 2),
        "transport_allowance": round(float(c.transport_allowance or 0.0), 2),
        "other_allowances": round(float(c.other_allowances or 0.0), 2),
        "probation_days": c.probation_days,
        "annual_leave_days": c.annual_leave_days,
        "file_url": c.file_url or "",
        "status": c.status,
        "created_at": _fmt_dt(c.created_at),
    }


@router.get("/contracts")
def list_contracts(request: Request, employee_id: str = Query("")):
    db = SessionLocal()
    try:
        user = require_user(db, request)
        q = db.query(EmployeeContract, Employee).join(Employee, Employee.id == EmployeeContract.employee_id).filter(EmployeeContract.user_id == user.id)
        if employee_id:
            q = q.filter(EmployeeContract.employee_id == employee_id)
        rows = q.order_by(EmployeeContract.created_at.desc()).all()
        items = []
        for c, e in rows:
            out = _contract_out(c)
            out["employee_name"] = e.name
            items.append(out)
        return {"items": items}
    finally:
        db.close()


@router.post("/employees/{employee_id}/contracts")
async def create_contract(
    employee_id: str,
    request: Request,
    contract_type: str = Form(""),
    start_date: str = Form(""),
    end_date: str = Form(""),
    salary: float = Form(0.0),
    housing_allowance: float = Form(0.0),
    transport_allowance: float = Form(0.0),
    other_allowances: float = Form(0.0),
    probation_days: int = Form(0),
    annual_leave_days: int = Form(21),
    file: Optional[UploadFile] = File(None),
):
    db = SessionLocal()
    try:
        require_csrf(request)
        user = require_user(db, request)
        employee = db.query(Employee).filter(Employee.user_id == user.id, Employee.id == employee_id).first()
        if not employee:
            raise HTTPException(404, "الموظف غير موجود")
        file_url = None
        if file is not None and getattr(file, "filename", ""):
            original = file.filename or "contract"
            lower = original.lower()
            if not any(lower.endswith(s) for s in _ALLOWED_DOC_SUFFIXES):
                raise HTTPException(400, "صيغة الملف يجب أن تكون PDF أو JPG أو PNG")
            content = await file.read()
            if len(content) > _MAX_DOC_MB * 1024 * 1024:
                raise HTTPException(400, f"حجم الملف أكبر من {_MAX_DOC_MB} ميجابايت")
            suffix = _Path(original).suffix.lower() or ".pdf"
            saved_name = f"{uuid.uuid4().hex}{suffix}"
            with open(_HR_CONTRACTS_DIR / saved_name, "wb") as f:
                f.write(content)
            file_url = f"/uploads/hr_contracts/{saved_name}"

        # عقد جديد ينهي أي عقد سابق ساري لنفس الموظف
        db.query(EmployeeContract).filter(
            EmployeeContract.employee_id == employee_id, EmployeeContract.status == "active"
        ).update({"status": "ended"})

        rec = EmployeeContract(
            id=uuid.uuid4().hex,
            user_id=user.id,
            employee_id=employee_id,
            contract_type=(contract_type or "").strip() or None,
            start_date=_parse_simple_date(start_date, "بداية العقد"),
            end_date=_parse_simple_date(end_date, "نهاية العقد"),
            salary=salary or 0.0,
            housing_allowance=housing_allowance or 0.0,
            transport_allowance=transport_allowance or 0.0,
            other_allowances=other_allowances or 0.0,
            probation_days=probation_days or 0,
            annual_leave_days=annual_leave_days or 21,
            file_url=file_url,
            status="active",
        )
        db.add(rec)

        # ترابط تلقائي: بدلات العقد تتحول لبدلات شهرية متكررة على الموظف دون إدخال يدوي
        db.query(AllowanceRecord).filter(
            AllowanceRecord.employee_id == employee_id, AllowanceRecord.recurring == 1, AllowanceRecord.applied == 0
        ).delete()
        for label, amount in (("بدل سكن (من العقد)", rec.housing_allowance), ("بدل نقل (من العقد)", rec.transport_allowance), ("بدلات أخرى (من العقد)", rec.other_allowances)):
            if amount and amount > 0:
                db.add(AllowanceRecord(
                    id=uuid.uuid4().hex, user_id=user.id, employee_id=employee_id,
                    allowance_type=label, amount=amount, recurring=1, date=dt.datetime.utcnow(), applied=0,
                ))

        if employee.basic_salary in (None, 0) and rec.salary:
            employee.basic_salary = rec.salary

        db.commit()
        log_event(db, "hr.contract.create", user.id, {"contract_id": rec.id, "employee_id": employee_id}, employee_id=employee_id)
        return _contract_out(rec)
    finally:
        db.close()


@router.delete("/contracts/{contract_id}")
def delete_contract(contract_id: str, request: Request):
    db = SessionLocal()
    try:
        require_csrf(request)
        user = require_user(db, request)
        rec = db.query(EmployeeContract).filter(EmployeeContract.user_id == user.id, EmployeeContract.id == contract_id).first()
        if not rec:
            raise HTTPException(404, "العقد غير موجود")
        has_payroll = db.query(Payroll).filter(Payroll.employee_id == rec.employee_id).first()
        if has_payroll:
            raise HTTPException(400, "لا يمكن حذف العقد لأن الموظف لديه سجل رواتب")
        db.delete(rec)
        db.commit()
        return {"deleted": True}
    finally:
        db.close()



# ============================================================
# السلف
# ============================================================

class AdvanceCreate(BaseModel):
    employee_id: str
    date: str = ""
    amount: float = Field(gt=0)
    reason: str = ""
    installments_count: int = Field(default=1, ge=1)


def _advance_out(a: Advance, employee_name: str = "") -> dict:
    return {
        "id": a.id, "employee_id": a.employee_id, "employee_name": employee_name,
        "date": _fmt_date(a.date), "amount": round(float(a.amount or 0.0), 2),
        "reason": a.reason or "", "installments_count": a.installments_count,
        "installment_amount": round(float(a.installment_amount or 0.0), 2),
        "remaining_amount": round(float(a.remaining_amount or 0.0), 2),
        "status": a.status,
    }


@router.get("/advances")
def list_advances(request: Request, employee_id: str = Query(""), status: str = Query("")):
    db = SessionLocal()
    try:
        user = require_user(db, request)
        q = db.query(Advance, Employee).join(Employee, Employee.id == Advance.employee_id).filter(Advance.user_id == user.id)
        if employee_id:
            q = q.filter(Advance.employee_id == employee_id)
        if status:
            q = q.filter(Advance.status == status)
        rows = q.order_by(Advance.created_at.desc()).all()
        return {"items": [_advance_out(a, e.name) for a, e in rows]}
    finally:
        db.close()


@router.post("/advances")
def create_advance(request: Request, body: AdvanceCreate = Body(...)):
    db = SessionLocal()
    try:
        require_csrf(request)
        user = require_user(db, request)
        employee = db.query(Employee).filter(Employee.user_id == user.id, Employee.id == body.employee_id).first()
        if not employee:
            raise HTTPException(404, "الموظف غير موجود")
        installment_amount = round(body.amount / max(body.installments_count, 1), 2)
        rec = Advance(
            id=uuid.uuid4().hex, user_id=user.id, employee_id=body.employee_id,
            date=_parse_simple_date(body.date, "التاريخ") or dt.datetime.utcnow(),
            amount=body.amount, reason=(body.reason or "").strip() or None,
            installments_count=body.installments_count, installment_amount=installment_amount,
            remaining_amount=body.amount, status="active",
        )
        db.add(rec)
        db.commit()
        log_event(db, "hr.advance.create", user.id, {"advance_id": rec.id}, employee_id=body.employee_id)
        return _advance_out(rec, employee.name)
    finally:
        db.close()


@router.delete("/advances/{advance_id}")
def delete_advance(advance_id: str, request: Request):
    db = SessionLocal()
    try:
        require_csrf(request)
        user = require_user(db, request)
        rec = db.query(Advance).filter(Advance.user_id == user.id, Advance.id == advance_id).first()
        if not rec:
            raise HTTPException(404, "السلفة غير موجودة")
        if rec.remaining_amount < rec.amount:
            raise HTTPException(400, "لا يمكن حذف سلفة تم خصم جزء منها بالفعل — يمكن إلغاؤها فقط")
        db.delete(rec)
        db.commit()
        return {"deleted": True}
    finally:
        db.close()


@router.post("/advances/{advance_id}/cancel")
def cancel_advance(advance_id: str, request: Request):
    db = SessionLocal()
    try:
        require_csrf(request)
        user = require_user(db, request)
        rec = db.query(Advance).filter(Advance.user_id == user.id, Advance.id == advance_id).first()
        if not rec:
            raise HTTPException(404, "السلفة غير موجودة")
        rec.status = "cancelled"
        db.commit()
        return _advance_out(rec)
    finally:
        db.close()


# ============================================================
# السحبيات (سحبية: مبلغ يُخصم دفعة واحدة من راتب الشهر القادم — بعكس السلفة التي تُقسّط)
# ============================================================

class WithdrawalCreate(BaseModel):
    employee_id: str
    date: str = ""
    amount: float = Field(gt=0)
    reason: str = ""


def _withdrawal_out(w: WithdrawalRecord, employee_name: str = "") -> dict:
    return {
        "id": w.id, "employee_id": w.employee_id, "employee_name": employee_name,
        "date": _fmt_date(w.date), "amount": round(float(w.amount or 0.0), 2),
        "reason": w.reason or "", "status": w.status,
        "status_label": {"pending": "معلّقة", "settled": "مُسوّاة", "cancelled": "ملغاة"}.get(w.status, w.status),
        "applied": bool(w.applied), "payroll_id": w.payroll_id,
    }


@router.get("/withdrawals")
def list_withdrawals(request: Request, employee_id: str = Query(""), status: str = Query("")):
    db = SessionLocal()
    try:
        user = require_user(db, request)
        q = db.query(WithdrawalRecord, Employee).join(Employee, Employee.id == WithdrawalRecord.employee_id).filter(WithdrawalRecord.user_id == user.id)
        if employee_id:
            q = q.filter(WithdrawalRecord.employee_id == employee_id)
        if status:
            q = q.filter(WithdrawalRecord.status == status)
        rows = q.order_by(WithdrawalRecord.created_at.desc()).all()
        return {"items": [_withdrawal_out(w, e.name) for w, e in rows]}
    finally:
        db.close()


@router.post("/withdrawals")
def create_withdrawal(request: Request, body: WithdrawalCreate = Body(...)):
    db = SessionLocal()
    try:
        require_csrf(request)
        user = require_user(db, request)
        employee = db.query(Employee).filter(Employee.user_id == user.id, Employee.id == body.employee_id).first()
        if not employee:
            raise HTTPException(404, "الموظف غير موجود")
        rec = WithdrawalRecord(
            id=uuid.uuid4().hex, user_id=user.id, employee_id=body.employee_id,
            date=_parse_simple_date(body.date, "التاريخ") or dt.datetime.utcnow(),
            amount=body.amount, reason=(body.reason or "").strip() or None, status="pending",
        )
        db.add(rec)
        db.commit()
        log_event(db, "hr.withdrawal.create", user.id, {"withdrawal_id": rec.id}, employee_id=body.employee_id)
        return _withdrawal_out(rec, employee.name)
    finally:
        db.close()


@router.delete("/withdrawals/{withdrawal_id}")
def delete_withdrawal(withdrawal_id: str, request: Request):
    db = SessionLocal()
    try:
        require_csrf(request)
        user = require_user(db, request)
        rec = db.query(WithdrawalRecord).filter(WithdrawalRecord.user_id == user.id, WithdrawalRecord.id == withdrawal_id).first()
        if not rec:
            raise HTTPException(404, "السحبية غير موجودة")
        if rec.applied:
            raise HTTPException(400, "لا يمكن حذف سحبية طُبّقت على راتب بالفعل — يمكن حذفها فقط قبل صرف الراتب المرتبط")
        db.delete(rec)
        db.commit()
        return {"deleted": True}
    finally:
        db.close()


@router.post("/withdrawals/{withdrawal_id}/cancel")
def cancel_withdrawal(withdrawal_id: str, request: Request):
    db = SessionLocal()
    try:
        require_csrf(request)
        user = require_user(db, request)
        rec = db.query(WithdrawalRecord).filter(WithdrawalRecord.user_id == user.id, WithdrawalRecord.id == withdrawal_id).first()
        if not rec:
            raise HTTPException(404, "السحبية غير موجودة")
        if rec.applied:
            raise HTTPException(400, "لا يمكن إلغاء سحبية طُبّقت على راتب بالفعل")
        rec.status = "cancelled"
        db.commit()
        return _withdrawal_out(rec)
    finally:
        db.close()


# ============================================================
# الاستقطاعات
# ============================================================

class DeductionCreate(BaseModel):
    employee_id: str
    deduction_type: str = ""
    amount: float = Field(gt=0)
    reason: str = ""
    date: str = ""


def _deduction_out(d: DeductionRecord, employee_name: str = "") -> dict:
    return {
        "id": d.id, "employee_id": d.employee_id, "employee_name": employee_name,
        "deduction_type": d.deduction_type or "", "amount": round(float(d.amount or 0.0), 2),
        "reason": d.reason or "", "date": _fmt_date(d.date),
        "applied": bool(d.applied), "payroll_id": d.payroll_id,
    }


@router.get("/deductions")
def list_deductions(request: Request, employee_id: str = Query(""), applied: str = Query("")):
    db = SessionLocal()
    try:
        user = require_user(db, request)
        q = db.query(DeductionRecord, Employee).join(Employee, Employee.id == DeductionRecord.employee_id).filter(DeductionRecord.user_id == user.id)
        if employee_id:
            q = q.filter(DeductionRecord.employee_id == employee_id)
        if applied == "0":
            q = q.filter(DeductionRecord.applied == 0)
        elif applied == "1":
            q = q.filter(DeductionRecord.applied == 1)
        rows = q.order_by(DeductionRecord.created_at.desc()).all()
        return {"items": [_deduction_out(d, e.name) for d, e in rows]}
    finally:
        db.close()


@router.post("/deductions")
def create_deduction(request: Request, body: DeductionCreate = Body(...)):
    db = SessionLocal()
    try:
        require_csrf(request)
        user = require_user(db, request)
        employee = db.query(Employee).filter(Employee.user_id == user.id, Employee.id == body.employee_id).first()
        if not employee:
            raise HTTPException(404, "الموظف غير موجود")
        rec = DeductionRecord(
            id=uuid.uuid4().hex, user_id=user.id, employee_id=body.employee_id,
            deduction_type=(body.deduction_type or "").strip() or None, amount=body.amount,
            reason=(body.reason or "").strip() or None,
            date=_parse_simple_date(body.date, "التاريخ") or dt.datetime.utcnow(), applied=0,
        )
        db.add(rec)
        db.commit()
        log_event(db, "hr.deduction.create", user.id, {"deduction_id": rec.id}, employee_id=body.employee_id)
        return _deduction_out(rec, employee.name)
    finally:
        db.close()


@router.delete("/deductions/{deduction_id}")
def delete_deduction(deduction_id: str, request: Request):
    db = SessionLocal()
    try:
        require_csrf(request)
        user = require_user(db, request)
        rec = db.query(DeductionRecord).filter(DeductionRecord.user_id == user.id, DeductionRecord.id == deduction_id).first()
        if not rec:
            raise HTTPException(404, "الاستقطاع غير موجود")
        if rec.applied:
            raise HTTPException(400, "لا يمكن حذف استقطاع تم تطبيقه على راتب بالفعل")
        db.delete(rec)
        db.commit()
        return {"deleted": True}
    finally:
        db.close()


# ============================================================
# البدلات
# ============================================================

class AllowanceCreate(BaseModel):
    employee_id: str
    allowance_type: str = ""
    amount: float = Field(gt=0)
    recurring: bool = False
    date: str = ""


def _allowance_out(a: AllowanceRecord, employee_name: str = "") -> dict:
    return {
        "id": a.id, "employee_id": a.employee_id, "employee_name": employee_name,
        "allowance_type": a.allowance_type or "", "amount": round(float(a.amount or 0.0), 2),
        "recurring": bool(a.recurring), "date": _fmt_date(a.date),
        "applied": bool(a.applied), "payroll_id": a.payroll_id,
    }


@router.get("/allowances")
def list_allowances(request: Request, employee_id: str = Query("")):
    db = SessionLocal()
    try:
        user = require_user(db, request)
        q = db.query(AllowanceRecord, Employee).join(Employee, Employee.id == AllowanceRecord.employee_id).filter(AllowanceRecord.user_id == user.id)
        if employee_id:
            q = q.filter(AllowanceRecord.employee_id == employee_id)
        rows = q.order_by(AllowanceRecord.created_at.desc()).all()
        return {"items": [_allowance_out(a, e.name) for a, e in rows]}
    finally:
        db.close()


@router.post("/allowances")
def create_allowance(request: Request, body: AllowanceCreate = Body(...)):
    db = SessionLocal()
    try:
        require_csrf(request)
        user = require_user(db, request)
        employee = db.query(Employee).filter(Employee.user_id == user.id, Employee.id == body.employee_id).first()
        if not employee:
            raise HTTPException(404, "الموظف غير موجود")
        rec = AllowanceRecord(
            id=uuid.uuid4().hex, user_id=user.id, employee_id=body.employee_id,
            allowance_type=(body.allowance_type or "").strip() or None, amount=body.amount,
            recurring=1 if body.recurring else 0,
            date=_parse_simple_date(body.date, "التاريخ") or dt.datetime.utcnow(), applied=0,
        )
        db.add(rec)
        db.commit()
        log_event(db, "hr.allowance.create", user.id, {"allowance_id": rec.id}, employee_id=body.employee_id)
        return _allowance_out(rec, employee.name)
    finally:
        db.close()


@router.delete("/allowances/{allowance_id}")
def delete_allowance(allowance_id: str, request: Request):
    db = SessionLocal()
    try:
        require_csrf(request)
        user = require_user(db, request)
        rec = db.query(AllowanceRecord).filter(AllowanceRecord.user_id == user.id, AllowanceRecord.id == allowance_id).first()
        if not rec:
            raise HTTPException(404, "البدل غير موجود")
        db.delete(rec)
        db.commit()
        return {"deleted": True}
    finally:
        db.close()


# ============================================================
# العمولات
# ============================================================

class CommissionCreate(BaseModel):
    employee_id: str
    sales_amount: float = Field(ge=0)
    percentage: Optional[float] = None
    date: str = ""


def _commission_out(c: CommissionRecord, employee_name: str = "") -> dict:
    return {
        "id": c.id, "employee_id": c.employee_id, "employee_name": employee_name,
        "sales_amount": round(float(c.sales_amount or 0.0), 2), "percentage": round(float(c.percentage or 0.0), 2),
        "commission_value": round(float(c.commission_value or 0.0), 2), "date": _fmt_date(c.date),
        "applied": bool(c.applied), "payroll_id": c.payroll_id,
    }


@router.get("/commissions")
def list_commissions(request: Request, employee_id: str = Query("")):
    db = SessionLocal()
    try:
        user = require_user(db, request)
        q = db.query(CommissionRecord, Employee).join(Employee, Employee.id == CommissionRecord.employee_id).filter(CommissionRecord.user_id == user.id)
        if employee_id:
            q = q.filter(CommissionRecord.employee_id == employee_id)
        rows = q.order_by(CommissionRecord.created_at.desc()).all()
        return {"items": [_commission_out(c, e.name) for c, e in rows]}
    finally:
        db.close()


@router.post("/commissions")
def create_commission(request: Request, body: CommissionCreate = Body(...)):
    db = SessionLocal()
    try:
        require_csrf(request)
        user = require_user(db, request)
        employee = db.query(Employee).filter(Employee.user_id == user.id, Employee.id == body.employee_id).first()
        if not employee:
            raise HTTPException(404, "الموظف غير موجود")
        pct = body.percentage if body.percentage is not None else (employee.commission_percentage or 0.0)
        value = round(body.sales_amount * (pct or 0.0) / 100.0, 2)
        rec = CommissionRecord(
            id=uuid.uuid4().hex, user_id=user.id, employee_id=body.employee_id,
            sales_amount=body.sales_amount, percentage=pct or 0.0, commission_value=value,
            date=_parse_simple_date(body.date, "التاريخ") or dt.datetime.utcnow(), applied=0,
        )
        db.add(rec)
        db.commit()
        log_event(db, "hr.commission.create", user.id, {"commission_id": rec.id}, employee_id=body.employee_id)
        return _commission_out(rec, employee.name)
    finally:
        db.close()


@router.delete("/commissions/{commission_id}")
def delete_commission(commission_id: str, request: Request):
    db = SessionLocal()
    try:
        require_csrf(request)
        user = require_user(db, request)
        rec = db.query(CommissionRecord).filter(CommissionRecord.user_id == user.id, CommissionRecord.id == commission_id).first()
        if not rec:
            raise HTTPException(404, "العمولة غير موجودة")
        db.delete(rec)
        db.commit()
        return {"deleted": True}
    finally:
        db.close()



# ============================================================
# الإجازات
# ============================================================

class LeaveCreate(BaseModel):
    employee_id: str
    leave_type: str = ""
    start_date: str
    end_date: str
    paid: bool = True


def _leave_out(l: LeaveRequest, employee_name: str = "") -> dict:
    return {
        "id": l.id, "employee_id": l.employee_id, "employee_name": employee_name,
        "leave_type": l.leave_type or "", "start_date": _fmt_date(l.start_date), "end_date": _fmt_date(l.end_date),
        "days_count": l.days_count, "paid": bool(l.paid), "status": l.status,
        "status_label": {"pending": "بانتظار الموافقة", "accepted": "مقبولة", "rejected": "مرفوضة"}.get(l.status, l.status),
        "applied": bool(l.applied),
    }


@router.get("/leaves")
def list_leaves(request: Request, employee_id: str = Query(""), status: str = Query("")):
    db = SessionLocal()
    try:
        user = require_user(db, request)
        q = db.query(LeaveRequest, Employee).join(Employee, Employee.id == LeaveRequest.employee_id).filter(LeaveRequest.user_id == user.id)
        if employee_id:
            q = q.filter(LeaveRequest.employee_id == employee_id)
        if status:
            q = q.filter(LeaveRequest.status == status)
        rows = q.order_by(LeaveRequest.created_at.desc()).all()
        return {"items": [_leave_out(l, e.name) for l, e in rows]}
    finally:
        db.close()


@router.post("/leaves")
def create_leave(request: Request, body: LeaveCreate = Body(...)):
    db = SessionLocal()
    try:
        require_csrf(request)
        user = require_user(db, request)
        employee = db.query(Employee).filter(Employee.user_id == user.id, Employee.id == body.employee_id).first()
        if not employee:
            raise HTTPException(404, "الموظف غير موجود")
        start = _parse_simple_date(body.start_date, "من تاريخ")
        end = _parse_simple_date(body.end_date, "إلى تاريخ")
        if not start or not end or end < start:
            raise HTTPException(400, "التواريخ غير صحيحة")
        days = (end - start).days + 1
        rec = LeaveRequest(
            id=uuid.uuid4().hex, user_id=user.id, employee_id=body.employee_id,
            leave_type=(body.leave_type or "").strip() or None, start_date=start, end_date=end,
            days_count=days, paid=1 if body.paid else 0, status="pending",
        )
        db.add(rec)
        db.commit()
        log_event(db, "hr.leave.create", user.id, {"leave_id": rec.id}, employee_id=body.employee_id)
        return _leave_out(rec, employee.name)
    finally:
        db.close()


@router.post("/leaves/{leave_id}/decision")
def decide_leave(leave_id: str, request: Request, decision: str = Body(..., embed=True)):
    db = SessionLocal()
    try:
        require_csrf(request)
        user = require_user(db, request)
        rec = db.query(LeaveRequest).filter(LeaveRequest.user_id == user.id, LeaveRequest.id == leave_id).first()
        if not rec:
            raise HTTPException(404, "طلب الإجازة غير موجود")
        if decision not in ("accepted", "rejected"):
            raise HTTPException(400, "قرار غير صالح")
        rec.status = decision
        db.commit()
        log_event(db, "hr.leave.decision", user.id, {"leave_id": rec.id, "decision": decision}, employee_id=rec.employee_id)
        return _leave_out(rec)
    finally:
        db.close()


@router.delete("/leaves/{leave_id}")
def delete_leave(leave_id: str, request: Request):
    db = SessionLocal()
    try:
        require_csrf(request)
        user = require_user(db, request)
        rec = db.query(LeaveRequest).filter(LeaveRequest.user_id == user.id, LeaveRequest.id == leave_id).first()
        if not rec:
            raise HTTPException(404, "طلب الإجازة غير موجود")
        db.delete(rec)
        db.commit()
        return {"deleted": True}
    finally:
        db.close()


# ============================================================
# الحضور والانصراف
# ============================================================

class AttendanceCreate(BaseModel):
    employee_id: str
    date: str
    check_in: str = ""
    check_out: str = ""
    late_minutes: int = 0
    overtime_minutes: int = 0
    is_absent: bool = False
    notes: str = ""


def _attendance_out(a: Attendance, employee_name: str = "") -> dict:
    return {
        "id": a.id, "employee_id": a.employee_id, "employee_name": employee_name,
        "date": _fmt_date(a.date), "check_in": a.check_in or "", "check_out": a.check_out or "",
        "late_minutes": a.late_minutes, "overtime_minutes": a.overtime_minutes,
        "is_absent": bool(a.is_absent), "notes": a.notes or "",
    }


@router.get("/attendance")
def list_attendance(request: Request, employee_id: str = Query(""), month: int = Query(0), year: int = Query(0)):
    db = SessionLocal()
    try:
        user = require_user(db, request)
        q = db.query(Attendance, Employee).join(Employee, Employee.id == Attendance.employee_id).filter(Attendance.user_id == user.id)
        if employee_id:
            q = q.filter(Attendance.employee_id == employee_id)
        rows = q.order_by(Attendance.date.desc()).all()
        if month:
            rows = [r for r in rows if r[0].date and r[0].date.month == month]
        if year:
            rows = [r for r in rows if r[0].date and r[0].date.year == year]
        return {"items": [_attendance_out(a, e.name) for a, e in rows]}
    finally:
        db.close()


@router.post("/attendance")
def create_attendance(request: Request, body: AttendanceCreate = Body(...)):
    db = SessionLocal()
    try:
        require_csrf(request)
        user = require_user(db, request)
        employee = db.query(Employee).filter(Employee.user_id == user.id, Employee.id == body.employee_id).first()
        if not employee:
            raise HTTPException(404, "الموظف غير موجود")
        day = _parse_simple_date(body.date, "التاريخ")
        rec = Attendance(
            id=uuid.uuid4().hex, user_id=user.id, employee_id=body.employee_id, date=day,
            check_in=(body.check_in or "").strip() or None, check_out=(body.check_out or "").strip() or None,
            late_minutes=body.late_minutes or 0, overtime_minutes=body.overtime_minutes or 0,
            is_absent=1 if body.is_absent else 0, notes=(body.notes or "").strip() or None,
        )
        db.add(rec)
        db.commit()
        return _attendance_out(rec, employee.name)
    finally:
        db.close()


@router.delete("/attendance/{attendance_id}")
def delete_attendance(attendance_id: str, request: Request):
    db = SessionLocal()
    try:
        require_csrf(request)
        user = require_user(db, request)
        rec = db.query(Attendance).filter(Attendance.user_id == user.id, Attendance.id == attendance_id).first()
        if not rec:
            raise HTTPException(404, "السجل غير موجود")
        db.delete(rec)
        db.commit()
        return {"deleted": True}
    finally:
        db.close()



# ============================================================
# الرواتب (مترابطة تلقائياً مع السلف/الاستقطاعات/البدلات/العمولات/الحضور/الإجازات)
# ============================================================

class LineItemIn(BaseModel):
    label: str = Field(min_length=1, max_length=120)
    amount: float = Field(ge=0)


class PayrollCreate(BaseModel):
    employee_id: str
    month: int = Field(ge=1, le=12)
    year: int = Field(ge=2000, le=2100)
    base_salary: Optional[float] = None
    extra_allowances: list[LineItemIn] = []
    extra_deductions: list[LineItemIn] = []
    notes: str = ""


def _payroll_out(db, p: Payroll, employee_name: str = "") -> dict:
    allowances = db.query(PayrollAllowance).filter(PayrollAllowance.payroll_id == p.id).all()
    deductions = db.query(PayrollDeduction).filter(PayrollDeduction.payroll_id == p.id).all()
    return {
        "id": p.id, "employee_id": p.employee_id, "employee_name": employee_name,
        "month": p.month, "year": p.year,
        "base_salary": round(float(p.base_salary or 0.0), 2),
        "total_allowances": round(float(p.total_allowances or 0.0), 2),
        "total_deductions": round(float(p.total_deductions or 0.0), 2),
        "gross_before_tax": round(float(p.gross_before_tax or 0.0), 2),
        "tax_amount": round(float(p.tax_amount or 0.0), 2),
        "net_salary": round(float(p.net_salary or 0.0), 2),
        "status": p.status,
        "paid_at": _fmt_dt(p.paid_at),
        "notes": p.notes or "",
        "allowances": [{"label": a.label, "amount": a.amount} for a in allowances],
        "deductions": [{"label": d.label, "amount": d.amount} for d in deductions],
    }


@router.get("/payrolls")
def list_payrolls(request: Request, month: int = Query(0), year: int = Query(0), status: str = Query(""), employee_id: str = Query("")):
    db = SessionLocal()
    try:
        user = require_user(db, request)
        q = db.query(Payroll, Employee).join(Employee, Employee.id == Payroll.employee_id).filter(Payroll.user_id == user.id)
        if month:
            q = q.filter(Payroll.month == month)
        if year:
            q = q.filter(Payroll.year == year)
        if status:
            q = q.filter(Payroll.status == status)
        if employee_id:
            q = q.filter(Payroll.employee_id == employee_id)
        rows = q.order_by(Payroll.year.desc(), Payroll.month.desc(), Payroll.created_at.desc()).all()
        return {"items": [_payroll_out(db, p, e.name) for p, e in rows]}
    finally:
        db.close()


@router.get("/payrolls/{payroll_id}")
def get_payroll(payroll_id: str, request: Request):
    db = SessionLocal()
    try:
        user = require_user(db, request)
        p = db.query(Payroll).filter(Payroll.user_id == user.id, Payroll.id == payroll_id).first()
        if not p:
            raise HTTPException(404, "الراتب غير موجود")
        e = db.query(Employee).filter(Employee.id == p.employee_id).first()
        return _payroll_out(db, p, e.name if e else "")
    finally:
        db.close()


def _build_payroll_record(db, user, employee, month: int, year: int, base_salary_override, extra_allowances: list, extra_deductions: list, notes: str):
    """يبني سجل راتب فعليًا من البيانات الحقيقية الحالية في القاعدة (بدون commit) —
    نفس منطق الحساب يُستخدم في الإنشاء وفي إعادة الحساب لاحقًا، بمصدر واحد للحقيقة
    لعملية البناء نفسها (لا يوجد منطقان مختلفان قد يتعارضان)."""
    base_salary = base_salary_override if base_salary_override is not None else (employee.basic_salary or 0.0)
    daily_rate = (base_salary / 30.0) if base_salary else 0.0

    allow_lines = []
    ded_lines = []

    # 1) البدلات (متكررة كل شهر + مرة واحدة غير مطبّقة)
    allowances = db.query(AllowanceRecord).filter(
        AllowanceRecord.user_id == user.id, AllowanceRecord.employee_id == employee.id,
    ).filter((AllowanceRecord.recurring == 1) | (AllowanceRecord.applied == 0)).all()
    consumed_allowance_ids = []
    for a in allowances:
        allow_lines.append((a.allowance_type or "بدل", a.amount))
        if not a.recurring:
            consumed_allowance_ids.append(a.id)

    # 2) العمولات غير المطبّقة
    commissions = db.query(CommissionRecord).filter(
        CommissionRecord.user_id == user.id, CommissionRecord.employee_id == employee.id, CommissionRecord.applied == 0,
    ).all()
    for c in commissions:
        allow_lines.append((f"عمولة ({c.percentage}% من {c.sales_amount})", c.commission_value))

    # 3) الاستقطاعات غير المطبّقة
    deductions = db.query(DeductionRecord).filter(
        DeductionRecord.user_id == user.id, DeductionRecord.employee_id == employee.id, DeductionRecord.applied == 0,
    ).all()
    for d in deductions:
        ded_lines.append((d.deduction_type or "استقطاع", d.amount))

    # 4) السلف السارية — خصم قسط واحد تلقائياً
    advances = db.query(Advance).filter(
        Advance.user_id == user.id, Advance.employee_id == employee.id, Advance.status == "active",
    ).all()
    for adv in advances:
        installment = min(adv.installment_amount, adv.remaining_amount)
        if installment > 0:
            ded_lines.append((f"قسط سلفة بتاريخ {_fmt_date(adv.date)}", installment))

    # 5) الغياب من سجل الحضور لنفس الشهر/السنة
    attendance_rows = db.query(Attendance).filter(
        Attendance.user_id == user.id, Attendance.employee_id == employee.id, Attendance.is_absent == 1,
    ).all()
    absent_days = sum(1 for a in attendance_rows if a.date and a.date.month == month and a.date.year == year)
    if absent_days and daily_rate:
        ded_lines.append((f"خصم غياب ({absent_days} يوم)", round(daily_rate * absent_days, 2)))

    # 6) الإجازات غير المدفوعة المقبولة ضمن نفس الشهر
    leaves = db.query(LeaveRequest).filter(
        LeaveRequest.user_id == user.id, LeaveRequest.employee_id == employee.id,
        LeaveRequest.status == "accepted", LeaveRequest.paid == 0, LeaveRequest.applied == 0,
    ).all()
    consumed_leave_ids = []
    for lv in leaves:
        in_month = (lv.start_date and lv.start_date.month == month and lv.start_date.year == year) or \
                   (lv.end_date and lv.end_date.month == month and lv.end_date.year == year)
        if in_month and daily_rate:
            ded_lines.append((f"خصم إجازة غير مدفوعة ({lv.days_count} يوم)", round(daily_rate * lv.days_count, 2)))
            consumed_leave_ids.append(lv.id)

    # 7) السحبيات المعلّقة — تُخصم بالكامل من راتب الشهر القادم
    withdrawals = db.query(WithdrawalRecord).filter(
        WithdrawalRecord.user_id == user.id, WithdrawalRecord.employee_id == employee.id, WithdrawalRecord.status == "pending",
    ).all()
    for w in withdrawals:
        ded_lines.append((f"سحبية بتاريخ {_fmt_date(w.date)}", w.amount))

    # 8) بنود يدوية إضافية من المستخدم
    for item in extra_allowances:
        allow_lines.append((item["label"], item["amount"]))
    for item in extra_deductions:
        ded_lines.append((item["label"], item["amount"]))

    total_allowances = round(sum(x[1] for x in allow_lines), 2)
    total_deductions_no_advance = round(sum(x[1] for x in ded_lines), 2)
    gross_before_tax = round(base_salary + total_allowances, 2)

    settings = _get_hr_settings(db)
    tax_amount = round(gross_before_tax * (settings.get("tax_percentage") or 0) / 100.0, 2) if settings.get("tax_enabled") else 0.0

    net_salary = round(gross_before_tax - tax_amount - total_deductions_no_advance, 2)

    payroll = Payroll(
        id=uuid.uuid4().hex, user_id=user.id, employee_id=employee.id,
        month=month, year=year, base_salary=base_salary,
        total_allowances=total_allowances, total_deductions=total_deductions_no_advance,
        gross_before_tax=gross_before_tax, tax_amount=tax_amount, net_salary=net_salary,
        status="unpaid", notes=(notes or "").strip() or None,
        extra_allowances_json=json.dumps(extra_allowances, ensure_ascii=False) if extra_allowances else None,
        extra_deductions_json=json.dumps(extra_deductions, ensure_ascii=False) if extra_deductions else None,
    )
    db.add(payroll)
    db.flush()

    for label, amount in allow_lines:
        db.add(PayrollAllowance(payroll_id=payroll.id, label=label, amount=amount))
    for label, amount in ded_lines:
        db.add(PayrollDeduction(payroll_id=payroll.id, label=label, amount=amount))

    # تحديث حالة السلف/الاستقطاعات/العمولات/البدلات/الإجازات المستهلكة
    for adv in advances:
        installment = min(adv.installment_amount, adv.remaining_amount)
        if installment > 0:
            adv.remaining_amount = round(adv.remaining_amount - installment, 2)
            if adv.remaining_amount <= 0.009:
                adv.remaining_amount = 0.0
                adv.status = "completed"
    for d in deductions:
        d.applied = 1
        d.payroll_id = payroll.id
    for c in commissions:
        c.applied = 1
        c.payroll_id = payroll.id
    for aid in consumed_allowance_ids:
        a = db.query(AllowanceRecord).get(aid)
        if a:
            a.applied = 1
            a.payroll_id = payroll.id
    for lid in consumed_leave_ids:
        lv = db.query(LeaveRequest).get(lid)
        if lv:
            lv.applied = 1
            lv.payroll_id = payroll.id
    for w in withdrawals:
        w.status = "settled"
        w.applied = 1
        w.payroll_id = payroll.id

    return payroll


@router.post("/payrolls")
def create_payroll(request: Request, body: PayrollCreate = Body(...)):
    db = SessionLocal()
    try:
        require_csrf(request)
        user = require_user(db, request)
        employee = db.query(Employee).filter(Employee.user_id == user.id, Employee.id == body.employee_id).first()
        if not employee:
            raise HTTPException(404, "الموظف غير موجود")
        existing = db.query(Payroll).filter(
            Payroll.employee_id == body.employee_id, Payroll.month == body.month, Payroll.year == body.year,
            Payroll.status != "cancelled",
        ).first()
        if existing:
            raise HTTPException(400, "يوجد بالفعل راتب لهذا الموظف لنفس الشهر والسنة")

        extra_allowances = [{"label": i.label, "amount": i.amount} for i in body.extra_allowances]
        extra_deductions = [{"label": i.label, "amount": i.amount} for i in body.extra_deductions]
        payroll = _build_payroll_record(
            db, user, employee, body.month, body.year, body.base_salary,
            extra_allowances, extra_deductions, body.notes,
        )
        db.commit()
        log_event(db, "hr.payroll.create", user.id, {"payroll_id": payroll.id}, employee_id=body.employee_id)
        return _payroll_out(db, payroll, employee.name)
    finally:
        db.close()


@router.post("/payrolls/{payroll_id}/recalculate")
def recalculate_payroll(payroll_id: str, request: Request):
    """يعيد بناء مسير الراتب من الصفر باستخدام أحدث بيانات حقيقية في القاعدة (سلف/استقطاعات/
    بدلات/عمولات/إجازات/سحبيات/حضور محدّثة) — وليس مجرد تعديل شكلي. يعمل فقط على راتب لم يُصرف
    بعد؛ البنود اليدوية التي أُدخلت عند الإنشاء تُعاد بدقّة من extra_allowances_json/extra_deductions_json."""
    db = SessionLocal()
    try:
        require_csrf(request)
        user = require_user(db, request)
        p = db.query(Payroll).filter(Payroll.user_id == user.id, Payroll.id == payroll_id).first()
        if not p:
            raise HTTPException(404, "الراتب غير موجود")
        if p.status == "paid":
            raise HTTPException(400, "لا يمكن إعادة حساب راتب تم صرفه بالفعل")
        employee = db.query(Employee).filter(Employee.id == p.employee_id).first()
        if not employee:
            raise HTTPException(404, "الموظف غير موجود")

        month, year, notes = p.month, p.year, p.notes
        extra_allowances = json.loads(p.extra_allowances_json) if p.extra_allowances_json else []
        extra_deductions = json.loads(p.extra_deductions_json) if p.extra_deductions_json else []

        _rollback_payroll_links(db, payroll_id)
        db.query(PayrollAllowance).filter(PayrollAllowance.payroll_id == payroll_id).delete()
        db.query(PayrollDeduction).filter(PayrollDeduction.payroll_id == payroll_id).delete()
        db.delete(p)
        db.flush()

        new_payroll = _build_payroll_record(
            db, user, employee, month, year, None, extra_allowances, extra_deductions, notes,
        )
        db.commit()
        log_event(db, "hr.payroll.recalculate", user.id, {"payroll_id": new_payroll.id}, employee_id=employee.id)
        return _payroll_out(db, new_payroll, employee.name)
    finally:
        db.close()


@router.post("/payrolls/{payroll_id}/pay")
def pay_payroll(payroll_id: str, request: Request):
    db = SessionLocal()
    try:
        require_csrf(request)
        user = require_user(db, request)
        p = db.query(Payroll).filter(Payroll.user_id == user.id, Payroll.id == payroll_id).first()
        if not p:
            raise HTTPException(404, "الراتب غير موجود")
        if p.status == "paid":
            raise HTTPException(400, "تم صرف هذا الراتب بالفعل")
        if p.status == "cancelled":
            raise HTTPException(400, "لا يمكن صرف راتب ملغى")
        p.status = "paid"
        p.paid_at = dt.datetime.utcnow()
        db.commit()
        log_event(db, "hr.payroll.pay", user.id, {"payroll_id": p.id}, employee_id=p.employee_id)
        e = db.query(Employee).filter(Employee.id == p.employee_id).first()
        return _payroll_out(db, p, e.name if e else "")
    finally:
        db.close()


def _rollback_payroll_links(db, payroll_id: str):
    for d in db.query(DeductionRecord).filter(DeductionRecord.payroll_id == payroll_id).all():
        d.applied = 0
        d.payroll_id = None
    for c in db.query(CommissionRecord).filter(CommissionRecord.payroll_id == payroll_id).all():
        c.applied = 0
        c.payroll_id = None
    for a in db.query(AllowanceRecord).filter(AllowanceRecord.payroll_id == payroll_id).all():
        a.applied = 0
        a.payroll_id = None
    for lv in db.query(LeaveRequest).filter(LeaveRequest.payroll_id == payroll_id).all():
        lv.applied = 0
        lv.payroll_id = None
    for w in db.query(WithdrawalRecord).filter(WithdrawalRecord.payroll_id == payroll_id).all():
        w.applied = 0
        w.payroll_id = None
        w.status = "pending"
    for line in db.query(PayrollDeduction).filter(PayrollDeduction.payroll_id == payroll_id).all():
        if line.label.startswith("قسط سلفة"):
            adv = db.query(Advance).filter(
                Advance.employee_id == db.query(Payroll).get(payroll_id).employee_id, Advance.status.in_(["active", "completed"]),
            ).order_by(Advance.created_at.desc()).first()
            if adv:
                adv.remaining_amount = round(min(adv.amount, adv.remaining_amount + line.amount), 2)
                if adv.remaining_amount > 0:
                    adv.status = "active"


@router.delete("/payrolls/{payroll_id}")
def delete_payroll(payroll_id: str, request: Request):
    db = SessionLocal()
    try:
        require_csrf(request)
        user = require_user(db, request)
        p = db.query(Payroll).filter(Payroll.user_id == user.id, Payroll.id == payroll_id).first()
        if not p:
            raise HTTPException(404, "الراتب غير موجود")
        if p.status == "paid":
            raise HTTPException(400, "لا يمكن حذف راتب تم صرفه بالفعل — يمكن إلغاؤه فقط إن لم يُصرف")
        emp_id_for_log = p.employee_id
        _rollback_payroll_links(db, payroll_id)
        db.query(PayrollAllowance).filter(PayrollAllowance.payroll_id == payroll_id).delete()
        db.query(PayrollDeduction).filter(PayrollDeduction.payroll_id == payroll_id).delete()
        db.delete(p)
        db.commit()
        log_event(db, "hr.payroll.delete", user.id, {"payroll_id": payroll_id}, employee_id=emp_id_for_log)
        return {"deleted": True}
    finally:
        db.close()


@router.post("/payrolls/{payroll_id}/cancel")
def cancel_payroll(payroll_id: str, request: Request):
    db = SessionLocal()
    try:
        require_csrf(request)
        user = require_user(db, request)
        p = db.query(Payroll).filter(Payroll.user_id == user.id, Payroll.id == payroll_id).first()
        if not p:
            raise HTTPException(404, "الراتب غير موجود")
        if p.status == "paid":
            raise HTTPException(400, "لا يمكن إلغاء راتب تم صرفه")
        _rollback_payroll_links(db, payroll_id)
        p.status = "cancelled"
        db.commit()
        e = db.query(Employee).filter(Employee.id == p.employee_id).first()
        return _payroll_out(db, p, e.name if e else "")
    finally:
        db.close()


PAYROLL_STATUS_LABELS = {"unpaid": "غير مصروف", "paid": "مصروف", "cancelled": "ملغى"}


def _payslip_rows(db, p: Payroll, employee_name: str):
    allowances = db.query(PayrollAllowance).filter(PayrollAllowance.payroll_id == p.id).all()
    deductions = db.query(PayrollDeduction).filter(PayrollDeduction.payroll_id == p.id).all()
    month_label = MONTHS_AR[p.month - 1] if 1 <= p.month <= 12 else str(p.month)
    rows = [
        ["الموظف", employee_name],
        ["الشهر", f"{month_label} {p.year}"],
        ["الراتب الأساسي", round(float(p.base_salary or 0.0), 2)],
    ]
    for a in allowances:
        rows.append([f"بدل: {a.label}", round(float(a.amount or 0.0), 2)])
    rows.append(["إجمالي البدلات", round(float(p.total_allowances or 0.0), 2)])
    rows.append(["الإجمالي قبل الضريبة", round(float(p.gross_before_tax or 0.0), 2)])
    if p.tax_amount:
        rows.append(["الضريبة", round(float(p.tax_amount or 0.0), 2)])
    for d in deductions:
        rows.append([f"استقطاع: {d.label}", round(float(d.amount or 0.0), 2)])
    rows.append(["إجمالي الاستقطاعات", round(float(p.total_deductions or 0.0), 2)])
    rows.append(["صافي الراتب", round(float(p.net_salary or 0.0), 2)])
    rows.append(["الحالة", PAYROLL_STATUS_LABELS.get(p.status, p.status)])
    if p.paid_at:
        rows.append(["تاريخ الصرف", _fmt_dt(p.paid_at)])
    if p.notes:
        rows.append(["ملاحظات", p.notes])
    return rows


@router.get("/payrolls/{payroll_id}/payslip/pdf")
def payslip_pdf(payroll_id: str, request: Request):
    db = SessionLocal()
    try:
        user = require_user(db, request)
        p = db.query(Payroll).filter(Payroll.user_id == user.id, Payroll.id == payroll_id).first()
        if not p:
            raise HTTPException(404, "الراتب غير موجود")
        e = db.query(Employee).filter(Employee.id == p.employee_id).first()
        employee_name = e.name if e else ""
        rows = _payslip_rows(db, p, employee_name)
        company_name, logo_url, generated_by = _pdf_meta(db, user)
        month_label = MONTHS_AR[p.month - 1] if 1 <= p.month <= 12 else str(p.month)
        title = f"قسيمة راتب — {employee_name} — {month_label} {p.year}"
        filename = f"قسيمة_راتب_{employee_name}_{p.month}_{p.year}.pdf"
        return generate_pdf_report(["البند", "القيمة"], rows, title, filename, company_name, logo_url, generated_by)
    finally:
        db.close()


@router.get("/payrolls/{payroll_id}/payslip/excel")
def payslip_excel(payroll_id: str, request: Request):
    db = SessionLocal()
    try:
        user = require_user(db, request)
        p = db.query(Payroll).filter(Payroll.user_id == user.id, Payroll.id == payroll_id).first()
        if not p:
            raise HTTPException(404, "الراتب غير موجود")
        e = db.query(Employee).filter(Employee.id == p.employee_id).first()
        employee_name = e.name if e else ""
        rows = _payslip_rows(db, p, employee_name)
        filename = f"قسيمة_راتب_{employee_name}_{p.month}_{p.year}.xlsx"
        return _xlsx_response(["البند", "القيمة"], rows, filename)
    finally:
        db.close()


@router.get("/dashboard-summary")
def hr_dashboard_summary(request: Request):
    db = SessionLocal()
    try:
        user = require_user(db, request)
        now = dt.datetime.utcnow()
        employees_count = db.query(Employee).filter(Employee.user_id == user.id, Employee.is_active == 1).count()
        rows = db.query(Payroll).filter(Payroll.user_id == user.id, Payroll.month == now.month, Payroll.year == now.year).all()
        total = round(sum(float(p.net_salary or 0) for p in rows), 2)
        paid = round(sum(float(p.net_salary or 0) for p in rows if p.status == "paid"), 2)
        unpaid = round(sum(float(p.net_salary or 0) for p in rows if p.status == "unpaid"), 2)
        advances_active = db.query(Advance).filter(Advance.user_id == user.id, Advance.status == "active").count()
        leaves_pending = db.query(LeaveRequest).filter(LeaveRequest.user_id == user.id, LeaveRequest.status == "pending").count()
        paid_employee_ids = {p.employee_id for p in rows if p.status == "paid"}
        active_employees = db.query(Employee).filter(Employee.user_id == user.id, Employee.status == "active").all()
        not_paid_yet = sum(1 for e in active_employees if e.id not in paid_employee_ids)
        absent_today = db.query(Attendance).filter(
            Attendance.user_id == user.id, Attendance.is_absent == 1,
            Attendance.date >= dt.datetime(now.year, now.month, now.day),
        ).count()
        active_emps = db.query(Employee).filter(Employee.user_id == user.id, Employee.is_active == 1).all()
        residency_expired_count = sum(1 for e in active_emps if (_days_left(e.residency_expiry) or 999) < 0)
        passport_expired_count = sum(1 for e in active_emps if (_days_left(e.passport_expiry) or 999) < 0)
        active_contracts = db.query(EmployeeContract).filter(EmployeeContract.user_id == user.id, EmployeeContract.status == "active").all()
        contracts_expiring_count = sum(
            1 for c in active_contracts if (_days_left(c.end_date) is not None and 0 <= _days_left(c.end_date) <= 30)
        )
        contracts_expired_count = sum(1 for c in active_contracts if (_days_left(c.end_date) or 999) < 0)
        today_start = dt.datetime(now.year, now.month, now.day)
        today_end = today_start + dt.timedelta(days=1)
        travel_active_count = db.query(TravelRecord).filter(
            TravelRecord.user_id == user.id, TravelRecord.status != "cancelled",
            TravelRecord.actual_return_date.is_(None),
            TravelRecord.departure_date <= now,
        ).count()
        leaves_today_count = db.query(LeaveRequest).filter(
            LeaveRequest.user_id == user.id, LeaveRequest.status == "accepted",
            LeaveRequest.start_date <= today_end, LeaveRequest.end_date >= today_start,
        ).count()
        advances_total = round(sum(
            float(a.remaining_amount or 0.0) for a in db.query(Advance).filter(Advance.user_id == user.id, Advance.status == "active").all()
        ), 2)
        withdrawals_total = round(sum(
            float(w.amount or 0.0) for w in db.query(WithdrawalRecord).filter(WithdrawalRecord.user_id == user.id, WithdrawalRecord.status == "pending").all()
        ), 2)
        allowances_total = round(sum(
            float(a.amount or 0.0) for a in db.query(AllowanceRecord).filter(AllowanceRecord.user_id == user.id, AllowanceRecord.applied == 0).all()
        ), 2)
        deductions_total = round(sum(
            float(d.amount or 0.0) for d in db.query(DeductionRecord).filter(DeductionRecord.user_id == user.id, DeductionRecord.applied == 0).all()
        ), 2)
        commissions_total = round(sum(
            float(c.commission_value or 0.0) for c in db.query(CommissionRecord).filter(CommissionRecord.user_id == user.id, CommissionRecord.applied == 0).all()
        ), 2)
        eos_total = round(sum(
            float(r.total_dues or 0.0) for r in db.query(EndOfService).filter(EndOfService.user_id == user.id).all()
        ), 2)
        return {
            "employees_count": employees_count,
            "total_this_month": total,
            "paid_amount": paid,
            "unpaid_amount": unpaid,
            "advances_active": advances_active,
            "leaves_pending": leaves_pending,
            "not_paid_yet": not_paid_yet,
            "absent_today": absent_today,
            "residency_expired_count": residency_expired_count,
            "passport_expired_count": passport_expired_count,
            "contracts_expiring_count": contracts_expiring_count,
            "contracts_expired_count": contracts_expired_count,
            "travel_active_count": travel_active_count,
            "leaves_today_count": leaves_today_count,
            "advances_total": advances_total,
            "withdrawals_total": withdrawals_total,
            "allowances_total": allowances_total,
            "deductions_total": deductions_total,
            "commissions_total": commissions_total,
            "eos_total": eos_total,
        }
    finally:
        db.close()



# ============================================================
# نهاية الخدمة
# ============================================================

class EndOfServiceCreate(BaseModel):
    employee_id: str
    end_date: str
    reason: str  # resigned|terminated|contract_end|retirement


EOS_REASON_LABELS = {"resigned": "مستقيل", "terminated": "مفصول", "contract_end": "انتهاء عقد", "retirement": "تقاعد"}


def _eos_out(r: EndOfService, employee_name: str = "") -> dict:
    return {
        "id": r.id, "employee_id": r.employee_id, "employee_name": employee_name,
        "end_date": _fmt_date(r.end_date), "reason": r.reason, "reason_label": EOS_REASON_LABELS.get(r.reason, r.reason),
        "remaining_salaries": round(float(r.remaining_salaries or 0.0), 2),
        "remaining_advances": round(float(r.remaining_advances or 0.0), 2),
        "deductions_total": round(float(r.deductions_total or 0.0), 2),
        "leave_balance_days": round(float(r.leave_balance_days or 0.0), 2),
        "leave_balance_value": round(float(r.leave_balance_value or 0.0), 2),
        "gratuity_amount": round(float(r.gratuity_amount or 0.0), 2),
        "total_dues": round(float(r.total_dues or 0.0), 2),
        "created_at": _fmt_dt(r.created_at),
    }


def _compute_gratuity(basic_salary: float, hire_date, end_date) -> float:
    """مكافأة نهاية خدمة مبسّطة على النمط السعودي: نصف شهر لكل سنة من أول 5 سنوات، وشهر كامل لكل سنة بعدها."""
    if not hire_date or not end_date or basic_salary <= 0:
        return 0.0
    years = max((end_date - hire_date).days / 365.25, 0)
    if years <= 5:
        return round(basic_salary * 0.5 * years, 2)
    return round(basic_salary * 0.5 * 5 + basic_salary * 1.0 * (years - 5), 2)


@router.get("/end-of-service")
def list_end_of_service(request: Request, employee_id: str = Query("")):
    db = SessionLocal()
    try:
        user = require_user(db, request)
        q = db.query(EndOfService, Employee).join(Employee, Employee.id == EndOfService.employee_id).filter(EndOfService.user_id == user.id)
        if employee_id:
            q = q.filter(EndOfService.employee_id == employee_id)
        rows = q.order_by(EndOfService.created_at.desc()).all()
        return {"items": [_eos_out(r, e.name) for r, e in rows]}
    finally:
        db.close()


@router.get("/employees/{employee_id}/end-of-service-preview")
def preview_end_of_service(employee_id: str, request: Request, end_date: str = Query(...)):
    db = SessionLocal()
    try:
        user = require_user(db, request)
        employee = db.query(Employee).filter(Employee.user_id == user.id, Employee.id == employee_id).first()
        if not employee:
            raise HTTPException(404, "الموظف غير موجود")
        end_dt = _parse_simple_date(end_date, "تاريخ نهاية الخدمة")

        unpaid_payrolls = db.query(Payroll).filter(Payroll.employee_id == employee_id, Payroll.status == "unpaid").all()
        remaining_salaries = round(sum(float(p.net_salary or 0) for p in unpaid_payrolls), 2)

        active_advances = db.query(Advance).filter(Advance.employee_id == employee_id, Advance.status == "active").all()
        remaining_advances = round(sum(float(a.remaining_amount or 0) for a in active_advances), 2)

        unapplied_deductions = db.query(DeductionRecord).filter(DeductionRecord.employee_id == employee_id, DeductionRecord.applied == 0).all()
        deductions_total = round(sum(float(d.amount or 0) for d in unapplied_deductions), 2)

        contract = db.query(EmployeeContract).filter(EmployeeContract.employee_id == employee_id, EmployeeContract.status == "active").order_by(EmployeeContract.created_at.desc()).first()
        annual_leave_days = contract.annual_leave_days if contract else 21
        hire_date = employee.hire_date
        years = max((end_dt - hire_date).days / 365.25, 0) if hire_date else 0
        earned_leave_days = round(annual_leave_days * years, 1)
        taken_leave_days = sum(
            l.days_count for l in db.query(LeaveRequest).filter(LeaveRequest.employee_id == employee_id, LeaveRequest.status == "accepted").all()
        )
        leave_balance_days = max(round(earned_leave_days - taken_leave_days, 1), 0)
        daily_rate = (employee.basic_salary or 0.0) / 30.0
        leave_balance_value = round(leave_balance_days * daily_rate, 2)

        gratuity = _compute_gratuity(employee.basic_salary or 0.0, hire_date, end_dt)

        total_dues = round(remaining_salaries - remaining_advances - deductions_total + leave_balance_value + gratuity, 2)

        return {
            "remaining_salaries": remaining_salaries, "remaining_advances": remaining_advances,
            "deductions_total": deductions_total, "leave_balance_days": leave_balance_days,
            "leave_balance_value": leave_balance_value, "gratuity_amount": gratuity, "total_dues": total_dues,
        }
    finally:
        db.close()


@router.post("/employees/{employee_id}/end-of-service")
def create_end_of_service(employee_id: str, request: Request, body: EndOfServiceCreate = Body(...)):
    db = SessionLocal()
    try:
        require_csrf(request)
        user = require_user(db, request)
        employee = db.query(Employee).filter(Employee.user_id == user.id, Employee.id == employee_id).first()
        if not employee:
            raise HTTPException(404, "الموظف غير موجود")
        existing_eos = db.query(EndOfService).filter(EndOfService.employee_id == employee_id).first()
        if existing_eos:
            raise HTTPException(400, "هذا الموظف لديه سجل نهاية خدمة مسبق بتاريخ " + _fmt_date(existing_eos.end_date) + " -- لا يمكن إنهاء الخدمة أكثر من مرة لنفس الموظف")
        end_dt = _parse_simple_date(body.end_date, "تاريخ نهاية الخدمة")
        unpaid_payrolls = db.query(Payroll).filter(Payroll.employee_id == employee_id, Payroll.status == "unpaid").all()
        remaining_salaries = round(sum(float(p.net_salary or 0) for p in unpaid_payrolls), 2)
        active_advances = db.query(Advance).filter(Advance.employee_id == employee_id, Advance.status == "active").all()
        remaining_advances = round(sum(float(a.remaining_amount or 0) for a in active_advances), 2)
        unapplied_deductions = db.query(DeductionRecord).filter(DeductionRecord.employee_id == employee_id, DeductionRecord.applied == 0).all()
        deductions_total = round(sum(float(d.amount or 0) for d in unapplied_deductions), 2)
        contract = db.query(EmployeeContract).filter(EmployeeContract.employee_id == employee_id, EmployeeContract.status == "active").order_by(EmployeeContract.created_at.desc()).first()
        annual_leave_days = contract.annual_leave_days if contract else 21
        hire_date = employee.hire_date
        years = max((end_dt - hire_date).days / 365.25, 0) if hire_date else 0
        earned_leave_days = round(annual_leave_days * years, 1)
        taken_leave_days = sum(l.days_count for l in db.query(LeaveRequest).filter(LeaveRequest.employee_id == employee_id, LeaveRequest.status == "accepted").all())
        leave_balance_days = max(round(earned_leave_days - taken_leave_days, 1), 0)
        daily_rate = (employee.basic_salary or 0.0) / 30.0
        leave_balance_value = round(leave_balance_days * daily_rate, 2)
        gratuity = _compute_gratuity(employee.basic_salary or 0.0, hire_date, end_dt)
        total_dues = round(remaining_salaries - remaining_advances - deductions_total + leave_balance_value + gratuity, 2)

        rec = EndOfService(
            id=uuid.uuid4().hex, user_id=user.id, employee_id=employee_id, end_date=end_dt, reason=body.reason,
            remaining_salaries=remaining_salaries, remaining_advances=remaining_advances,
            deductions_total=deductions_total, leave_balance_days=leave_balance_days,
            leave_balance_value=leave_balance_value, gratuity_amount=gratuity, total_dues=total_dues,
        )
        db.add(rec)
        employee.status = "terminated"
        employee.is_active = 0
        db.commit()
        log_event(db, "hr.end_of_service.create", user.id, {"employee_id": employee_id}, employee_id=employee_id)
        return _eos_out(rec, employee.name)
    finally:
        db.close()

def _doc_status(expiry_date) -> dict:
    if not expiry_date:
        return {"status": "active", "status_label": "سارية", "severity": "none", "days_left": None}
    today = dt.datetime.utcnow().date()
    exp = expiry_date.date() if hasattr(expiry_date, "date") else expiry_date
    days_left = (exp - today).days
    if days_left < 0:
        return {"status": "expired", "status_label": "منتهية", "severity": "expired", "days_left": days_left}
    if days_left <= 7:
        return {"status": "expiring_7", "status_label": f"تنتهي خلال {days_left} يوم", "severity": "critical", "days_left": days_left}
    if days_left <= 15:
        return {"status": "expiring_15", "status_label": f"تنتهي خلال {days_left} يوم", "severity": "high", "days_left": days_left}
    if days_left <= 30:
        return {"status": "expiring_30", "status_label": f"تنتهي خلال {days_left} يوم", "severity": "medium", "days_left": days_left}
    if days_left <= 60:
        return {"status": "expiring_60", "status_label": f"تنتهي خلال {days_left} يوم", "severity": "notice", "days_left": days_left}
    return {"status": "active", "status_label": "سارية", "severity": "none", "days_left": days_left}


_RESIDENCY_DOC_MARKERS = ("إقامة",)
_PASSPORT_DOC_MARKERS = ("جواز",)


def _sync_identity_docs_from_documents(db, employee_id: str):
    """الوثائق هي المصدر الوحيد للحقيقة لبيانات الإقامة/الجواز. هذه الدالة تُحدّث
    أعمدة الموظف (المستخدمة في التنبيهات والتقارير السريعة) تلقائيًا من أحدث وثيقة
    محفوظة من نوع (إقامة/جواز) — بحيث لا يحتاج المستخدم لإدخال نفس البيانات مرتين،
    ولا يمكن أن تتعارض التنبيهات مع ما هو مسجّل فعليًا في تبويب الوثائق."""
    employee = db.query(Employee).filter(Employee.id == employee_id).first()
    if not employee:
        return
    docs = db.query(EmployeeDocument).filter(EmployeeDocument.employee_id == employee_id).all()

    residency_docs = [d for d in docs if any(m in (d.doc_type or "") for m in _RESIDENCY_DOC_MARKERS)]
    latest_residency = max(residency_docs, key=lambda d: (d.expiry_date or dt.datetime.min), default=None)
    employee.residency_expiry = latest_residency.expiry_date if latest_residency else None

    passport_docs = [d for d in docs if any(m in (d.doc_type or "") for m in _PASSPORT_DOC_MARKERS)]
    latest_passport = max(passport_docs, key=lambda d: (d.expiry_date or dt.datetime.min), default=None)
    employee.passport_expiry = latest_passport.expiry_date if latest_passport else None
    employee.passport_number = latest_passport.doc_number if latest_passport else None

    db.commit()


def _upsert_identity_document(db, user_id: str, employee_id: str, doc_type_marker: str, doc_number: str, expiry_date):
    """يُستخدم من مسارات الإدخال الجماعي (استيراد Excel) — بدل ما يكتب مباشرة على
    عمود الموظف، ينشئ/يحدّث وثيقة حقيقية في جدول الوثائق (نفس مصدر تبويب الوثائق)،
    ثم تتم المزامنة التلقائية لعمود الموظف من نفس الدالة أعلاه."""
    if not expiry_date and not (doc_number or "").strip():
        return
    existing = (
        db.query(EmployeeDocument)
        .filter(EmployeeDocument.employee_id == employee_id, EmployeeDocument.doc_type.contains(doc_type_marker))
        .order_by(EmployeeDocument.created_at.desc())
        .first()
    )
    if existing:
        if expiry_date:
            existing.expiry_date = expiry_date
        if (doc_number or "").strip():
            existing.doc_number = doc_number.strip()
    else:
        default_type = "إقامة" if doc_type_marker == "إقامة" else "جواز سفر"
        db.add(EmployeeDocument(
            id=uuid.uuid4().hex, user_id=user_id, employee_id=employee_id,
            doc_type=default_type, doc_number=(doc_number or "").strip() or None,
            expiry_date=expiry_date,
        ))


def _doc_out(d: EmployeeDocument) -> dict:
    out = {
        "id": d.id,
        "employee_id": d.employee_id,
        "doc_type": d.doc_type,
        "doc_number": d.doc_number or "",
        "issue_date": d.issue_date.strftime("%Y-%m-%d") if d.issue_date else "",
        "expiry_date": d.expiry_date.strftime("%Y-%m-%d") if d.expiry_date else "",
        "issuing_authority": d.issuing_authority or "",
        "notes": d.notes or "",
        "file_url": d.file_url or "",
        "updated_at": d.updated_at.strftime("%Y-%m-%d %H:%M") if d.updated_at else "",
    }
    out.update(_doc_status(d.expiry_date))
    return out

@router.get("/employees/{employee_id}/documents")
def list_employee_documents(employee_id: str, request: Request):
    db = SessionLocal()
    try:
        user = require_user(db, request)
        employee = db.query(Employee).filter(Employee.user_id == user.id, Employee.id == employee_id).first()
        if not employee:
            raise HTTPException(404, "الموظف غير موجود")
        rows = (
            db.query(EmployeeDocument)
            .filter(EmployeeDocument.user_id == user.id, EmployeeDocument.employee_id == employee_id)
            .order_by(EmployeeDocument.created_at.desc())
            .all()
        )
        items = [_doc_out(r) for r in rows]
        total = len(items)
        active = sum(1 for x in items if x["status"] == "active")
        expiring_30 = sum(1 for x in items if x["days_left"] is not None and 0 <= x["days_left"] <= 30)
        expired = sum(1 for x in items if x["status"] == "expired")
        last_updated = max((x["updated_at"] for x in items), default="")
        return {
            "items": items,
            "stats": {
                "total": total,
                "active": active,
                "expiring_30": expiring_30,
                "expired": expired,
                "last_updated": last_updated,
            },
        }
    finally:
        db.close()


@router.post("/employees/{employee_id}/documents")
async def create_employee_document(
    employee_id: str,
    request: Request,
    doc_type: str = Form(...),
    doc_number: str = Form(""),
    issue_date: str = Form(""),
    expiry_date: str = Form(""),
    issuing_authority: str = Form(""),
    notes: str = Form(""),
    file: Optional[UploadFile] = File(None),
):
    db = SessionLocal()
    try:
        require_csrf(request)
        user = require_user(db, request)
        employee = db.query(Employee).filter(Employee.user_id == user.id, Employee.id == employee_id).first()
        if not employee:
            raise HTTPException(404, "الموظف غير موجود")
        if not (doc_type or "").strip():
            raise HTTPException(400, "نوع الوثيقة مطلوب")
        file_url = None
        if file is not None and getattr(file, "filename", ""):
            original = file.filename or "doc"
            lower = original.lower()
            if not any(lower.endswith(s) for s in _ALLOWED_DOC_SUFFIXES):
                raise HTTPException(400, "صيغة الملف يجب أن تكون PDF أو JPG أو PNG")
            content = await file.read()
            if len(content) > _MAX_DOC_MB * 1024 * 1024:
                raise HTTPException(400, f"حجم الملف أكبر من {_MAX_DOC_MB} ميجابايت")
            suffix = _Path(original).suffix.lower() or ".pdf"
            saved_name = f"{uuid.uuid4().hex}{suffix}"
            with open(_HR_DOCS_DIR / saved_name, "wb") as f:
                f.write(content)
            file_url = f"/uploads/hr_documents/{saved_name}"
        rec = EmployeeDocument(
            id=uuid.uuid4().hex,
            user_id=user.id,
            employee_id=employee_id,
            doc_type=doc_type.strip(),
            doc_number=(doc_number or "").strip() or None,
            issue_date=_parse_simple_date(issue_date, "تاريخ الإصدار"),
            expiry_date=_parse_simple_date(expiry_date, "تاريخ الانتهاء"),
            issuing_authority=(issuing_authority or "").strip() or None,
            notes=(notes or "").strip() or None,
            file_url=file_url,
        )
        db.add(rec)
        db.commit()
        _sync_identity_docs_from_documents(db, employee_id)
        log_event(db, "hr.document.create", user.id, {"document_id": rec.id, "employee_id": employee_id, "doc_type": rec.doc_type}, employee_id=employee_id)
        return _doc_out(rec)
    finally:
        db.close()


class DocumentUpdate(BaseModel):
    doc_type: Optional[str] = None
    doc_number: Optional[str] = None
    issue_date: Optional[str] = None
    expiry_date: Optional[str] = None
    issuing_authority: Optional[str] = None
    notes: Optional[str] = None


@router.patch("/documents/{document_id}")
def update_document(document_id: str, request: Request, body: DocumentUpdate = Body(...)):
    db = SessionLocal()
    try:
        require_csrf(request)
        user = require_user(db, request)
        rec = db.query(EmployeeDocument).filter(EmployeeDocument.user_id == user.id, EmployeeDocument.id == document_id).first()
        if not rec:
            raise HTTPException(404, "الوثيقة غير موجودة")
        if body.doc_type is not None and body.doc_type.strip():
            rec.doc_type = body.doc_type.strip()
        if body.doc_number is not None:
            rec.doc_number = body.doc_number.strip() or None
        if body.issue_date is not None:
            rec.issue_date = _parse_simple_date(body.issue_date, "تاريخ الإصدار")
        if body.expiry_date is not None:
            rec.expiry_date = _parse_simple_date(body.expiry_date, "تاريخ الانتهاء")
        if body.issuing_authority is not None:
            rec.issuing_authority = body.issuing_authority.strip() or None
        if body.notes is not None:
            rec.notes = body.notes.strip() or None
        db.commit()
        _sync_identity_docs_from_documents(db, rec.employee_id)
        return _doc_out(rec)
    finally:
        db.close()


@router.post("/documents/{document_id}/file")
async def upload_document_file(document_id: str, request: Request, file: UploadFile = File(...)):
    db = SessionLocal()
    try:
        require_csrf(request)
        user = require_user(db, request)
        rec = db.query(EmployeeDocument).filter(EmployeeDocument.user_id == user.id, EmployeeDocument.id == document_id).first()
        if not rec:
            raise HTTPException(404, "الوثيقة غير موجودة")
        original = file.filename or "doc"
        lower = original.lower()
        if not any(lower.endswith(s) for s in _ALLOWED_DOC_SUFFIXES):
            raise HTTPException(400, "صيغة الملف يجب أن تكون PDF أو JPG أو PNG")
        content = await file.read()
        if len(content) > _MAX_DOC_MB * 1024 * 1024:
            raise HTTPException(400, f"حجم الملف أكبر من {_MAX_DOC_MB} ميجابايت")
        suffix = _Path(original).suffix.lower() or ".pdf"
        saved_name = f"{uuid.uuid4().hex}{suffix}"
        with open(_HR_DOCS_DIR / saved_name, "wb") as f:
            f.write(content)
        rec.file_url = f"/uploads/hr_documents/{saved_name}"
        db.commit()
        log_event(db, "hr.document.file_upload", user.id, {"document_id": rec.id}, employee_id=rec.employee_id)
        return _doc_out(rec)
    finally:
        db.close()


@router.delete("/documents/{document_id}")
def delete_document(document_id: str, request: Request):
    db = SessionLocal()
    try:
        require_csrf(request)
        user = require_user(db, request)
        rec = db.query(EmployeeDocument).filter(EmployeeDocument.user_id == user.id, EmployeeDocument.id == document_id).first()
        if not rec:
            raise HTTPException(404, "الوثيقة غير موجودة")
        emp_id_for_sync = rec.employee_id
        db.query(EmployeeDocumentRenewal).filter(EmployeeDocumentRenewal.document_id == rec.id).delete()
        db.delete(rec)
        db.commit()
        _sync_identity_docs_from_documents(db, emp_id_for_sync)
        log_event(db, "hr.document.delete", user.id, {"document_id": document_id}, employee_id=emp_id_for_sync)
        return {"deleted": True}
    finally:
        db.close()


class DocumentRenew(BaseModel):
    new_issue_date: str = ""
    new_expiry_date: str = ""
    notes: str = ""


@router.post("/documents/{document_id}/renew")
def renew_document(document_id: str, request: Request, body: DocumentRenew = Body(...)):
    db = SessionLocal()
    try:
        require_csrf(request)
        user = require_user(db, request)
        rec = db.query(EmployeeDocument).filter(EmployeeDocument.user_id == user.id, EmployeeDocument.id == document_id).first()
        if not rec:
            raise HTTPException(404, "الوثيقة غير موجودة")
        new_issue = _parse_simple_date(body.new_issue_date, "تاريخ الإصدار الجديد")
        new_expiry = _parse_simple_date(body.new_expiry_date, "تاريخ الانتهاء الجديد")
        if not new_expiry:
            raise HTTPException(400, "تاريخ الانتهاء الجديد مطلوب")
        history = EmployeeDocumentRenewal(
            id=uuid.uuid4().hex,
            document_id=rec.id,
            old_issue_date=rec.issue_date,
            old_expiry_date=rec.expiry_date,
            renewed_by_name=user.username,
            notes=(body.notes or "").strip() or None,
        )
        db.add(history)
        if new_issue:
            rec.issue_date = new_issue
        rec.expiry_date = new_expiry
        db.commit()
        _sync_identity_docs_from_documents(db, rec.employee_id)
        log_event(db, "hr.document.renew", user.id, {"document_id": rec.id, "doc_type": rec.doc_type}, employee_id=rec.employee_id)
        return _doc_out(rec)
    finally:
        db.close()


@router.get("/documents/{document_id}/history")
def document_history(document_id: str, request: Request):
    db = SessionLocal()
    try:
        user = require_user(db, request)
        rec = db.query(EmployeeDocument).filter(EmployeeDocument.user_id == user.id, EmployeeDocument.id == document_id).first()
        if not rec:
            raise HTTPException(404, "الوثيقة غير موجودة")
        rows = (
            db.query(EmployeeDocumentRenewal)
            .filter(EmployeeDocumentRenewal.document_id == document_id)
            .order_by(EmployeeDocumentRenewal.renewed_at.desc())
            .all()
        )
        return {
            "items": [
                {
                    "old_issue_date": r.old_issue_date.strftime("%Y-%m-%d") if r.old_issue_date else "",
                    "old_expiry_date": r.old_expiry_date.strftime("%Y-%m-%d") if r.old_expiry_date else "",
                    "renewed_at": r.renewed_at.strftime("%Y-%m-%d %H:%M") if r.renewed_at else "",
                    "renewed_by_name": r.renewed_by_name or "",
                    "notes": r.notes or "",
                }
                for r in rows
            ]
        }
    finally:
        db.close()


@router.get("/documents/alerts")
def documents_alerts(
    request: Request,
    doc_type: str = Query(""),
    department: str = Query(""),
    branch_name: str = Query(""),
    employee_id: str = Query(""),
    status: str = Query(""),
    q: str = Query(""),
    only_attention: bool = Query(False),
    limit: int = Query(500, ge=1, le=2000),
):
    db = SessionLocal()
    try:
        user = require_user(db, request)
        query = (
            db.query(EmployeeDocument, Employee)
            .join(Employee, Employee.id == EmployeeDocument.employee_id)
            .filter(EmployeeDocument.user_id == user.id)
        )
        if (doc_type or "").strip():
            query = query.filter(EmployeeDocument.doc_type == doc_type.strip())
        if (department or "").strip():
            query = query.filter(Employee.department == department.strip())
        if (branch_name or "").strip():
            query = query.filter(Employee.branch_name == branch_name.strip())
        if (employee_id or "").strip():
            query = query.filter(EmployeeDocument.employee_id == employee_id.strip())
        rows = query.order_by(EmployeeDocument.expiry_date.asc().nullslast()).limit(limit).all()
        qn = (q or "").strip().lower()
        items = []
        for doc, emp in rows:
            out = _doc_out(doc)
            out["employee_name"] = emp.name
            out["department"] = emp.department or ""
            out["branch_name"] = emp.branch_name or ""
            if qn and qn not in (f"{emp.name} {doc.doc_number or ''} {doc.doc_type}").lower():
                continue
            if status and out["status"] != status:
                continue
            if only_attention and out["status"] == "active":
                continue
            items.append(out)
        return {"items": items}
    finally:
        db.close()



# ============================================================
# السفر
# ============================================================

TRAVEL_TYPE_LABELS = {"domestic": "داخل المملكة", "international": "خارج المملكة"}
TRAVEL_STATUS_LABELS = {"scheduled": "مجدول", "ongoing": "قيد السفر", "completed": "مكتمل", "cancelled": "ملغى"}


class TravelCreate(BaseModel):
    employee_id: str
    travel_type: str = "domestic"
    destination: str = ""
    departure_date: str
    return_date: str = ""
    purpose: str = ""
    notes: str = ""


def _travel_status(t: TravelRecord) -> str:
    if t.status == "cancelled":
        return "cancelled"
    if t.actual_return_date:
        return "completed"
    today = dt.datetime.utcnow()
    if t.departure_date and today < t.departure_date:
        return "scheduled"
    if t.return_date and today > t.return_date:
        return "completed"
    return "ongoing"


def _travel_out(t: TravelRecord, employee_name: str = "") -> dict:
    status = _travel_status(t)
    return {
        "id": t.id, "employee_id": t.employee_id, "employee_name": employee_name,
        "travel_type": t.travel_type, "travel_type_label": TRAVEL_TYPE_LABELS.get(t.travel_type, t.travel_type),
        "destination": t.destination or "", "departure_date": _fmt_date(t.departure_date),
        "return_date": _fmt_date(t.return_date), "actual_return_date": _fmt_date(t.actual_return_date),
        "purpose": t.purpose or "", "notes": t.notes or "",
        "status": status, "status_label": TRAVEL_STATUS_LABELS.get(status, status),
    }


@router.get("/travel")
def list_travel(request: Request, employee_id: str = Query(""), status: str = Query("")):
    db = SessionLocal()
    try:
        user = require_user(db, request)
        q = db.query(TravelRecord, Employee).join(Employee, Employee.id == TravelRecord.employee_id).filter(TravelRecord.user_id == user.id)
        if employee_id:
            q = q.filter(TravelRecord.employee_id == employee_id)
        rows = q.order_by(TravelRecord.departure_date.desc()).all()
        items = [_travel_out(t, e.name) for t, e in rows]
        if status:
            items = [i for i in items if i["status"] == status]
        return {"items": items}
    finally:
        db.close()


@router.post("/travel")
def create_travel(request: Request, body: TravelCreate = Body(...)):
    db = SessionLocal()
    try:
        require_csrf(request)
        user = require_user(db, request)
        employee = db.query(Employee).filter(Employee.user_id == user.id, Employee.id == body.employee_id).first()
        if not employee:
            raise HTTPException(404, "الموظف غير موجود")
        if body.travel_type not in ("domestic", "international"):
            raise HTTPException(400, "نوع السفر غير صالح")
        departure = _parse_simple_date(body.departure_date, "تاريخ السفر")
        ret = _parse_simple_date(body.return_date, "تاريخ العودة") if body.return_date else None
        if not departure:
            raise HTTPException(400, "تاريخ السفر مطلوب")
        if ret and ret < departure:
            raise HTTPException(400, "تاريخ العودة لا يمكن أن يسبق تاريخ السفر")
        rec = TravelRecord(
            id=uuid.uuid4().hex, user_id=user.id, employee_id=body.employee_id,
            travel_type=body.travel_type, destination=(body.destination or "").strip() or None,
            departure_date=departure, return_date=ret,
            purpose=(body.purpose or "").strip() or None, notes=(body.notes or "").strip() or None,
            status="scheduled",
        )
        db.add(rec)
        db.commit()
        log_event(db, "hr.travel.create", user.id, {"travel_id": rec.id}, employee_id=body.employee_id)
        return _travel_out(rec, employee.name)
    finally:
        db.close()


@router.post("/travel/{travel_id}/return")
def mark_travel_returned(travel_id: str, request: Request):
    db = SessionLocal()
    try:
        require_csrf(request)
        user = require_user(db, request)
        rec = db.query(TravelRecord).filter(TravelRecord.user_id == user.id, TravelRecord.id == travel_id).first()
        if not rec:
            raise HTTPException(404, "سجل السفر غير موجود")
        if rec.status == "cancelled":
            raise HTTPException(400, "لا يمكن تسجيل عودة لرحلة ملغاة")
        rec.actual_return_date = dt.datetime.utcnow()
        rec.status = "completed"
        db.commit()
        log_event(db, "hr.travel.return", user.id, {"travel_id": rec.id}, employee_id=rec.employee_id)
        return _travel_out(rec)
    finally:
        db.close()


@router.post("/travel/{travel_id}/cancel")
def cancel_travel(travel_id: str, request: Request):
    db = SessionLocal()
    try:
        require_csrf(request)
        user = require_user(db, request)
        rec = db.query(TravelRecord).filter(TravelRecord.user_id == user.id, TravelRecord.id == travel_id).first()
        if not rec:
            raise HTTPException(404, "سجل السفر غير موجود")
        rec.status = "cancelled"
        db.commit()
        return _travel_out(rec)
    finally:
        db.close()


@router.delete("/travel/{travel_id}")
def delete_travel(travel_id: str, request: Request):
    db = SessionLocal()
    try:
        require_csrf(request)
        user = require_user(db, request)
        rec = db.query(TravelRecord).filter(TravelRecord.user_id == user.id, TravelRecord.id == travel_id).first()
        if not rec:
            raise HTTPException(404, "سجل السفر غير موجود")
        db.delete(rec)
        db.commit()
        return {"deleted": True}
    finally:
        db.close()


# ============================================================
# العهد (عهد الشركة لدى الموظف)
# ============================================================

CUSTODY_STATUS_LABELS = {"assigned": "بعهدة الموظف", "returned": "تم الإرجاع", "lost": "مفقودة", "damaged": "تالفة"}


class CustodyCreate(BaseModel):
    employee_id: str
    item_name: str = Field(min_length=1)
    item_type: str = ""
    serial_number: str = ""
    value: Optional[float] = None
    assigned_date: str = ""
    expected_return_date: str = ""
    notes: str = ""


def _custody_out(c: CustodyRecord, employee_name: str = "") -> dict:
    return {
        "id": c.id, "employee_id": c.employee_id, "employee_name": employee_name,
        "item_name": c.item_name, "item_type": c.item_type or "", "serial_number": c.serial_number or "",
        "value": round(float(c.value), 2) if c.value is not None else None,
        "assigned_date": _fmt_date(c.assigned_date), "expected_return_date": _fmt_date(c.expected_return_date),
        "actual_return_date": _fmt_date(c.actual_return_date),
        "status": c.status, "status_label": CUSTODY_STATUS_LABELS.get(c.status, c.status),
        "notes": c.notes or "",
    }


@router.get("/custody")
def list_custody(request: Request, employee_id: str = Query(""), status: str = Query("")):
    db = SessionLocal()
    try:
        user = require_user(db, request)
        q = db.query(CustodyRecord, Employee).join(Employee, Employee.id == CustodyRecord.employee_id).filter(CustodyRecord.user_id == user.id)
        if employee_id:
            q = q.filter(CustodyRecord.employee_id == employee_id)
        if status:
            q = q.filter(CustodyRecord.status == status)
        rows = q.order_by(CustodyRecord.created_at.desc()).all()
        return {"items": [_custody_out(c, e.name) for c, e in rows]}
    finally:
        db.close()


@router.post("/custody")
def create_custody(request: Request, body: CustodyCreate = Body(...)):
    db = SessionLocal()
    try:
        require_csrf(request)
        user = require_user(db, request)
        employee = db.query(Employee).filter(Employee.user_id == user.id, Employee.id == body.employee_id).first()
        if not employee:
            raise HTTPException(404, "الموظف غير موجود")
        rec = CustodyRecord(
            id=uuid.uuid4().hex, user_id=user.id, employee_id=body.employee_id,
            item_name=body.item_name.strip(), item_type=(body.item_type or "").strip() or None,
            serial_number=(body.serial_number or "").strip() or None, value=body.value,
            assigned_date=_parse_simple_date(body.assigned_date, "تاريخ التسليم") or dt.datetime.utcnow(),
            expected_return_date=_parse_simple_date(body.expected_return_date, "تاريخ الإرجاع المتوقع"),
            notes=(body.notes or "").strip() or None, status="assigned",
        )
        db.add(rec)
        db.commit()
        log_event(db, "hr.custody.create", user.id, {"custody_id": rec.id, "item_name": rec.item_name}, employee_id=body.employee_id)
        return _custody_out(rec, employee.name)
    finally:
        db.close()


@router.post("/custody/{custody_id}/return")
def return_custody(custody_id: str, request: Request, condition: str = Body("returned", embed=True)):
    db = SessionLocal()
    try:
        require_csrf(request)
        user = require_user(db, request)
        rec = db.query(CustodyRecord).filter(CustodyRecord.user_id == user.id, CustodyRecord.id == custody_id).first()
        if not rec:
            raise HTTPException(404, "سجل العهدة غير موجود")
        if condition not in ("returned", "lost", "damaged"):
            raise HTTPException(400, "حالة غير صالحة")
        rec.status = condition
        rec.actual_return_date = dt.datetime.utcnow()
        db.commit()
        log_event(db, "hr.custody.return", user.id, {"custody_id": rec.id, "condition": condition}, employee_id=rec.employee_id)
        return _custody_out(rec)
    finally:
        db.close()


@router.delete("/custody/{custody_id}")
def delete_custody(custody_id: str, request: Request):
    db = SessionLocal()
    try:
        require_csrf(request)
        user = require_user(db, request)
        rec = db.query(CustodyRecord).filter(CustodyRecord.user_id == user.id, CustodyRecord.id == custody_id).first()
        if not rec:
            raise HTTPException(404, "سجل العهدة غير موجود")
        db.delete(rec)
        db.commit()
        return {"deleted": True}
    finally:
        db.close()


# ============================================================
# التقارير (Excel export لكل تقرير)
# ============================================================

MONTHS_AR = ["يناير","فبراير","مارس","أبريل","مايو","يونيو","يوليو","أغسطس","سبتمبر","أكتوبر","نوفمبر","ديسمبر"]


@router.get("/reports/payroll/export")
def report_payroll_export(request: Request, month: int = Query(0), year: int = Query(0)):
    db = SessionLocal()
    try:
        user = require_user(db, request)
        q = db.query(Payroll, Employee).join(Employee, Employee.id == Payroll.employee_id).filter(Payroll.user_id == user.id)
        if month:
            q = q.filter(Payroll.month == month)
        if year:
            q = q.filter(Payroll.year == year)
        rows = q.order_by(Payroll.year.desc(), Payroll.month.desc()).all()
        data = [[e.name, MONTHS_AR[p.month-1], p.year, p.base_salary, p.total_allowances, p.total_deductions, p.tax_amount, p.net_salary, p.status] for p, e in rows]
        return _xlsx_response(["الموظف","الشهر","السنة","الأساسي","البدلات","الاستقطاعات","الضريبة","الصافي","الحالة"], data, "رواتب.xlsx")
    finally:
        db.close()


@router.get("/reports/advances/export")
def report_advances_export(request: Request):
    db = SessionLocal()
    try:
        user = require_user(db, request)
        rows = db.query(Advance, Employee).join(Employee, Employee.id == Advance.employee_id).filter(Advance.user_id == user.id).all()
        data = [[e.name, _fmt_date(a.date), a.amount, a.reason or "", a.installments_count, a.remaining_amount, a.status] for a, e in rows]
        return _xlsx_response(["الموظف","التاريخ","المبلغ","السبب","عدد الأقساط","المتبقي","الحالة"], data, "السلف.xlsx")
    finally:
        db.close()


@router.get("/reports/deductions/export")
def report_deductions_export(request: Request):
    db = SessionLocal()
    try:
        user = require_user(db, request)
        rows = db.query(DeductionRecord, Employee).join(Employee, Employee.id == DeductionRecord.employee_id).filter(DeductionRecord.user_id == user.id).all()
        data = [[e.name, d.deduction_type or "", d.amount, d.reason or "", _fmt_date(d.date), "مطبّق" if d.applied else "بالانتظار"] for d, e in rows]
        return _xlsx_response(["الموظف","النوع","القيمة","السبب","التاريخ","الحالة"], data, "الاستقطاعات.xlsx")
    finally:
        db.close()


@router.get("/reports/commissions/export")
def report_commissions_export(request: Request):
    db = SessionLocal()
    try:
        user = require_user(db, request)
        rows = db.query(CommissionRecord, Employee).join(Employee, Employee.id == CommissionRecord.employee_id).filter(CommissionRecord.user_id == user.id).all()
        data = [[e.name, c.sales_amount, c.percentage, c.commission_value, _fmt_date(c.date), "مطبّقة" if c.applied else "بالانتظار"] for c, e in rows]
        return _xlsx_response(["الموظف","المبيعات","النسبة %","قيمة العمولة","التاريخ","الحالة"], data, "العمولات.xlsx")
    finally:
        db.close()


@router.get("/reports/attendance/export")
def report_attendance_export(request: Request, month: int = Query(0), year: int = Query(0)):
    db = SessionLocal()
    try:
        user = require_user(db, request)
        rows = db.query(Attendance, Employee).join(Employee, Employee.id == Attendance.employee_id).filter(Attendance.user_id == user.id).all()
        if month:
            rows = [r for r in rows if r[0].date and r[0].date.month == month]
        if year:
            rows = [r for r in rows if r[0].date and r[0].date.year == year]
        data = [[e.name, _fmt_date(a.date), a.check_in or "", a.check_out or "", a.late_minutes, a.overtime_minutes, "غياب" if a.is_absent else "حضور"] for a, e in rows]
        return _xlsx_response(["الموظف","التاريخ","الدخول","الخروج","التأخير(دقيقة)","الإضافي(دقيقة)","الحالة"], data, "الحضور.xlsx")
    finally:
        db.close()


@router.get("/reports/leaves/export")
def report_leaves_export(request: Request):
    db = SessionLocal()
    try:
        user = require_user(db, request)
        rows = db.query(LeaveRequest, Employee).join(Employee, Employee.id == LeaveRequest.employee_id).filter(LeaveRequest.user_id == user.id).all()
        data = [[e.name, l.leave_type or "", _fmt_date(l.start_date), _fmt_date(l.end_date), l.days_count, "مدفوعة" if l.paid else "غير مدفوعة", l.status] for l, e in rows]
        return _xlsx_response(["الموظف","النوع","من","إلى","الأيام","مدفوعة؟","الحالة"], data, "الإجازات.xlsx")
    finally:
        db.close()


@router.get("/reports/residency-expiry/export")
def report_residency_expiry_export(request: Request):
    db = SessionLocal()
    try:
        user = require_user(db, request)
        rows = db.query(Employee).filter(Employee.user_id == user.id, Employee.residency_expiry.isnot(None)).all()
        data = [[e.name, e.national_id or "", _fmt_date(e.residency_expiry), _days_left(e.residency_expiry)] for e in rows]
        return _xlsx_response(["الموظف","رقم الهوية/الإقامة","تاريخ الانتهاء","الأيام المتبقية"], data, "انتهاء_الاقامات.xlsx")
    finally:
        db.close()


@router.get("/reports/passport-expiry/export")
def report_passport_expiry_export(request: Request):
    db = SessionLocal()
    try:
        user = require_user(db, request)
        rows = db.query(Employee).filter(Employee.user_id == user.id, Employee.passport_expiry.isnot(None)).all()
        data = [[e.name, e.passport_number or "", _fmt_date(e.passport_expiry), _days_left(e.passport_expiry)] for e in rows]
        return _xlsx_response(["الموظف","رقم الجواز","تاريخ الانتهاء","الأيام المتبقية"], data, "انتهاء_الجوازات.xlsx")
    finally:
        db.close()


@router.get("/reports/contracts-expiry/export")
def report_contracts_expiry_export(request: Request):
    db = SessionLocal()
    try:
        user = require_user(db, request)
        rows = db.query(EmployeeContract, Employee).join(Employee, Employee.id == EmployeeContract.employee_id).filter(EmployeeContract.user_id == user.id, EmployeeContract.status == "active").all()
        data = [[e.name, c.contract_type or "", _fmt_date(c.end_date), _days_left(c.end_date)] for c, e in rows]
        return _xlsx_response(["الموظف","نوع العقد","تاريخ الانتهاء","الأيام المتبقية"], data, "انتهاء_العقود.xlsx")
    finally:
        db.close()


@router.get("/reports/employee-cost/export")
def report_employee_cost_export(request: Request, month: int = Query(0), year: int = Query(0)):
    db = SessionLocal()
    try:
        user = require_user(db, request)
        q = db.query(Payroll, Employee).join(Employee, Employee.id == Payroll.employee_id).filter(Payroll.user_id == user.id)
        if month:
            q = q.filter(Payroll.month == month)
        if year:
            q = q.filter(Payroll.year == year)
        rows = q.all()
        by_emp = {}
        for p, e in rows:
            by_emp.setdefault(e.name, 0.0)
            by_emp[e.name] += float(p.net_salary or 0.0)
        data = [[name, round(total, 2)] for name, total in sorted(by_emp.items(), key=lambda x: -x[1])]
        return _xlsx_response(["الموظف","إجمالي التكلفة"], data, "تكلفة_الموظفين.xlsx")
    finally:
        db.close()


@router.get("/reports/employees-by-branch/export")
def report_employees_by_branch_export(request: Request):
    db = SessionLocal()
    try:
        user = require_user(db, request)
        rows = db.query(Employee).filter(Employee.user_id == user.id).all()
        by_branch = {}
        for e in rows:
            b = e.branch_name or "بدون فرع"
            by_branch.setdefault(b, 0)
            by_branch[b] += 1
        data = [[b, c] for b, c in sorted(by_branch.items(), key=lambda x: -x[1])]
        return _xlsx_response(["الفرع","عدد الموظفين"], data, "الموظفين_حسب_الفرع.xlsx")
    finally:
        db.close()


@router.get("/reports/employees-by-department/export")
def report_employees_by_department_export(request: Request):
    db = SessionLocal()
    try:
        user = require_user(db, request)
        rows = db.query(Employee).filter(Employee.user_id == user.id).all()
        by_dept = {}
        for e in rows:
            d = e.department or "بدون قسم"
            by_dept.setdefault(d, 0)
            by_dept[d] += 1
        data = [[d, c] for d, c in sorted(by_dept.items(), key=lambda x: -x[1])]
        return _xlsx_response(["القسم","عدد الموظفين"], data, "الموظفين_حسب_القسم.xlsx")
    finally:
        db.close()


# ============================================================
# نفس التقارير أعلاه — نسخة PDF احترافية (شعار + اسم الشركة + تاريخ + ترقيم صفحات + دعم عربي كامل)
# ============================================================

@router.get("/reports/payroll/pdf")
def report_payroll_pdf(request: Request, month: int = Query(0), year: int = Query(0)):
    db = SessionLocal()
    try:
        user = require_user(db, request)
        q = db.query(Payroll, Employee).join(Employee, Employee.id == Payroll.employee_id).filter(Payroll.user_id == user.id)
        if month:
            q = q.filter(Payroll.month == month)
        if year:
            q = q.filter(Payroll.year == year)
        rows = q.order_by(Payroll.year.desc(), Payroll.month.desc()).all()
        data = [[e.name, MONTHS_AR[p.month-1], p.year, p.base_salary, p.total_allowances, p.total_deductions, p.tax_amount, p.net_salary, p.status] for p, e in rows]
        company_name, logo_url, generated_by = _pdf_meta(db, user)
        return generate_pdf_report(["الموظف","الشهر","السنة","الأساسي","البدلات","الاستقطاعات","الضريبة","الصافي","الحالة"], data, "تقرير الرواتب", "رواتب.pdf", company_name, logo_url, generated_by)
    finally:
        db.close()


@router.get("/reports/advances/pdf")
def report_advances_pdf(request: Request):
    db = SessionLocal()
    try:
        user = require_user(db, request)
        rows = db.query(Advance, Employee).join(Employee, Employee.id == Advance.employee_id).filter(Advance.user_id == user.id).all()
        data = [[e.name, _fmt_date(a.date), a.amount, a.reason or "", a.installments_count, a.remaining_amount, a.status] for a, e in rows]
        company_name, logo_url, generated_by = _pdf_meta(db, user)
        return generate_pdf_report(["الموظف","التاريخ","المبلغ","السبب","عدد الأقساط","المتبقي","الحالة"], data, "تقرير السلف", "السلف.pdf", company_name, logo_url, generated_by)
    finally:
        db.close()


@router.get("/reports/deductions/pdf")
def report_deductions_pdf(request: Request):
    db = SessionLocal()
    try:
        user = require_user(db, request)
        rows = db.query(DeductionRecord, Employee).join(Employee, Employee.id == DeductionRecord.employee_id).filter(DeductionRecord.user_id == user.id).all()
        data = [[e.name, d.deduction_type or "", d.amount, d.reason or "", _fmt_date(d.date), "مطبّق" if d.applied else "بالانتظار"] for d, e in rows]
        company_name, logo_url, generated_by = _pdf_meta(db, user)
        return generate_pdf_report(["الموظف","النوع","القيمة","السبب","التاريخ","الحالة"], data, "تقرير الاستقطاعات", "الاستقطاعات.pdf", company_name, logo_url, generated_by)
    finally:
        db.close()


@router.get("/reports/commissions/pdf")
def report_commissions_pdf(request: Request):
    db = SessionLocal()
    try:
        user = require_user(db, request)
        rows = db.query(CommissionRecord, Employee).join(Employee, Employee.id == CommissionRecord.employee_id).filter(CommissionRecord.user_id == user.id).all()
        data = [[e.name, c.sales_amount, c.percentage, c.commission_value, _fmt_date(c.date), "مطبّقة" if c.applied else "بالانتظار"] for c, e in rows]
        company_name, logo_url, generated_by = _pdf_meta(db, user)
        return generate_pdf_report(["الموظف","المبيعات","النسبة %","قيمة العمولة","التاريخ","الحالة"], data, "تقرير العمولات", "العمولات.pdf", company_name, logo_url, generated_by)
    finally:
        db.close()


@router.get("/reports/attendance/pdf")
def report_attendance_pdf(request: Request, month: int = Query(0), year: int = Query(0)):
    db = SessionLocal()
    try:
        user = require_user(db, request)
        rows = db.query(Attendance, Employee).join(Employee, Employee.id == Attendance.employee_id).filter(Attendance.user_id == user.id).all()
        if month:
            rows = [r for r in rows if r[0].date and r[0].date.month == month]
        if year:
            rows = [r for r in rows if r[0].date and r[0].date.year == year]
        data = [[e.name, _fmt_date(a.date), a.check_in or "", a.check_out or "", a.late_minutes, a.overtime_minutes, "غياب" if a.is_absent else "حضور"] for a, e in rows]
        company_name, logo_url, generated_by = _pdf_meta(db, user)
        return generate_pdf_report(["الموظف","التاريخ","الدخول","الخروج","التأخير(دقيقة)","الإضافي(دقيقة)","الحالة"], data, "تقرير الحضور والانصراف", "الحضور.pdf", company_name, logo_url, generated_by)
    finally:
        db.close()


@router.get("/reports/leaves/pdf")
def report_leaves_pdf(request: Request):
    db = SessionLocal()
    try:
        user = require_user(db, request)
        rows = db.query(LeaveRequest, Employee).join(Employee, Employee.id == LeaveRequest.employee_id).filter(LeaveRequest.user_id == user.id).all()
        data = [[e.name, l.leave_type or "", _fmt_date(l.start_date), _fmt_date(l.end_date), l.days_count, "مدفوعة" if l.paid else "غير مدفوعة", l.status] for l, e in rows]
        company_name, logo_url, generated_by = _pdf_meta(db, user)
        return generate_pdf_report(["الموظف","النوع","من","إلى","الأيام","مدفوعة؟","الحالة"], data, "تقرير الإجازات", "الإجازات.pdf", company_name, logo_url, generated_by)
    finally:
        db.close()


@router.get("/reports/residency-expiry/pdf")
def report_residency_expiry_pdf(request: Request):
    db = SessionLocal()
    try:
        user = require_user(db, request)
        rows = db.query(Employee).filter(Employee.user_id == user.id, Employee.residency_expiry.isnot(None)).all()
        data = [[e.name, e.national_id or "", _fmt_date(e.residency_expiry), _days_left(e.residency_expiry)] for e in rows]
        company_name, logo_url, generated_by = _pdf_meta(db, user)
        return generate_pdf_report(["الموظف","رقم الهوية/الإقامة","تاريخ الانتهاء","الأيام المتبقية"], data, "تقرير انتهاء الإقامات", "انتهاء_الاقامات.pdf", company_name, logo_url, generated_by)
    finally:
        db.close()


@router.get("/reports/passport-expiry/pdf")
def report_passport_expiry_pdf(request: Request):
    db = SessionLocal()
    try:
        user = require_user(db, request)
        rows = db.query(Employee).filter(Employee.user_id == user.id, Employee.passport_expiry.isnot(None)).all()
        data = [[e.name, e.passport_number or "", _fmt_date(e.passport_expiry), _days_left(e.passport_expiry)] for e in rows]
        company_name, logo_url, generated_by = _pdf_meta(db, user)
        return generate_pdf_report(["الموظف","رقم الجواز","تاريخ الانتهاء","الأيام المتبقية"], data, "تقرير انتهاء الجوازات", "انتهاء_الجوازات.pdf", company_name, logo_url, generated_by)
    finally:
        db.close()


@router.get("/reports/contracts-expiry/pdf")
def report_contracts_expiry_pdf(request: Request):
    db = SessionLocal()
    try:
        user = require_user(db, request)
        rows = db.query(EmployeeContract, Employee).join(Employee, Employee.id == EmployeeContract.employee_id).filter(EmployeeContract.user_id == user.id, EmployeeContract.status == "active").all()
        data = [[e.name, c.contract_type or "", _fmt_date(c.end_date), _days_left(c.end_date)] for c, e in rows]
        company_name, logo_url, generated_by = _pdf_meta(db, user)
        return generate_pdf_report(["الموظف","نوع العقد","تاريخ الانتهاء","الأيام المتبقية"], data, "تقرير انتهاء العقود", "انتهاء_العقود.pdf", company_name, logo_url, generated_by)
    finally:
        db.close()


@router.get("/reports/employee-cost/pdf")
def report_employee_cost_pdf(request: Request, month: int = Query(0), year: int = Query(0)):
    db = SessionLocal()
    try:
        user = require_user(db, request)
        q = db.query(Payroll, Employee).join(Employee, Employee.id == Payroll.employee_id).filter(Payroll.user_id == user.id)
        if month:
            q = q.filter(Payroll.month == month)
        if year:
            q = q.filter(Payroll.year == year)
        rows = q.all()
        by_emp = {}
        for p, e in rows:
            by_emp.setdefault(e.name, 0.0)
            by_emp[e.name] += float(p.net_salary or 0.0)
        data = [[name, round(total, 2)] for name, total in sorted(by_emp.items(), key=lambda x: -x[1])]
        company_name, logo_url, generated_by = _pdf_meta(db, user)
        return generate_pdf_report(["الموظف","إجمالي التكلفة"], data, "تقرير تكلفة الموظفين", "تكلفة_الموظفين.pdf", company_name, logo_url, generated_by)
    finally:
        db.close()


@router.get("/reports/employees-by-branch/pdf")
def report_employees_by_branch_pdf(request: Request):
    db = SessionLocal()
    try:
        user = require_user(db, request)
        rows = db.query(Employee).filter(Employee.user_id == user.id).all()
        by_branch = {}
        for e in rows:
            b = e.branch_name or "بدون فرع"
            by_branch.setdefault(b, 0)
            by_branch[b] += 1
        data = [[b, c] for b, c in sorted(by_branch.items(), key=lambda x: -x[1])]
        company_name, logo_url, generated_by = _pdf_meta(db, user)
        return generate_pdf_report(["الفرع","عدد الموظفين"], data, "تقرير الموظفين حسب الفرع", "الموظفين_حسب_الفرع.pdf", company_name, logo_url, generated_by)
    finally:
        db.close()


@router.get("/reports/employees-by-department/pdf")
def report_employees_by_department_pdf(request: Request):
    db = SessionLocal()
    try:
        user = require_user(db, request)
        rows = db.query(Employee).filter(Employee.user_id == user.id).all()
        by_dept = {}
        for e in rows:
            d = e.department or "بدون قسم"
            by_dept.setdefault(d, 0)
            by_dept[d] += 1
        data = [[d, c] for d, c in sorted(by_dept.items(), key=lambda x: -x[1])]
        company_name, logo_url, generated_by = _pdf_meta(db, user)
        return generate_pdf_report(["القسم","عدد الموظفين"], data, "تقرير الموظفين حسب القسم", "الموظفين_حسب_القسم.pdf", company_name, logo_url, generated_by)
    finally:
        db.close()



# ============================================================
# لوحة التنبيهات الموحّدة (أعلى الصفحة)
# ============================================================

def _tier_for_days(days_left, expired_label="منتهية"):
    if days_left is None:
        return None
    if days_left < 0:
        return ("critical", expired_label)
    if days_left <= 30:
        return ("high", f"تنتهي خلال {days_left} يوم")
    if days_left <= 90:
        return ("medium", f"تنتهي خلال {days_left} يوم")
    return None


@router.get("/alerts/feed")
def alerts_feed(request: Request):
    db = SessionLocal()
    try:
        user = require_user(db, request)
        employees = db.query(Employee).filter(Employee.user_id == user.id, Employee.is_active == 1).all()
        items = []
        today = dt.datetime.utcnow().date()

        for e in employees:
            # الإقامة
            tier = _tier_for_days(_days_left(e.residency_expiry), "الإقامة منتهية")
            if tier:
                sev, label = tier
                items.append({"type": "residency", "severity": sev, "employee_id": e.id, "employee_name": e.name, "message": f"إقامة {e.name} {label}"})
            # الجواز
            tier = _tier_for_days(_days_left(e.passport_expiry), "الجواز منتهي")
            if tier:
                sev, label = tier
                items.append({"type": "passport", "severity": sev, "employee_id": e.id, "employee_name": e.name, "message": f"جواز {e.name} {label}"})
            # عيد الميلاد (خلال 7 أيام القادمة)
            if e.birth_date:
                try:
                    next_bd = e.birth_date.replace(year=today.year)
                except ValueError:
                    next_bd = e.birth_date.replace(year=today.year, day=28)
                if next_bd.date() < today:
                    try:
                        next_bd = next_bd.replace(year=today.year + 1)
                    except ValueError:
                        next_bd = next_bd.replace(year=today.year + 1, day=28)
                bd_days = (next_bd.date() - today).days
                if 0 <= bd_days <= 7:
                    items.append({"type": "birthday", "severity": "notice", "employee_id": e.id, "employee_name": e.name, "message": f"🎂 عيد ميلاد {e.name} بعد {bd_days} يوم" if bd_days else f"🎂 اليوم عيد ميلاد {e.name}"})

        contracts = db.query(EmployeeContract, Employee).join(Employee, Employee.id == EmployeeContract.employee_id).filter(
            EmployeeContract.user_id == user.id, EmployeeContract.status == "active",
        ).all()
        for c, e in contracts:
            tier = _tier_for_days(_days_left(c.end_date), "العقد منتهي")
            if tier:
                sev, label = tier
                items.append({"type": "contract", "severity": sev, "employee_id": e.id, "employee_name": e.name, "message": f"عقد {e.name} {label}"})
            if c.start_date and c.probation_days:
                probation_end = c.start_date + dt.timedelta(days=c.probation_days)
                p_days = (probation_end.date() - today).days
                if 0 <= p_days <= 7:
                    items.append({"type": "probation", "severity": "notice", "employee_id": e.id, "employee_name": e.name, "message": f"تنتهي فترة تجربة {e.name} بعد {p_days} يوم"})

        if today.day >= 25:
            now = dt.datetime.utcnow()
            paid_ids = {
                p.employee_id for p in db.query(Payroll).filter(
                    Payroll.user_id == user.id, Payroll.month == now.month, Payroll.year == now.year, Payroll.status == "paid",
                ).all()
            }
            for e in employees:
                if e.id not in paid_ids:
                    items.append({"type": "salary_unpaid", "severity": "medium", "employee_id": e.id, "employee_name": e.name, "message": f"{e.name} لم يستلم راتب هذا الشهر بعد"})

        # الموظفون الموجودون في سفر حالياً
        emp_by_id = {e.id: e for e in employees}
        travels = db.query(TravelRecord).filter(TravelRecord.user_id == user.id, TravelRecord.status != "cancelled", TravelRecord.actual_return_date.is_(None)).all()
        now_dt = dt.datetime.utcnow()
        for t in travels:
            e = emp_by_id.get(t.employee_id)
            if not e or not t.departure_date or t.departure_date > now_dt:
                continue
            items.append({"type": "travel", "severity": "notice", "employee_id": t.employee_id, "employee_name": e.name, "message": f"✈️ {e.name} في سفر حالياً" + (f" (متوقع العودة {_fmt_date(t.return_date)})" if t.return_date else "")})

        # سلف قائمة (نشطة) لم تُسدَّد بعد بالكامل
        active_advances = db.query(Advance).filter(Advance.user_id == user.id, Advance.status == "active").all()
        for a in active_advances:
            e = emp_by_id.get(a.employee_id)
            if not e:
                continue
            items.append({"type": "advance_active", "severity": "notice", "employee_id": a.employee_id, "employee_name": e.name, "message": f"💰 {e.name} لديه سلفة قائمة (المتبقي {round(float(a.remaining_amount or 0), 2)})"})

        # سحبيات معلّقة لم تُسوَّ بعد
        pending_withdrawals = db.query(WithdrawalRecord).filter(WithdrawalRecord.user_id == user.id, WithdrawalRecord.status == "pending").all()
        for w in pending_withdrawals:
            e = emp_by_id.get(w.employee_id)
            if not e:
                continue
            items.append({"type": "withdrawal_pending", "severity": "notice", "employee_id": w.employee_id, "employee_name": e.name, "message": f"🧾 سحبية معلّقة لـ{e.name} بقيمة {round(float(w.amount or 0), 2)}"})

        # الغياب اليوم
        today_start = dt.datetime(today.year, today.month, today.day)
        absentees = db.query(Attendance).filter(Attendance.user_id == user.id, Attendance.is_absent == 1, Attendance.date >= today_start).all()
        for a in absentees:
            e = emp_by_id.get(a.employee_id)
            if not e:
                continue
            items.append({"type": "absent_today", "severity": "medium", "employee_id": a.employee_id, "employee_name": e.name, "message": f"🚫 {e.name} غائب اليوم"})

        order = {"critical": 0, "high": 1, "medium": 2, "notice": 3}
        items.sort(key=lambda x: order.get(x["severity"], 9))
        return {"items": items, "count": len(items)}
    finally:
        db.close()

