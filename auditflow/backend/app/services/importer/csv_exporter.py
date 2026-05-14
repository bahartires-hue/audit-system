from __future__ import annotations

import os
import re
from pathlib import Path
from typing import Any, Dict, List, Optional

import pandas as pd
from openpyxl import load_workbook
from copy import copy

_DEFAULT_SALLA_COLUMNS = [
    "النوع ",
    "أسم المنتج",
    "تصنيف المنتج",
    "صورة المنتج",
    "وصف صورة المنتج",
    "نوع المنتج",
    "سعر المنتج",
    "الوصف",
    "هل يتطلب شحن؟",
    "رمز المنتج sku",
    "سعر التكلفة",
    "السعر المخفض",
    "تاريخ بداية التخفيض",
    "تاريخ نهاية التخفيض",
    "اقصي كمية لكل عميل",
    "إخفاء خيار تحديد الكمية",
    "اضافة صورة عند الطلب",
    "الوزن",
    "وحدة الوزن",
    "الماركة",
    "العنوان الترويجي",
    "تثبيت المنتج",
    "الباركود",
    "السعرات الحرارية",
    "MPN",
    "GTIN",
    "خاضع للضريبة ؟",
    "سبب عدم الخضوع للضريبة",
    "[1] الاسم",
    "[1] النوع",
    "[1] القيمة",
    "[1] الصورة / اللون",
    "[2] الاسم",
    "[2] النوع",
    "[2] القيمة",
    "[2] الصورة / اللون",
    "[3] الاسم",
    "[3] النوع",
    "[3] القيمة",
    "[3] الصورة / اللون",
]

def safe_set(row: Dict[str, Any], columns: List[str], col_name: str, value: Any) -> None:
    if col_name in columns:
        row[col_name] = value if value is not None else ""


def _resolve_col(columns: List[str], variants: tuple[str, ...]) -> Optional[str]:
    """أول عمود موجود في القالب يطابق أحد المتغيرات (مع تجاهل فراغات زائدة)."""
    colset = {c for c in columns if c}
    for v in variants:
        if v in colset:
            return v
    stripped = {c.strip(): c for c in columns if c}
    for v in variants:
        vs = v.strip()
        if vs in stripped:
            return stripped[vs]
    return None


def _strip_html_simple(value: Any) -> str:
    """إزالة وسوم HTML فقط؛ يُبقى المقاس (مثل 185/65R15) كما هو."""
    s = str(value or "")
    s = re.sub(r"<[^>]*>", " ", s)
    return re.sub(r"\s+", " ", s).strip()


def _salla_plain_text(value: Any, max_len: int = 220) -> str:
    """
    نص آمن لحقول سلة التي ترفض أحرفًا خاصة (مثل وصف الصورة): عربي/لاتيني وأرقام ومسافة وشرطة فقط.
    """
    s = str(value or "")
    s = re.sub(r"<[^>]*>", " ", s)
    for ent, ch in (
        ("&nbsp;", " "),
        ("&amp;", " "),
        ("&quot;", " "),
        ("&#x27;", ""),
        ("&apos;", ""),
    ):
        s = s.replace(ent, ch)
    s = re.sub(r"[\r\n\t]+", " ", s)
    out: List[str] = []
    for ch in s:
        o = ord(ch)
        if ch in "- " or ch.isdigit():
            out.append(ch)
        elif "A" <= ch <= "Z" or "a" <= ch <= "z":
            out.append(ch)
        elif 0x0600 <= o <= 0x06FF:
            out.append(ch)
    result = re.sub(r"\s+", " ", "".join(out)).strip()
    if len(result) > max_len:
        cut = result[:max_len]
        result = cut.rsplit(" ", 1)[0].strip() if " " in cut else cut.strip()
    return result


def _set_resolved(row: Dict[str, Any], columns: List[str], variants: tuple[str, ...], value: Any) -> None:
    cn = _resolve_col(columns, variants)
    if cn:
        row[cn] = value if value is not None else ""


def build_public_image_url(filename: str) -> str:
    base = (os.getenv("PUBLIC_BASE_URL") or "").strip().rstrip("/")
    safe = (filename or "").strip().replace("\\", "/").split("/")[-1]
    if not safe:
        return ""
    if base:
        return f"{base}/uploads/products/{safe}"
    return ""


def _to_public_image_value(p: Dict[str, Any]) -> str:
    cloud = str(p.get("image_cloudinary") or "").strip()
    if cloud.startswith("https://res.cloudinary.com/"):
        return cloud
    return ""


def _resolve_template_path(base_dir: Path) -> Path | None:
    candidates = [
        Path(__file__).resolve().parent / "templates" / "Salla Products Template (3).xlsx",
        Path(__file__).resolve().parent / "templates" / "Salla Products Template.xlsx",
        base_dir / "Salla Products Template (3).xlsx",
        base_dir / "Salla Products Template.xlsx",
        base_dir.parent / "Salla Products Template (3).xlsx",
        base_dir.parent / "Salla Products Template.xlsx",
    ]
    for p in candidates:
        if p.exists() and p.is_file():
            return p
    return None


def export_to_salla_template(products: List[Dict[str, Any]], template_path: Path | None, output_path: Path) -> Path:
    if not template_path:
        raise ValueError("تعذر العثور على ملف قالب سلة الأصلي.")

    wb = load_workbook(template_path)
    ws = wb.active
    header_row = 1
    for r in range(1, min(20, ws.max_row) + 1):
        c1 = str(ws.cell(row=r, column=1).value or "").strip()
        c2 = str(ws.cell(row=r, column=2).value or "").strip()
        if c1 in {"النوع", "النوع "} and c2 == "أسم المنتج":
            header_row = r
            break
    columns = [str(ws.cell(row=header_row, column=c).value or "").strip() for c in range(1, ws.max_column + 1)]
    header_map = {name: idx + 1 for idx, name in enumerate(columns) if name}
    output_rows: List[Dict[str, Any]] = []

    for p in products:
        image_value = _to_public_image_value(p)
        if not str(p.get("brand", "")).strip():
            continue
        if not str(p.get("size", "")).strip():
            continue
        try:
            price_num = float(str(p.get("price", "")).replace(",", ""))
        except Exception:
            continue
        if price_num <= 0:
            continue
        brand = str(p.get("brand", "")).strip()
        size = str(p.get("size", "")).strip()
        price = str(p.get("price", "")).strip()
        title = _strip_html_simple(p.get("product_title", "")) or _strip_html_simple(p.get("name", ""))
        title = title.strip() or f"{brand} {size}".strip()
        row = {col: "" for col in columns}
        promo_bits = []
        year = str(p.get("year", "") or "").strip()
        country = str(p.get("country", "") or "").strip()
        if year and country:
            promo_bits.append(f"سنة الصنع {year} - بلد الصنع {country}")
        elif year:
            promo_bits.append(f"سنة الصنع {year}")
        elif country:
            promo_bits.append(f"بلد الصنع {country}")
        if p.get("warranty"):
            promo_bits.append(f"الضمان {p.get('warranty')}")
        if not image_value:
            promo_bits.append("needs_image")
        promo = " - ".join(promo_bits)
        _set_resolved(row, columns, ("النوع ", "النوع"), "منتج")
        _set_resolved(row, columns, ("أسم المنتج", "اسم المنتج"), title)
        safe_set(row, columns, "تصنيف المنتج", "قسم الإطارات")
        safe_set(row, columns, "صورة المنتج", image_value)
        alt_src = p.get("image_alt_text") or f"كفر {brand} {size}"
        alt_plain = _salla_plain_text(alt_src) or _salla_plain_text(f"كفر {brand} {size}")
        _set_resolved(row, columns, ("وصف صورة المنتج",), alt_plain)
        _set_resolved(row, columns, ("نوع المنتج",), "منتج جاهز")
        safe_set(row, columns, "سعر المنتج", price)
        safe_set(row, columns, "الوصف", p.get("description", ""))
        _set_resolved(row, columns, ("هل يتطلب شحن؟", "يتطلب شحن", "هل يتطلب شحن"), "نعم")
        safe_set(row, columns, "الوزن", 25)
        safe_set(row, columns, "وحدة الوزن", "kg")
        safe_set(row, columns, "الماركة", brand)
        safe_set(row, columns, "العنوان الترويجي", promo)
        safe_set(row, columns, "تثبيت المنتج", "لا")
        safe_set(row, columns, "خاضع للضريبة ؟", "نعم")

        safe_set(row, columns, "[1] الاسم", "مقاس الإطار")
        safe_set(row, columns, "[1] النوع", "نص")
        safe_set(row, columns, "[1] القيمة", size)
        safe_set(row, columns, "[2] الاسم", "")
        safe_set(row, columns, "[2] النوع", "")
        safe_set(row, columns, "[2] القيمة", "")
        safe_set(row, columns, "[2] الصورة / اللون", "")
        safe_set(row, columns, "[3] الاسم", "")
        safe_set(row, columns, "[3] النوع", "")
        safe_set(row, columns, "[3] القيمة", "")
        safe_set(row, columns, "[3] الصورة / اللون", "")
        output_rows.append(row)

    start_row = header_row + 1
    style_row = start_row if ws.max_row >= start_row else header_row
    base_height = ws.row_dimensions[style_row].height
    base_styles = {c: copy(ws.cell(row=style_row, column=c)._style) for c in range(1, ws.max_column + 1)}
    max_clear_row = max(ws.max_row, start_row + len(output_rows) - 1)

    for r in range(start_row, max_clear_row + 1):
        if base_height is not None:
            ws.row_dimensions[r].height = base_height
        for c in range(1, ws.max_column + 1):
            cell = ws.cell(row=r, column=c)
            cell._style = copy(base_styles[c])
            cell.value = None

    for idx, row in enumerate(output_rows, start=start_row):
        for k, v in row.items():
            col = header_map.get(k)
            if col:
                ws.cell(row=idx, column=col).value = v

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    return output_path


def export_products_files(products: List[Dict[str, Any]], csv_path: Path, xlsx_path: Path) -> Dict[str, str]:
    csv_path.parent.mkdir(parents=True, exist_ok=True)
    xlsx_path.parent.mkdir(parents=True, exist_ok=True)
    df = pd.DataFrame(products)
    df.to_csv(csv_path, index=False, encoding="utf-8-sig")
    df.to_excel(xlsx_path, index=False)
    salla_xlsx_path = xlsx_path.parent / "salla_products_ready.xlsx"
    template_path = _resolve_template_path(xlsx_path.parent)
    out_path = export_to_salla_template(products, template_path, salla_xlsx_path)
    return {
        "csv_path": str(csv_path),
        "xlsx_path": str(xlsx_path),
        "salla_csv_path": "",
        "salla_xlsx_path": str(out_path),
    }

