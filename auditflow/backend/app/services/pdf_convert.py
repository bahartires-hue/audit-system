from __future__ import annotations

import io
import re
import unicodedata
from pathlib import Path
from typing import Any

import pandas as pd
import pdfplumber

from .analyzer import (
    _dataframe_from_pdf_grid,
    _is_balance_column_name,
    _normalize_arabic_digits,
    _promote_ledger_header_row,
    _trim_pdf_table_df,
    detect_columns,
    detect_document_type_column,
    infer_document_kind_from_narrative,
    read_pdf,
    resolve_document_columns,
    safe,
)


def _has_arabic_script(s: str) -> bool:
    for c in s:
        o = ord(c)
        if 0x0600 <= o <= 0x06FF or 0xFB50 <= o <= 0xFDFF or 0xFE70 <= o <= 0xFEFF:
            return True
    return False


def _excel_arabic_display(value: Any) -> Any:
    """تطبيع خفيف للنص العربي دون قلب ترتيب الأرقام داخل Excel."""
    if value is None:
        return value
    if isinstance(value, float) and pd.isna(value):
        return value
    if not isinstance(value, str):
        return value
    s = value.strip()
    if not s or not _has_arabic_script(s):
        return value
    try:
        # NFKC يحول presentation forms لنص عربي قياسي بدون تغيير ترتيب الكلمات/الأرقام.
        return unicodedata.normalize("NFKC", s)
    except Exception:
        return value

# تاريخ في السطر (كشوف عربية غالبًا يوم/شهر/سنة)
_DATE_IN_ROW = re.compile(
    r"(\d{4}\s*[/\-\.]\s*\d{1,2}\s*[/\-\.]\s*\d{1,2}|\d{1,2}\s*[/\-\.]\s*\d{1,2}\s*[/\-\.]\s*\d{2,4})"
)


def _row_text_blob(row: pd.Series, columns) -> str:
    parts: list[str] = []
    for c in columns:
        v = row.get(c)
        if pd.notna(v) and str(v).strip():
            parts.append(str(v).strip())
    return _normalize_arabic_digits(" ".join(parts))


def _row_has_currency_amount(row: pd.Series, columns) -> bool:
    for c in columns:
        v = safe(row.get(c))
        if v is not None and abs(float(v)) > 0.005:
            return True
    return False


def _trim_table_for_excel(df: pd.DataFrame) -> pd.DataFrame:
    """إزالة صفوف ترويسة الشركة/التقرير؛ البدء من أول صف يشبه حركة (تاريخ + مبلغ)."""
    dfc = df.copy()
    dfc.columns = dfc.columns.astype(str).str.strip()
    dfc = _promote_ledger_header_row(dfc)
    dfc.columns = dfc.columns.astype(str).str.strip()
    if dfc.empty:
        return dfc

    for idx in range(len(dfc)):
        row = dfc.iloc[idx]
        if row.isna().all():
            continue
        blob = _row_text_blob(row, dfc.columns)
        if not _DATE_IN_ROW.search(blob):
            continue
        if _row_has_currency_amount(row, dfc.columns):
            return dfc.iloc[idx:].reset_index(drop=True)

    try:
        detected_cols = detect_columns(dfc.copy())
        debit_col, credit_col = detected_cols.get("debit"), detected_cols.get("credit")
    except Exception:
        return dfc

    for idx in range(len(dfc)):
        row = dfc.iloc[idx]
        if row.isna().all():
            continue
        deb = safe(row[debit_col]) if debit_col and debit_col in dfc.columns else None
        cre = safe(row[credit_col]) if credit_col and credit_col in dfc.columns else None
        if deb is not None or cre is not None:
            return dfc.iloc[idx:].reset_index(drop=True)

    return dfc


def _trim_text_style_for_excel(df: pd.DataFrame) -> pd.DataFrame:
    """جدول مستخرج من النص (أعمدة التاريخ/مدين/دائن)."""
    if "التاريخ" not in df.columns:
        return df
    start = 0
    for i in range(len(df)):
        row = df.iloc[i]
        val = row.get("التاريخ")
        if pd.isna(val):
            start = i + 1
            continue
        s = _normalize_arabic_digits(str(val).strip())
        if not _DATE_IN_ROW.search(s):
            start = i + 1
            continue
        if _row_has_currency_amount(row, df.columns):
            start = i
            break
        if pd.notna(row.get("بيان")) and str(row.get("بيان")).strip():
            start = i
            break
    return df.iloc[start:].reset_index(drop=True) if len(df) else df


def _prepare_dataframe_for_excel_export(df: pd.DataFrame) -> pd.DataFrame:
    if df is None or df.empty:
        return df
    cols = [str(c).strip() for c in df.columns]
    if "التاريخ" in cols and "مدين" in cols:
        return _trim_text_style_for_excel(df)
    return _trim_table_for_excel(df)


def _narrative_blob(row: pd.Series, df: pd.DataFrame) -> str:
    """نص للاستدلال على نوع المستند من أعمدة البيان/المستند."""
    parts: list[str] = []
    for name in ("بيان", "المستند", "مستند", "الشرح", "الوصف"):
        if name not in df.columns:
            continue
        v = row.get(name)
        if pd.notna(v) and str(v).strip():
            parts.append(str(v).strip())
    return " ".join(parts)


def _find_doc_number_column(df: pd.DataFrame) -> str | None:
    for c in df.columns:
        n = str(c).strip().lower()
        if "رقم السند" in n or "رقم سند" in n:
            return c
        if "السند" in n and "رقم" in n:
            return c
    return None


def _extract_doc_number_from_text(text: str) -> str:
    s = _normalize_arabic_digits(str(text or ""))
    if not s.strip():
        return ""
    nums = re.findall(r"\b\d{3,}\b", s)
    return nums[-1] if nums else ""


def _normalize_debit_credit_cells(deb: Any, cre: Any) -> tuple[Any, Any]:
    """إظهار 0 في الطرف المقابل إذا وُجد مبلغ حقيقي في طرف واحد."""
    d = safe(deb)
    c = safe(cre)
    d_has = d is not None and abs(float(d)) > 0.0001
    c_has = c is not None and abs(float(c)) > 0.0001
    if d_has and not c_has:
        return round(abs(float(d)), 2), 0
    if c_has and not d_has:
        return 0, round(abs(float(c)), 2)
    if d is None and c is None:
        return "", ""
    out_d = round(abs(float(d)), 2) if d is not None else 0
    out_c = round(abs(float(c)), 2) if c is not None else 0
    return out_d, out_c


def _looks_like_seq_column(series: pd.Series) -> bool:
    vals = pd.to_numeric(series, errors="coerce").dropna()
    if len(vals) < 3:
        return False
    ints = vals.astype(int)
    if ((vals - ints).abs() > 0.0001).any():
        return False
    uniq = sorted(set(ints.tolist()))
    if len(uniq) < 3:
        return False
    # 1..N متتابعة غالباً تسلسل صفوف وليس مبلغاً.
    return uniq[0] in (0, 1) and all((b - a) == 1 for a, b in zip(uniq[:-1], uniq[1:]))


def _looks_like_doc_number_column(series: pd.Series) -> bool:
    """استبعاد عمود رقم السند في الجداول المجهولة (عادة أرقام صحيحة فريدة بلا كسور/أصفار)."""
    nums = pd.to_numeric(series, errors="coerce").dropna()
    if len(nums) < 4:
        return False
    frac_ratio = float(((nums - nums.round()).abs() > 0.0001).mean())
    unique_ratio = float(nums.nunique()) / float(len(nums))
    zero_ratio = float((nums.abs() <= 0.0001).mean())
    return frac_ratio <= 0.02 and unique_ratio >= 0.85 and zero_ratio <= 0.05


def _reframe_anonymous_pdf_table(df: pd.DataFrame) -> pd.DataFrame | None:
    """جداول PDF ذات أعمدة _c* بدون رؤوس واضحة."""
    if df is None or df.empty:
        return None
    col_names = [str(c) for c in df.columns]
    if not any(n.startswith("_c") for n in col_names):
        return None
    if any(k in " ".join(col_names) for k in ("التاريخ", "مدين", "دائن")):
        return None

    # أعمدة رقمية مرشحة (مع استبعاد تسلسل الصف).
    numeric_cols: list[str] = []
    numeric_stats: list[tuple[str, float, int, int]] = []  # (col, zero_ratio, valid_count, col_idx)
    for c in df.columns:
        s = df[c]
        nums = pd.to_numeric(s, errors="coerce").dropna()
        if len(nums) < max(2, int(len(df) * 0.35)):
            continue
        if _looks_like_seq_column(s):
            continue
        if _looks_like_doc_number_column(s):
            continue
        if nums.max() <= 0:
            continue
        name = str(c)
        numeric_cols.append(name)
        non_null = len(nums)
        zeros = int((nums.abs() <= 0.0001).sum())
        zero_ratio = float(zeros) / float(non_null) if non_null else 0.0
        numeric_stats.append((name, zero_ratio, non_null, list(df.columns).index(c)))
    if not numeric_cols:
        return None

    doc_no_col: str | None = None
    for c in df.columns:
        if _looks_like_doc_number_column(df[c]):
            doc_no_col = str(c)
            break

    # في كشوف الحساب: عمودا الحركة (مدين/دائن) غالباً فيهما أصفار كثيرة.
    movement_cols = [x[0] for x in sorted(numeric_stats, key=lambda t: (-t[1], -t[2], t[3]))[:2]]
    if len(movement_cols) < 2:
        movement_cols = numeric_cols[:2]

    # عمود الرصيد النصي غالباً يحتوي "مدين/دائن" بشكل متكرر.
    balance_text_col: str | None = None
    best_hits = 0
    for c in df.columns:
        hits = 0
        for v in df[c].tolist():
            if pd.isna(v):
                continue
            s = _normalize_arabic_digits(str(v))
            if ("مدين" in s) or ("دائن" in s) or ("نيدم" in s) or ("نئاد" in s):
                hits += 1
        if hits > best_hits:
            best_hits = hits
            balance_text_col = str(c)
    # إن لم يظهر عمود رصيد نصي واضح، نُفضّل عموداً رقمياً ثالثاً (غالباً الرصيد الجاري).
    balance_numeric_col: str | None = None
    movement_set = set(movement_cols)
    numeric_candidates = [name for name in numeric_cols if name not in movement_set]
    if numeric_candidates:
        # نُفضّل عموداً ذا سلوك مالي (كسور/فواصل عشرية) لتفادي التقاط أرقام السند.
        def _money_like_score(col_name: str) -> tuple[float, int]:
            src = df[col_name].astype(str).fillna("")
            has_decimal_hint = float(src.str.contains(r"[\.٫]", regex=True).mean())
            vals = pd.to_numeric(df[col_name], errors="coerce").dropna()
            frac_ratio = float(((vals - vals.round()).abs() > 0.0001).mean()) if len(vals) else 0.0
            return (max(has_decimal_hint, frac_ratio), list(df.columns).index(col_name))

        balance_numeric_col = max(
            numeric_candidates,
            key=_money_like_score,
        )

    date_pat = re.compile(r"\d{4}[-/]\d{2}[-/]\d{2}")
    out_rows: list[dict[str, Any]] = []
    for _, row in df.iterrows():
        blob_parts: list[str] = []
        for c in df.columns:
            v = row.get(c)
            if pd.notna(v) and str(v).strip():
                blob_parts.append(str(v).strip())
        blob = _normalize_arabic_digits(" ".join(blob_parts))
        dm = date_pat.search(blob)
        if not dm:
            continue
        date_s = dm.group(0).replace("/", "-")
        doc_t = infer_document_kind_from_narrative(blob) or _infer_doc_type_fallback(blob)

        row_vals: list[float] = []
        for c in movement_cols:
            v = safe(row.get(c))
            if v is None:
                continue
            x = abs(float(v))
            if x <= 0.0001:
                continue
            row_vals.append(x)
        if not row_vals:
            continue
        amt = max(row_vals)

        ns = unicodedata.normalize("NFKC", blob)
        has_credit = ("دائن" in ns) or ("نئاد" in ns) or ("ﺩﺍﺋﻥ" in blob)
        has_debit = ("مدين" in ns) or ("نيدم" in ns) or ("ﻣﺩﻳﻥ" in blob)
        debit_val: Any = ""
        credit_val: Any = ""
        kind = (doc_t or "")
        if "توريد" in kind:
            credit_val = round(amt, 2)
        elif "صرف" in kind:
            debit_val = round(amt, 2)
        elif has_credit and not has_debit:
            credit_val = round(amt, 2)
        elif has_debit and not has_credit:
            debit_val = round(amt, 2)
        else:
            debit_val = round(amt, 2)
        debit_val, credit_val = _normalize_debit_credit_cells(debit_val, credit_val)

        bal_val: Any = ""
        if balance_numeric_col and balance_numeric_col in df.columns:
            raw_bal = safe(row.get(balance_numeric_col))
            if raw_bal is not None:
                bal_val = round(abs(float(raw_bal)), 2)
        if balance_text_col and balance_text_col in df.columns:
            bv = row.get(balance_text_col)
            if pd.notna(bv) and str(bv).strip():
                text_bal = _extract_balance_from_text(str(bv))
                if text_bal != "":
                    bal_val = text_bal
        if bal_val == "":
            bal_val = _extract_balance_from_text(blob)

        doc_no = ""
        if doc_no_col and doc_no_col in df.columns:
            doc_no = _extract_doc_number_from_text(row.get(doc_no_col))
        if not doc_no:
            doc_no = _extract_doc_number_from_text(blob)

        out_rows.append(
            {
                "رقم السند": doc_no,
                "التاريخ": date_s,
                "نوع المستند": doc_t or "",
                "مدين": debit_val,
                "دائن": credit_val,
                "الرصيد": bal_val,
            }
        )

    if len(out_rows) < 2:
        return None
    return pd.DataFrame(out_rows)


def _infer_doc_type_fallback(narrative: str) -> str:
    """كلمات شائعة في كشوف العملاء إذا لم يُستنتج النوع من القاموس الرئيسي."""
    if not narrative:
        return ""
    s = unicodedata.normalize("NFKC", _normalize_arabic_digits(narrative))
    if "افتتاح" in s or "الافتتاحي" in s or "ﺍﻺﻔﺘﺗﺎﺤﻳ" in narrative:
        return "الرصيد الافتتاحي"
    if "سند قبض" in s or "ﻕﺑﻗ ﺩﻧﺳ" in narrative:
        if "بنك" in s or "ﻲﻛﻧﺑ" in narrative:
            return "سند قبض بنكي"
        return "سند قبض"
    if "سند صرف" in s:
        return "سند صرف"
    if ("توريد" in s and "مخزن" in s) or ("ديروت" in s and "ينزخم" in s) or ("ﻲﻧزﺧﻣ" in narrative and "دﯾروﺗ" in narrative):
        return "توريد مخزني"
    if ("صرف" in s and "مخزن" in s) or ("فرص" in s and "ينزخم" in s) or ("ﻲﻧزﺧﻣ" in narrative and "فرﺻ" in narrative):
        return "صرف مخزني"
    if "تحويل" in s and "مخزن" in s:
        return "تحويل مخزني"
    if "فاتورة مبيعات" in s or ("مبيعات" in s and "فاتور" in s):
        if "آجل" in s or "ﺝﻵﺍ" in narrative:
            return "فاتورة مبيعات آجلة"
        return "فاتورة مبيعات"
    if "فاتورة مشتريات" in s or ("مشتريات" in s and "فاتور" in s):
        return "فاتورة مشتريات"
    if "مردود" in s:
        return "مردود"
    if "دفع" in s or "ﻊﻓﺩ" in narrative:
        return "دفعة / تسوية"
    return ""


_NUM_TOKEN = re.compile(r"[-+]?\d[\d,٬٫\.]*")


def _extract_amount_candidates(text: str) -> list[float]:
    vals: list[float] = []
    s = _normalize_arabic_digits(text or "")
    for tok in _NUM_TOKEN.findall(s):
        v = safe(tok)
        if v is None:
            continue
        x = float(v)
        # تجاهل الأرقام الصغيرة غالباً (رقم تسلسل/يوم)
        if abs(x) < 1:
            continue
        vals.append(abs(x))
    return vals


def _extract_balance_from_text(text: str) -> Any:
    s0 = _normalize_arabic_digits(text or "")
    s = unicodedata.normalize("NFKC", s0)
    s = s.replace("ی", "ي").replace("ى", "ي").replace("ک", "ك")
    m = re.search(r"(?:مدين|دائن|نيدم|نئاد|نئد|نادئ)\s*([-+]?\d[\d,٬٫\.]*)", s)
    if not m:
        m = re.search(r"([-+]?\d[\d,٬٫\.]*)\s*(?:مدين|دائن|نيدم|نئاد|نئد|نادئ)", s)
    if not m:
        return ""
    v = safe(m.group(1))
    if v is None:
        return ""
    return round(abs(float(v)), 2)


def _extract_statement_rows_from_pdf_text(file_path: str) -> pd.DataFrame | None:
    """
    Fallback لكشوف الحساب التي تفشل extract_tables فيها.
    يلتقط أسطر الحركة بصيغة: [رقم سند] [نوع] [تاريخ] [مدين] [دائن] [رصيد...]
    """
    date_pat = re.compile(
        r"(\d{4}\s*\D\s*\d{2}\s*\D\s*\d{2})|(\d{2}\s*\D\s*\d{2}\s*\D\s*\d{4})"
    )
    rows: list[dict[str, Any]] = []
    try:
        with pdfplumber.open(file_path) as pdf:
            for pg in pdf.pages:
                txt = pg.extract_text() or ""
                for raw in txt.splitlines():
                    line = _normalize_arabic_digits((raw or "").replace("\x00", " "))
                    line = re.sub(r"\s+", " ", line).strip()
                    if not line:
                        continue
                    dm = date_pat.search(line)
                    if not dm:
                        continue
                    date_s = dm.group(0).replace("/", "-").replace(".", "-")
                    after = line[dm.end() :].strip()
                    raw_nums = _NUM_TOKEN.findall(after)
                    nums: list[str] = []
                    for tok in raw_nums:
                        v = safe(tok)
                        if v is None:
                            continue
                        t = tok.strip()
                        # تجاهل شظايا التاريخ مثل -04 و -01، والإبقاء على مبالغ/أصفار الحركة.
                        looks_amount = ("." in t or "٫" in t or abs(float(v)) >= 50.0 or abs(float(v)) <= 0.0001)
                        if looks_amount:
                            nums.append(t)
                    if len(nums) < 2:
                        continue
                    deb = safe(nums[0])
                    cre = safe(nums[1])
                    if deb is None and cre is None:
                        continue
                    if (deb is None or abs(float(deb)) <= 0.0001) and (cre is None or abs(float(cre)) <= 0.0001):
                        continue
                    bal = safe(nums[2]) if len(nums) >= 3 else None
                    rows.append(
                        {
                            "التاريخ": date_s,
                            "مدين": float(deb) if deb is not None else "",
                            "دائن": float(cre) if cre is not None else "",
                            "الرصيد": float(bal) if bal is not None else "",
                            "بيان": raw.strip(),
                            "مستند": "",
                        }
                    )
    except Exception:
        return None
    if len(rows) < 2:
        return None
    return pd.DataFrame(rows)


def _extract_table_df_direct(file_path: str) -> pd.DataFrame | None:
    """استخراج مباشر من جداول pdfplumber لاستخدامه كـ fallback عند فشل read_pdf في اختيار المرشح الأفضل."""
    try:
        with pdfplumber.open(file_path) as pdf:
            tables = [tb for pg in pdf.pages for tb in (pg.extract_tables() or [])]
            grid_rows = [rw for tb in tables for rw in (tb or []) if rw and any(c is not None for c in rw)]
        if not grid_rows:
            return None
        df = _dataframe_from_pdf_grid(grid_rows)
        if df is None or df.empty:
            return None
        return _trim_pdf_table_df(df.reset_index(drop=True))
    except Exception:
        return None


def _normalize_debit_credit_by_context(doc_t: str, narrative: str, deb: Any, cre: Any) -> tuple[Any, Any]:
    """تصحيح انحرافات استخراج مدين/دائن عندما يظهر المبلغ الحقيقي داخل البيان فقط."""
    d = safe(deb)
    c = safe(cre)
    d0 = abs(float(d)) if d is not None else 0.0
    c0 = abs(float(c)) if c is not None else 0.0

    nums = _extract_amount_candidates(narrative)
    major = max(nums) if nums else 0.0
    if major <= 0:
        return deb, cre

    kind = (doc_t or "").strip()
    ns = unicodedata.normalize("NFKC", _normalize_arabic_digits(narrative or ""))
    has_credit_hint = ("دائن" in ns) or ("نئاد" in ns) or ("ﺩﺍﺋﻥ" in narrative)
    has_debit_hint = ("مدين" in ns) or ("نيدم" in ns) or ("ﻣﺩﻳﻥ" in narrative)
    # حالة شائعة في كشوف المخازن: عمود الحركة يلتقط المبلغ في المدين رغم أن النص يصرّح "دائن".
    if "توريد مخزني" in kind and has_credit_hint and not has_debit_hint and d0 > 0 and c0 <= 0.0001:
        return "", round(max(major, d0), 2)

    # إذا المبلغ المستخرج في الخانات أصغر بكثير من المبلغ الموجود في البيان، نستخدم الأكبر.
    weak_detect = (max(d0, c0) <= 0) or (major >= max(d0, c0) * 1.8 and major >= 50)
    if not weak_detect:
        return deb, cre

    if "قبض" in kind:
        return "", round(major, 2)
    if "صرف" in kind:
        return round(major, 2), ""
    if "توريد مخزني" in kind:
        return "", round(major, 2)
    if "صرف مخزني" in kind:
        return round(major, 2), ""
    if "مبيعات" in kind or "افتتاح" in kind:
        return round(major, 2), ""
    if "مشتريات" in kind:
        return "", round(major, 2)
    # في النوع غير المعروف فقط: استخدم التلميح الصريح دائن/مدين.
    if has_credit_hint and not has_debit_hint:
        return "", round(major, 2)
    if has_debit_hint and not has_credit_hint:
        return round(major, 2), ""
    # افتراضي: حافظ على الجانب الأقوى إن وجد، وإلا ضعها مدين
    if c0 > d0:
        return "", round(major, 2)
    return round(major, 2), ""


def _is_noise_ledger_row(dt: Any, doc_t: str, deb: Any, cre: Any, bal: Any) -> bool:
    """حذف صفوف ضجيج شائعة (مثل رقم 1 القادم من الترويسة) قبل التصدير."""
    d = safe(deb)
    c = safe(cre)
    b = safe(bal)
    d0 = abs(float(d)) if d is not None else 0.0
    c0 = abs(float(c)) if c is not None else 0.0
    b0 = abs(float(b)) if b is not None else 0.0
    has_doc = bool((doc_t or "").strip())
    has_date = pd.notna(dt) and str(dt).strip() not in ("", "NaT", "nat")
    # صف بتاريخ + مبلغ صغير جداً + بدون نوع مستند/رصيد => غالباً ليس حركة.
    return has_date and not has_doc and b0 <= 0.0001 and max(d0, c0) <= 1.0


def _reframe_ledger_columns_clean(df: pd.DataFrame) -> pd.DataFrame:
    """
    صيغة موحّدة للتصدير: التاريخ، نوع المستند، مدين، دائن، الرصيد (فارغ إن لم يُستخرج عمود رصيد).
    """
    if df is None or df.empty:
        return df
    anon = _reframe_anonymous_pdf_table(df)
    if anon is not None and not anon.empty:
        return anon

    detected_cols = detect_columns(df.copy())
    debit_col = detected_cols.get("debit")
    credit_col = detected_cols.get("credit")
    date_col = detected_cols.get("date")
    if "التاريخ" in df.columns:
        date_col = date_col or "التاريخ"
    if debit_col is None and "مدين" in df.columns:
        debit_col = "مدين"

    if date_col is None or debit_col is None or date_col not in df.columns or debit_col not in df.columns:
        return df

    doc_no_col = _find_doc_number_column(df)
    # بعض ملفات PDF تُسند رقم السند خطأً إلى دائن/مدين عبر detect_columns.
    if doc_no_col is not None:
        if debit_col == doc_no_col:
            debit_col = "مدين" if "مدين" in df.columns else None
        if credit_col == doc_no_col:
            credit_col = "دائن" if "دائن" in df.columns else None
    if debit_col is None and "مدين" in df.columns:
        debit_col = "مدين"
    if credit_col is None and "دائن" in df.columns:
        credit_col = "دائن"

    if date_col is None or debit_col is None:
        return df

    balance_col = None
    for c in df.columns:
        if _is_balance_column_name(c):
            balance_col = c
            break

    doc_type_col = detect_document_type_column(df)
    if not doc_type_col:
        prim, _fb = resolve_document_columns(df)
        doc_type_col = prim

    rows: list[dict[str, Any]] = []
    for _, row in df.iterrows():
        dt = row.get(date_col)
        deb = row.get(debit_col)
        cre = row.get(credit_col) if credit_col and credit_col in df.columns else None
        bal = row.get(balance_col) if balance_col and balance_col in df.columns else None
        doc_no = ""
        if doc_no_col and doc_no_col in df.columns:
            doc_no = _extract_doc_number_from_text(row.get(doc_no_col))

        if doc_type_col and doc_type_col in df.columns:
            raw_dt = row.get(doc_type_col)
            doc_t = str(raw_dt).strip() if pd.notna(raw_dt) and str(raw_dt).strip() else ""
        else:
            doc_t = ""

        if not doc_t:
            nar = _narrative_blob(row, df)
            doc_t = infer_document_kind_from_narrative(nar) or _infer_doc_type_fallback(nar)
        else:
            nar = _narrative_blob(row, df)

        deb, cre = _normalize_debit_credit_by_context(doc_t, nar, deb, cre)
        deb, cre = _normalize_debit_credit_cells(deb, cre)

        def _blank_num(x: Any) -> Any:
            if x is None:
                return ""
            if isinstance(x, float) and pd.isna(x):
                return ""
            return x

        if _is_noise_ledger_row(dt, doc_t, deb, cre, bal):
            continue

        rows.append(
            {
                "رقم السند": doc_no,
                "التاريخ": dt,
                "نوع المستند": doc_t or "",
                "مدين": deb,
                "دائن": _blank_num(cre),
                "الرصيد": _blank_num(bal),
            }
        )

    return pd.DataFrame(rows)


def pdf_to_excel_bytes(file_path: str) -> bytes:
    df = read_pdf(file_path)
    if df is not None and len(df) <= 2:
        table_df = _extract_table_df_direct(file_path)
        if table_df is not None and len(table_df) > len(df):
            df = table_df
        else:
            fallback_df = _extract_statement_rows_from_pdf_text(file_path)
            if fallback_df is not None and len(fallback_df) > len(df):
                df = fallback_df
    if df is None or df.empty:
        raise ValueError(
            "تعذر استخراج بيانات من PDF. جرّب ملفاً نصياً (وليس صورة ممسوحة)، أو صدّره من البرنامج كـ PDF، أو استخدم Excel/CSV."
        )
    raw = df.copy()
    df = _prepare_dataframe_for_excel_export(df)
    # إن أزالت إزالة الترويسة كل الصفوف، نصدّر الجدول الخام كما استُخرج (أفضل من فشل كامل).
    if df is None or df.empty:
        df = raw

    df = _reframe_ledger_columns_clean(df)

    df = _apply_arabic_for_excel_export(df)

    out = io.BytesIO()
    with pd.ExcelWriter(out, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="converted")
        ws = writer.book["converted"]
        ws.sheet_view.rightToLeft = True
        _format_converted_excel_sheet(ws)
    return out.getvalue()


def _format_converted_excel_sheet(ws: Any) -> None:
    """عرض أوضح: تفاف نص، عرض أعمدة يمنع قص التواريخ/العربي في الواجهة."""
    from openpyxl.styles import Alignment
    from openpyxl.utils import get_column_letter

    body = Alignment(wrap_text=True, vertical="top")
    header = Alignment(wrap_text=True, vertical="center", horizontal="center")

    for cell in ws[1]:
        cell.alignment = header
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for cell in row:
            cell.alignment = body

    max_scan = min(ws.max_row, 800)
    for c in range(1, ws.max_column + 1):
        max_len = 14
        for r in range(1, max_scan + 1):
            val = ws.cell(row=r, column=c).value
            if val is None:
                continue
            s = str(val)
            max_len = max(max_len, min(len(s), 120))
        # عربي يحتاج عرضاً أكبر من عدد الحروف؛ حد أعلى معقول لـ Excel
        width = min(max(max_len * 1.12 + 2, 14), 78)
        ws.column_dimensions[get_column_letter(c)].width = width


def _apply_arabic_for_excel_export(df: pd.DataFrame) -> pd.DataFrame:
    """يُطبَّق على الأعمدة والخلايا النصية فقط؛ الأرقام تبقى أرقاماً."""
    out = df.copy()
    out.columns = [_excel_arabic_display(str(c)) for c in out.columns]
    for col in out.columns:
        out[col] = out[col].map(lambda x: _excel_arabic_display(x) if isinstance(x, str) else x)
    return out


def pdf_file_to_excel_bytes(pdf_path: str | Path) -> bytes:
    """للاختبار أو الاستدعاء من مسارات ملف."""
    return pdf_to_excel_bytes(str(pdf_path))
