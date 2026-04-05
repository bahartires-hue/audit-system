from __future__ import annotations

import csv
import io
import os
import platform
from pathlib import Path
from typing import Any, Dict, List


def mismatches_to_csv_bytes(entries: List[Dict[str, Any]]) -> bytes:
    """
    Return CSV bytes encoded so that Excel opens Arabic correctly.
    """
    output = io.StringIO()
    writer = csv.writer(output)

    writer.writerow(["الفرع", "المبلغ", "نوع العملية", "التاريخ", "المستند", "السبب"])
    for e in entries:
        writer.writerow(
            [
                e.get("branch", ""),
                e.get("amount", ""),
                e.get("type", ""),
                e.get("date", "") or "",
                e.get("doc", "") or "",
                e.get("reason", "") or "",
            ]
        )

    # utf-8-sig helps Excel detect Arabic correctly
    return output.getvalue().encode("utf-8-sig")


def mismatches_to_excel_bytes(entries: List[Dict[str, Any]]) -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

    headers = ["الفرع", "المبلغ", "التاريخ", "المستند", "السبب"]
    rows = []
    for e in entries:
        rows.append(
            [
                e.get("branch", "") or "",
                e.get("amount", ""),
                e.get("date", "") or "",
                e.get("doc", "") or "",
                e.get("reason", "") or "",
            ]
        )

    wb = Workbook()
    ws = wb.active
    ws.title = "errors"
    ws.sheet_view.rightToLeft = True

    header_fill = PatternFill(fill_type="solid", start_color="FFF200", end_color="FFF200")
    header_font = Font(color="FF0000", bold=True, size=12)
    body_font = Font(color="000000", size=11)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin = Side(style="thin", color="000000")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for col, h in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center
        cell.border = border

    for r_idx, r in enumerate(rows, start=2):
        for c_idx, val in enumerate(r, start=1):
            cell = ws.cell(row=r_idx, column=c_idx, value=val)
            cell.font = body_font
            cell.alignment = center
            cell.border = border

    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 16
    ws.column_dimensions["D"].width = 20
    ws.column_dimensions["E"].width = 48
    ws.freeze_panes = "A2"

    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()


def _register_arabic_pdf_font() -> str:
    """تسجيل خط يدعم العربية؛ يعيد اسماً مسجلاً في ReportLab."""
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.ttfonts import TTFont

    name = "OmArabic"
    if name in pdfmetrics.getRegisteredFontNames():
        return name

    path = (os.environ.get("AUDITFLOW_ARABIC_FONT") or "").strip()
    if not path or not Path(path).is_file():
        if platform.system() == "Windows":
            wind = os.environ.get("WINDIR", r"C:\Windows")
            cand = Path(wind) / "Fonts" / "arial.ttf"
            if cand.is_file():
                path = str(cand)
    if path and Path(path).is_file():
        pdfmetrics.registerFont(TTFont(name, path))
        return name
    return "Helvetica"


def _shape_pdf_cell(text: str, font_name: str) -> str:
    s = str(text or "")
    if font_name == "Helvetica":
        return s
    try:
        import arabic_reshaper
        from bidi.algorithm import get_display

        reshaped = arabic_reshaper.reshape(s)
        return get_display(reshaped)
    except Exception:
        return s


def mismatches_to_pdf_bytes(entries: List[Dict[str, Any]]) -> bytes:
    try:
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import A4
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle
    except ImportError as e:
        raise RuntimeError("تصدير PDF يحتاج: pip install reportlab") from e

    font = _register_arabic_pdf_font()
    header_font = "Helvetica-Bold" if font == "Helvetica" else font

    out = io.BytesIO()
    doc = SimpleDocTemplate(
        out,
        pagesize=A4,
        rightMargin=24,
        leftMargin=24,
        topMargin=24,
        bottomMargin=24,
    )

    headers = ["الفرع", "المبلغ", "التاريخ", "المستند", "السبب"]
    data: List[List[str]] = [[_shape_pdf_cell(h, font) for h in headers]]
    for e in entries:
        data.append(
            [
                _shape_pdf_cell(str(e.get("branch", "") or ""), font),
                _shape_pdf_cell(str(e.get("amount", "") or ""), font),
                _shape_pdf_cell(str(e.get("date", "") or ""), font),
                _shape_pdf_cell(str(e.get("doc", "") or ""), font),
                _shape_pdf_cell(str(e.get("reason", "") or ""), font),
            ]
        )

    table = Table(data, colWidths=[90, 60, 80, 90, 210], repeatRows=1)
    table.setStyle(
        TableStyle(
            [
                ("GRID", (0, 0), (-1, -1), 0.8, colors.black),
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#FFF200")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.red),
                ("FONTNAME", (0, 0), (-1, 0), header_font),
                ("FONTSIZE", (0, 0), (-1, 0), 11),
                ("ALIGN", (0, 0), (-1, -1), "CENTER"),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                ("FONTSIZE", (0, 1), (-1, -1), 9),
                ("FONTNAME", (0, 1), (-1, -1), font),
            ]
        )
    )

    doc.build([table])
    return out.getvalue()

