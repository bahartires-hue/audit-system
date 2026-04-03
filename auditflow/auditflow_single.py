from __future__ import annotations

"""
AuditFlow (single-file version)
--------------------------------
Backend + Frontend in ONE Python file.

Run:
  pip install fastapi uvicorn[standard] sqlalchemy pandas pdfplumber openpyxl python-multipart
  uvicorn auditflow_single:app --host 127.0.0.1 --port 8001

Open:
  http://127.0.0.1:8001/
"""

import csv
import datetime as dt
import hashlib
import hmac
import io
import json
import os
import re
import secrets
import shutil
import unicodedata
import uuid
import zipfile
from difflib import SequenceMatcher
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

import pandas as pd
import pdfplumber
from fastapi import FastAPI, File, Form, HTTPException, Query, Request, UploadFile
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import FileResponse, HTMLResponse, RedirectResponse, Response
from sqlalchemy import JSON, Column, DateTime, ForeignKey, Integer, String, create_engine, text
from sqlalchemy.orm import Session, declarative_base, sessionmaker

try:
    from reportlab.lib.pagesizes import A4
    from reportlab.pdfgen import canvas
except Exception:
    A4 = None
    canvas = None


# =========================
# CONFIG / PATHS
# =========================
BASE_DIR = Path(__file__).resolve().parent
DB_PATH = Path(os.getenv("AUDITFLOW_DB", str(BASE_DIR / "auditflow.db")))
UPLOAD_DIR = BASE_DIR / "uploads"
BACKUP_DIR = BASE_DIR / "backups"


# =========================
# DB
# =========================
engine = create_engine(
    f"sqlite:///{DB_PATH}",
    connect_args={"check_same_thread": False},
)
SessionLocal = sessionmaker(autocommit=False, autoflush=False, bind=engine)
Base = declarative_base()


class AnalysisReport(Base):
    __tablename__ = "analysis_reports"

    id = Column(String, primary_key=True)  # uuid4 hex
    user_id = Column(String, ForeignKey("users.id"), nullable=True, index=True)
    title = Column(String, nullable=True)

    branch1_name = Column(String, nullable=False)
    branch2_name = Column(String, nullable=False)

    file1_original = Column(String, nullable=True)
    file2_original = Column(String, nullable=True)
    file1_path = Column(String, nullable=True)
    file2_path = Column(String, nullable=True)

    status = Column(String, default="completed", nullable=False)
    created_at = Column(DateTime, default=dt.datetime.utcnow, nullable=False)

    total_ops = Column(Integer, nullable=False, default=0)
    matched_ops = Column(Integer, nullable=False, default=0)
    mismatch_ops = Column(Integer, nullable=False, default=0)
    errors_count = Column(Integer, nullable=False, default=0)
    warnings_count = Column(Integer, nullable=False, default=0)

    stats_json = Column(JSON, nullable=False, default=dict)
    analysis_json = Column(JSON, nullable=False, default=dict)


class User(Base):
    __tablename__ = "users"

    id = Column(String, primary_key=True)
    username = Column(String, unique=True, nullable=False, index=True)
    password_hash = Column(String, nullable=False)
    failed_attempts = Column(Integer, nullable=False, default=0)
    locked_until = Column(DateTime, nullable=True)
    created_at = Column(DateTime, default=dt.datetime.utcnow, nullable=False)


class AuditLog(Base):
    __tablename__ = "audit_logs"

    id = Column(String, primary_key=True)
    user_id = Column(String, ForeignKey("users.id"), nullable=True, index=True)
    action = Column(String, nullable=False)
    meta_json = Column(JSON, nullable=False, default=dict)
    created_at = Column(DateTime, default=dt.datetime.utcnow, nullable=False)


class UserSession(Base):
    __tablename__ = "user_sessions"

    token = Column(String, primary_key=True)
    user_id = Column(String, ForeignKey("users.id"), nullable=False, index=True)
    created_at = Column(DateTime, default=dt.datetime.utcnow, nullable=False)
    expires_at = Column(DateTime, nullable=False)


Base.metadata.create_all(bind=engine)


def run_migrations() -> None:
    """Best-effort SQLite migrations for existing local DB."""
    with engine.connect() as conn:
        cols = [r[1] for r in conn.execute(text("PRAGMA table_info(analysis_reports)")).fetchall()]
        if "user_id" not in cols:
            conn.execute(text("ALTER TABLE analysis_reports ADD COLUMN user_id VARCHAR"))
            conn.commit()

        cols = [r[1] for r in conn.execute(text("PRAGMA table_info(users)")).fetchall()]
        if "failed_attempts" not in cols:
            conn.execute(text("ALTER TABLE users ADD COLUMN failed_attempts INTEGER DEFAULT 0"))
            conn.commit()
        if "locked_until" not in cols:
            conn.execute(text("ALTER TABLE users ADD COLUMN locked_until DATETIME"))
            conn.commit()


def db_session() -> Session:
    return SessionLocal()


# =========================
# AUTH
# =========================
SESSION_COOKIE = "auditflow_session"
CSRF_COOKIE = "auditflow_csrf"
SESSION_DAYS = 14
MAX_UPLOAD_SIZE_MB = 15
LOCK_MINUTES = 15


def hash_password(password: str) -> str:
    salt = secrets.token_bytes(16)
    dk = hashlib.pbkdf2_hmac("sha256", password.encode("utf-8"), salt, 200000)
    return f"pbkdf2_sha256${salt.hex()}${dk.hex()}"


def verify_password(password: str, stored: str) -> bool:
    # Backward compatibility for old sha256 hashes.
    if "$" not in stored:
        return hashlib.sha256(password.encode("utf-8")).hexdigest() == stored
    try:
        algo, salt_hex, hash_hex = stored.split("$", 2)
        if algo != "pbkdf2_sha256":
            return False
        salt = bytes.fromhex(salt_hex)
        expected = bytes.fromhex(hash_hex)
        dk = hashlib.pbkdf2_hmac("sha256", password.encode("utf-8"), salt, 200000)
        return hmac.compare_digest(dk, expected)
    except Exception:
        return False


def create_session(db: Session, user_id: str) -> str:
    token = secrets.token_urlsafe(40)
    now = dt.datetime.utcnow()
    session = UserSession(
        token=token,
        user_id=user_id,
        created_at=now,
        expires_at=now + dt.timedelta(days=SESSION_DAYS),
    )
    db.add(session)
    db.commit()
    return token


def issue_csrf_token() -> str:
    return secrets.token_urlsafe(24)


def require_csrf(request: Request) -> None:
    cookie_token = request.cookies.get(CSRF_COOKIE, "")
    header_token = request.headers.get("x-csrf-token", "")
    if not cookie_token or not header_token or not hmac.compare_digest(cookie_token, header_token):
        raise HTTPException(403, "CSRF token غير صالح")


def current_user_from_request(db: Session, request: Request) -> Optional[User]:
    token = request.cookies.get(SESSION_COOKIE)
    if not token:
        return None
    s = db.query(UserSession).filter(UserSession.token == token).first()
    if not s:
        return None
    if s.expires_at < dt.datetime.utcnow():
        db.delete(s)
        db.commit()
        return None
    return db.query(User).filter(User.id == s.user_id).first()


def require_user(db: Session, request: Request) -> User:
    u = current_user_from_request(db, request)
    if not u:
        raise HTTPException(401, "يرجى تسجيل الدخول أولاً")
    return u


def log_event(db: Session, action: str, user_id: Optional[str] = None, meta: Optional[Dict[str, Any]] = None) -> None:
    db.add(
        AuditLog(
            id=uuid.uuid4().hex,
            user_id=user_id,
            action=action,
            meta_json=meta or {},
        )
    )
    db.commit()


# =========================
# UTILS (storage)
# =========================
def ensure_dir(p: Path) -> None:
    p.mkdir(parents=True, exist_ok=True)


def save_upload_file(upload: UploadFile, dest_dir: Path) -> Tuple[str, str]:
    ensure_dir(dest_dir)
    original = upload.filename or "upload"
    suffix = Path(original).suffix
    saved_name = f"{uuid.uuid4().hex}{suffix}"
    saved_path = dest_dir / saved_name

    content = upload.file.read()
    max_bytes = MAX_UPLOAD_SIZE_MB * 1024 * 1024
    if len(content) > max_bytes:
        raise HTTPException(400, f"حجم الملف أكبر من {MAX_UPLOAD_SIZE_MB}MB")

    lower = original.lower()
    if lower.endswith(".pdf") and not content.startswith(b"%PDF"):
        raise HTTPException(400, "ملف PDF غير صالح")
    if lower.endswith(".xlsx") and not content.startswith(b"PK"):
        raise HTTPException(400, "ملف Excel (.xlsx) غير صالح")
    if lower.endswith(".xls") and not content.startswith(b"\xD0\xCF\x11\xE0"):
        raise HTTPException(400, "ملف Excel (.xls) غير صالح")

    allowed = (".pdf", ".xlsx", ".xls", ".csv", ".xlsm", ".xlsb")
    if not lower.endswith(allowed):
        raise HTTPException(400, "نوع الملف غير مدعوم. المسموح: Excel/PDF")

    with open(saved_path, "wb") as f:
        f.write(content)

    return str(saved_path), original


# =========================
# ANALYZER (local)
# =========================
def _is_plausible_currency_amount(x: float) -> bool:
    """Reject tax IDs / long reference numbers mistaken for amounts (e.g. 311378677200003)."""
    try:
        xf = float(x)
    except (TypeError, ValueError):
        return False
    if pd.isna(xf) or xf != xf:
        return False
    ax = abs(xf)
    if ax < 1e-9:
        return True
    if ax >= 1e11:
        return False
    if float(x) == int(x) and ax >= 1e6:
        nd = len(str(int(ax)))
        if nd >= 12:
            return False
        if nd >= 10 and ax >= 1e9:
            return False
    return True


def _normalize_doc_text(s: Any) -> str:
    t = unicodedata.normalize("NFKC", str(s or ""))
    t = re.sub(r"\s+", " ", t).strip()
    return t


def _arabic_letters_for_match(s: str) -> str:
    """توحید أشكال الألف (ی/ي، ة/ه) لتحسین مطابقة کلمات مستخرجة من PDF."""
    t = unicodedata.normalize("NFKC", s or "")
    for a, b in (
        ("ى", "ي"),
        ("ی", "ي"),
        ("ئ", "ي"),
        ("ة", "ه"),
        ("ۀ", "ه"),
        ("ٱ", "ا"),
        ("أ", "ا"),
        ("إ", "ا"),
        ("آ", "ا"),
    ):
        t = t.replace(a, b)
    return t.lower()


_DOC_KIND_SPECS: List[Tuple[str, str]] = [
    ("مردود مشتريات", "مردود مشتريات"),
    ("مردود مبيعات", "مردود مبيعات"),
    # كلمات شائعة بخطأ اتجاه/عكس حروف في مستخرج PDF
    ("هروتاف", "فاتورة"),
    ("تاعیبم", "مبيعات"),
    ("تاعيبم", "مبيعات"),
    ("تایرتشم", "مشتريات"),
    ("تارتشم", "مشتريات"),
    ("تايرتشم", "مشتريات"),
    ("مشتريات", "مشتريات"),
    ("مبیعات", "مبيعات"),
    ("مبيعات", "مبيعات"),
    ("فاتوره", "فاتورة"),
    ("فاتورة مبيعات", "فاتورة مبيعات"),
    ("فاتورة مشتريات", "فاتورة مشتريات"),
    ("فاتورة", "فاتورة"),
    ("سند قبض", "سند قبض"),
    ("سند صرف", "سند صرف"),
    ("قيد يومية", "قيد يومية"),
    ("طلب شراء", "مشتريات"),
    ("طلب بيع", "مبيعات"),
]


def _has_arabic(s: str) -> bool:
    return any("\u0600" <= c <= "\u06ff" for c in (s or ""))


def _expand_text_for_doc_kind(s: str) -> str:
    t = _normalize_doc_text(s)
    if not t:
        return ""
    parts = [t]
    words = t.split()
    rev = []
    for w in words:
        if _has_arabic(w) and len(w) >= 3:
            rev.append(w[::-1])
        else:
            rev.append(w)
    parts.append(" ".join(rev))
    return " \n ".join(parts)


def infer_document_kind_from_narrative(text: Optional[str]) -> Optional[str]:
    """استنتاج نوع المستند للعرض والمطابقة (مبيعات/مشتريات/...) من نص مشوّه أو RTL."""
    if not text or not str(text).strip():
        return None
    hay_raw = _expand_text_for_doc_kind(str(text))
    hay = _arabic_letters_for_match(hay_raw)
    if "هروتاف" in hay and "تاعيبم" in hay:
        return "فاتورة مبيعات"
    if "هروتاف" in hay and "تايرتشم" in hay:
        return "فاتورة مشتريات"
    for needle, label in _DOC_KIND_SPECS:
        n = _arabic_letters_for_match(needle)
        if n in hay or needle in hay_raw:
            return label
    return None


def _enrich_doc_field(doc_out: Optional[str], narrative: str) -> Optional[str]:
    kind = infer_document_kind_from_narrative(narrative) or infer_document_kind_from_narrative(doc_out or "")
    if kind:
        return kind
    return doc_out


def _dedupe_extracted_rows(rows: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    """دمج صفوف متطابقة (فرع + مبلغ + نوع + یوم) غالبًا بسبب تكرار سطر PDF."""
    order: List[Tuple[Any, ...]] = []
    by_key: Dict[Tuple[Any, ...], Dict[str, Any]] = {}

    def score_doc(d: Any) -> int:
        s = str(d or "")
        if not s:
            return 0
        if s in (
            "مبيعات",
            "مشتريات",
            "مردود مبيعات",
            "مردود مشتريات",
            "فاتورة",
            "فاتورة مبيعات",
            "فاتورة مشتريات",
            "سند قبض",
            "سند صرف",
            "قيد يومية",
        ) or s.startswith("فاتورة "):
            return 100
        return max(0, 80 - len(s))

    for r in rows:
        k = (
            r.get("branch"),
            round(float(r["amount"]), 2),
            r.get("type"),
            r.get("date") or "",
        )
        if k not in by_key:
            by_key[k] = r
            order.append(k)
        elif score_doc(r.get("doc")) > score_doc(by_key[k].get("doc")):
            by_key[k] = r

    return [by_key[k] for k in order]


def _parse_currency_numbers_from_narrative(narrative: str) -> List[float]:
    s = _normalize_arabic_digits(narrative or "")
    tokens = re.findall(r"[-+]?\d[\d,٬٫\.]*", s)
    vals: List[float] = []
    for tok in tokens:
        v = _parse_number_token(tok)
        if v is None or not _is_plausible_currency_amount(v):
            continue
        vals.append(float(v))
    if len(vals) >= 2:
        vals = [v for v in vals if not (v == int(v) and (1 <= abs(v) <= 31 or 1900 <= abs(v) <= 2100))] or vals
    return vals


def _looks_like_serial_voucher_amount(x: float) -> bool:
    if x is None or (isinstance(x, float) and pd.isna(x)):
        return False
    xf = float(x)
    if abs(xf - int(xf)) > 0.0001:
        return False
    ax = abs(xf)
    if 1900 <= ax <= 2100:
        return False
    return 100 <= ax <= 49999


def _replace_voucher_with_ledger_from_narrative(column_amount: float, narrative: str) -> float:
    narrative = _normalize_doc_text(narrative)
    if not narrative:
        return column_amount
    vals = _parse_currency_numbers_from_narrative(narrative)
    if not vals:
        return column_amount
    decimals = [v for v in vals if abs(v - int(v)) > 0.0001 and abs(v) >= 0.0001]
    for d in decimals:
        if abs(d - column_amount) < 0.02:
            return column_amount
    if _looks_like_serial_voucher_amount(column_amount) and decimals:
        return round(decimals[0], 2)
    for v in vals:
        if abs(v - column_amount) < 0.02:
            return column_amount
    if decimals:
        return round(decimals[0], 2)
    return column_amount


def _row_narrative_for_amounts(
    row: pd.Series,
    df: pd.DataFrame,
    doc_col: Optional[str],
    doc_fb: Optional[str],
    debit_col: Optional[str],
    credit_col: Optional[str],
    date_col: Optional[str],
) -> str:
    seen: set[Any] = set()
    parts: List[str] = []
    priority: List[Any] = []
    if doc_fb and doc_fb in df.columns:
        priority.append(doc_fb)
    if doc_col and doc_col in df.columns:
        priority.append(doc_col)
    for col in priority:
        if col in seen:
            continue
        seen.add(col)
        v = row.get(col)
        if pd.notna(v):
            t = str(v).strip()
            if t:
                parts.append(t)
    for col in df.columns:
        if col in (debit_col, credit_col, date_col) or col in seen:
            continue
        n = str(col).lower()
        if "ضريبي" in n and "بيان" not in n:
            continue
        if "بيان" in n or "مستند" in n or "نوع" in n or "فاتورة" in n or "وصف" in n or "تفاصيل" in n:
            seen.add(col)
            v = row.get(col)
            if pd.notna(v):
                t = str(v).strip()
                if t:
                    parts.append(t)
    return _normalize_doc_text(" ".join(parts))


def _side_hint_from_narrative(narrative: str) -> Optional[str]:
    hay = _arabic_letters_for_match(_expand_text_for_doc_kind(narrative or ""))
    has_debit = "مدين" in hay
    has_credit = ("دائن" in hay) or ("داين" in hay)
    if has_debit and not has_credit:
        return "debit"
    if has_credit and not has_debit:
        return "credit"
    return None


def _extract_best_amount_from_text(s: str) -> Optional[float]:
    vals = _parse_currency_numbers_from_narrative(s or "")
    if not vals:
        return None
    non_zero = [v for v in vals if abs(v) >= 0.0001]
    if not non_zero:
        return 0.0
    decimals = [v for v in non_zero if abs(v - int(v)) > 0.0001]
    if decimals:
        return round(decimals[-1], 2)
    return round(non_zero[-1], 2)


def safe(v: Any) -> Optional[float]:
    try:
        if v is None:
            return None
        if isinstance(v, (int, float)) and not isinstance(v, bool):
            if isinstance(v, float) and pd.isna(v):
                return None
            x = float(v)
            return round(x, 2) if _is_plausible_currency_amount(x) else None
        s = str(v).replace(",", "").replace("٬", "").strip()
        if not s or s.lower() in ("nan", "none", "-"):
            return None
        s = re.sub(r"(?i)\b(debit|credit|مدين|دائن|dr|cr)\b", "", s)
        s = s.strip()
        if re.fullmatch(r"[-+]?\d+(?:\.\d+)?", s):
            x = float(s)
        else:
            x = _extract_best_amount_from_text(s)
            if x is None:
                return None
        if not _is_plausible_currency_amount(x):
            return None
        return round(x, 2)
    except Exception:
        return None


def _column_name_excludes_from_amount(col: Any) -> bool:
    """Do not use سند / رصيد / تسلسل as مدين or دائن."""
    name = str(col).lower().strip()
    if not name or name.isdigit():
        return False
    if name in ("السند", "سند"):
        return True
    block = (
        "رقم السند",
        "رقم سند",
        "الرصيد",
        "رصيد",
        "balance",
        "تسلسل",
        "مسلسل",
        "seq",
        "serial",
        "ضريبي",
        "هوية",
        "الرقم الضريبي",
        "السجل التجاري",
    )
    for b in block:
        if b in name:
            return True
    if name in ("#", "م", "no", "no."):
        return True
    return False


def _column_values_look_like_ids(series: pd.Series) -> bool:
    nums = pd.to_numeric(series, errors="coerce").dropna()
    if len(nums) == 0:
        return False
    mx = float(nums.max())
    if mx >= 1e11:
        return True
    if mx == int(mx) and mx >= 1e9:
        return len(str(int(mx))) >= 10
    return False


def _is_balance_column_name(col: Any) -> bool:
    n = str(col).lower()
    return "رصيد" in n or "balance" in n


def _currency_amount_rank(series: pd.Series) -> float:
    """تفضيل أعمدة فيها كسور (مبالغ) وتخفيف أعمدة أرقام صحيحة تشبه أرقام السندات."""
    nums = pd.to_numeric(series, errors="coerce").dropna()
    if len(nums) == 0:
        return float("-inf")
    mean_v = float(nums.mean())
    max_v = float(nums.max())
    frac_decimal = ((nums % 1).abs() > 0.001).sum() / len(nums)
    score = mean_v + frac_decimal * min(500_000.0, mean_v * 2.0 + 1.0)
    if frac_decimal < 0.08 and max_v <= 99_999_999:
        score -= min(400_000.0, mean_v * 1.5 + 100_000.0)
    return score


def _column_header_indicates_debit(col: Any) -> bool:
    """الاعتماد الأساسي على اسم العمود: مدين / debit (وليس دائن)."""
    name = str(col).lower().strip()
    if _column_name_excludes_from_amount(col):
        return False
    if "دائن" in name or "credit" in name:
        return False
    if "مدين" in name or "debit" in name:
        return True
    if re.fullmatch(r"dr|d\.r\.?", name) or re.search(r"(^|[^a-z])dr([^a-z]|$)", name):
        return True
    return False


def _column_header_indicates_credit(col: Any) -> bool:
    """الاعتماد الأساسي على اسم العمود: دائن / credit (وليس مدين)."""
    name = str(col).lower().strip()
    if _column_name_excludes_from_amount(col):
        return False
    if "مدين" in name or "debit" in name:
        return False
    if "دائن" in name or "credit" in name:
        return True
    if re.fullmatch(r"cr\.?", name):
        return True
    return False


def _promote_ledger_header_row(df: pd.DataFrame) -> pd.DataFrame:
    """
    جداول PDF غالبًا بلا عناوين أعمدة حقيقية؛ إن وُجد صف فيه كلمتا مدين ودائن نستخدمه كرأس.
    """
    if df is None or df.empty:
        return df
    dfc = df.copy()
    dfc.columns = dfc.columns.astype(str).str.strip()
    hdr = " ".join(str(c).lower() for c in dfc.columns)
    if "مدين" in hdr and "دائن" in hdr:
        return dfc
    ncols = len(dfc.columns)
    for idx in range(min(25, len(dfc))):
        cells_raw = dfc.iloc[idx].tolist()
        cell_texts: List[str] = []
        for v in cells_raw:
            if pd.isna(v):
                continue
            t = str(v).strip().lower()
            if t:
                cell_texts.append(t)
        if len(cell_texts) < 2:
            continue
        has_deb = any("مدين" in c and "دائن" not in c for c in cell_texts)
        has_cred = any("دائن" in c and "مدين" not in c for c in cell_texts)
        if not (has_deb and has_cred):
            continue
        new_cols: List[str] = []
        for j in range(ncols):
            v = cells_raw[j] if j < len(cells_raw) else None
            if v is None or (isinstance(v, float) and pd.isna(v)) or not str(v).strip():
                new_cols.append(f"_c{j}")
            else:
                new_cols.append(str(v).strip())
        dfc.columns = new_cols
        dfc = dfc.iloc[idx + 1 :].dropna(how="all").reset_index(drop=True)
        return dfc

    return dfc


def detect_columns(df: pd.DataFrame) -> Tuple[Optional[str], Optional[str], Optional[str]]:
    df.columns = df.columns.astype(str).str.strip()

    debit_col = None
    credit_col = None
    date_col = None

    for col in df.columns:
        name = str(col).lower()
        if date_col is None and any(x in name for x in ["تاريخ", "التاريخ", "التأريخ", "date"]):
            date_col = col
            continue
        if debit_col is None and _column_header_indicates_debit(col):
            debit_col = col
            continue
        if credit_col is None and _column_header_indicates_credit(col):
            credit_col = col

    if debit_col and not credit_col:
        for col in df.columns:
            if col == debit_col or _column_name_excludes_from_amount(col):
                continue
            if _column_header_indicates_credit(col):
                credit_col = col
                break
    if credit_col and not debit_col:
        for col in df.columns:
            if col == credit_col or _column_name_excludes_from_amount(col):
                continue
            if _column_header_indicates_debit(col):
                debit_col = col
                break

    named_debit = debit_col is not None
    named_credit = credit_col is not None

    numeric_cols: List[Tuple[str, float, float]] = []
    for col in df.columns:
        if _column_name_excludes_from_amount(col):
            continue
        if _is_balance_column_name(col):
            continue
        nums = pd.to_numeric(df[col], errors="coerce")
        valid = nums.dropna()
        frac = 0.3 if len(df) >= 12 else 0.15
        need = max(1, int(len(df) * frac + 0.5))
        if len(df) >= 4:
            need = max(2, need)
        if len(valid) < need:
            continue
        if _column_values_look_like_ids(df[col]):
            continue
        mean_val = valid.mean()
        max_val = float(valid.max())
        # Ignore tiny index columns; keep columns that have at least one plausible amount.
        if mean_val < 10 and max_val < 10:
            continue
        if not _is_plausible_currency_amount(max_val):
            continue
        numeric_cols.append((col, float(mean_val), _currency_amount_rank(df[col])))

    ranked_by_score = sorted(numeric_cols, key=lambda x: x[2], reverse=True)

    if not named_debit and not named_credit:
        ordered = sorted(
            numeric_cols,
            key=lambda t: list(df.columns).index(t[0]),
        )
        r_tail = max(ordered[-2][2], ordered[-1][2]) if len(ordered) >= 2 else (ordered[0][2] if ordered else 0.0)
        while len(ordered) > 2 and r_tail > 0 and ordered[0][2] < r_tail * 0.2:
            ordered.pop(0)
        if len(ordered) >= 2:
            debit_col, credit_col = ordered[-2][0], ordered[-1][0]
        elif len(ordered) == 1:
            debit_col = ordered[0][0]
    else:
        if not debit_col and len(ranked_by_score) >= 1:
            for col, _, _ in ranked_by_score:
                if col != credit_col:
                    debit_col = col
                    break
        if not credit_col and len(ranked_by_score) >= 1:
            for col, _, _ in ranked_by_score:
                if col != debit_col:
                    credit_col = col
                    break

    if not date_col:
        for col in df.columns:
            parsed = pd.to_datetime(df[col], errors="coerce")
            if parsed.notna().sum() > len(df) * 0.5:
                date_col = col
                break

    return debit_col, credit_col, date_col


def detect_document_type_column(df: pd.DataFrame) -> Optional[str]:
    """
    عمود نوع المستند فقط (مثل: فاتورة مبيعات آجل).
    لا نستخدم رقم السند، العنوان، الرقم الضريبي، أو عمود البيان الطويل.
    """
    for col in df.columns:
        n = str(col).strip().lower()
        if "رقم السند" in n or "رقم سند" in n or n in ("#", "م"):
            continue
        if n in ("العنوان", "عنوان", "العميل", "اسم العميل", "المورد"):
            continue
        if "ضريبي" in n or "سجل التجاري" in n:
            continue
        if "نوع المستند" in n:
            return col
    for col in df.columns:
        n = str(col).strip().lower()
        if "رقم السند" in n or "رقم سند" in n:
            continue
        if n == "المستند" or n.endswith(" المستند"):
            return col
    for col in df.columns:
        n = str(col).strip().lower()
        if "رقم السند" in n or "رقم سند" in n:
            continue
        if n == "نوع" or (n.startswith("نوع") and "عميل" not in n and "مورد" not in n and "حساب" not in n):
            return col
    for col in df.columns:
        n = str(col).strip().lower()
        if "رقم السند" in n or "رقم سند" in n:
            continue
        for key in ("نوع القيد", "نوع الحركة", "طبيعة القيد", "تصنيف الحركة"):
            if key in n:
                return col
    return None


def resolve_document_columns(df: pd.DataFrame) -> Tuple[Optional[str], Optional[str]]:
    """عمود نوع المستند، وإن لم يوجد نستخدم البيان كملخص نصي (ليس رقم السند)."""
    primary = detect_document_type_column(df)
    fallback: Optional[str] = None
    if primary is None:
        for col in df.columns:
            n = str(col).strip().lower()
            if "رقم السند" in n or "رقم سند" in n or n in ("#", "م"):
                continue
            if "بيان" in n and "ضريبي" not in n:
                fallback = col
                break
    return primary, fallback


def read_excel(file_path: str) -> Optional[pd.DataFrame]:
    df = pd.read_excel(file_path)
    if df is None or df.empty:
        return None
    return df.dropna(how="all")


def _normalize_arabic_digits(text: str) -> str:
    trans = str.maketrans("٠١٢٣٤٥٦٧٨٩", "0123456789")
    return text.translate(trans)


def _parse_number_token(token: str) -> Optional[float]:
    t = _normalize_arabic_digits(str(token))
    t = t.replace("٬", "").replace(",", "")
    t = t.replace("٫", ".")
    t = re.sub(r"[^\d\.\-+]", "", t)
    if not t:
        return None
    try:
        return float(t)
    except Exception:
        return None


def _debit_credit_from_tail_numbers(numbers: List[str]) -> Tuple[Optional[float], Optional[float]]:
    """
    Map trailing numeric tokens to مدين/دائن. Many ERP PDFs end with: ... debit, credit, رصيد.
    Some lines are: م | رقم_سند | ... | تاريخ | مدين | دائن | رصيد → use last 3 for deb/cred/bal.
    """
    if not numbers:
        return None, None
    vals: List[Optional[float]] = [_parse_number_token(n) for n in numbers]
    vals = [v for v in vals if v is not None]
    if not vals:
        return None, None
    while len(vals) >= 4 and vals[0] is not None:
        v0 = vals[0]
        if v0 == int(v0) and 1 <= abs(v0) <= 999:
            vals = vals[1:]
            continue
        if (
            v0 == int(v0)
            and 1_000 <= abs(v0) <= 9_999_999
            and vals[1] is not None
            and abs(vals[1]) > abs(v0) * 3
        ):
            vals = vals[1:]
            continue
        break
    if len(vals) == 1:
        return vals[0], None
    if len(vals) >= 5:
        return vals[-3], vals[-2]
    if len(vals) == 4:
        first = vals[0]
        if first is not None and first == int(first) and 1 <= abs(first) <= 999:
            return vals[-2], vals[-1]
        return vals[-3], vals[-2]
    if len(vals) == 3:
        third = vals[0]
        if third is not None and third == int(third) and abs(third) >= 10000:
            return vals[-2], vals[-1]
        return vals[-3], vals[-2]
    return vals[-2], vals[-1]


_LETTERHEAD_MARKERS = (
    "الرقم الضريبي",
    "السجل التجاري",
    "كشف حساب",
    "اسم العميل",
)


def _skip_statement_letterhead_lines(lines: List[str]) -> List[str]:
    """Drop company letterhead / meta block before جدول الحركة (كما في كشوف الموردين)."""
    start = 0
    for i, raw in enumerate(lines):
        s = raw.strip()
        if not s:
            continue
        low = s.lower()
        if "مدين" in s and "دائن" in s:
            start = i + 1
            break
        if "مدين" in s and "التاريخ" in s:
            start = i + 1
            break
        if "الرصيد" in s and "مدين" in s:
            start = i + 1
            break
        if re.search(r"\d{4}\s*[-/.]\s*\d{1,2}\s*[-/.]\s*\d{1,2}", _normalize_arabic_digits(s)):
            if any(m in s for m in _LETTERHEAD_MARKERS):
                continue
            start = i
            break
    return lines[start:] if start else lines


def _extract_pdf_rows_from_text(raw_text: str) -> List[Dict[str, Any]]:
    # Dates with / - . and optional spaces (common in bank/extract_text layouts).
    date_pat = re.compile(
        r"(\d{4}\s*[/\-\.]\s*\d{1,2}\s*[/\-\.]\s*\d{1,2}|\d{1,2}\s*[/\-\.]\s*\d{1,2}\s*[/\-\.]\s*\d{2,4})"
    )
    num_pat = re.compile(r"[-+]?\d[\d,٬٫\.]*")

    body_lines = _skip_statement_letterhead_lines((raw_text or "").splitlines())
    rows: List[Dict[str, Any]] = []
    last_date: Optional[str] = None
    for raw_line in body_lines:
        line = _normalize_arabic_digits(raw_line).strip()
        if len(line) < 2:
            continue
        if any(m in raw_line for m in _LETTERHEAD_MARKERS) and "مدين" not in raw_line:
            continue

        date_m = date_pat.search(line)
        if date_m:
            last_date = date_m.group(1).strip()

        # Strip the date token from the line so 2024/01/15 does not become extra "amounts".
        if date_m:
            work_line = (line[: date_m.start()] + " " + line[date_m.end() :]).strip()
        else:
            work_line = line

        effective_date = (date_m.group(1).strip() if date_m else None) or last_date
        if not effective_date:
            continue

        numbers = [
            n
            for n in num_pat.findall(work_line)
            if (pv := _parse_number_token(n)) is not None and _is_plausible_currency_amount(pv)
        ]
        if not numbers:
            continue

        debit_val, credit_val = _debit_credit_from_tail_numbers(numbers)

        if (debit_val is None or abs(debit_val) < 0.0001) and (credit_val is None or abs(credit_val) < 0.0001):
            continue

        doc_text = _normalize_doc_text(work_line)
        if numbers:
            fm = num_pat.search(work_line)
            if fm:
                doc_text = _normalize_doc_text(work_line[: fm.start()])
        if len(doc_text) > 160:
            doc_text = doc_text[:160] + "..."

        rows.append(
            {
                "التاريخ": effective_date,
                "مدين": debit_val if debit_val is not None else "",
                "دائن": credit_val if credit_val is not None else "",
                "بيان": raw_line.strip(),
                "مستند": doc_text or "",
            }
        )

    return rows


def _trim_pdf_table_df(df: pd.DataFrame) -> pd.DataFrame:
    """Remove header/letter rows merged into pdfplumber table (before first real movement row)."""
    if df is None or df.empty:
        return df
    dfc = df.copy()
    dfc.columns = dfc.columns.astype(str).str.strip()

    for idx in range(min(8, len(dfc))):
        parts = [str(x) for x in dfc.iloc[idx].tolist() if pd.notna(x) and str(x).strip()]
        row_txt = " ".join(parts)
        if "مدين" in row_txt and "دائن" in row_txt:
            return dfc.iloc[idx + 1 :].reset_index(drop=True)

    try:
        _, _, date_col = detect_columns(dfc.copy())
    except Exception:
        return dfc

    if date_col and date_col in dfc.columns:
        for idx, row in dfc.iterrows():
            val = row.get(date_col)
            if pd.isna(val):
                continue
            parsed = pd.to_datetime(val, errors="coerce", dayfirst=True)
            if pd.notna(parsed):
                return dfc.iloc[int(idx) :].reset_index(drop=True)
    return dfc


def _estimate_extractable_rows(df: Optional[pd.DataFrame]) -> int:
    """How many rows look like ledger lines (for picking table vs text PDF parse)."""
    if df is None or df.empty:
        return 0
    try:
        dfc = df.copy()
        dfc.columns = dfc.columns.astype(str).str.strip()
        dfc = _promote_ledger_header_row(dfc)
        dfc.columns = dfc.columns.astype(str).str.strip()
        debit_col, credit_col, _ = detect_columns(dfc)
    except Exception:
        return 0

    n = 0
    for _, row in dfc.iterrows():
        if row.isna().all():
            continue
        deb = safe(row[debit_col]) if debit_col and debit_col in dfc.columns else None
        cre = safe(row[credit_col]) if credit_col and credit_col in dfc.columns else None
        if deb is not None or cre is not None:
            n += 1
    return n


def read_pdf(file_path: str) -> Optional[pd.DataFrame]:
    table_df: Optional[pd.DataFrame] = None
    all_text = ""

    try:
        with pdfplumber.open(file_path) as pdf:
            grid_rows: List[List[Any]] = []
            text_parts: List[str] = []
            for page in pdf.pages:
                text_parts.append(page.extract_text() or "")
                tables = page.extract_tables()
                if not tables:
                    continue
                for table in tables:
                    for row in table:
                        if row and any(cell is not None for cell in row):
                            grid_rows.append(row)

            all_text = "\n".join(text_parts)

            if grid_rows:
                df = pd.DataFrame(grid_rows).dropna(how="all")
                if len(df) >= 2:
                    df.columns = df.iloc[0]
                    df = df[1:].dropna(how="all")
                    if not df.empty:
                        table_df = _trim_pdf_table_df(df.reset_index(drop=True))
    except Exception:
        return None

    parsed_rows = _extract_pdf_rows_from_text(all_text)
    text_df: Optional[pd.DataFrame] = None
    if parsed_rows:
        text_df = pd.DataFrame(parsed_rows).dropna(how="all")

    candidates: List[pd.DataFrame] = [d for d in (table_df, text_df) if d is not None and not d.empty]
    if not candidates:
        return None

    scored = [(d, _estimate_extractable_rows(d)) for d in candidates]
    best_score = max(sc for _, sc in scored)
    if best_score > 0:
        for d, sc in scored:
            if sc == best_score:
                return d
    return max(candidates, key=len)


def read_any(file_path: str, filename: str) -> pd.DataFrame:
    name = (filename or "").lower()
    if name.endswith(".xlsx") or name.endswith(".xls"):
        out = read_excel(file_path)
        if out is None:
            raise ValueError("Excel file has no readable data")
        return out
    if name.endswith(".csv"):
        try:
            out = pd.read_csv(file_path, encoding="utf-8-sig")
        except UnicodeDecodeError:
            out = pd.read_csv(file_path, encoding="cp1256")
        if out is None or out.empty:
            raise ValueError("ملف CSV فارغ أو غير مقروء")
        return out.dropna(how="all")
    if name.endswith(".pdf"):
        out = read_pdf(file_path)
        if out is None:
            raise ValueError(
                "لم يُستخرج من PDF جداول أو أسطر حركة واضحة. جرّب: ملف Excel، أو PDF نصّي (وليس صورة ممسوحة)، أو تأكد أن أعمدة المدين/الدائن ظاهرة في النص."
            )
        return out
    raise ValueError("نوع الملف غير مدعوم")


doc_map: Dict[str, str] = {
    "مردود مبيعات": "مردود مشتريات",
    "مردود مشتريات": "مردود مبيعات",
    "سند قبض": "سند صرف",
    "سند صرف": "سند قبض",
    "تحويل مخزني": "تحويل مخزني",
    "توريد مخزني": "صرف مخزني",
    "صرف مخزني": "توريد مخزني",
    "قيد يومية": "قيد يومية",
    "قيد افتتاحي": "قيد افتتاحي",
    "مبيعات": "مشتريات",
    "مشتريات": "مبيعات",
    "فاتورة مبيعات": "فاتورة مشتريات",
    "فاتورة مشتريات": "فاتورة مبيعات",
}


def clean(s: Any) -> str:
    if not s:
        return ""
    s = str(s).lower().strip()
    for w in ["رقم", "no", "doc", "ref"]:
        s = s.replace(w, "")
    s = re.sub(r"\d+", "", s)
    for ch in [" ", "-", "_", "/", "\\", ".", ","]:
        s = s.replace(ch, "")
    return s


def _is_voucher_number_string(s: Any) -> bool:
    """رقم سند / مرجع أرقام فقط — لا يُعتبر نوع مستند ولا يُستخدم في المطابقة."""
    if s is None or (isinstance(s, float) and pd.isna(s)):
        return False
    t = _normalize_arabic_digits(str(s).strip()).replace(",", "").replace("٬", "").replace(" ", "")
    if not t or "." in t or "٫" in str(s):
        return False
    return bool(re.fullmatch(r"\d{3,20}", t))


def _doc_for_matching(d: Any) -> Optional[str]:
    if d is None or (isinstance(d, str) and not str(d).strip()):
        return None
    if _is_voucher_number_string(d):
        return None
    return str(d).strip()


def match_doc(d1: Any, d2: Any) -> bool:
    if not d1 and not d2:
        return True
    if not d1 or not d2:
        return False
    if _is_voucher_number_string(d1) or _is_voucher_number_string(d2):
        return False

    d1 = clean(d1)
    d2 = clean(d2)

    if not d1 and not d2:
        return True
    if not d1 or not d2:
        return False
    if d1 == d2:
        return True
    if len(d1) > 3 and len(d2) > 3:
        if d1 in d2 or d2 in d1:
            return True

    for key, val in doc_map.items():
        k = clean(key)
        v = clean(val)
        if (k in d1 and v in d2) or (v in d1 and k in d2):
            return True

    return SequenceMatcher(None, d1, d2).ratio() > 0.7


def date_diff_days(d1: Any, d2: Any) -> Optional[int]:
    try:
        dd1 = pd.to_datetime(d1, errors="coerce")
        dd2 = pd.to_datetime(d2, errors="coerce")
        if pd.isna(dd1) or pd.isna(dd2):
            return None
        return abs((dd1 - dd2).days)
    except Exception:
        return None


def extract_row_date_doc(
    row: pd.Series,
    df: pd.DataFrame,
    date_col: Optional[str],
    doc_col: Optional[str],
    doc_fallback_col: Optional[str] = None,
) -> Tuple[Optional[str], Optional[str]]:
    date_out: Optional[str] = None
    if date_col and date_col in df.columns:
        try:
            val = row[date_col]
            if pd.isna(val):
                date_out = None
            else:
                d = pd.to_datetime(val, errors="coerce", dayfirst=False)
                date_out = None if pd.isna(d) else d.strftime("%Y-%m-%d")
        except Exception:
            date_out = str(row[date_col])

    doc_out: Optional[str] = None
    for col in (doc_col, doc_fallback_col):
        if not col or col not in df.columns:
            continue
        val = row[col]
        if pd.isna(val):
            continue
        raw = _normalize_doc_text(val)
        if not raw:
            continue
        if _is_voucher_number_string(raw):
            continue
        doc_out = raw if len(raw) <= 200 else (raw[:200] + "...")
        break

    return date_out, doc_out


def _skip_likely_pdf_line_index_row(
    amount: float, debit: Optional[float], credit: Optional[float], both_positive: bool
) -> bool:
    """لا نستبعد المبالغ الصغيرة تلقائياً (قد تكون فواتير فعلية)."""
    return False


def process(file_path: str, filename: str, branch: str) -> List[Dict[str, Any]]:
    df = read_any(file_path, filename)
    if df is None or len(df) == 0:
        return []

    df.columns = df.columns.astype(str).str.strip()
    df = _promote_ledger_header_row(df)
    df.columns = df.columns.astype(str).str.strip()
    debit_col, credit_col, date_col = detect_columns(df)
    doc_col, doc_fb = resolve_document_columns(df)

    data: List[Dict[str, Any]] = []
    for _, row in df.iterrows():
        if row.isna().all():
            continue

        narrative = _row_narrative_for_amounts(row, df, doc_col, doc_fb, debit_col, credit_col, date_col)

        debit = safe(row[debit_col]) if debit_col and debit_col in df.columns else None
        credit = safe(row[credit_col]) if credit_col and credit_col in df.columns else None

        decs = [
            v
            for v in _parse_currency_numbers_from_narrative(narrative)
            if abs(v - int(v)) > 0.0001 and abs(v) >= 0.0001
        ]
        if debit is not None and credit is not None and debit > 0 and credit > 0:
            if (
                _looks_like_serial_voucher_amount(debit)
                and _looks_like_serial_voucher_amount(credit)
                and len(decs) >= 2
            ):
                debit, credit = round(decs[0], 2), round(decs[1], 2)
            else:
                debit = _replace_voucher_with_ledger_from_narrative(debit, narrative)
                credit = _replace_voucher_with_ledger_from_narrative(credit, narrative)
        else:
            if debit is not None:
                debit = _replace_voucher_with_ledger_from_narrative(debit, narrative)
            if credit is not None:
                credit = _replace_voucher_with_ledger_from_narrative(credit, narrative)

        if debit and credit and debit > 0 and credit > 0:
            side_hint = _side_hint_from_narrative(narrative)
            if side_hint == "debit":
                credit = None
            elif side_hint == "credit":
                debit = None

        if debit is None and credit is None:
            continue

        if debit and credit and debit > 0 and credit > 0:
            date_out, doc_out = extract_row_date_doc(row, df, date_col, doc_col, doc_fb)
            doc_out = _enrich_doc_field(doc_out, narrative)
            amount = max(debit, credit)
            t = "credit" if credit >= debit else "debit"
            data.append(
                {"amount": float(amount), "type": t, "branch": branch, "date": date_out, "doc": doc_out}
            )
            continue

        if credit and credit > 0:
            amount = credit
            t = "credit"
        elif debit and debit > 0:
            amount = debit
            t = "debit"
        else:
            continue

        if _skip_likely_pdf_line_index_row(amount, debit, credit, False):
            continue

        date_out, doc_out = extract_row_date_doc(row, df, date_col, doc_col, doc_fb)
        doc_out = _enrich_doc_field(doc_out, narrative)
        data.append({"amount": float(amount), "type": t, "branch": branch, "date": date_out, "doc": doc_out})

    return _dedupe_extracted_rows(data)


def analyze(
    d1: List[Dict[str, Any]],
    d2: List[Dict[str, Any]],
    *,
    allow_same_direction: bool = True,
) -> Tuple[List[Dict[str, Any]], Dict[str, int]]:
    res: List[Dict[str, Any]] = []
    used = [False] * len(d2)
    counts: Dict[str, int] = {}

    def is_sale_return_pair(doc1: Any, doc2: Any) -> bool:
        """
        Treat (sales, sales return) and (purchases, purchases return) as valid internal neutralizing pairs.
        """
        c1 = clean(doc1)
        c2 = clean(doc2)
        if not c1 or not c2:
            return False

        pairs = [
            (clean("مبيعات"), clean("مردود مبيعات")),
            (clean("مشتريات"), clean("مردود مشتريات")),
        ]
        for base, ret in pairs:
            if (base in c1 and ret in c2) or (ret in c1 and base in c2):
                return True
        return False

    def remove_reversals(data: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
        cleaned: List[Dict[str, Any]] = []
        used_local = [False] * len(data)
        for i, x1 in enumerate(data):
            if used_local[i]:
                continue
            found = False
            for j, x2 in enumerate(data):
                if i == j or used_local[j]:
                    continue
                if x1["branch"] != x2["branch"]:
                    continue
                if abs(x1["amount"] - x2["amount"]) > 0.01:
                    continue
                days = date_diff_days(x1["date"], x2["date"])
                if days is None:
                    continue

                dm1, dm2 = _doc_for_matching(x1.get("doc")), _doc_for_matching(x2.get("doc"))
                # Rule 1: classic reversal (opposite direction, same amount, close date).
                classic_reversal = (
                    x1["type"] != x2["type"]
                    and days <= 1
                    and (not (dm1 and dm2) or match_doc(dm1, dm2))
                )

                # Rule 2: same-branch sales/purchase with their return in the SAME day.
                sales_return_same_day = (
                    days == 0 and is_sale_return_pair(x1.get("doc"), x2.get("doc"))
                )

                if not classic_reversal and not sales_return_same_day:
                    continue

                used_local[i] = True
                used_local[j] = True
                found = True
                break
            if not found:
                cleaned.append(x1)
        return cleaned

    d1 = remove_reversals(d1)
    d2 = remove_reversals(d2)

    if not d2:
        for x in d1:
            res.append({**x, "reason": "لا يوجد مقابل ❌ (الفرع الثاني فارغ)"})
            b = x.get("branch") or "unknown"
            counts[b] = counts.get(b, 0) + 1
        return res, counts

    def match_score(x1: Dict[str, Any], x2: Dict[str, Any]) -> Tuple[int, List[str]]:
        score = 0
        reasons: List[str] = []

        diff = abs(x1["amount"] - x2["amount"])
        if diff < 0.01:
            score += 50
            reasons.append("نفس المبلغ")
        elif diff < 1:
            score += 30
            reasons.append("مبلغ قريب")
        else:
            reasons.append("فرق مبلغ (غير مانع)")

        if (x1["type"] == "credit" and x2["type"] == "debit") or (x1["type"] == "debit" and x2["type"] == "credit"):
            score += 30
            reasons.append("اتجاه عكسي صحيح")
        elif allow_same_direction:
            score += 15
            reasons.append("نفس الاتجاه بين الملفين")
        else:
            return 0, ["نفس الاتجاه"]

        d1m = _doc_for_matching(x1.get("doc"))
        d2m = _doc_for_matching(x2.get("doc"))
        if d1m and d2m:
            if not match_doc(d1m, d2m):
                return 0, ["اختلاف نوع المستند"]
            score += 20
            reasons.append("نوع مستند مطابق")

        both_no_doc = not d1m and not d2m
        days = date_diff_days(x1["date"], x2["date"])
        if days is None:
            score -= 5
            reasons.append("تاريخ غير واضح")
        elif days == 0:
            score += 20
            reasons.append("نفس اليوم")
        elif days <= 7:
            score += 10
            reasons.append("تاريخ قريب")
        elif days <= 45:
            reasons.append("فارق تاريخ ضمن المدى")
        elif both_no_doc:
            reasons.append("فارق تاريخ (بدون بيان مستند)")
        else:
            score -= 5
            reasons.append("تاريخ بعيد جدا مع اختلاف بيان المستند")

        return score, reasons

    for x1 in d1:
        if x1.get("type") == "error":
            res.append(x1)
            b = x1.get("branch") or "unknown"
            counts[b] = counts.get(b, 0) + 1
            continue

        best_i = -1
        best_score = -1
        best_reason: List[str] = []

        for i, x2 in enumerate(d2):
            if used[i]:
                continue
            if x2.get("type") == "error":
                continue
            score, reasons = match_score(x1, x2)
            if score > best_score:
                best_score = score
                best_i = i
                best_reason = reasons

        if best_score >= 60 and best_i != -1:
            used[best_i] = True
        elif best_score >= 45 and best_i != -1:
            res.append({**x1, "reason": f"تطابق ضعيف ⚠️ | score={best_score} | {' , '.join(best_reason)}"})
            used[best_i] = True
        else:
            res.append({**x1, "reason": f"لا يوجد مقابل ❌ | score={best_score} | {' , '.join(best_reason)}"})
            b = x1.get("branch") or "unknown"
            counts[b] = counts.get(b, 0) + 1

    for i, x in enumerate(d2):
        if not used[i]:
            if x.get("type") == "error":
                res.append(x)
                b = x.get("branch") or "unknown"
                counts[b] = counts.get(b, 0) + 1
                continue
            res.append({**x, "reason": "لا يوجد مقابل ❌ (من الفرع الآخر)"})
            b = x.get("branch") or "unknown"
            counts[b] = counts.get(b, 0) + 1

    return res, counts


def compute_summary(d1: List[Dict[str, Any]], d2: List[Dict[str, Any]], mismatches: List[Dict[str, Any]]) -> Dict[str, int]:
    total_ops = len(d1) + len(d2)
    mismatch_ops = len(mismatches)
    matched_ops = max(0, total_ops - mismatch_ops)
    errors = 0
    warnings = 0
    for e in mismatches:
        reason = e.get("reason") or ""
        if e.get("type") == "error" or "❌" in reason or "لا يوجد مقابل" in reason:
            errors += 1
        elif "⚠️" in reason:
            warnings += 1
    return {
        "total_ops": total_ops,
        "matched_ops": matched_ops,
        "mismatch_ops": mismatch_ops,
        "errors_count": errors,
        "warnings_count": warnings,
    }


def mismatches_to_csv_bytes(entries: List[Dict[str, Any]]) -> bytes:
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(["الفرع", "المبلغ", "نوع العملية", "التاريخ", "المستند", "السبب"])
    for e in entries:
        writer.writerow(
            [
                e.get("branch", ""),
                e.get("amount", ""),
                e.get("type", ""),
                e.get("date", "") or "",
                e.get("doc", "") or "",
                e.get("reason", "") or "",
            ]
        )
    return output.getvalue().encode("utf-8-sig")


def mismatches_to_excel_bytes(entries: List[Dict[str, Any]]) -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

    # Ordered to match requested visual layout (RTL) with branch restored
    headers = ["الفرع", "المبلغ", "التاريخ", "المستند", "السبب"]
    rows = []
    for e in entries:
        rows.append(
            [
                e.get("branch", "") or "",
                e.get("amount", ""),
                e.get("date", "") or "",
                e.get("doc", "") or "",
                e.get("reason", "") or "",
            ]
        )

    wb = Workbook()
    ws = wb.active
    ws.title = "errors"
    ws.sheet_view.rightToLeft = True

    header_fill = PatternFill(fill_type="solid", start_color="FFF200", end_color="FFF200")
    header_font = Font(color="FF0000", bold=True, size=12)
    body_font = Font(color="000000", size=11)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin = Side(style="thin", color="000000")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Header
    for col, h in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center
        cell.border = border

    # Body
    for r_idx, r in enumerate(rows, start=2):
        for c_idx, val in enumerate(r, start=1):
            cell = ws.cell(row=r_idx, column=c_idx, value=val)
            cell.font = body_font
            cell.alignment = center
            cell.border = border

    # Widths close to screenshot proportions
    ws.column_dimensions["A"].width = 20  # الفرع
    ws.column_dimensions["B"].width = 14  # المبلغ
    ws.column_dimensions["C"].width = 16  # التاريخ
    ws.column_dimensions["D"].width = 20  # المستند
    ws.column_dimensions["E"].width = 48  # السبب

    ws.freeze_panes = "A2"

    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()


def mismatches_to_pdf_bytes(entries: List[Dict[str, Any]]) -> bytes:
    if canvas is None or A4 is None:
        raise HTTPException(500, "PDF export يحتاج تثبيت reportlab: pip install reportlab")

    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle

    out = io.BytesIO()
    doc = SimpleDocTemplate(
        out,
        pagesize=A4,
        rightMargin=24,
        leftMargin=24,
        topMargin=24,
        bottomMargin=24,
    )

    # Keep same order/shape as Excel export
    data = [["الفرع", "المبلغ", "التاريخ", "المستند", "السبب"]]
    for e in entries:
        data.append(
            [
                str(e.get("branch", "") or ""),
                str(e.get("amount", "") or ""),
                str(e.get("date", "") or ""),
                str(e.get("doc", "") or ""),
                str(e.get("reason", "") or ""),
            ]
        )

    table = Table(data, colWidths=[90, 60, 80, 90, 210], repeatRows=1)
    table.setStyle(
        TableStyle(
            [
                ("GRID", (0, 0), (-1, -1), 0.8, colors.black),
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#FFF200")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.red),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("FONTSIZE", (0, 0), (-1, 0), 11),
                ("ALIGN", (0, 0), (-1, -1), "CENTER"),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                ("FONTSIZE", (0, 1), (-1, -1), 9),
            ]
        )
    )

    doc.build([table])
    return out.getvalue()


# =========================
# FRONTEND (HTML + JS)
# =========================
APP_JS = r"""function syncCsrfFromCookie() {
  const key = "auditflow_csrf=";
  const i = document.cookie.indexOf(key);
  if (i === -1) return;
  let v = document.cookie.slice(i + key.length).split(";")[0] || "";
  try {
    v = decodeURIComponent(v);
  } catch (e) {}
  if (v) try { localStorage.setItem("csrf_token", v); } catch (e) {}
}
syncCsrfFromCookie();

async function readErrorMessage(res) {
  const raw = await res.text().catch(() => "");
  const ct = (res.headers.get("content-type") || "").toLowerCase();
  if (ct.includes("application/json") && raw) {
    try {
      const j = JSON.parse(raw);
      const d = j.detail;
      if (typeof d === "string") return d;
      if (Array.isArray(d))
        return d
          .map((x) => (x && typeof x === "object" && x.msg ? String(x.msg) : JSON.stringify(x)))
          .join("; ");
    } catch (e) {}
  }
  return raw || `HTTP ${res.status}`;
}

function qs(name) {
  return new URLSearchParams(window.location.search).get(name);
}

async function apiGet(url) {
  const res = await fetch(url, { headers: { Accept: "application/json" }, credentials: "include" });
  if (!res.ok) throw new Error(await readErrorMessage(res));
  return res.json();
}

async function apiPostForm(url, formData) {
  syncCsrfFromCookie();
  const csrf = localStorage.getItem("csrf_token") || "";
  const res = await fetch(url, {
    method: "POST",
    body: formData,
    headers: { Accept: "application/json", "X-CSRF-Token": csrf },
    credentials: "include",
  });
  if (!res.ok) throw new Error(await readErrorMessage(res));
  return res.json();
}

async function apiPostJson(url, body) {
  syncCsrfFromCookie();
  const csrf = localStorage.getItem("csrf_token") || "";
  const res = await fetch(url, {
    method: "POST",
    credentials: "include",
    headers: { Accept: "application/json", "Content-Type": "application/json", "X-CSRF-Token": csrf },
    body: JSON.stringify(body || {}),
  });
  if (!res.ok) throw new Error(await readErrorMessage(res));
  return res.json();
}

async function apiDelete(url) {
  syncCsrfFromCookie();
  const csrf = localStorage.getItem("csrf_token") || "";
  const res = await fetch(url, { method: "DELETE", headers: { Accept: "application/json", "X-CSRF-Token": csrf }, credentials: "include" });
  if (!res.ok) throw new Error(await readErrorMessage(res));
  return res.json();
}

function showToast(msg, color = "#10b981") {
  const t = document.getElementById("toast");
  if (!t) return;
  t.innerText = msg;
  t.style.background = color;
  t.classList.remove("hidden");
  setTimeout(() => t.classList.add("hidden"), 3000);
}

function setLoading(btn, loading, text) {
  if (!btn) return;
  if (loading) {
    btn.disabled = true;
    btn.dataset.oldText = btn.innerText;
    btn.innerHTML = text || "جارٍ التحليل ...";
  } else {
    btn.disabled = false;
    btn.innerText = btn.dataset.oldText || (text || "ابدأ التحليل");
  }
}

function renderReportRow(item) {
  const li = document.createElement("div");
  li.className = "bg-white rounded-xl border border-slate-200 p-4 flex flex-col gap-2";
  li.innerHTML = `
    <div class="flex items-start justify-between gap-4">
      <div class="min-w-0">
        <div class="font-extrabold text-slate-900 truncate">
          ${item.title ? item.title : "تقرير بدون عنوان"}
        </div>
        <div class="text-sm text-slate-600 mt-1">
          ${item.branch1_name} مقابل ${item.branch2_name}
        </div>
      </div>
      <a class="px-3 py-1.5 rounded-lg bg-slate-900 text-white text-sm font-extrabold" href="/report?id=${item.id}">عرض</a>
    </div>
    <div class="flex gap-3 flex-wrap">
      <div class="text-sm text-slate-700"><span class="font-extrabold">متطابق:</span> ${item.stats.matched_ops}</div>
      <div class="text-sm text-slate-700"><span class="font-extrabold">أخطاء:</span> ${item.stats.errors_count}</div>
      <div class="text-sm text-slate-700"><span class="font-extrabold">تحذيرات:</span> ${item.stats.warnings_count}</div>
    </div>
    <button class="self-end px-3 py-1.5 rounded-lg border border-rose-200 text-rose-600 text-sm font-extrabold hover:bg-rose-50" onclick="deleteReport('${item.id}')">حذف</button>
  `;
  return li;
}

async function deleteReport(id) {
  if (!confirm("هل تريد حذف هذا التقرير؟")) return;
  await apiDelete(`/reports?id=${encodeURIComponent(id)}`);
  showToast("تم الحذف ✔️", "#10b981");
  await loadReports();
}

async function loadReports() {
  const host = document.getElementById("reportsHost");
  if (!host) return;
  host.innerHTML = `
    <div class="text-slate-600 text-center py-10">جارٍ تحميل التقارير ...</div>
  `;
  const data = await apiGet("/reports");
  const items = data.items || [];
  host.innerHTML = "";
  if (!items.length) {
    host.innerHTML = `<div class="text-slate-600 text-center py-10">لا توجد تقارير بعد.</div>`;
    return;
  }
  for (const item of items) {
    host.appendChild(renderReportRow(item));
  }
}

function renderMismatchTable(entries, host) {
  const rows = entries
    .map((e) => {
      const reason = e.reason || "";
      const severity = e.type === "error" || reason.includes("❌") ? "error" : reason.includes("⚠️") ? "warning" : "mismatch";
      const sevColor = severity === "error" ? "bg-rose-50 text-rose-700 border-rose-200" : severity === "warning" ? "bg-amber-50 text-amber-700 border-amber-200" : "bg-slate-50 text-slate-700 border-slate-200";
      const sevText = severity === "error" ? "خطأ" : severity === "warning" ? "تحذير" : "مخالفة";
      return `
        <tr class="border-b border-slate-200">
          <td class="px-3 py-3 text-sm text-slate-800">${e.branch || "-"}</td>
          <td class="px-3 py-3 text-sm text-slate-800">${e.amount ?? "-"}</td>
          <td class="px-3 py-3 text-sm text-slate-800">${e.type || "-"}</td>
          <td class="px-3 py-3 text-sm text-slate-800">${e.date || "-"}</td>
          <td class="px-3 py-3 text-sm text-slate-800">${e.doc || "-"}</td>
          <td class="px-3 py-3 text-sm">
            <span class="inline-flex items-center px-2 py-1 rounded-full border ${sevColor} text-xs font-extrabold">${sevText}</span>
          </td>
          <td class="px-3 py-3 text-sm text-slate-700">${reason || "-"}</td>
        </tr>
      `;
    })
    .join("");

  host.innerHTML = `
    <table class="w-full text-right table-fixed">
      <thead class="bg-slate-50">
        <tr>
          <th class="px-3 py-2 text-xs text-slate-600 font-extrabold w-[120px]">الفرع</th>
          <th class="px-3 py-2 text-xs text-slate-600 font-extrabold w-[110px]">المبلغ</th>
          <th class="px-3 py-2 text-xs text-slate-600 font-extrabold w-[90px]">نوع</th>
          <th class="px-3 py-2 text-xs text-slate-600 font-extrabold w-[110px]">التاريخ</th>
          <th class="px-3 py-2 text-xs text-slate-600 font-extrabold">المستند</th>
          <th class="px-3 py-2 text-xs text-slate-600 font-extrabold w-[110px]">الحالة</th>
          <th class="px-3 py-2 text-xs text-slate-600 font-extrabold">السبب</th>
        </tr>
      </thead>
      <tbody>
        ${rows || `<tr><td colspan="7" class="px-3 py-6 text-center text-slate-600">لا توجد بيانات</td></tr>`}
      </tbody>
    </table>
  `;
}

function applyTableFilters(entries) {
  const host = document.getElementById("mismatchTableHost");
  if (!host) return;

  const fDoc = (document.getElementById("filterDoc")?.value || "").toLowerCase().trim();
  const fAmount = (document.getElementById("filterAmount")?.value || "").trim();
  const fType = (document.getElementById("filterType")?.value || "").trim();

  let filtered = entries;
  if (fDoc) filtered = filtered.filter((x) => (x.doc || "").toLowerCase().includes(fDoc));
  if (fAmount) filtered = filtered.filter((x) => String(x.amount ?? "") === fAmount);
  if (fType) {
    filtered = filtered.filter((x) => (x.reason || "").includes(fType));
  }
  renderMismatchTable(filtered, host);
}

async function loadReportDetail() {
  const reportId = qs("id");
  if (!reportId) {
    showToast("معرّف التقرير غير موجود", "#ef4444");
    return;
  }
  const data = await apiGet(`/report?id=${encodeURIComponent(reportId)}`);

  document.getElementById("reportTitle").innerText = data.title || "تقرير بدون عنوان";
  document.getElementById("reportBranches").innerText = `${data.branch1_name} مقابل ${data.branch2_name}`;

  const stats = data.stats;
  document.getElementById("statTotal").innerText = String(stats.total_ops);
  document.getElementById("statMatched").innerText = String(stats.matched_ops);
  document.getElementById("statErrors").innerText = String(stats.errors_count);
  document.getElementById("statWarnings").innerText = String(stats.warnings_count);

  const analysis = data.analysis_json || {};
  const mismatches = analysis.mismatches || [];
  const counts = analysis.counts || {};
  const statsJson = data.stats_json || {};
  const b1Total = Number(statsJson.branch1_total || 0);
  const b2Total = Number(statsJson.branch2_total || 0);
  const b1Err = Number(counts[data.branch1_name] || 0);
  const b2Err = Number(counts[data.branch2_name] || 0);
  const b1Rate = b1Total > 0 ? ((b1Err / b1Total) * 100).toFixed(1) : "0.0";
  const b2Rate = b2Total > 0 ? ((b2Err / b2Total) * 100).toFixed(1) : "0.0";

  const b1ErrEl = document.getElementById("branch1Errors");
  const b2ErrEl = document.getElementById("branch2Errors");
  const b1RateEl = document.getElementById("branch1Rate");
  const b2RateEl = document.getElementById("branch2Rate");
  const b1Label = document.getElementById("branch1Label");
  const b2Label = document.getElementById("branch2Label");
  if (b1ErrEl) b1ErrEl.innerText = String(b1Err);
  if (b2ErrEl) b2ErrEl.innerText = String(b2Err);
  if (b1RateEl) b1RateEl.innerText = `${b1Rate}%`;
  if (b2RateEl) b2RateEl.innerText = `${b2Rate}%`;
  if (b1Label) b1Label.innerText = data.branch1_name;
  if (b2Label) b2Label.innerText = data.branch2_name;

  window.__MISMATCHES__ = mismatches;
  renderMismatchTable(mismatches, document.getElementById("mismatchTableHost"));
}

function downloadErrors(id, format) {
  const fmt = format || "excel";
  window.location.href = `/download?id=${encodeURIComponent(id)}&format=${encodeURIComponent(fmt)}`;
}

async function startAnalyze() {
  const btn = document.getElementById("startBtn");
  setLoading(btn, true, "جارٍ التحليل ...");
  try {
    const file1 = document.getElementById("file1").files?.[0] || null;
    const file2 = document.getElementById("file2").files?.[0] || null;
    const b1 = (document.getElementById("b1").value || "").trim();
    const b2 = (document.getElementById("b2").value || "").trim();
    const title = document.getElementById("title").value || null;

    if (!file1 || !file2) {
      showToast("اختَر الملفين أولاً", "#ef4444");
      return;
    }
    if (!b1 || !b2) {
      showToast("اكتب اسم الفرع الأول والثاني", "#ef4444");
      return;
    }

    const fd = new FormData();
    fd.append("file1", file1);
    fd.append("file2", file2);
    fd.append("b1", b1);
    fd.append("b2", b2);
    if (title) fd.append("title", title);

    const data = await apiPostForm("/analyze", fd);
    const id = data.reportId;
    showToast("تم التحليل ✔️", "#10b981");
    window.location.href = `/report?id=${encodeURIComponent(id)}`;
  } catch (e) {
    showToast(e.message || "فشل التحليل", "#ef4444");
  } finally {
    setLoading(btn, false, "ابدأ التحليل");
  }
}

function initAnalyzePage() {
  document.getElementById("startBtn")?.addEventListener("click", () => startAnalyze());
}

async function initAuthUI() {
  const host = document.getElementById("authArea");
  if (!host) return;

  function ensureAuthModal() {
    let modal = document.getElementById("authModal");
    if (modal) return modal;
    modal = document.createElement("div");
    modal.id = "authModal";
    modal.className = "hidden fixed inset-0 z-[70] bg-black/50 items-center justify-center p-4";
    modal.innerHTML = `
      <div class="w-full max-w-md rounded-2xl bg-white border border-slate-200 shadow-2xl p-5">
        <div class="flex items-center justify-between mb-4">
          <h3 id="authModalTitle" class="text-xl font-extrabold text-slate-900"></h3>
          <button id="authCloseBtn" class="px-2 py-1 rounded-lg border border-slate-200 text-slate-600 hover:bg-slate-50">✕</button>
        </div>
        <div class="space-y-3">
          <div>
            <label class="block text-sm font-extrabold text-slate-700 mb-1">اسم المستخدم</label>
            <input id="authUsername" class="w-full rounded-xl border border-slate-200 bg-white px-3 py-2 outline-none focus:ring-2 focus:ring-slate-900/10" />
          </div>
          <div>
            <label class="block text-sm font-extrabold text-slate-700 mb-1">كلمة المرور</label>
            <input id="authPassword" type="password" class="w-full rounded-xl border border-slate-200 bg-white px-3 py-2 outline-none focus:ring-2 focus:ring-slate-900/10" />
          </div>
        </div>
        <div class="mt-4 flex items-center justify-end gap-2">
          <button id="authCancelBtn" class="px-4 py-2 rounded-xl border border-slate-200 text-sm font-extrabold hover:bg-slate-50">إلغاء</button>
          <button id="authSubmitBtn" class="px-4 py-2 rounded-xl bg-slate-900 text-white text-sm font-extrabold hover:bg-slate-800"></button>
        </div>
      </div>
    `;
    document.body.appendChild(modal);
    return modal;
  }

  function openAuthModal(mode) {
    const modal = ensureAuthModal();
    const title = document.getElementById("authModalTitle");
    const submit = document.getElementById("authSubmitBtn");
    const cancel = document.getElementById("authCancelBtn");
    const close = document.getElementById("authCloseBtn");
    const u = document.getElementById("authUsername");
    const p = document.getElementById("authPassword");
    if (!title || !submit || !cancel || !close || !u || !p) return;

    title.innerText = mode === "register" ? "إنشاء حساب جديد" : "تسجيل الدخول";
    submit.innerText = mode === "register" ? "تسجيل" : "دخول";
    u.value = "";
    p.value = "";
    modal.classList.remove("hidden");
    modal.classList.add("flex");
    u.focus();

    const closeModal = () => {
      modal.classList.add("hidden");
      modal.classList.remove("flex");
    };

    close.onclick = closeModal;
    cancel.onclick = closeModal;
    modal.onclick = (e) => {
      if (e.target === modal) closeModal();
    };
    submit.onclick = async () => {
      const username = (u.value || "").trim();
      const password = (p.value || "").trim();
      if (!username || !password) {
        showToast("أدخل اسم المستخدم وكلمة المرور", "#ef4444");
        return;
      }
      try {
        if (mode === "register") {
          await apiPostJson("/auth/register", { username, password });
          showToast("تم إنشاء الحساب وتسجيل الدخول ✔️");
        } else {
          await apiPostJson("/auth/login", { username, password });
          showToast("تم تسجيل الدخول ✔️");
        }
        closeModal();
        window.location.reload();
      } catch (e) {
        showToast(e.message || "فشل العملية", "#ef4444");
      }
    };
  }

  async function render() {
    try {
      const me = await apiGet("/auth/me");
      if (me?.csrf_token) localStorage.setItem("csrf_token", me.csrf_token);
      const username = me?.username || "";
      if (username) {
        host.innerHTML = `
          <div class="flex items-center gap-2">
            <span class="text-sm font-extrabold text-slate-700">مرحباً ${username}</span>
            <button id="logoutBtn" class="px-3 py-1.5 rounded-xl border border-slate-200 text-sm font-extrabold hover:bg-slate-50">خروج</button>
          </div>
        `;
        document.getElementById("logoutBtn")?.addEventListener("click", async () => {
          await apiPostJson("/auth/logout", {});
          showToast("تم تسجيل الخروج");
          window.location.reload();
        });
        return;
      }
    } catch (_) {
      // not logged in
    }

    host.innerHTML = `
      <div class="flex items-center gap-2">
        <button id="loginBtn" class="px-3 py-1.5 rounded-xl border border-slate-200 text-sm font-extrabold hover:bg-slate-50">تسجيل دخول</button>
        <button id="registerBtn" class="px-3 py-1.5 rounded-xl bg-slate-900 text-white text-sm font-extrabold hover:bg-slate-800">تسجيل</button>
      </div>
    `;

    document.getElementById("registerBtn")?.addEventListener("click", () => {
      openAuthModal("register");
    });

    document.getElementById("loginBtn")?.addEventListener("click", () => {
      openAuthModal("login");
    });
  }

  await render();
}

window.deleteReport = deleteReport;
window.initAuthUI = initAuthUI;
"""

INDEX_HTML = r"""<!doctype html>
<html lang="ar" dir="rtl">
  <head>
    <meta charset="UTF-8" />
    <meta name="viewport" content="width=device-width, initial-scale=1.0" />
    <title>AuditFlow</title>
    <link
      href="https://fonts.googleapis.com/css2?family=IBM+Plex+Sans+Arabic:wght@300;400;600;700;800&display=swap"
      rel="stylesheet"
    />
    <link rel="stylesheet" href="/static/tailwind.css" />
    <style>
      body { font-family: "IBM Plex Sans Arabic", system-ui, -apple-system, "Segoe UI", Roboto, Arial, sans-serif; }
    </style>
  </head>
  <body class="bg-slate-50 text-slate-900">
    <div id="toast" class="hidden fixed bottom-5 left-5 z-50 text-white px-4 py-2 rounded-xl font-extrabold"></div>

    <header class="sticky top-0 bg-white/90 backdrop-blur border-b border-slate-200 z-40">
      <div class="max-w-6xl mx-auto px-4 py-4 flex items-center justify-between gap-3">
        <div class="font-extrabold text-slate-900 text-lg">AuditFlow | نظام التدقيق</div>
        <nav class="flex gap-3">
          <a class="px-3 py-2 rounded-xl font-extrabold text-sm bg-slate-900 text-white" href="/">لوحة التحكم</a>
          <a class="px-3 py-2 rounded-xl font-extrabold text-sm bg-white border border-slate-200 hover:bg-slate-50" href="/analyze">تحليل</a>
          <a class="px-3 py-2 rounded-xl font-extrabold text-sm bg-white border border-slate-200 hover:bg-slate-50" href="/reports">التقارير</a>
        </nav>
        <div id="authArea"></div>
      </div>
    </header>

    <main class="max-w-6xl mx-auto px-4 py-8">
      <h1 class="text-3xl font-extrabold text-center">نظام المطابقة المالية</h1>
      <p class="text-center text-slate-600 mt-2">ارفع ملفي Excel / PDF وقارن العمليات تلقائياً.</p>

      <div class="grid md:grid-cols-3 gap-4 mt-8">
        <div class="bg-white border border-slate-200 rounded-2xl p-5 shadow-sm">
          <div class="text-slate-500 font-extrabold text-sm">آخر التقارير</div>
          <div id="dashTotalReports" class="text-3xl font-extrabold mt-2">0</div>
        </div>
        <div class="bg-white border border-slate-200 rounded-2xl p-5 shadow-sm">
          <div class="text-slate-500 font-extrabold text-sm">إجمالي الأخطاء (آخر تقرير)</div>
          <div id="dashErrors" class="text-3xl font-extrabold mt-2">0</div>
        </div>
        <div class="bg-white border border-slate-200 rounded-2xl p-5 shadow-sm">
          <div class="text-slate-500 font-extrabold text-sm">إجمالي التحذيرات (آخر تقرير)</div>
          <div id="dashWarnings" class="text-3xl font-extrabold mt-2">0</div>
        </div>
      </div>

      <section class="mt-8">
        <div class="flex items-center justify-between gap-3">
          <h2 class="text-xl font-extrabold">آخر التقارير</h2>
          <a class="text-sm font-extrabold text-slate-900 hover:underline" href="/reports">عرض الكل</a>
        </div>
        <div id="dashReportsHost" class="mt-4 grid gap-3 md:grid-cols-2"></div>
      </section>
    </main>
    <footer class="max-w-6xl mx-auto px-4 pb-8 text-center text-slate-500 text-sm font-extrabold">
      تطوير الموقع: محمد علي السوداني
    </footer>

    <script src="/static/app.js"></script>
    <script>
      (async function () {
        try {
          await initAuthUI();
          const data = await apiGet("/reports");
          const items = data.items || [];
          document.getElementById("dashTotalReports").innerText = String(items.length);
          if (items.length) {
            document.getElementById("dashErrors").innerText = String(items[0].stats.errors_count);
            document.getElementById("dashWarnings").innerText = String(items[0].stats.warnings_count);
          }

          const host = document.getElementById("dashReportsHost");
          host.innerHTML = "";
          (items.slice(0, 6) || []).forEach((item) => {
            const card = document.createElement("div");
            card.className = "bg-white border border-slate-200 rounded-2xl p-5 shadow-sm flex flex-col gap-3";
            card.innerHTML = `
              <div class="flex items-start justify-between gap-4">
                <div class="min-w-0">
                  <div class="font-extrabold truncate">${item.title ? item.title : "تقرير بدون عنوان"}</div>
                  <div class="text-sm text-slate-600 mt-1">${item.branch1_name} مقابل ${item.branch2_name}</div>
                </div>
                <a class="px-3 py-1.5 rounded-lg bg-slate-900 text-white text-sm font-extrabold" href="/report?id=${item.id}">عرض</a>
              </div>
              <div class="flex gap-3 flex-wrap text-sm">
                <div class="font-extrabold text-slate-800">أخطاء: <span class="text-rose-600">${item.stats.errors_count}</span></div>
                <div class="font-extrabold text-slate-800">تحذيرات: <span class="text-amber-600">${item.stats.warnings_count}</span></div>
              </div>
            `;
            host.appendChild(card);
          });
        } catch (e) {
          console.error(e);
        }
      })();
    </script>
  </body>
</html>
"""

ANALYZE_HTML = r"""<!doctype html>
<html lang="ar" dir="rtl">
  <head>
    <meta charset="UTF-8" />
    <meta name="viewport" content="width=device-width, initial-scale=1.0" />
    <title>تحليل - AuditFlow</title>
    <link
      href="https://fonts.googleapis.com/css2?family=IBM+Plex+Sans+Arabic:wght@300;400;600;700;800&display=swap"
      rel="stylesheet"
    />
    <link rel="stylesheet" href="/static/tailwind.css" />
    <style>
      body { font-family: "IBM Plex Sans Arabic", system-ui, -apple-system, "Segoe UI", Roboto, Arial, sans-serif; }
    </style>
  </head>
  <body class="bg-slate-50 text-slate-900">
    <div id="toast" class="hidden fixed bottom-5 left-5 z-50 text-white px-4 py-2 rounded-xl font-extrabold"></div>

    <header class="sticky top-0 bg-white/90 backdrop-blur border-b border-slate-200 z-40">
      <div class="max-w-6xl mx-auto px-4 py-4 flex items-center justify-between gap-3">
        <div class="font-extrabold text-slate-900 text-lg">AuditFlow | نظام التدقيق</div>
        <nav class="flex gap-3">
          <a class="px-3 py-2 rounded-xl font-extrabold text-sm bg-white border border-slate-200 hover:bg-slate-50" href="/">لوحة التحكم</a>
          <a class="px-3 py-2 rounded-xl font-extrabold text-sm bg-slate-900 text-white" href="/analyze">تحليل</a>
          <a class="px-3 py-2 rounded-xl font-extrabold text-sm bg-white border border-slate-200 hover:bg-slate-50" href="/reports">التقارير</a>
        </nav>
        <div id="authArea"></div>
      </div>
    </header>

    <main class="max-w-6xl mx-auto px-4 py-8">
      <div class="bg-white border border-slate-200 rounded-3xl p-6 md:p-8 shadow-sm">
        <h1 class="text-2xl md:text-3xl font-extrabold text-center">تحليل المطابقة المالية</h1>
        <p class="text-center text-slate-600 mt-2">ارفع ملفي الفرع الأول والثاني (Excel/PDF) ثم شغّل التحليل.</p>

        <div class="grid lg:grid-cols-2 gap-4 mt-6">
          <section class="bg-slate-50 border border-slate-200 rounded-2xl p-4">
            <div class="font-extrabold text-slate-900 mb-3">الفرع الأول</div>

            <label class="block text-sm font-extrabold text-slate-700 mb-1">اسم الفرع</label>
            <input id="b1" class="w-full rounded-xl border border-slate-200 bg-white px-3 py-2 outline-none focus:ring-2 focus:ring-slate-900/10" placeholder="اكتب اسم الفرع الأول هنا" />

            <div class="flex gap-2 justify-center mt-3 mb-3">
              <button type="button" id="b1_excel" class="px-3 py-1.5 rounded-full border border-slate-200 text-slate-600 font-extrabold text-xs active:bg-emerald-50 active:border-emerald-200 active:text-emerald-700" onclick="setType(1,'excel')">Excel</button>
              <button type="button" id="b1_pdf" class="px-3 py-1.5 rounded-full border border-slate-200 text-slate-600 font-extrabold text-xs" onclick="setType(1,'pdf')">PDF</button>
            </div>

            <input type="file" id="file1" class="hidden" />
            <div id="dz1" class="dropzone border-2 border-dashed border-slate-300 rounded-2xl bg-white h-28 flex items-center justify-center flex-col gap-1 cursor-pointer hover:border-slate-400" draggable="false">
              <div class="text-sm font-extrabold text-slate-900">اسحب وأفلت</div>
              <div class="text-xs text-slate-500">أو اضغط للاختيار</div>
            </div>
            <div id="fileName1" class="text-xs text-slate-500 mt-2 min-h-4"></div>
          </section>

          <section class="bg-slate-50 border border-slate-200 rounded-2xl p-4">
            <div class="font-extrabold text-slate-900 mb-3">الفرع الثاني</div>

            <label class="block text-sm font-extrabold text-slate-700 mb-1">اسم الفرع</label>
            <input id="b2" class="w-full rounded-xl border border-slate-200 bg-white px-3 py-2 outline-none focus:ring-2 focus:ring-slate-900/10" placeholder="اكتب اسم الفرع الثاني هنا" />

            <div class="flex gap-2 justify-center mt-3 mb-3">
              <button type="button" id="b2_excel" class="px-3 py-1.5 rounded-full border border-slate-200 text-slate-600 font-extrabold text-xs" onclick="setType(2,'excel')">Excel</button>
              <button type="button" id="b2_pdf" class="px-3 py-1.5 rounded-full border border-slate-200 text-slate-600 font-extrabold text-xs" onclick="setType(2,'pdf')">PDF</button>
            </div>

            <input type="file" id="file2" class="hidden" />
            <div id="dz2" class="dropzone border-2 border-dashed border-slate-300 rounded-2xl bg-white h-28 flex items-center justify-center flex-col gap-1 cursor-pointer hover:border-slate-400" draggable="false">
              <div class="text-sm font-extrabold text-slate-900">اسحب وأفلت</div>
              <div class="text-xs text-slate-500">أو اضغط للاختيار</div>
            </div>
            <div id="fileName2" class="text-xs text-slate-500 mt-2 min-h-4"></div>
          </section>
        </div>

        <div class="mt-6 flex items-center justify-center gap-3 flex-wrap">
          <button id="startBtn" class="px-6 py-3 bg-slate-900 text-white rounded-2xl font-extrabold hover:bg-slate-800" onclick="startAnalyze()">ابدأ التحليل</button>
          <input id="title" class="w-72 max-w-full rounded-xl border border-slate-200 bg-white px-3 py-2 outline-none" placeholder="عنوان التقرير (اختياري)" />
        </div>
      </div>
    </main>
    <footer class="max-w-6xl mx-auto px-4 pb-8 text-center text-slate-500 text-sm font-extrabold">
      تطوير الموقع: محمد علي السوداني
    </footer>

    <script src="/static/app.js"></script>
    <script>
      initAuthUI();
      let type1 = "excel";
      let type2 = "pdf";

      function setType(branch, type) {
        if (branch === 1) type1 = type;
        if (branch === 2) type2 = type;

        const b1e = document.getElementById("b1_excel");
        const b1p = document.getElementById("b1_pdf");
        const b2e = document.getElementById("b2_excel");
        const b2p = document.getElementById("b2_pdf");

        if (b1e) b1e.className = "px-3 py-1.5 rounded-full border border-slate-200 font-extrabold text-xs" + (type1 === "excel" ? " bg-emerald-50 border-emerald-200 text-emerald-700" : " text-slate-600");
        if (b1p) b1p.className = "px-3 py-1.5 rounded-full border border-slate-200 font-extrabold text-xs" + (type1 === "pdf" ? " bg-blue-50 border-blue-200 text-blue-700" : " text-slate-600");
        if (b2e) b2e.className = "px-3 py-1.5 rounded-full border border-slate-200 font-extrabold text-xs" + (type2 === "excel" ? " bg-emerald-50 border-emerald-200 text-emerald-700" : " text-slate-600");
        if (b2p) b2p.className = "px-3 py-1.5 rounded-full border border-slate-200 font-extrabold text-xs" + (type2 === "pdf" ? " bg-blue-50 border-blue-200 text-blue-700" : " text-slate-600");
      }

      function validate(file, expected) {
        if (!file) return false;
        const n = (file.name || "").toLowerCase();
        if (expected === "excel") return /\.(xlsx|xls|xlsm|xlsb|csv)$/.test(n);
        if (expected === "pdf") return n.endsWith(".pdf");
        return true;
      }

      function bindDZ(dzId, inputId, expectedGetter, fileNameId) {
        const dz = document.getElementById(dzId);
        const inp = document.getElementById(inputId);
        const fileName = document.getElementById(fileNameId);
        if (!dz || !inp) return;

        dz.addEventListener("click", () => inp.click());
        dz.addEventListener("dragover", (e) => { e.preventDefault(); dz.classList.add("border-slate-500"); });
        dz.addEventListener("dragleave", () => dz.classList.remove("border-slate-500"));
        dz.addEventListener("drop", (e) => {
          e.preventDefault();
          dz.classList.remove("border-slate-500");
          const file = e.dataTransfer?.files?.[0];
          if (!file || !validate(file, expectedGetter())) {
            showToast("نوع الملف غير صحيح", "#ef4444");
            return;
          }
          inp.files = e.dataTransfer.files;
          if (fileName) fileName.innerText = "تم اختيار: " + file.name;
        });

        inp.addEventListener("change", () => {
          const file = inp.files?.[0] || null;
          if (file && !validate(file, expectedGetter())) {
            showToast("نوع الملف غير صحيح", "#ef4444");
            inp.value = "";
            if (fileName) fileName.innerText = "";
            return;
          }
          if (fileName) fileName.innerText = file ? ("تم اختيار: " + file.name) : "";
        });
      }

      bindDZ("dz1", "file1", () => type1, "fileName1");
      bindDZ("dz2", "file2", () => type2, "fileName2");

      setType(1, "excel");
      setType(2, "pdf");
    </script>
  </body>
</html>
"""

REPORTS_HTML = r"""<!doctype html>
<html lang="ar" dir="rtl">
  <head>
    <meta charset="UTF-8" />
    <meta name="viewport" content="width=device-width, initial-scale=1.0" />
    <title>التقارير - AuditFlow</title>
    <link
      href="https://fonts.googleapis.com/css2?family=IBM+Plex+Sans+Arabic:wght@300;400;600;700;800&display=swap"
      rel="stylesheet"
    />
    <link rel="stylesheet" href="/static/tailwind.css" />
    <style>
      body { font-family: "IBM Plex Sans Arabic", system-ui, -apple-system, "Segoe UI", Roboto, Arial, sans-serif; }
    </style>
  </head>
  <body class="bg-slate-50 text-slate-900">
    <div id="toast" class="hidden fixed bottom-5 left-5 z-50 text-white px-4 py-2 rounded-xl font-extrabold"></div>

    <header class="sticky top-0 bg-white/90 backdrop-blur border-b border-slate-200 z-40">
      <div class="max-w-6xl mx-auto px-4 py-4 flex items-center justify-between gap-3">
        <div class="font-extrabold text-slate-900 text-lg">AuditFlow | نظام التدقيق</div>
        <nav class="flex gap-3">
          <a class="px-3 py-2 rounded-xl font-extrabold text-sm bg-white border border-slate-200 hover:bg-slate-50" href="/">لوحة التحكم</a>
          <a class="px-3 py-2 rounded-xl font-extrabold text-sm bg-white border border-slate-200 hover:bg-slate-50" href="/analyze">تحليل</a>
          <a class="px-3 py-2 rounded-xl font-extrabold text-sm bg-slate-900 text-white" href="/reports">التقارير</a>
        </nav>
        <div id="authArea"></div>
      </div>
    </header>

    <main class="max-w-6xl mx-auto px-4 py-8">
      <div class="bg-white border border-slate-200 rounded-3xl p-6 md:p-8 shadow-sm">
        <div class="flex items-center justify-between gap-3">
          <h1 class="text-2xl font-extrabold">التقارير</h1>
          <a href="/analyze" class="px-4 py-2 rounded-2xl bg-slate-900 text-white font-extrabold hover:bg-slate-800">تحليل جديد</a>
        </div>

        <div id="reportsHost" class="mt-6 grid gap-4 md:grid-cols-2"></div>
      </div>
    </main>
    <footer class="max-w-6xl mx-auto px-4 pb-8 text-center text-slate-500 text-sm font-extrabold">
      تطوير الموقع: محمد علي السوداني
    </footer>

    <script src="/static/app.js"></script>
    <script>
      initAuthUI();
      loadReports().catch((e) => {
        console.error(e);
        showToast("فشل تحميل التقارير", "#ef4444");
      });
    </script>
  </body>
</html>
"""

REPORT_HTML = r"""<!doctype html>
<html lang="ar" dir="rtl">
  <head>
    <meta charset="UTF-8" />
    <meta name="viewport" content="width=device-width, initial-scale=1.0" />
    <title>تقرير - AuditFlow</title>
    <link
      href="https://fonts.googleapis.com/css2?family=IBM+Plex+Sans+Arabic:wght@300;400;600;700;800&display=swap"
      rel="stylesheet"
    />
    <link rel="stylesheet" href="/static/tailwind.css" />
    <style>
      body { font-family: "IBM Plex Sans Arabic", system-ui, -apple-system, "Segoe UI", Roboto, Arial, sans-serif; }
    </style>
  </head>
  <body class="bg-slate-50 text-slate-900">
    <div id="toast" class="hidden fixed bottom-5 left-5 z-50 text-white px-4 py-2 rounded-xl font-extrabold"></div>

    <header class="sticky top-0 bg-white/90 backdrop-blur border-b border-slate-200 z-40">
      <div class="max-w-6xl mx-auto px-4 py-4 flex items-center justify-between gap-3">
        <div class="font-extrabold text-slate-900 text-lg">AuditFlow | نظام التدقيق</div>
        <nav class="flex gap-3">
          <a class="px-3 py-2 rounded-xl font-extrabold text-sm bg-white border border-slate-200 hover:bg-slate-50" href="/">لوحة التحكم</a>
          <a class="px-3 py-2 rounded-xl font-extrabold text-sm bg-white border border-slate-200 hover:bg-slate-50" href="/analyze">تحليل</a>
          <a class="px-3 py-2 rounded-xl font-extrabold text-sm bg-white border border-slate-200 hover:bg-slate-50" href="/reports">التقارير</a>
        </nav>
        <div id="authArea"></div>
      </div>
    </header>

    <main class="max-w-6xl mx-auto px-4 py-8">
      <div class="bg-white border border-slate-200 rounded-3xl p-6 md:p-8 shadow-sm">
        <div class="flex items-start justify-between gap-4 flex-wrap">
          <div>
            <h1 id="reportTitle" class="text-2xl font-extrabold">تقرير</h1>
            <div id="reportBranches" class="text-slate-600 font-extrabold mt-2"></div>
          </div>
          <div class="flex gap-2">
            <button
              id="downloadExcelBtn"
              class="px-4 py-2 rounded-2xl bg-emerald-500 text-white font-extrabold hover:bg-emerald-600"
              onclick="downloadErrors(qs('id'),'excel')"
            >
              تنزيل الأخطاء Excel
            </button>
            <button
              id="downloadPdfBtn"
              class="px-4 py-2 rounded-2xl bg-blue-500 text-white font-extrabold hover:bg-blue-600"
              onclick="downloadErrors(qs('id'),'pdf')"
            >
              تنزيل الأخطاء PDF
            </button>
          </div>
        </div>

        <div class="grid md:grid-cols-4 gap-4 mt-6">
          <div class="bg-slate-50 border border-slate-200 rounded-2xl p-4">
            <div class="text-slate-500 font-extrabold text-sm">الإجمالي</div>
            <div id="statTotal" class="text-3xl font-extrabold mt-2">0</div>
          </div>
          <div class="bg-emerald-50 border border-emerald-200 rounded-2xl p-4">
            <div class="text-emerald-700 font-extrabold text-sm">متطابق</div>
            <div id="statMatched" class="text-3xl font-extrabold mt-2 text-emerald-800">0</div>
          </div>
          <div class="bg-rose-50 border border-rose-200 rounded-2xl p-4">
            <div class="text-rose-700 font-extrabold text-sm">أخطاء</div>
            <div id="statErrors" class="text-3xl font-extrabold mt-2 text-rose-800">0</div>
          </div>
          <div class="bg-amber-50 border border-amber-200 rounded-2xl p-4">
            <div class="text-amber-700 font-extrabold text-sm">تحذيرات</div>
            <div id="statWarnings" class="text-3xl font-extrabold mt-2 text-amber-800">0</div>
          </div>
        </div>

        <div class="grid md:grid-cols-2 gap-4 mt-4">
          <div class="bg-rose-50 border border-rose-200 rounded-2xl p-4">
            <div class="text-rose-700 font-extrabold text-sm">أخطاء <span id="branch1Label">الفرع الأول</span></div>
            <div id="branch1Errors" class="text-3xl font-extrabold mt-2 text-rose-800">0</div>
            <div class="text-sm text-rose-700 mt-1">نسبة الخطأ: <span id="branch1Rate" class="font-extrabold">0.0%</span></div>
          </div>
          <div class="bg-rose-50 border border-rose-200 rounded-2xl p-4">
            <div class="text-rose-700 font-extrabold text-sm">أخطاء <span id="branch2Label">الفرع الثاني</span></div>
            <div id="branch2Errors" class="text-3xl font-extrabold mt-2 text-rose-800">0</div>
            <div class="text-sm text-rose-700 mt-1">نسبة الخطأ: <span id="branch2Rate" class="font-extrabold">0.0%</span></div>
          </div>
        </div>

        <div class="mt-6 bg-slate-50 border border-slate-200 rounded-2xl p-4">
          <div class="flex items-center justify-between gap-3 flex-wrap">
            <h2 class="font-extrabold">فلترة الأخطاء</h2>
            <div class="flex gap-2 flex-wrap">
              <input id="filterDoc" placeholder="نوع المستند" class="rounded-xl border border-slate-200 bg-white px-3 py-2 text-sm outline-none" />
              <input id="filterAmount" placeholder="المبلغ" class="rounded-xl border border-slate-200 bg-white px-3 py-2 text-sm outline-none w-32" />
              <input id="filterType" placeholder="نوع الخطأ (❌ أو ⚠️)" class="rounded-xl border border-slate-200 bg-white px-3 py-2 text-sm outline-none w-44" />
              <button class="px-4 py-2 rounded-xl bg-slate-900 text-white font-extrabold text-sm" onclick="applyTableFilters(window.__MISMATCHES__ || [])">تطبيق</button>
            </div>
          </div>
        </div>

        <div id="mismatchTableHost" class="mt-4 overflow-auto rounded-2xl border border-slate-200"></div>
      </div>
    </main>
    <footer class="max-w-6xl mx-auto px-4 pb-8 text-center text-slate-500 text-sm font-extrabold">
      تطوير الموقع: محمد علي السوداني
    </footer>

    <script src="/static/app.js"></script>
    <script>
      window.addEventListener("DOMContentLoaded", () => {
        initAuthUI();
        loadReportDetail().catch((e) => {
          console.error(e);
          showToast("فشل تحميل التقرير", "#ef4444");
        });
      });
    </script>
  </body>
</html>
"""

SETTINGS_HTML = r"""<!doctype html>
<html lang="ar" dir="rtl">
  <head>
    <meta charset="UTF-8" />
    <meta name="viewport" content="width=device-width, initial-scale=1.0" />
    <title>الإعدادات - AuditFlow</title>
    <link href="https://fonts.googleapis.com/css2?family=IBM+Plex+Sans+Arabic:wght@300;400;600;700;800&display=swap" rel="stylesheet" />
    <link rel="stylesheet" href="/static/tailwind.css" />
    <style>body { font-family: "IBM Plex Sans Arabic", system-ui, sans-serif; }</style>
  </head>
  <body class="bg-slate-50 text-slate-900">
    <div id="toast" class="hidden fixed bottom-5 left-5 z-50 text-white px-4 py-2 rounded-xl font-extrabold"></div>
    <header class="sticky top-0 bg-white/90 backdrop-blur border-b border-slate-200 z-40">
      <div class="max-w-4xl mx-auto px-4 py-4 flex items-center justify-between gap-3">
        <div class="font-extrabold text-slate-900 text-lg">AuditFlow | نظام التدقيق</div>
        <nav class="flex gap-3">
          <a class="px-3 py-2 rounded-xl font-extrabold text-sm bg-white border border-slate-200 hover:bg-slate-50" href="/">لوحة التحكم</a>
          <a class="px-3 py-2 rounded-xl font-extrabold text-sm bg-slate-900 text-white" href="/settings">الإعدادات</a>
        </nav>
        <div id="authArea"></div>
      </div>
    </header>
    <main class="max-w-4xl mx-auto px-4 py-8">
      <div class="bg-white border border-slate-200 rounded-3xl p-6 shadow-sm">
        <h1 class="text-2xl font-extrabold">إعدادات الحساب</h1>
        <p class="text-slate-600 mt-2">يمكنك تغيير كلمة المرور وإنشاء نسخة احتياطية.</p>
        <div class="mt-6 grid gap-3">
          <input id="oldPass" type="password" placeholder="كلمة المرور الحالية" class="rounded-xl border border-slate-200 px-3 py-2" />
          <input id="newPass" type="password" placeholder="كلمة المرور الجديدة" class="rounded-xl border border-slate-200 px-3 py-2" />
          <button id="changePassBtn" class="px-4 py-2 rounded-xl bg-slate-900 text-white font-extrabold w-fit">تغيير كلمة المرور</button>
        </div>
        <div class="mt-8">
          <button onclick="window.location.href='/backup'" class="px-4 py-2 rounded-xl bg-emerald-600 text-white font-extrabold">تنزيل نسخة احتياطية</button>
        </div>
      </div>
    </main>
    <footer class="max-w-4xl mx-auto px-4 pb-8 text-center text-slate-500 text-sm font-extrabold">تطوير الموقع: محمد علي السوداني</footer>
    <script src="/static/app.js"></script>
    <script>
      initAuthUI();
      document.getElementById("changePassBtn").addEventListener("click", async () => {
        const old_password = document.getElementById("oldPass").value || "";
        const new_password = document.getElementById("newPass").value || "";
        try {
          await apiPostJson("/auth/change-password", { old_password, new_password });
          showToast("تم تغيير كلمة المرور ✔️");
          document.getElementById("oldPass").value = "";
          document.getElementById("newPass").value = "";
        } catch (e) {
          showToast(e.message || "فشل تغيير كلمة المرور", "#ef4444");
        }
      });
    </script>
  </body>
</html>
"""

LOGIN_HTML = r"""<!doctype html>
<html lang="ar" dir="rtl">
  <head>
    <meta charset="UTF-8" />
    <meta name="viewport" content="width=device-width, initial-scale=1.0" />
    <title>تسجيل الدخول - AuditFlow</title>
    <link href="https://fonts.googleapis.com/css2?family=IBM+Plex+Sans+Arabic:wght@300;400;600;700;800&display=swap" rel="stylesheet" />
    <link rel="stylesheet" href="/static/tailwind.css" />
    <style>body { font-family: "IBM Plex Sans Arabic", system-ui, sans-serif; }</style>
  </head>
  <body class="bg-slate-50 text-slate-900 min-h-screen">
    <div id="toast" class="hidden fixed bottom-5 left-5 z-50 text-white px-4 py-2 rounded-xl font-extrabold"></div>
    <main class="max-w-xl mx-auto px-4 py-16">
      <div class="bg-white border border-slate-200 rounded-3xl p-8 shadow-sm text-center">
        <h1 class="text-3xl font-extrabold">AuditFlow</h1>
        <p class="text-slate-600 mt-3">يجب تسجيل الدخول أو إنشاء حساب أولًا للوصول إلى التحليل والتقارير والتنزيل.</p>
        <div id="authArea" class="mt-6 flex justify-center"></div>
      </div>
    </main>
    <footer class="max-w-xl mx-auto px-4 pb-8 text-center text-slate-500 text-sm font-extrabold">تطوير الموقع: محمد علي السوداني</footer>
    <script src="/static/app.js"></script>
    <script>initAuthUI();</script>
  </body>
</html>
"""


def wants_html(request: Request) -> bool:
    accept = request.headers.get("accept", "").lower()
    if "application/json" in accept:
        return False
    return ("text/html" in accept) or (accept == "" or "*/*" in accept)


# =========================
# APP
# =========================
app = FastAPI(title="AuditFlow (single file)")
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)
run_migrations()


@app.get("/", response_class=HTMLResponse)
def ui_home(request: Request):
    db = db_session()
    try:
        _ = require_user(db, request)
        return HTMLResponse(INDEX_HTML)
    except HTTPException:
        return RedirectResponse(url="/login", status_code=302)
    finally:
        db.close()


@app.get("/analyze", response_class=HTMLResponse)
def ui_analyze(request: Request):
    db = db_session()
    try:
        _ = require_user(db, request)
        return HTMLResponse(ANALYZE_HTML)
    except HTTPException:
        return RedirectResponse(url="/login", status_code=302)
    finally:
        db.close()


@app.get("/settings", response_class=HTMLResponse)
def ui_settings(request: Request):
    db = db_session()
    try:
        _ = require_user(db, request)
        return HTMLResponse(SETTINGS_HTML)
    except HTTPException:
        return RedirectResponse(url="/login", status_code=302)
    finally:
        db.close()


@app.get("/login", response_class=HTMLResponse)
def ui_login(request: Request):
    db = db_session()
    try:
        u = current_user_from_request(db, request)
        if u:
            return RedirectResponse(url="/", status_code=302)
        return HTMLResponse(LOGIN_HTML)
    finally:
        db.close()


@app.get("/static/app.js")
def ui_js():
    return Response(content=APP_JS.encode("utf-8"), media_type="application/javascript; charset=utf-8")


TAILWIND_CSS_PATH = Path(__file__).resolve().parent / "frontend" / "tailwind.css"


@app.get("/static/tailwind.css")
def ui_tailwind_css():
    if not TAILWIND_CSS_PATH.is_file():
        return Response(
            content=b"/* Missing auditflow/frontend/tailwind.css - run npm run build:css */\n",
            media_type="text/css; charset=utf-8",
            status_code=503,
        )
    return FileResponse(TAILWIND_CSS_PATH, media_type="text/css; charset=utf-8")


@app.get("/auth/me")
def auth_me(request: Request):
    db = db_session()
    try:
        u = current_user_from_request(db, request)
        csrf = request.cookies.get(CSRF_COOKIE) or issue_csrf_token()
        username = u.username if u else None
        res = Response(
            content=json.dumps({"username": username, "csrf_token": csrf}),
            media_type="application/json",
        )
        res.set_cookie(
            key=CSRF_COOKIE,
            value=csrf,
            httponly=False,
            samesite="lax",
            max_age=SESSION_DAYS * 24 * 60 * 60,
        )
        return res
    finally:
        db.close()


@app.post("/auth/register")
async def auth_register(request: Request):
    payload = await request.json()
    username = str((payload or {}).get("username", "")).strip()
    password = str((payload or {}).get("password", "")).strip()
    if len(username) < 3:
        raise HTTPException(400, "اسم المستخدم قصير")
    if len(password) < 4:
        raise HTTPException(400, "كلمة المرور قصيرة")

    db = db_session()
    try:
        require_csrf(request)
        exists = db.query(User).filter(User.username == username).first()
        if exists:
            raise HTTPException(400, "اسم المستخدم موجود بالفعل")
        user = User(id=uuid.uuid4().hex, username=username, password_hash=hash_password(password))
        db.add(user)
        db.commit()

        token = create_session(db, user.id)
        csrf = issue_csrf_token()
        log_event(db, "auth.register", user.id, {"username": username})
        res = Response(content='{"ok":true}', media_type="application/json")
        res.set_cookie(
            key=SESSION_COOKIE,
            value=token,
            httponly=True,
            samesite="lax",
            max_age=SESSION_DAYS * 24 * 60 * 60,
        )
        res.set_cookie(key=CSRF_COOKIE, value=csrf, httponly=False, samesite="lax", max_age=SESSION_DAYS * 24 * 60 * 60)
        return res
    finally:
        db.close()


@app.post("/auth/login")
async def auth_login(request: Request):
    payload = await request.json()
    username = str((payload or {}).get("username", "")).strip()
    password = str((payload or {}).get("password", "")).strip()

    db = db_session()
    try:
        require_csrf(request)
        user = db.query(User).filter(User.username == username).first()
        if not user:
            raise HTTPException(401, "بيانات الدخول غير صحيحة")
        if user.locked_until and user.locked_until > dt.datetime.utcnow():
            raise HTTPException(429, "الحساب مقفل مؤقتاً. حاول لاحقاً")
        if not verify_password(password, user.password_hash):
            user.failed_attempts = int(user.failed_attempts or 0) + 1
            if user.failed_attempts >= 5:
                user.locked_until = dt.datetime.utcnow() + dt.timedelta(minutes=LOCK_MINUTES)
                user.failed_attempts = 0
            db.commit()
            raise HTTPException(401, "بيانات الدخول غير صحيحة")

        if "$" not in user.password_hash:
            user.password_hash = hash_password(password)
        user.failed_attempts = 0
        user.locked_until = None
        db.commit()

        token = create_session(db, user.id)
        csrf = issue_csrf_token()
        log_event(db, "auth.login", user.id, {"username": username})
        res = Response(content='{"ok":true}', media_type="application/json")
        res.set_cookie(
            key=SESSION_COOKIE,
            value=token,
            httponly=True,
            samesite="lax",
            max_age=SESSION_DAYS * 24 * 60 * 60,
        )
        res.set_cookie(key=CSRF_COOKIE, value=csrf, httponly=False, samesite="lax", max_age=SESSION_DAYS * 24 * 60 * 60)
        return res
    finally:
        db.close()


@app.post("/auth/logout")
def auth_logout(request: Request):
    db = db_session()
    try:
        require_csrf(request)
        user = current_user_from_request(db, request)
        token = request.cookies.get(SESSION_COOKIE)
        if token:
            s = db.query(UserSession).filter(UserSession.token == token).first()
            if s:
                db.delete(s)
                db.commit()
        if user:
            log_event(db, "auth.logout", user.id)
        res = Response(content='{"ok":true}', media_type="application/json")
        res.delete_cookie(SESSION_COOKIE)
        res.delete_cookie(CSRF_COOKIE)
        return res
    finally:
        db.close()


@app.post("/auth/change-password")
async def auth_change_password(request: Request):
    require_csrf(request)
    payload = await request.json()
    old_password = str((payload or {}).get("old_password", ""))
    new_password = str((payload or {}).get("new_password", ""))
    if len(new_password) < 8:
        raise HTTPException(400, "كلمة المرور الجديدة يجب أن تكون 8 أحرف على الأقل")
    db = db_session()
    try:
        user = require_user(db, request)
        if not verify_password(old_password, user.password_hash):
            raise HTTPException(400, "كلمة المرور الحالية غير صحيحة")
        user.password_hash = hash_password(new_password)
        db.commit()
        log_event(db, "auth.change_password", user.id)
        return {"ok": True}
    finally:
        db.close()


@app.post("/analyze")
def analyze_api(
    request: Request,
    file1: UploadFile = File(...),
    file2: UploadFile = File(...),
    b1: str = Form(...),
    b2: str = Form(...),
    title: Optional[str] = Form(None),
    strict_mirror_types: bool = Form(False),
):
    require_csrf(request)
    db = db_session()
    try:
        user = require_user(db, request)
    finally:
        db.close()

    report_id = uuid.uuid4().hex
    saved1, original1 = save_upload_file(file1, UPLOAD_DIR / report_id / "file1")
    saved2, original2 = save_upload_file(file2, UPLOAD_DIR / report_id / "file2")

    try:
        d1 = process(saved1, original1, b1)
        d2 = process(saved2, original2, b2)
        mismatches, counts = analyze(d1, d2, allow_same_direction=not strict_mirror_types)
    except HTTPException:
        raise
    except Exception as e:
        msg = str(e).strip() or e.__class__.__name__
        raise HTTPException(400, f"تعذّر تحليل الملفات: {msg}")

    summary = compute_summary(d1, d2, mismatches)

    title_eff = (title or "").strip() or f"{b1.strip()} مقابل {b2.strip()}"

    created = AnalysisReport(
        id=report_id,
        user_id=user.id,
        title=title_eff,
        branch1_name=b1,
        branch2_name=b2,
        file1_original=original1,
        file2_original=original2,
        file1_path=saved1,
        file2_path=saved2,
        total_ops=summary["total_ops"],
        matched_ops=summary["matched_ops"],
        mismatch_ops=summary["mismatch_ops"],
        errors_count=summary["errors_count"],
        warnings_count=summary["warnings_count"],
        stats_json={
            "counts": counts,
            "branch1_total": len(d1),
            "branch2_total": len(d2),
        },
        analysis_json={
            "extracted_branch1": d1,
            "extracted_branch2": d2,
            "mismatches": mismatches,
            "counts": counts,
        },
    )

    db = db_session()
    try:
        db.add(created)
        db.commit()
        log_event(db, "analysis.create", user.id, {"report_id": report_id})
    finally:
        db.close()

    return {"reportId": report_id}


@app.get("/reports")
def reports(request: Request):
    if wants_html(request):
        db = db_session()
        try:
            _ = require_user(db, request)
            return HTMLResponse(REPORTS_HTML)
        except HTTPException:
            return RedirectResponse(url="/login", status_code=302)
        finally:
            db.close()

    db = db_session()
    try:
        user = require_user(db, request)
        rows: List[AnalysisReport] = (
            db.query(AnalysisReport)
            .filter(AnalysisReport.user_id == user.id)
            .order_by(AnalysisReport.created_at.desc())
            .limit(200)
            .all()
        )
        items = []
        for r in rows:
            items.append(
                {
                    "id": r.id,
                    "title": r.title,
                    "branch1_name": r.branch1_name,
                    "branch2_name": r.branch2_name,
                    "status": r.status,
                    "created_at": r.created_at,
                    "stats": {
                        "total_ops": r.total_ops,
                        "matched_ops": r.matched_ops,
                        "mismatch_ops": r.mismatch_ops,
                        "errors_count": r.errors_count,
                        "warnings_count": r.warnings_count,
                    },
                }
            )
        return {"items": items}
    finally:
        db.close()


@app.get("/report")
def report(request: Request, id: str = Query(...)):
    if wants_html(request):
        db = db_session()
        try:
            _ = require_user(db, request)
            return HTMLResponse(REPORT_HTML)
        except HTTPException:
            return RedirectResponse(url="/login", status_code=302)
        finally:
            db.close()

    db = db_session()
    try:
        user = require_user(db, request)
        r: AnalysisReport | None = (
            db.query(AnalysisReport)
            .filter(AnalysisReport.id == id, AnalysisReport.user_id == user.id)
            .first()
        )
        if not r:
            raise HTTPException(404, "Report not found")

        return {
            "id": r.id,
            "title": r.title,
            "branch1_name": r.branch1_name,
            "branch2_name": r.branch2_name,
            "status": r.status,
            "created_at": r.created_at,
            "stats": {
                "total_ops": r.total_ops,
                "matched_ops": r.matched_ops,
                "mismatch_ops": r.mismatch_ops,
                "errors_count": r.errors_count,
                "warnings_count": r.warnings_count,
            },
            "file1_original": r.file1_original,
            "file2_original": r.file2_original,
            "stats_json": r.stats_json,
            "analysis_json": r.analysis_json,
        }
    finally:
        db.close()


@app.get("/download")
def download(request: Request, id: str = Query(...), format: str = Query("excel")):
    db = db_session()
    try:
        user = require_user(db, request)
        r: AnalysisReport | None = (
            db.query(AnalysisReport)
            .filter(AnalysisReport.id == id, AnalysisReport.user_id == user.id)
            .first()
        )
        if not r:
            raise HTTPException(404, "Report not found")
        mismatches = (r.analysis_json or {}).get("mismatches", []) or []
        fmt = (format or "excel").lower().strip()

        if fmt in ("excel", "xlsx"):
            excel_bytes = mismatches_to_excel_bytes(mismatches)
            filename = f"report_{r.id}.xlsx"
            log_event(db, "download.excel", user.id, {"report_id": r.id})
            return Response(
                content=excel_bytes,
                media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                headers={"Content-Disposition": f'attachment; filename=\"{filename}\"'},
            )

        if fmt == "pdf":
            pdf_bytes = mismatches_to_pdf_bytes(mismatches)
            filename = f"report_{r.id}.pdf"
            log_event(db, "download.pdf", user.id, {"report_id": r.id})
            return Response(
                content=pdf_bytes,
                media_type="application/pdf",
                headers={"Content-Disposition": f'attachment; filename=\"{filename}\"'},
            )

        if fmt == "csv":
            csv_bytes = mismatches_to_csv_bytes(mismatches)
            filename = f"report_{r.id}.csv"
            return Response(
                content=csv_bytes,
                media_type="text/csv; charset=utf-8",
                headers={"Content-Disposition": f'attachment; filename=\"{filename}\"'},
            )

        raise HTTPException(400, "format يجب أن يكون: excel أو pdf")
    finally:
        db.close()


@app.delete("/reports")
def delete_report(request: Request, id: str = Query(...)):
    require_csrf(request)
    db = db_session()
    try:
        user = require_user(db, request)
        r: AnalysisReport | None = (
            db.query(AnalysisReport)
            .filter(AnalysisReport.id == id, AnalysisReport.user_id == user.id)
            .first()
        )
        if not r:
            raise HTTPException(404, "Report not found")
        db.delete(r)
        db.commit()
        log_event(db, "report.delete", user.id, {"report_id": id})
        return {"deleted": True}
    finally:
        db.close()


@app.get("/backup")
def download_backup(request: Request):
    db = db_session()
    try:
        user = require_user(db, request)
    finally:
        db.close()
    ensure_dir(BACKUP_DIR)
    stamp = dt.datetime.utcnow().strftime("%Y%m%d_%H%M%S")
    zip_path = BACKUP_DIR / f"auditflow_backup_{stamp}.zip"
    with zipfile.ZipFile(zip_path, "w", compression=zipfile.ZIP_DEFLATED) as zf:
        if DB_PATH.exists():
            zf.write(DB_PATH, arcname="auditflow.db")
        if UPLOAD_DIR.exists():
            for root, _, files in os.walk(UPLOAD_DIR):
                for f in files:
                    abs_f = Path(root) / f
                    rel_f = abs_f.relative_to(BASE_DIR)
                    zf.write(abs_f, arcname=str(rel_f))
    db2 = db_session()
    try:
        log_event(db2, "backup.download", user.id, {"zip": zip_path.name})
    finally:
        db2.close()
    with open(zip_path, "rb") as fh:
        content = fh.read()
    return Response(
        content=content,
        media_type="application/zip",
        headers={"Content-Disposition": f'attachment; filename="{zip_path.name}"'},
    )

